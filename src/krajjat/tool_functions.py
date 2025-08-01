"""Functions to perform simple tasks that can be used throughout the toolbox and beyond."""
import collections
import datetime
import glob
import numbers
import os
import os.path as op
import random

import math
import json
import string
import warnings
from ast import literal_eval

import chardet
import numpy as np
import pandas as pd
from scipy.interpolate import CubicSpline, PchipInterpolator, Akima1DInterpolator, interp1d
from scipy.io import loadmat
from scipy.signal import savgol_filter

import openpyxl as xl

from krajjat.classes.exceptions import ModuleNotFoundException, InvalidParameterValueException, NotASubPathException, \
    InvalidPathException, DifferentSequencesLengthsException, MissingRecordingDateException
from krajjat.classes.graph_element import Graph
from krajjat.classes.joint import Joint

# Constants
UNITS = {"ns": 1000000000, "1ns": 1000000000, "10ns": 100000000, "100ns": 10000000,
         "µs": 1000000, "1µs": 1000000, "10µs": 100000, "100µs": 10000,
         "ms": 1000, "1ms": 1000, "10ms": 100, "100ms": 10,
         "s": 1, "sec": 1, "1s": 1, "min": 1 / 60, "mn": 1 / 60, "h": 1 / 3600, "hr": 1 / 3600,
         "d": 1 / 86400, "day": 1 / 86400}

CLEAN_DERIV_NAMES = {"x-axis": "x", "x_coord": "x", "coord_x": "x", "x_coordinate": "x", "x": "x",
                     "y-axis": "y", "y_coord": "y", "coord_y": "y", "y_coordinate": "y", "y": "y",
                     "z-axis": "z", "z_coord": "z", "coord_z": "z", "z_coordinate": "z", "z": "z",
                     "coord": "coordinates", "coords": "coordinates", "coordinate": "coordinates", "xyz": "coordinates",
                     "coordinates": "coordinates",
                     "d": "distance", "distances": "distance", "dist": "distance", 0: "distance", "0": "distance",
                     "distance": "distance",
                     "dx": "distance x", "distance_x": "distance x", "x_distance": "distance x", "dist_x": "distance x",
                     "x_dist": "distance x", "distance x": "distance x",
                     "dy": "distance y", "distance_y": "distance y", "y_distance": "distance y", "dist_y": "distance y",
                     "y_dist": "distance y", "distance y": "distance y",
                     "dz": "distance z", "distance_z": "distance z", "z_distance": "distance z", "dist_z": "distance z",
                     "z_dist": "distance z", "distance z": "distance z",
                     "v": "velocity", "vel": "velocity", "velocities": "velocity", "speed": "velocity", 1: "velocity",
                     "1": "velocity", "velocity": "velocity",
                     "a": "acceleration", "acc": "acceleration", "accelerations": "acceleration", 2: "acceleration",
                     "2": "acceleration", "acceleration": "acceleration",
                     "j": "jerk", 3: "jerk", "3": "jerk", "jerk": "jerk",
                     "s": "snap", 4: "snap", "4": "snap", "snap": "snap",
                     "c": "crackle", 5: "crackle", "5": "crackle", "crackle": "crackle",
                     "p": "pop", 6: "pop", "6": "pop", "pop": "pop"}

MODULE_DIR = op.dirname(__file__)

DEFAULT_FONT_PATH = op.join(MODULE_DIR, "res", "junction_bold.otf")
SILHOUETTE_PATH = op.join(MODULE_DIR, "res", "silhouette_cropped.png")


# === Folder and path functions ===
def compute_subpath(path1, path2, separator=os.sep, remove_leading_separators=True, remove_ending_separators=True):
    """Considering two absolute paths path1 and path2, where one is a parent path of the other, this function
    returns the subdirectories absent from the other.

    .. versionadded:: 2.0

    Parameters
    ----------
    path1: str
        An absolute path.
    path2: str
        An absolute path.
    separator: str
        The separator symbol between the elements of the path (default: ``os.sep``).
    remove_leading_separators: bool
        If `True` (default), the function will remove the leading separators from the subpath (e.g. ``"/docs/file.txt"``
        will become ``"docs/file.txt"``).
    remove_ending_separators: bool
        If `True` (default), the function will remove the ending separators from the subpath (e.g. ``"/docs/files/"``
        will become ``"/docs/files"``).

    Returns
    -------
    str
        The subdirectories of one of the two paths that is absent from the other.

    Example
    -------
    >>> path1 = "C:/Users/Shawn/"
    >>> path2 = "C:/Users/Shawn/Music/Coldplay/A Rush of Blood to the Head/"
    >>> print(compute_subpath(path1, path2))
    Music/Coldplay/A Rush of Blood to the Head
    """

    path1 = op.normpath(path1)
    path2 = op.normpath(path2)

    if len(path1) > len(path2):
        if path2 not in path1:
            raise NotASubPathException(path2, path1)
        else:
            sub_path = path1.replace(path2, "")
    else:
        if path1 not in path2:
            raise NotASubPathException(path2, path1)
        else:
            sub_path = path2.replace(path1, "")

    if remove_leading_separators:
        while sub_path.startswith(separator):
            sub_path = sub_path[len(separator):]

    if remove_ending_separators:
        while sub_path.endswith(separator):
            sub_path = sub_path[:-len(separator)]

    return sub_path


def get_difference_paths(*paths, separator="/"):
    """Returns, for each path, a list of the subfolders that are not common with all the other paths.

    .. versionadded:: 2.0

    Parameters
    ----------
    paths: str
        The paths, or a list of paths preceded by an asterisk ``*``.
    separator: str
        The separator symbol between the elements of the path (default: ``"/"``).

    Returns
    -------
    list(list(str))
        A list containing a list for each path, which itself contains the subfolders that are not common with the other
        paths.

    Example
    -------
    >>> path1 = "C:/Sequences/Raw/Subject1/Feb1/Sequence1"
    >>> path2 = "C:/Sequences/Resampled/Subject1/Mar1/Sequence1"
    >>> get_difference_paths(path1, path2)
    [['Raw', 'Feb1'], ['Original', 'Mar1']]
    >>> get_difference_paths(*[path1, path2])
    [['Raw', 'Feb1'], ['Original', 'Mar1']]
    """

    paths_split = []
    min_len = None

    # We split each path in lists, and get the smallest path length
    for path in paths:
        path_split = path.split(separator)
        paths_split.append(path_split)
        length_path = len(path_split)
        if min_len is None:
            min_len = length_path
        elif min_len > length_path:
            min_len = length_path

    new_paths = [[] for _ in range(len(paths))]

    # We calculate which subpaths are not the same
    if min_len > 1:
        for i in range(min_len):
            keep = False
            for p in range(len(paths_split)):
                if paths_split[p][i] != paths_split[0][i]:
                    keep = True
            if keep:
                for p in range(len(paths_split)):
                    new_paths[p].append(paths_split[p][i])

        for p in range(len(paths_split)):
            if len(paths_split[p]) > min_len:
                for i in range(min_len, len(paths_split[p])):
                    new_paths[p].append(paths_split[p][i])

    else:
        new_paths = paths_split

    return new_paths


def get_objects_paths(*objects):
    """Returns a list of the attributes ``path`` of the objects passed as parameter (Sequence or Audio instances).

    .. versionadded:: 2.0

    Parameters
    ----------
    objects: Sequence|Audio
        One or many Sequence or Audio instances, or a list of these objects preceded by an asterisk ``*``.

    Returns
    -------
    list(str)
        A list of paths.

    Example
    -------
    >>> sequence1 = Sequence("C:/Documents/Sequences/sequence1.json")
    >>> sequence2 = Sequence("C:/Documents/Sequences/sequence2.json")
    >>> get_objects_paths(sequence1, sequence2)
    ["C:/Documents/Sequences/sequence1.json", "C:/Documents/Sequences/sequence2.json"]
    >>> get_objects_paths(*[sequence1, sequence2])
    """
    paths = []
    for element in objects:
        paths.append(element.get_path())
    return paths


def sort_files_trailing_index(path, accepted_extensions=None, separator="_", ignore_files_without_index=True,
                              error_if_missing_indices=True, error_if_multiple_extensions=True, object_type="sequence",
                              verbosity=1, add_tabs=0):
    """Assuming a list of files where each name contains a trailing index (e.g. filename_42.ext), returns the list
    of files sorted by their trailing index.

    Note
    ----
    This function allows to order the files consistently across different file systems despite the lack of leading zeros
    in the index: some systems would place ``pose_11.json`` alphabetically before ``pose_2.json`` (1 comes before 2),
    while some other systems place it after as 11 is greater than 2. In order to avoid these inconsistencies, the
    function converts the number after the underscore into an integer to place it properly according to its index.

    .. versionadded:: 2.0

    Parameters
    ----------
    path: str
        A path containing files where each name contains a trailing index (e.g. filename_42.ext).

    accepted_extensions: str|list|None
        An extension, or a list containing the extensions accepted. Any file with a different extension will be ignored
        by the function. If set on `None`, all the extensions are accepted. Note that having multiple extensions in the
        target folder may raise an exception - see the parameter `error_if_multiple_extensions`.

    separator: str
        A character, or series of characters to look for in the name of the file that separate the name of the file
        from its index (default: underscore ``"_"``). If the separator is blank, the function tries to find the trailing
        number.

    ignore_files_without_index: bool
        If set on `False`, the function returns an error when finding a file that does not contain the separator symbol
        and hence, cannot find the index. If set on `True` (default), the files without indices are simply ignored.

    error_if_missing_indices: bool
        If set on `True` (default), the function checks, after sorting the files, if some expected indices are missing.
        For example, if a `file_23.ext` and a `file_25.ext` are found, but no `file_24.ext`, the function would return
        an error. If set on `False`, the function does not check for missing indices and returns the sorted files.

    error_if_multiple_extensions: bool
        If set on `True`, raises an error if more than one of the accepted extensions are found in the folder. If set on
        `False`, does not raise an error - the function will, however, return an error if two files have the same
        index but with different extensions.

    object_type: str
        The type of object that will be created from opening this file. This parameter only exists in order to return
        exceptions.

    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    add_tabs: int, optional
        Adds the specified amount of tabulations to the verbosity outputs. This parameter may be used by other
        functions to encapsulate the verbosity outputs by indenting them. In a normal use, it shouldn't be set
        by the user.

    Example
    -------
    >>> # Let's consider a directory with the following files: `sequence_1.txt`, `sequence_10.txt`,
    >>> # `sequence_100.txt`, `sequence_11.txt` and `sequence_2.txt`.
    >>> print(sort_files_trailing_index("C:/Sequences"))
    ["sequence_1.txt", "sequence_2.txt", "sequence_10.txt", "sequence_11.txt", "sequence_100.txt"]
    """

    t = add_tabs * "\t"

    if "*" in path:
        files = [os.path.split(file)[1] for file in glob.glob(path)]
        path = os.path.split(path)[0]
    else:
        files = os.listdir(path)  # List all the files in the folder
    files_indices = {}
    # ordered_files = ["" for _ in range(len(files))]  # Create an empty list the length of the files
    extensions_found = []  # Save the valid extensions found in the folder

    accept_any_extension = False
    if accepted_extensions is None:
        accept_any_extension = True

    # Here, we find all the files that have an index in the folder.
    for f in files:

        # We check if the element is a file
        if op.isfile(op.join(path, f)):

            # We get the file extension - if there is no dot, there is no extension
            if "." in f:
                extension = f.split(".")[-1]
            else:
                extension = ""

            if accept_any_extension or extension in accepted_extensions:

                if verbosity > 1:
                    print(t+"\tAdding file: " + str(f))

                # We try to find the index
                if separator in f:

                    if len(separator) > 0:
                        file_index = f.split(".")[0].split(separator)[-1]
                    else:
                        file_index = ""
                        for i in range(len(f.split(".")[0])-1, 0, -1):
                            if f[i] in string.digits:
                                file_index = f[i] + file_index
                            else:
                                break

                    if len(file_index) > 0:
                        if file_index.isdigit():
                            files_indices[f] = int(file_index)
                        else:
                            if ignore_files_without_index:
                                if verbosity > 1:
                                    print(t + f"\tIgnoring file (no index detected in the name): " + str(f))
                            else:
                                raise Exception(f"A file without an index was found: ({f}).")

                    elif not ignore_files_without_index:
                        raise Exception(f"A file without an index was found: ({f}).")

                    if extension not in extensions_found:
                        extensions_found.append(extension)

                        if len(extensions_found) > 1 and error_if_multiple_extensions:
                            raise InvalidPathException(path, object_type, "More than one of the accepted extensions "
                                                       "has been found in the provided folder: " +
                                                       f"{extensions_found[0]} and {extensions_found[1]}")

                else:

                    if ignore_files_without_index:
                        if verbosity > 1:
                            print(t+f"""\tIgnoring file (no "{separator}" detected in the name): """ + str(f))

                    else:
                        raise Exception(f"A file without an index was found: ({f}).")

            else:
                if verbosity > 1:
                    print(t+"\tIgnoring file (not an accepted filetype): " + str(f))

        else:
            if verbosity > 1:
                print(t+"\tIgnoring folder: " + str(f))

    # Reverse dictionary
    indices_files = {v: k for k, v in files_indices.items()}
    indices = list(indices_files.keys())
    indices.sort()

    if len(indices_files) != len(files_indices):
        values = {i: 0 for i in indices}
        for k in files_indices.keys():
            values[files_indices[k]] += 1
            if values[files_indices[k]] == 2:
                raise InvalidPathException(path, object_type, f"At least two files have the same index ( " +
                                           f"{files_indices[k]}) in the target folder.")

    ordered_files = []
    last_index = None
    for index in indices:
        ordered_files.append(indices_files[index])

        if last_index is not None and index != last_index + 1 and error_if_missing_indices:
            raise InvalidPathException(path, object_type, f"At least one of the files is missing (index " +
                                       f"{last_index + 1}).")
        last_index = index

    if verbosity > 0:
        print(str(len(ordered_files)) + " file(s) found.")

    return ordered_files


# === Name functions ===
def get_objects_names(*objects):
    """Returns a list of the attributes ``name`` of the objects passed as parameter (Sequence or Audio instances).

    .. versionadded:: 2.0

    Parameters
    ----------
    objects: Sequence|Audio
        One or many Sequence or Audio instances, or a list of these objects preceded by an asterisk ``*``.

    Returns
    -------
    list(str)
        A list of names.

    Example
    -------
    >>> sequence1 = Sequence("C:/Documents/Sequences/sequence1.json")
    >>> sequence2 = Sequence("C:/Documents/Sequences/sequence2.json", name="seq2")
    >>> get_objects_names(sequence1, sequence2)
    ["sequence1.json", "seq2"]
    >>> get_objects_names(*[sequence1, sequence2])
    ["sequence1.json", "seq2"]
    """
    names = []
    for element in objects:
        names.append(element.get_name())
    return names


# === Sequences functions ===
def compute_different_joints(sequence1, sequence2, acceptable_rounding_decimal=5, verbosity=1):
    """Returns the number of joints having different coordinates between two sequences having the same amount of poses.
    This function can be used to check how many joints have been modified by a processing function.

    .. versionadded:: 2.0

    Note
    ----
    To discard the rounding errors due to other functions, this function rounds the coordinates to the fifth decimal.

    Parameters
    ----------
    sequence1: Sequence
        A Sequence instance.
    sequence2: Sequence
        A Sequence instance.
    acceptable_rounding_decimal: int
        The decimal at which to round the values to account for rounding errors (default: 5).
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    int
        The number of different joints between the two sequences.
    int
        The total number of joints in each of the sequences.

    Example
    -------
    >>> sequence_1 = Sequence("sequence_1.json")
    >>> sequence_1_corrected = sequence_1.correct_jitter(0.1, 2)
    >>> print(compute_different_joints(sequence_1, sequence_1_corrected))
    (1, 9)
    """
    if len(sequence1) != len(sequence2):
        raise DifferentSequencesLengthsException(sequence1, sequence2)

    different_joints = 0
    total_joints = 0

    rd = acceptable_rounding_decimal

    for p in range(len(sequence1.poses)):  # For all the poses
        for j in sequence1.poses[p].joints.keys():  # For all the joints

            # We round the coordinates to 5 significant digits in order to account for rounding errors
            if round(sequence1.poses[p].joints[j].x, rd) == round(sequence2.poses[p].joints[j].x, rd) and \
                    round(sequence1.poses[p].joints[j].y, rd) == round(sequence2.poses[p].joints[j].y, rd) and \
                    round(sequence1.poses[p].joints[j].z, rd) == round(sequence2.poses[p].joints[j].z, rd):
                total_joints += 1

            else:
                # If verbosity, we show the number of the pose, the name of the joint, and the coordinates of the
                # two sequences
                if verbosity > 0:
                    print("Pose n°" + str(p) + ", " + str(j))
                    print("X: " + str(sequence1.poses[p].joints[j].x) +
                          "; Y: " + str(sequence1.poses[p].joints[j].y) +
                          "; Z: " + str(sequence1.poses[p].joints[j].z))
                    print("X: " + str(sequence2.poses[p].joints[j].x) +
                          "; Y: " + str(sequence2.poses[p].joints[j].y) +
                          "; Z: " + str(sequence2.poses[p].joints[j].z))

                different_joints += 1
                total_joints += 1

    # We print the result
    if verbosity > 0:
        print("Different joints: " + str(different_joints) + "/" + str(total_joints) + " (" + str(round(
            different_joints * 100 / total_joints, 2)) + "%).")

    return different_joints, total_joints


def align_two_sequences(sequence1, sequence2, verbosity=1):
    """Checks if one sequence is a subset of another; if true, the function returns the indices at which the two
    sequences start synchronizing. Otherwise, it returns ``False``.

    .. versionadded:: 2.0

    Parameters
    ----------
    sequence1: Sequence
        A Sequence instance.
    sequence2: Sequence
        A Sequence instance.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    tuple or None
        If one of the sequences is a subsequence of the other, the function returns a tuple of two integers,
        corresponding to the indices of the first and the second sequences where the two sequences synchronize,
        respectively. Otherwise, this function returns `None`.

    Example
    -------
    >>> sequence_1 = Sequence("C:/Users/Shawn/sequence_1.tsv")  # This sequence is sampled at 10 Hz
    >>> sequence_2 = sequence_1.trim(start = 1)  # We trim at 1s, meaning we remove the first 10 poses
    >>> print(align_two_sequences(sequence_1, sequence_2))
    (0, 10)
    """

    indices = None

    # Define longest sequence
    if len(sequence1) > len(sequence2):
        longest_sequence = 1
    else:
        longest_sequence = 2

    joint_labels_1 = sequence1.get_joint_labels()
    joint_labels_2 = sequence2.get_joint_labels()

    # If the joints are not the same between the two sequences, we return False.
    if joint_labels_1 != joint_labels_2:
        if verbosity > 0:
            print("No alignment found between the two sequences: the joint labels are different.")
        return None

    # We compare for the x coordinate of the first joint
    subsequence1 = []
    for p in range(0, len(sequence1.poses)):
        subsequence1.append(sequence1.poses[p].joints[joint_labels_1[0]].x)

    subsequence2 = []
    for p in range(0, len(sequence2.poses)):
        subsequence2.append(sequence2.poses[p].joints[joint_labels_2[0]].x)

    if longest_sequence == 2:
        for i in range(0, len(subsequence2) - len(subsequence1) + 1):
            if subsequence1 == subsequence2[i:i + len(subsequence1)]:
                indices = (0, i)
    else:
        for i in range(0, len(subsequence1) - len(subsequence2) + 1):
            if subsequence2 == subsequence1[i:i + len(subsequence2)]:
                indices = (i, 0)

    # If we didn't find any subsequence, we return False
    if indices is None:
        if verbosity > 0:
            print("No alignment found between the two sequences: no match found for the x coordinates.")
        return None

    # Otherwise, we check for every other coordinate
    for joint_label in joint_labels_1:

        for coordinate in range(0, 3):

            subsequence1 = []
            for p in range(0, len(sequence1.poses)):
                subsequence1.append(sequence1.poses[p].joints[joint_label].get_position()[coordinate])
            subsequence2 = []
            for p in range(0, len(sequence2.poses)):
                subsequence2.append(sequence2.poses[p].joints[joint_label].get_position()[coordinate])

            if longest_sequence == 2:
                if subsequence1 != subsequence2[indices[1]:indices[1] + len(subsequence1)]:
                    if verbosity > 0:
                        print("No alignment found between the two sequences: mismatch for the values.")
                    return None
            else:
                if subsequence2 != subsequence1[indices[0]:indices[0] + len(subsequence2)]:
                    if verbosity > 0:
                        print("No alignment found between the two sequences: mismatch for the values.")
                    return None

    if verbosity > 0:
        print(f"Alignment found between the two sequences, at indices {indices[0]} and {indices[1]}.")

    return indices


def align_multiple_sequences(*sequences, verbosity=1):
    """Returns the synchronized timestamps of multiple sequences. If one or more sequences passed as parameter are
    not subsets of other sequences or do not contain other sequences, their original timestamps are returned.

    .. versionadded:: 2.0

    Parameters
    ----------
    sequences: Sequence
        One or more Sequence instances.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    list(list(float))
        A list of synchronized timestamps arrays, matching the order the Sequence instances were passed as parameters.

    Example
    -------
    >>> sequence_1 = Sequence("C:/Users/Irving/sequence_1.tsv")  # This sequence is sampled at 10 Hz with 5 samples
    >>> sequence_2 = sequence_1.trim(start = 0.1)  # We trim at 0.1s, meaning we remove the first pose
    >>> sequence_3 = sequence_1.trim(start = 0.3)  # We trim at 0.3s, meaning we remove the first 3 poses
    >>> print(align_multiple_sequences(sequence_1, sequence_2, sequence_3))
    [[0.0, 0.1, 0.2, 0.3, 0.4], [0.1, 0.2, 0.3, 0.4], [0.3, 0.4]]
    """

    # Get the time array
    timestamps = []

    for sequence_index in range(len(sequences)):
        sequence = sequences[sequence_index]
        aligned = False

        # We compare the current sequence to every previous sequence
        for prior_sequence_index in range(sequence_index):
            prior_sequence = sequences[prior_sequence_index]

            # Returns False if the two sequence don't align, returns two indices otherwise
            alignment_indices = align_two_sequences(sequence, prior_sequence, verbosity)

            # If the sequences do align
            if alignment_indices:

                # Case where the current sequence is a subsequence of the prior
                if alignment_indices[0] == 0:
                    if verbosity > 1:
                        print("\tAligning sequence " + str(sequence_index + 1) + " to sequence " +
                              str(prior_sequence_index + 1) + ".")
                    timestamps.append(np.array(timestamps[prior_sequence_index][alignment_indices[1]:min(alignment_indices[1] +
                                                                                                len(sequence),
                                                                                                len(prior_sequence))]))

                # Case where the prior sequence is a subsequence of the current
                else:
                    if verbosity > 1:
                        print("\tAligning sequence " + str(prior_sequence_index + 1) + " to sequence " +
                              str(sequence_index + 1) + ".")
                    timestamps.append(np.array(sequence.get_timestamps()))
                    timestamps[prior_sequence_index] = timestamps[sequence_index][alignment_indices[0]:
                                                                                  min(alignment_indices[0] +
                                                                                  len(prior_sequence), len(sequence))]

                aligned = True
                break

        if not aligned:
            timestamps.append(np.array(sequence.get_timestamps()))

    return timestamps


def sort_sequences_by_date(*sequences):
    """Takes multiple sequences as a parameter, and returns an array containing the same sequences, ordered by
    recording date. If at least one sequence of the array has an attribute :attr:`Sequence.date_recording` set on
    `None`, the function will throw an error.

    .. versionadded:: 2.0

    Parameters
    ----------
    sequences: Sequence
        A list of Sequence instances.

    Returns
    -------
    list(Sequence)
        A list of Sequence instances, ordered by date of recording.

    Raises
    ------
    MissingRecordingDateException
        If the recording date is missing from one of the sequences.

    Example
    -------
    >>> sequence_1 = Sequence("C:/Users/Walter/sequence_1.tsv", name="seq_1")  # Sequence recorded 2000-01-01
    >>> sequence_2 = Sequence("C:/Users/Walter/sequence_1.tsv", name="seq_2")  # Sequence recorded 2000-01-03
    >>> sequence_3 = Sequence("C:/Users/Walter/sequence_1.tsv", name="seq_3")  # Sequence recorded 2000-01-02
    >>> print(sort_sequences_by_date(sequence_1, sequence_2, sequence_3))
    [seq_1, seq_3, seq_2]
    """
    dates = {}

    for sequence in sequences:
        if sequence.get_date_recording() is None:
            raise MissingRecordingDateException(sequence)
        else:
            dates[sequence.get_date_recording()] = sequence

    dates_sorted = sorted(dates.keys())
    sequences_ordered = [dates[s] for s in dates_sorted]
    return sequences_ordered


# === File reading functions ===
def get_system_csv_separator():
    """Returns the separator (comma or semicolon) of the regional settings of the system; if it can't access it,
    returns a comma by default.

    .. versionadded:: 2.0

    Note
    ----
    This function detects the local decimal separator, and sets the csv delimiter on a semicolon (``";"``) if the local
    decimal separator is a comma (``","``). In any other case, the csv delimiter is set on a comma (``","``). Note that
    in certain cases, the result of this function may not be ideal (e.g. if the comma can be used as a symbol in the
    values of the csv file). In that case, you can force the csv delimiter to be the symbol of your choice by writing
    ``"csv,"`` or ``"csv;"`` when asked for a file extension.

    Returns
    -------
    str
        The character used as separator in csv files on the current system.
    """
    try:
        import locale
        locale.setlocale(locale.LC_ALL, '')
        dec_pt_chr = locale.localeconv()['decimal_point']
        if dec_pt_chr == ",":
            separator = ";"
        else:
            separator = ","
    except OSError:
        print("Impossible to get the regional list separator, using default comma.")
        separator = ","

    return separator


# Constant
SEPARATOR = get_system_csv_separator()


def get_filetype_separator(extension):
    """Returns the separator used in specific text format files (comma, semicolon or tab). For csv, returns the
    separator used on the current system. For txt or tsv files (or any other extension), returns a tabulation symbol.

    .. versionadded:: 2.0

    Parameters
    ----------
    extension: str
        A file extension (csv, tsv or txt).

    Returns
    -------
    str
         The character used as separator between values in the target file extension.

    Examples
    --------
    >>> get_filetype_separator("csv")
    ","
    >>> get_filetype_separator("tsv")
    "\t"
    """
    if extension == "csv":  # Get the separator from local user (to get , or ;)
        separator = SEPARATOR
    else:  # For text or tsv files, tab separator
        separator = "\t"
    return separator


def read_json(path):
    """Detects the encoding, loads and returns the content of a `.json` file.

    .. versionadded:: 2.0

    Parameters
    ----------
    path: str
        The path to a json file.

    Returns
    -------
    list|dict
        The content of the json file.

    Example
    -------
    >>> read_json("winning_lottery_numbers.json")
    {"January 1": [4, 8, 15, 16, 23, 42], "January 2": [1, 2, 3, 5, 8, 13]}
    """

    with open(path, "rb") as f:
        rawdata = f.read()
        encoding = chardet.detect(rawdata)['encoding']

    with open(path, "r", encoding=encoding) as f:
        content = f.read()

    return json.loads(content)


def read_text_table(path, convert_strings=True, separator=None, read_metadata=True, standardize_labels="auto",
                    keyword_data_start="Timestamp", keyword_data_in_table=True, verbosity=1):
    """Detects the encoding, loads and returns the content of a `.csv`, `.tsv` or `.txt` file containing a table.
    The function also returns a dictionary containing the metadata contained in the file. The metadata is any line
    from the beginning of the file before the appearance of `Timestamp` or `MARKER_NAME` (for Qualisys files).
    Each line of the metadata must contain two elements: a key and a value, separated by one of the three accepted
    separators (tab, ``","`` or ``";"``).

    .. versionadded:: 2.0

    Parameters
    ----------
    path: str
        The path to a text file.
    convert_strings: bool, optional
        If `True` (default), the function will try to convert the values from the table to other types (int, float,
        boolean).
    separator: str|None, optional
        If `None`, the symbol separating each cell on a row in the file will be deduced from the extension, using
        :func:`get_filetype_separator`. Otherwise, the value of this parameter will be used (e.g. ``"\\t"``, ``","``,
        ``";"``).
    read_metadata: bool, optional
        If set on `True` (default), the function tries to find metadata leading the file, before the keyword
        "Timestamps" starts a new line.
    keyword_data_start: str, optional
        The keyword that marks when to stop saving metadata and start saving data (default: ``"Timestamp"``).
    keyword_data_in_table: bool, optional
        If `True` (default), the line starting with the `keyword_data_start` is included in the table. Otherwise,
        the table starts at the next available line.
    standardize_labels: bool|str, optional
        If set on ``"auto"`` (default), renames the Qualisys labels (if ``"HeadR"`` is detected in the joint labels -
        see :ref:`Qualisys to Kualisys conversion<qualisys_to_kualisys>`). Enforce the renaming by setting
        this parameter on `True`.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    list(list)
        The content of the text file, with each sub-list containing the elements of a row of the file.
    dict
        The metadata of the recording, if contained in the text file. This parameter will return an empty dictionary
        if the file does not contain metadata.
    """

    try:
        import chardet
    except ImportError:
        raise ModuleNotFoundException("chardet", "read a table from a text file.")

    # Read the file with the proper encoding
    with open(path, "rb") as f:
        rawdata = f.read()
        encoding = chardet.detect(rawdata)['encoding']

    with open(path, "r", encoding=encoding) as f:
        data = f.read().split("\n")

    # If the data is exported from QTM (Qualisys), convert the tsv to manageable data
    if path.split(".")[-1] == "tsv":
        if data[0].split("\t")[0] == "NO_OF_FRAMES":
            return convert_data_from_qtm(data, standardize_labels, verbosity)

    if separator is None:
        separator = get_filetype_separator(path.split(".")[-1])

    new_data = []
    metadata = {}
    save_data = not read_metadata
    save_data_next_line = False

    for line in data:
        elements = line.split(separator)
        if not save_data:
            if save_data_next_line:
                save_data = True
            elif elements[0] == keyword_data_start or (elements[0] == "" and keyword_data_start == "\n"):
                if keyword_data_in_table:
                    save_data = True
                else:
                    save_data_next_line = True
            else:
                if len(elements) == 2:
                    metadata[elements[0]] = elements[1]
                else:
                    metadata[elements[0]] = elements[1:]

        if save_data and len(line) != 0:
            new_data.append(line.split(separator))

    if convert_strings:
        for i in range(len(new_data)):
            for j in range(len(new_data[i])):
                try:
                    new_data[i][j] = literal_eval(new_data[i][j])
                except ValueError:
                    pass

        for key in metadata.keys():
            try:
                metadata[key] = literal_eval(metadata[key])
            except ValueError:
                pass
            except SyntaxError:
                pass

    return new_data, metadata


def read_xlsx(path, sheet=0, read_metadata=True, metadata_sheet=1, verbosity=1):
    """Loads and returns the content of a ``.xlsx`` (Excel) file.

    .. versionadded:: 2.0

    Note
    ----
    If the Excel file is currently opened, Openpyxl will return a PermissionError. This function catches it and asks
    the user to press the Enter key once the Excel file is closed. This allows to resume a long processing without
    having to restart it from scratch.

    Parameters
    ----------
    path: str
        The path to an Excel file.
    sheet: int|str, optional
        The index of the Excel sheet (starts at 0, which is also the default value), or its name.
    read_metadata: bool, optional
        If `True`, tries to find metadata in the sheet indicated by `sheet_metadata`. If `False` (default), the
        function returns only the first output.
    metadata_sheet: int|str, optional
        The index of the Excel sheet containing the metadata (default: 1), or its name. If no second sheet is present,
        the function returns an empty dict (``{}``) as its second parameter.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    list(list)
        The content of the Excel file, with each element of the list being a row of the Excel file.
    dict
        The metadata of the recording, if contained in the Excel file; ``{}`` otherwise.
    """
    data = None
    metadata = None

    opened = False
    has_metadata = False
    while not opened:
        try:
            workbook = xl.load_workbook(path)
            if len(workbook.sheetnames) > 1:
                has_metadata = True

            if type(sheet) is str:
                data = workbook[sheet]
            elif type(sheet) is int:
                data = workbook[workbook.sheetnames[sheet]]
            else:
                raise Exception(f"The parameter sheet ({sheet}) should be either int or str (not {type(sheet)}).")

            if has_metadata and read_metadata:
                if verbosity > 0:
                    print(f"Metadata found in sheet {workbook.sheetnames[metadata_sheet]}.")

                if type(metadata_sheet) is str:
                    metadata = workbook[metadata_sheet]
                elif type(metadata_sheet) is int:
                    metadata = workbook[workbook.sheetnames[metadata_sheet]]
                else:
                    raise Exception(f"The parameter metadata_sheet ({metadata_sheet}) should be either int or str " +
                                    f"(not {type(metadata_sheet)}).")
            workbook.close()
            opened = True
        except PermissionError:
            print("The Excel file is currently opened.")
            _ = input("Press Enter to try again.")

    # Get the headers (timestamp and joint labels) from the first row
    if verbosity > 1:
        print(f"Reading Excel data...")

    table = []
    for i in range(1, len(data["A"]) + 1):
        row = []
        for j in range(1, len(data["1"]) + 1):
            row.append(data.cell(i, j).value)
        table.append(row)

    metadata_dict = {}
    if has_metadata and read_metadata:
        if verbosity > 1:
            print(f"Reading Excel metadata...")
        width = len(metadata["1"])
        for i in range(1, len(metadata["A"]) + 1):
            values = []
            for j in range(2, width + 1):
                value = metadata.cell(i, j).value
                if value is not None:
                    values.append(value)
            if len(values) == 1:
                values = values[0]
            metadata_dict[metadata.cell(i, 1).value] = values

    return table, metadata_dict


def read_pandas_dataframe(path, verbosity=1):
    """Reads a file, and turns it into a Pandas dataframe.

    .. versionadded:: 2.0

    Note
    ----
    If the Excel file is currently opened, pandas will return a PermissionError. This function catches it and asks
    the user to press the Enter key once the Excel file is closed. This allows to resume a long processing without
    having to restart it from scratch.

    Parameters
    ----------
    path: str
        The path to the file containing data to turn to a pandas dataframe.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    `pandas.Dataframe <https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.html>`_
        A pandas dataframe.
    """

    if path.split(".")[-1] == "json":
        if verbosity > 0:
            print("Reading a dataframe contained in a JSON file...")
        return pd.read_json(path)
    elif path.split(".")[-1] == "csv":
        if verbosity > 0:
            print("Reading a dataframe contained in a CSV file...")
        return pd.read_csv(path, sep=SEPARATOR)
    elif path.split(".")[-1] == "xlsx":
        if verbosity > 0:
            print("Reading a dataframe contained in an Excel file...")
        while True:
            try:
                return pd.read_excel(path)
            except PermissionError:
                print("The Excel file is currently opened.")
                _ = input("Press Enter to try again.")
    elif path.split(".")[-1] == "mat":
        if verbosity > 0:
            print("Reading a dataframe contained in a MAT file...")
        data = loadmat(path)["data"]
        headers = data.dtype
        ndata = {}
        for header in headers.names:
            ndata[header] = data[header][0, 0].flatten().flatten()
            for i in range(len(ndata[header])):
                if isinstance(ndata[header][i], np.ndarray):
                    ndata[header][i] = ndata[header][i][0]
        df = pd.DataFrame(ndata)
        return df
    elif path.split(".")[-1] == "pkl":
        if verbosity > 0:
            print("Reading a dataframe contained in an Pickle file...")
        return pd.read_pickle(path)
    elif path.split(".")[-1] == "gzip":
        if verbosity > 0:
            print("Reading a dataframe contained in an GZIP file...")
        return pd.read_parquet(path)
    elif path.split(".")[-1] in ["tsv", "txt"]:
        if verbosity > 0:
            print("Reading a dataframe contained in an text file...")
        return pd.read_table(path, sep="\t")
    return None


# === Conversion functions ===
def convert_data_from_qtm(data, standardize_labels="auto", verbosity=1):
    """Processes and converts the data from a ``.tsv`` file produced by QTM, by stripping the header data,
    standardizing the name of the joint labels, and converting the distance unit from mm to m. This function then
    returns the loaded table, along with a dictionary containing the metadata of the recording.

    .. versionadded:: 2.0

    Parameters
    ----------
    data: list(str)
        The data from a ``.tsv`` QTM file with each line separated as the elements of a list.
    standardize_labels: bool|str, optional
        If set on ``"auto"`` (default), renames the labels if ``"HeadR"`` is detected in the joint labels
        (see :ref:`Qualisys to Kualisys conversion<qualisys_to_kualisys>`). Enforce the renaming by setting
        this parameter on `True`.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    list(list(str))
        A table containing the QTM data, with each sub-list containing the elements of a row of the file.
    dict
        A dictionary containing the metadata from the recording contained in the file.
    """
    new_data = []
    metadata = {"ORIGIN": "Qualisys"}
    save_data = False
    save_metadata = True

    if verbosity > 0:
        print("\n\tConverting data from Qualisys...")

    joints_conversions = load_qualisys_joint_label_conversion()

    # First, we find where the data starts
    perc = 10
    for i in range(len(data)):
        elements = data[i].split("\t")

        if elements[0] == "MARKER_NAMES":
            save_metadata = False
            header = ["Timestamp"]

            if standardize_labels == "auto" and set(elements[1:]).issubset(set(joints_conversions.keys())):
                standardize_labels = True
            elif standardize_labels is not True:
                standardize_labels = False

            if standardize_labels:
                if verbosity == 1:
                    print("\t\tStandardizing joints labels...", end=" ")
                elif verbosity > 1:
                    print("\t\tStandardizing joints labels...")

            for j in range(1, len(elements)):

                if standardize_labels:
                    # We remove the prefix by looking for a joint name in the label:
                    if elements[j] not in joints_conversions.keys():
                        for label in joints_conversions.keys():
                            if label in elements[j]:
                                if verbosity > 1:
                                    print("\t\t\tChanging label name " + elements[j] + " to " +
                                          joints_conversions[label] + ".")
                                elements[j] = label
                                break

                    header.append(joints_conversions[elements[j]] + "_X")
                    header.append(joints_conversions[elements[j]] + "_Y")
                    header.append(joints_conversions[elements[j]] + "_Z")

                else:
                    header.append(elements[j] + "_X")
                    header.append(elements[j] + "_Y")
                    header.append(elements[j] + "_Z")

            new_data.append(header)

            if standardize_labels:
                if verbosity > 0:
                    print("\t\tLabels standardized.")

        if save_metadata:
            if len(elements) == 2:
                try:
                    metadata[elements[0]] = literal_eval(elements[1])
                except ValueError:
                    pass
                except SyntaxError:
                    metadata[elements[0]] = elements[1]

            else:
                metadata[elements[0]] = elements[1:]

        if elements[0] == "1":
            save_data = True
            if verbosity > 0:
                print("\t\tConverting coordinates from mm to m...", end=" ")

        if save_data:

            perc = show_progression(verbosity, i, len(data), perc)
            line = []

            for j in range(1, len(elements)):

                # We keep the timestamp in ms
                if j == 1:
                    line.append(float(elements[j]))

                # We convert the joints coordinates in m
                else:
                    line.append(float(elements[j]) / 1000)

            if line:
                new_data.append(line)

    if verbosity > 0:
        print("100% - Done.")

    return new_data, metadata


# === File saving functions ===
def write_text_table(table, separator, path, metadata=None, separator_metadata=None, encoding="utf-8", verbosity=1):
    """Converts a table to a string, where elements on the same row are separated by a defined separator, and each row
    is separated by a line break.

    .. versionadded:: 2.0

    Parameters
    ----------
    table: list(list)
        A two-dimensional list where each element is a table row.
    separator: str
        The character used to separate elements on the same row.
    path: str
        The complete path of the text file to save.
    metadata: dict|None, optional
        The metadata of the file, to write as the leading data in the text file.
    separator_metadata: str|None, optional
        If not `None` (default), defines the content of the line separating the metadata from the table.
    encoding: str, optional
        The encoding of the file to save. By default, the file is saved in UTF-8 encoding. This input can take any of
        the `official Python accepted formats <https://docs.python.org/3/library/codecs.html#standard-encodings>`_.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    str
        The table in string version.
    """

    text = ""
    perc = 10

    if metadata is not None and len(metadata) > 0:
        if verbosity > 1:
            print("Writing metadata...", end=" ")
        for key in metadata.keys():
            text += str(key) + separator + str(metadata[key]) + "\n"
        if verbosity > 1:
            print("Done.")

    if separator_metadata is not None:
        if separator_metadata == "\n":
            text += "\n"
        else:
            text += separator_metadata + "\n"

    for i in range(len(table)):
        perc = show_progression(verbosity, i, len(table), perc)
        for j in range(len(table[i])):
            text += str(table[i][j])
            if j != len(table[i]) - 1:
                text += separator
        if i != len(table) - 1:
            text += "\n"

    with open(path, 'w', encoding=encoding) as f:
        f.write(text)


def write_xlsx(table, path, sheet_name=None, metadata=None, metadata_sheet_name=None, verbosity=1):
    """Saves a table in a ``.xlsx`` file.

    .. versionadded:: 2.0

    Parameters
    ----------
    table: list(list)
        A list where each sublist is a row of the table.
    path: str
        The complete path of where to store the Excel file.
    sheet_name: str|None, optional
        The name of the sheet containing the data. If `None`, a default name will be attributed to the sheet
        (``"Sheet"``).
    metadata: dict|None, optional
        A dictionary containing metadata keys and values. These will be saved in a separate second sheet, with each
        key/value pair on a different row.
    metadata_sheet_name: str|None, optional
        The name of the sheet containing the metadata. If `None`, a default name will be attributed to the sheet
        (``"Sheet1"``).
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.
    """
    workbook_out = xl.Workbook()
    sheet_data = workbook_out.active
    if sheet_name is not None:
        sheet_data.title = "Sheet"

    if verbosity > 0:
        print("Writing the data...", end=" ")

    perc = 10
    for i in range(len(table)):
        perc = show_progression(verbosity, i, len(table), perc)
        for j in range(len(table[i])):
            sheet_data.cell(i + 1, j + 1, table[i][j])

    if verbosity > 0:
        print("Done.")

    if metadata is not None:

        if verbosity > 0:
            print("Writing the metadata...")

        sheet_metadata = workbook_out.create_sheet()
        if metadata_sheet_name is not None:
            sheet_metadata.title = metadata_sheet_name
        i = 1
        for key in metadata.keys():
            sheet_metadata.cell(i, 1, key)
            if type(metadata[key]) is list:
                for j in range(len(metadata[key])):
                    sheet_metadata.cell(i, j + 2, metadata[key][j])
            else:
                sheet_metadata.cell(i, 2, metadata[key])
            i += 1

        if verbosity > 0:
            print("Done.")

    workbook_out.save(path)
    workbook_out.close()


# === Calculation functions ===
def find_closest_value_index(array, value, rtol=1e-5, atol=1e-8):
    """Find the index of the value in the array that is closest to the given value within a specified tolerance.

    .. versionadded:: 2.0

    Parameters
    ----------
    array: numpy.ndarray
        The array to search.
    value: float
        The value to find in the array.
    rtol: float
        The relative tolerance parameter for numpy.isclose.
    atol: float
        The absolute tolerance parameter for numpy.isclose.

    Returns
    -------
    int|None
        The index of the closest value if found, otherwise None.
    """

    # Calculate absolute differences between the value and each element in the array
    differences = np.abs(array - value)

    # Use np.isclose to find indices of values close to the given value
    close_indices = np.isclose(array, value, rtol=rtol, atol=atol)

    if np.any(close_indices):
        # Find the index of the minimum difference among the close values
        closest_index = np.argmin(differences[close_indices])
        return np.where(close_indices)[0][closest_index]
    else:
        return None

def resample_data(array, original_timestamps, resampling_frequency, window_size=1e7, overlap_ratio=0.5,
                  method="cubic", time_unit="s", verbosity=1):
    """Resamples an array to the `resampling_frequency` parameter. It first creates a new set of timestamps at the
    desired frequency, and then interpolates the original data to the new timestamps.

    Parameters
    ----------
    array: list or np.ndarray
        An array of samples.

    original_timestamps: list(float) or numpy.ndarray(float)
        A list or an array of the time points corresponding to the values of the data.

    resampling_frequency: int or float
        The frequency at which you want to resample the array, in Hz. A frequency of 4 will return samples
        at 0.25 s intervals.

    window_size: int, optional
        The size of the windows in which to cut the elements to perform the resampling. Cutting long arrays
        in windows allows to speed up the computation. If this parameter is set on `None`, the window size will be
        set on the number of elements. A good value for this parameter is generally 10 million (1e7). If this
        parameter is set on 0, on None or on a number of samples bigger than the amount of elements in the array, the
        window size is set on the length of the array.

    overlap_ratio: float or None, optional
        The ratio of samples overlapping between each window. If this parameter is not `None`, each window will
        overlap with the previous (and, logically, the next) for an amount of samples equal to the number of samples in
        a window times the overlap ratio. Then, only the central values of each window will be preserved and
        concatenated; this allows to discard any "edge" effect due to the windowing. If the parameter is set on `None`
        or 0, the windows will not overlap.

    method: str, optional
        This parameter allows for various values:

        • ``"linear"`` performs a linear
          `numpy.interp <https://numpy.org/devdocs/reference/generated/numpy.interp.html>`_ interpolation. This method,
          though simple, may not be very precise for upsampling naturalistic stimuli.
        • ``"cubic"`` performs a cubic interpolation via `scipy.interpolate.CubicSpline
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.CubicSpline.html>`_. This method,
          while smoother than the linear interpolation, may lead to unwanted oscillations nearby strong variations in
          the data.
        • ``"pchip"`` performs a monotonic cubic spline interpolation (Piecewise Cubic Hermite Interpolating
          Polynomial) via `scipy.interpolate.PchipInterpolator
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.PchipInterpolator.html>`_.
        • ``"akima"`` performs another type of monotonic cubic spline interpolation, using
          `scipy.interpolate.Akima1DInterpolator
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.Akima1DInterpolator.html>`_.
        • ``"take"`` keeps one out of n samples from the original array. While being the fastest computation, it will
          be prone to imprecision if the downsampling factor is not an integer divider of the original frequency.
        • ``"interp1d_XXX"`` uses the function `scipy.interpolate.interp1d
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`. The XXX part of the
          parameter can be replaced by ``"linear"``, ``"nearest"``, ``"nearest-up"``, ``"zero"``, "slinear"``,
          ``"quadratic"``, ``"cubic"``, ``"previous"``, and ``"next"`` (see the documentation of this function for
          specifics).

    time_unit: str, optional
        The time unit of the time points. This parameter can take the following values: "ns", "1ns", "10ns", "100ns",
        "µs", "1µs", "10µs", "100µs", "ms", "1ms", "10ms", "100ms", "s" (default), "sec", "1s", "min", "mn", "h", "hr",
        "d", "day".

    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    numpy.ndarray(float)
        The resampled values.
    numpy.ndarray(float)
        The resampled time points, at a fixed frequency.

    Warning
    -------
    This function allows both the **upsampling** and the **downsampling** of arrays. However, during any of
    these operations, the algorithm only **estimates** the real values of the samples. You should then consider
    the upsampling (and the downsampling, to a lesser extent) with care.

    Example
    -------
    >>> array = np.linspace(1, 0, 11)
    >>> original_timestamps = np.linspace(0, 2, 11)
    >>> resampling_frequency = 4
    >>> resample_data(array, original_timestamps, resampling_frequency, method="linear", verbosity=0)
    ([1, 0.875, 0.75, 0.625, 0.5, 0.375, 0.25, 0.125, 0], [0, 0.25, 0.5, 0.75, 1, 1.25, 1.5, 1.75, 2])
    """

    if len(array) != len(original_timestamps):
        raise Exception("The length of the data array (" + str(len(array)) + ") is inconsistent with the length of " +
                        "the time points (" + str(len(original_timestamps)) + ").")

    if verbosity > 0:
        print("\tResampling the array at " + str(resampling_frequency) + " Hz (mode: " + str(method) + ")...")
        if verbosity > 1:
            print("\t\tPerforming the resampling...")
        else:
            print("\t\tPerforming the resampling...", end=" ")

    step = (1 / resampling_frequency) * UNITS[time_unit]

    # Define the timestamps
    start_timestamp = np.round(original_timestamps[0], 6)
    end_timestamp = np.round(original_timestamps[-1], 6)
    resampled_timestamps = np.arange(start_timestamp, end_timestamp + (1 / resampling_frequency),
                                     1 / resampling_frequency)

    # If the last timestamp is over the last original timestamp, we remove it
    if np.round(resampled_timestamps[-1], 6) > end_timestamp:
        resampled_timestamps = resampled_timestamps[:-1]

    # Settings
    if overlap_ratio is None:
        overlap_ratio = 0

    window_size = int(window_size)
    overlap = int(np.ceil(overlap_ratio * window_size))
    number_of_windows = get_number_of_windows(len(array), window_size, overlap_ratio, True)

    if verbosity > 1 and number_of_windows != 1:
        print("\t\t\tCreating " + str(number_of_windows) + " window(s), each containing " + str(window_size) +
              " samples, with a " + str(round(overlap_ratio * 100, 2)) + " % overlap (" + str(overlap) +
              " samples).")

    resampled_array = np.zeros(np.size(resampled_timestamps))
    j = 0
    next_percentage = 10

    for i in range(number_of_windows):

        window_start_original = i * (window_size - overlap)
        window_end_original = np.min([(i + 1) * window_size - i * overlap, len(original_timestamps) - 1])
        window_start_resampled = int(np.round(original_timestamps[window_start_original] * resampling_frequency))
        window_end_resampled = int(np.round(original_timestamps[window_end_original] * resampling_frequency))

        if verbosity == 1:
            while i / number_of_windows > next_percentage / 100:
                print(str(next_percentage) + "%", end=" ")
                next_percentage += 10

        if verbosity > 1:
            print("\t\t\tGetting samples from window " + str(i + 1) + "/" + str(number_of_windows) + ".")

        resampled_window = resample_window(array, original_timestamps, resampled_timestamps, window_start_original,
                                           window_end_original, window_start_resampled, window_end_resampled,
                                           method, verbosity)

        if verbosity > 1:
            print("Done.\n\t\t\t\tThe resampled window contains " + str(np.size(resampled_window)) + " sample(s).")

        # Keep only the center values
        if i == 0:
            window_slice_start = 0
            resampled_slice_start = 0
        else:
            window_slice_start = (j - window_start_resampled) // 2
            resampled_slice_start = window_start_resampled + window_slice_start

        preserved_samples = resampled_window[window_slice_start:]

        if verbosity > 1:
            print("\t\t\tKeeping the samples from " + str(resampled_slice_start) + " to " +
                  str(resampled_slice_start + len(preserved_samples) - 1) + " (in the window, samples " +
                  str(window_slice_start) + " to " + str(len(resampled_window) - 1) + ")", end="")
            if resampled_slice_start != j:
                print(" · Rewriting samples " + str(resampled_slice_start) + " to " + str(j - 1) + "...", end=" ")
            else:
                print("...", end=" ")

        resampled_array[resampled_slice_start:resampled_slice_start + len(preserved_samples)] = preserved_samples
        j = resampled_slice_start + len(preserved_samples)

        if verbosity > 1:
            print("Done.")

    if verbosity == 1:
        print("100% - Done.")
    elif verbosity > 1:
        print("\t\tResampling done.")

    if verbosity > 0:
        print("\t\tThe original array had " + str(len(array)) + " samples.")
        print("\t\tThe new array has " + str(len(resampled_array)) + " samples.")

    return resampled_array, resampled_timestamps


def resample_window(array, original_timestamps, resampled_timestamps, index_start_original, index_end_original,
                    index_start_resampled, index_end_resampled, method="cubic", verbosity=1):
    """Performs and returns the resampling on a subarray of samples.

    Parameters
    ----------
    array: list or np.ndarray
        An array of samples.

    original_timestamps: list or np.ndarray
        An array containing the timestamps for each sample of the original array.

    resampled_timestamps: list or np.ndarray
        An array containing the timestamps for each desired sample in the resampled array.

    index_start_original: int
        The index in the array where the window starts.

    index_end_original: int
        The index in the array where the window ends.

    index_start_resampled: int
        The index in the resampled array where the window starts.

    index_end_resampled: int
        The index in the resampled array where the window ends.

    method: str, optional
        This parameter allows for various values:

        • ``"linear"`` performs a linear
          `numpy.interp <https://numpy.org/devdocs/reference/generated/numpy.interp.html>`_ interpolation. This method,
          though simple, may not be very precise for upsampling naturalistic stimuli.
        • ``"cubic"`` performs a cubic interpolation via `scipy.interpolate.CubicSpline
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.CubicSpline.html>`_. This method,
          while smoother than the linear interpolation, may lead to unwanted oscillations nearby strong variations in
          the data.
        • ``"pchip"`` performs a monotonic cubic spline interpolation (Piecewise Cubic Hermite Interpolating
          Polynomial) via `scipy.interpolate.PchipInterpolator
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.PchipInterpolator.html>`_.
        • ``"akima"`` performs another type of monotonic cubic spline interpolation, using
          `scipy.interpolate.Akima1DInterpolator
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.Akima1DInterpolator.html>`_.
        • ``"take"`` keeps one out of n samples from the original array. While being the fastest computation, it will
          be prone to imprecision if the downsampling factor is not an integer divider of the original frequency.
        • ``"interp1d_XXX"`` uses the function `scipy.interpolate.interp1d
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`. The XXX part of the
          parameter can be replaced by ``"linear"``, ``"nearest"``, ``"nearest-up"``, ``"zero"``, "slinear"``,
          ``"quadratic"``, ``"cubic"``, ``"previous"``, and ``"next"`` (see the documentation of this function for
          specifics).

    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    np.array
        The envelope of the original array.
    """

    array_window = array[index_start_original:index_end_original + 1]
    original_timestamps_window = original_timestamps[index_start_original:index_end_original + 1]
    resampled_timestamps_window = resampled_timestamps[index_start_resampled:index_end_resampled + 1]

    if verbosity > 1:
        print("\t\t\t\tIn the original array, the window contains samples " + str(index_start_original) + " to " +
              str(index_end_original) + " (from timestamps " + str(original_timestamps_window[0]) + " to " +
              str(original_timestamps_window[-1]) + ").")
        print("\t\t\t\tIn the new array, the window contains samples " + str(index_start_resampled) + " to " +
              str(index_end_resampled) + " (from timestamps " + str(resampled_timestamps_window[0]) + " to " +
              str(resampled_timestamps_window[-1]) + ").")
        print("\t\t\t\tInterpolating the data...", end=" ")

    if np.size(array_window) == 1:
        raise Exception("Only one sample is present in the current window. Please select a larger window size.")

    if method == "linear":
        return np.interp(resampled_timestamps_window, original_timestamps_window, array_window)
    elif method == "cubic":
        interp = CubicSpline(original_timestamps_window, array_window)
        return interp(resampled_timestamps_window)
    elif method == "pchip":
        interp = PchipInterpolator(original_timestamps_window, array_window)
        return interp(resampled_timestamps_window)
    elif method == "akima":
        interp = Akima1DInterpolator(original_timestamps_window, array_window)
        return interp(resampled_timestamps_window)
    elif method.startswith("interp1d"):
        interp = interp1d(original_timestamps, array, kind=method.split("_")[1])
        return interp(resampled_timestamps_window)
    else:
        raise Exception("Invalid resampling method: " + str(method) + ".")


def interpolate_data(data, time_points_data, time_points_interpolation, method="linear"):
    """Interpolates incomplete data to a linear array of values.

    .. versionadded:: 2.0

    Important
    ---------
    This function is dependent of the modules `numpy <https://numpy.org/>`_ and `scipy <https://scipy.org/>`_.

    Parameters
    ----------
    data: list(float) or numpy.ndarray(float)
        A list or an array of values.
    time_points_data: list(float) or numpy.ndarray(float)
        A list or an array of the time points corresponding to the values of the data.
    time_points_interpolation: list(float) or numpy.ndarray(float)
        A list or an array of the time points to which interpolate the data.
    method: str, optional
        This parameter allows for various values:

        • ``"linear"`` performs a linear
          `numpy.interp <https://numpy.org/devdocs/reference/generated/numpy.interp.html>`_ interpolation. This method,
          though simple, may not be very precise for upsampling naturalistic stimuli.
        • ``"cubic"`` performs a cubic interpolation via `scipy.interpolate.CubicSpline
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.CubicSpline.html>`_. This method,
          while smoother than the linear interpolation, may lead to unwanted oscillations nearby strong variations in
          the data.
        • ``"pchip"`` performs a monotonic cubic spline interpolation (Piecewise Cubic Hermite Interpolating
          Polynomial) via `scipy.interpolate.PchipInterpolator
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.PchipInterpolator.html>`_.
        • ``"akima"`` performs another type of monotonic cubic spline interpolation, using
          `scipy.interpolate.Akima1DInterpolator
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.Akima1DInterpolator.html>`_.
        • ``"take"`` keeps one out of n samples from the original array. While being the fastest computation, it will
          be prone to imprecision if the downsampling factor is not an integer divider of the original frequency.
        • ``"interp1d_XXX"`` uses the function `scipy.interpolate.interp1d
          <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`. The XXX part of the
          parameter can be replaced by ``"linear"``, ``"nearest"``, ``"nearest-up"``, ``"zero"``, "slinear"``,
          ``"quadratic"``, ``"cubic"``, ``"previous"``, and ``"next"`` (see the documentation of this function for
          specifics).

    Returns
    -------
    numpy.ndarray(float)
        The interpolated values.
    numpy.ndarray(float)
        The interpolated time points (same as the parameter ``time_points_interpolation``).

    Example
    -------
    >>> array = np.linspace(1, 0, 11)
    >>> original_timestamps = np.linspace(0, 1, 11)
    >>> interpolation_timestamps = np.linspace(0.05, 0.95, 10)
    >>> interpolate_data(array, original_timestamps, interpolation_timestamps, "linear")
    ([0.95 0.85 0.75 0.65 0.55 0.45 0.35 0.25 0.15 0.05], [0.05 0.15 0.25 0.35 0.45 0.55 0.65 0.75 0.85 0.95])
    """

    if len(data) != len(time_points_data):
        raise Exception("The length of the data array (" + str(len(data)) + ") is inconsistent with the length of " +
                        "the time points (" + str(len(time_points_data)) + ").")

    np_data = np.array(data)
    np_time_points = np.array(time_points_data)
    np_time_points_complete = np.array(time_points_interpolation)

    #print(time_points_data)

    if method == "linear":
        return np.interp(np_time_points_complete, np_time_points, np_data), np_time_points_complete
    elif method == "cubic":
        interp = CubicSpline(np_time_points, np_data)
        return interp(np_time_points_complete), np_time_points_complete
    elif method == "pchip":
        for t in range(1, len(time_points_data)):
            if time_points_data[t] <= time_points_data[t-1]:
                print(t, t-1, time_points_data[t], time_points_data[t-1])
        interp = PchipInterpolator(np_time_points, np_data)
        return interp(np_time_points_complete), np_time_points_complete
    elif method == "akima":
        interp = Akima1DInterpolator(np_time_points, np_data)
        return interp(np_time_points_complete), np_time_points_complete
    elif method.startswith("interp1d"):
        interp = interp1d(np_time_points, np_data, kind=method.split("_")[1])
        return interp(np_time_points_complete), np_time_points_complete
    else:
        raise Exception("Invalid resampling method: " + str(method) + ".")

def pad(data, time_points_data, time_points_padding, padding_value=0, verbosity=1):
    """Given an array of values (``data``) and its corresponding ``time_points_data``, and a larger array
    ``time_points_padding``, pads the data array with the value specified in ``padding_value`` for the time points
    present in ``time_points_padding`` that are absent from ``time_points_data``.

    .. versionadded:: 2.0

    Parameters
    ----------
    data: list(float) or numpy.ndarray(float)
        A list or an array of values.
    time_points_data: list(float) or numpy.ndarray(float)
        A list or an array of the time points corresponding to the values of the data.
    time_points_padding: list(float) or numpy.ndarray(float)
        A list or an array of time points containing the values from ``time_points_data`` (or values equal up to the
        fifth decimal), with additional values at the beginning and/or at the end.
    padding_value: int|numpy.nan|float|str
        The value with which to pad the data (default: 0), or ``"edge"`` to copy the value present at the edges of the
        data.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    np.ndarray
        The original data array, padded with zeros.
    list
        The array ``time_points_padding``, passed as parameter.

    Example
    -------
    >>> data = [5, 4, 3, 2, 1]
    >>> time_points_data = [1, 2, 3, 4, 5]
    >>> time_points_padding = [0, 1, 2, 3, 4, 5, 6, 7]
    >>> pad(data, time_points_data, time_points_padding, 0, verbosity=0)
    ([0 5 4 3 2 1 0 0], [0, 1, 2, 3, 4, 5, 6, 7])
    """

    if type(data) == list:
        data = np.array(data)

    if type(time_points_data) == list:
        time_points_data = np.array(time_points_data)

    if type(time_points_padding) == list:
        time_points_padding = np.array(time_points_padding)

    if data.size != time_points_data.size:
        raise Exception("The length of the data array (" + str(data.size) + ") is inconsistent with the length of " +
                        "the time points (" + str(time_points_data.size) + ").")

    if time_points_data[0] < time_points_padding[0]:
        raise Exception("The first point of the time points padding array (" + str(time_points_padding[0]) + ") must " +
                        "be inferior or equal to the first time point of the data (" + str(time_points_data[0]) + ").")

    if time_points_data[-1] > time_points_padding[-1]:
        raise Exception("The last point of the time points padding array (" + str(time_points_padding[-1]) + ") must " +
                        "be superior or equal to the last time point of the data (" + str(time_points_data[-1]) + ").")

    # Get how much to pad at the beginning and at the end
    pad_start = np.argmin(np.abs(time_points_padding - time_points_data[0]))
    pad_end = len(time_points_padding) - (pad_start + len(time_points_data))

    if not np.allclose(time_points_padding[pad_start:pad_start + len(time_points_data)], time_points_data):
        raise Exception("The array time_points_padding must contain the timestamps of time_points_data.")

    if verbosity > 1:
        print(f"Adding {pad_start} samples to the beginning and {pad_end} samples to the end.")

    if padding_value == "edge":
        padded_data = np.pad(data, (pad_start, pad_end), "edge")
    else:
        padded_data = np.pad(data, (pad_start, pad_end), "constant", constant_values=padding_value)

    return padded_data, time_points_padding


def add_delay(timestamps, delay):
    """Given an array of timestamps, adds or removes a delay to each timestamp.

    .. versionadded:: 2.0

    Parameters
    ----------
    timestamps: numpy.array|list(float)
        An array of timestamps.
    delay: float
        The delay, positive or negative to add to each timestamp.

    Returns
    -------
    list(float)
        An array of timestamps with the delay specified as parameter.

    Example
    -------
    >>> timestamps = [1, 2, 3, 4, 5]
    >>> add_delay(timestamps, 5)
    [6, 7, 8, 9, 10]
    """

    if type(timestamps) is list:
        timestamps = np.array(timestamps)

    return timestamps + delay


def calculate_distance(joint1, joint2, axis=None):
    """Uses the Euclidian formula to calculate the distance between two joints. This can be used to calculate
    the distance travelled by one joint between two poses, or the distance between two joints on the same pose.
    If an axis is specified, the distance is calculated on this axis only.

    .. versionadded:: 2.0

    Parameters
    ----------
    joint1: Joint
        The first Joint object.
    joint2: Joint
        The second Joint object.
    axis: str|None, optional
        If specified, the returned distances travelled will be calculated on a single axis. If ``None`` (default),
        the distance travelled will be calculated based on the 3D coordinates.

    Returns
    -------
    float
        The absolute distance, in meters, between the two joints.

    Example
    -------
    >>> sequence_1 = Sequence("sequences/leela/sequence_1.json", verbosity=0)
    >>> joint_1 = sequence_1.poses[0].joints["Head"]
    >>> joint_2 = sequence_1.poses[1].joints["Head"]
    >>> calculate_distance(joint_1, joint_2)
    1.6885703516949238
    """
    if axis == "x":
        return abs(joint2.x - joint1.x)
    elif axis == "y":
        return abs(joint2.y - joint1.y)
    elif axis == "z":
        return abs(joint2.z - joint1.z)
    elif axis is None:
        x = (joint2.x - joint1.x) ** 2
        y = (joint2.y - joint1.y) ** 2
        z = (joint2.z - joint1.z) ** 2
        return math.sqrt(x + y + z)
    else:
        raise InvalidParameterValueException("axis", axis, ["x", "y", "z"])

def calculate_consecutive_distances(array):
    """Given an array of values, calculates and returns the absolute distance travelled between each point and the
    previous. The length of the returned array will be one element smaller than the original array.

    .. versionadded:: 2.0

    Parameters
    ----------
    array: list(float)|numpy.array
        A list or array containing values.

    Returns
    -------
    numpy.array
        An array of absolute distances, with a length of n-1 elements compared to the array parameter.

    Example
    -------
    >>> calculate_consecutive_distances([4, 8, 15, 16, 23, 42])
    [4, 7, 1, 7, 19]
    """
    if type(array) is list:
        array = np.array(array)

    distance_vector = np.abs(array[1:] - array[:-1])
    return distance_vector

def calculate_euclidian_distances(x_array, y_array, z_array=None):
    """
    Given two or three arrays of values, returns an array of euclidian distances between consecutive values.

    .. versionadded:: 2.0

    Parameters
    ----------
    x_array: list(float)|numpy.array
        A list or array containing positions on the x-axis.
    y_array: list(float)|numpy.array
        A list or array containing positions on the y-axis.
    z_array: list(float)|numpy.array|None
        A list or array containing positions on the z-axis (optional).

    Example
    -------
    >>> calculate_euclidian_distances([3, 4, 5], [5, 12, 13], [8, 15, 17])
    [9.94987437 2.44948974]
    """
    distances_x = np.power(calculate_consecutive_distances(x_array), 2)
    distances_y = np.power(calculate_consecutive_distances(y_array), 2)
    if z_array is None:
        return np.sqrt(distances_x + distances_y)
    else:
        distances_z = np.power(calculate_consecutive_distances(z_array), 2)
        return np.sqrt(distances_x + distances_y + distances_z)

def calculate_derivative(array, derivative="velocity", window_length=7, poly_order=None, freq=1):
    """

    .. versionadded:: 2.0

    Parameters
    ----------
    array: numpy.array|list
        An array of numbers representing the distance travelled for each pair of timestamps.

    derivative: str|int, optional
        The derivative to calculate, as an integer or as a string, among:

        • ``"velocity"``, ``"speed"``, or ``1`` (default)
        • ``"acceleration"`` or ``2``
        • ``"jerk"`` or ``3``
        • ``"snap"``, ``"jounce"``, or ``4``
        • ``"crackle"`` or ``5``
        • ``"pop"`` or ``6``
        • Though probably unnecessary, any higher derivative can be referred to by an integer.

    window_length: int, optional
        The length of the window for the Savitzky–Golay filter (default: 7). See
        `scipy.signal.savgol_filter <https://docs.scipy.org/doc/scipy/reference/generated/scipy.signal.savgol_filter.html>`_
        for more info.

    poly_order: int|None, optional
        The order of the polynomial for the Savitzky–Golay filter. If set on None, the polynomial will be set to one
        over the derivative rank. See
        `scipy.signal.savgol_filter <https://docs.scipy.org/doc/scipy/reference/generated/scipy.signal.savgol_filter.html>`_
        for more info.

    freq: int|float, optional
        The frequency of the original data, in Hz (default: 1).
    """

    derivatives = {"velocity": 1, "speed": 1, "acceleration": 2, "jerk": 3, "jounce": 4, "snap": 4, "crackle": 5,
                   "pop": 6}

    if type(derivative) is str:

        if derivative in derivatives.keys():
            derivative = derivatives[derivative]
        else:
            raise InvalidParameterValueException("derivative", derivative, list(derivatives.keys()))

    if poly_order is None:
        poly_order = derivative + 1
        if poly_order >= window_length:
            poly_order = window_length - 1

    return savgol_filter(array, window_length, poly_order, derivative, mode="interp") * np.power(freq, derivative)


def calculate_delay(pose1, pose2, absolute=False):
    """Returns the delay between two poses, in seconds.

    .. versionadded:: 2.0

    Parameters
    ----------
    pose1: Pose
        The first Pose object.
    pose2: Pose
        The second Pose object.
    absolute: bool
        If set on ``True``, returns a positive value no matter the order of the poses (default: ``False``).

    Returns
    -------
    float
        The time, in seconds, between the two poses.

    Note
    ----
    By default, the time returned by the function is **not** an absolute value. In other words, if ``pose1`` has a
    higher timestamp than ``pose2``, the returned value will be negative.

    Example
    -------
    >>> sequence_1 = Sequence("C:/Users/Adrian/sequence_1.tsv")  # Sampled at 10 Hz, each pose lasts 0.1 s
    >>> calculate_delay(sequence_1.poses[5], sequence_1.poses[6])
    0.1
    >>> calculate_delay(sequence_1.poses[6], sequence_1.poses[5])
    -0.1
    >>> calculate_delay(sequence_1.poses[6], sequence_1.poses[5], absolute=True)
    0.1
    """
    time_a = pose1.get_relative_timestamp()
    time_b = pose2.get_relative_timestamp()

    if absolute:
        return np.abs(time_b - time_a)
    else:
        return time_b - time_a


def generate_random_joints(number_of_joints, x_scale=0.2, y_scale=0.3, z_scale=0.5):
    """Creates and returns a list of Joint objects with random coordinates. The coordinates are generated following
    a uniform distribution centered around 0 and with limits defined by ``x_scale``, ``y_scale`` and ``z_scale``.

    .. versionadded:: 2.0

    Parameters
    ----------
    number_of_joints: int
        The number of random joints to generate.
    x_scale: int|float
        The absolute maximum value of the random uniform distribution on the x-axis.
    y_scale: int|float
        The absolute maximum value of the random uniform distribution on the y-axis.
    z_scale: int|float
        The absolute maximum value of the random uniform distribution on the z axis.

    Returns
    -------
    list(Joint)
        A list of joints with randomized coordinates.
    """
    random_joints = []
    for i in range(number_of_joints):
        x = np.random.uniform(-x_scale, x_scale)
        y = np.random.uniform(-y_scale, y_scale)
        z = np.random.uniform(-z_scale, z_scale)
        j = Joint(None, x, y, z)
        random_joints.append(j)
    return random_joints


def get_number_of_windows(array_length_or_array, window_size, overlap_ratio=0, add_incomplete_window=True):
    """Given an array, calculates how many windows from the defined `window_size` can be created, with or
    without overlap.

    .. versionadded:: 2.0

    Parameters
    ----------
    array_length_or_array: list(int or float) or np.array(int or float) or int
        An array of numerical values, or its length.
    window_size: int
        The number of array elements in each window.
    overlap_ratio: int or float
        The ratio of array elements overlapping between each window and the next. In the case where the overlap converts
        to a non-integer amount of values, the closest upper integer is used (e.g., if the array is of length 10, the
        window size of length 5 and the overlap ratio of 0.5, the amount of values in the overlap will not be 2.5 but
        3).
    add_incomplete_window: bool
        If set on ``True``, the last window will be included even if its size is smaller than ``window_size``.
        Otherwise, it will be ignored.

    Returns
    -------
    int
        The number of windows than can be created from the array.

    Examples
    --------
    >>> get_number_of_windows(100, 10, 0, True)
    10
    >>>  get_number_of_windows(100, 10, 0.5, True)
    19
    """

    if not isinstance(array_length_or_array, int):
        array_length = len(array_length_or_array)
    else:
        array_length = array_length_or_array

    overlap = int(np.ceil(overlap_ratio * window_size))

    if overlap_ratio >= 1:
        raise Exception("The size of the overlap (" + str(overlap) + ") cannot be bigger than or equal to the " +
                        "size of the window (" + str(window_size) + ").")
    if window_size > array_length:
        window_size = array_length

    number_of_windows = (array_length - overlap) / (window_size - overlap)

    if add_incomplete_window and array_length + (overlap * (window_size - 1)) % window_size != 0:
        return int(np.ceil(number_of_windows))
    return int(number_of_windows)


def get_window_length(array_length_or_array, number_of_windows, overlap_ratio, round_up=True):
    """Given an array to be split in a given overlapping number of windows, calculates the number of elements in each
    window.

    .. versionadded:: 2.0

    Parameters
    ----------
    array_length_or_array: list, np.ndarray or int
        An array of numerical values, or its length.
    number_of_windows: int
        The number of windows to split the array in.
    overlap_ratio: float
        The ratio of overlapping elements between each window.
    round_up: bool, optional
        Specifies if the result should be returned rounded up to the nearest integer (default) or not.

    Returns
    -------
    int|float
        The number of elements in each window. Note: the last window may have fewer elements than the others if the
        number of windows does not divide the result of :math:`array_length + (number_of_windows - 1) × overlap`.

    Examples
    --------
    >>> get_window_length(100, 10, 0)
    10
    >>> get_window_length(10, 8, 0.5)
    3
    >>> get_window_length(10, 8, 0.5, False)
    2.22222222
    """

    if not isinstance(array_length_or_array, int):
        array_length = len(array_length_or_array)
    else:
        array_length = array_length_or_array

    if overlap_ratio >= 1 or overlap_ratio < 0:
        raise Exception("The size of the overlap ratio (" + str(overlap_ratio) + ") must be superior or equal to 0, " +
                        "and strictly inferior to 1.")

    if round_up:
        return np.ceil(array_length / (number_of_windows + overlap_ratio - number_of_windows * overlap_ratio))
    return array_length / (number_of_windows + overlap_ratio - number_of_windows * overlap_ratio)


def divide_in_windows(array, window_size, overlap=0, add_incomplete_window=True):
    """Given an array of values, divides the array in windows of a size ``window_size``, with or without an overlap.

    .. versionadded:: 2.0

    Parameters
    ----------
    array: list(int or float) or np.array(int or float)
        An array of numerical values.
    window_size: int
        The number of array elements in each window.
    overlap: int
        The number of array elements overlapping in each window.
    add_incomplete_window: bool
        If set on ``True``, the last window will be added even if its size is smaller than ``window_size``. Otherwise,
        it will be ignored.

    Returns
    -------
    list(list(int or float))
        A list where each element is a window of size ``window_size``.

    Example
    -------
    >>> array = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
    >>> divide_in_windows(array, 5, 0)
    [[1, 2, 3, 4, 5], [6, 7, 8, 9, 10], [11, 12, 13, 14, 15], [16, 17, 18, 19, 20]]
    """
    if overlap >= window_size:
        raise Exception("The size of the overlap (" + str(overlap) + ") cannot be bigger than or equal to the size " +
                        "of the window (" + str(window_size) + ").")
    if overlap > len(array) or window_size > len(array):
        raise Exception("The size of the window (" + str(window_size) + ") or the overlap (" + str(overlap) + ") " +
                        "cannot be bigger than the size of the array (" + str(len(array)) + ").")

    window_start = 0
    windows = []

    while window_start + window_size <= len(array):
        windows.append(array[window_start : window_start + window_size])
        window_start += window_size - overlap

    if add_incomplete_window:
        if array[window_start:]:
            windows.append(array[window_start:])

    return windows


# === Color functions ===
def load_color_names():
    """Returns a dictionary containing the accepted color names, and a tuple of their 256-level RGBA codes as values.
    The colors are contained in ``res/color_codes.txt``. The color names follow the `140 HTML/CSS color names
    <https://en.wikipedia.org/wiki/X11_color_names>`_), and also contain all the accepted colors from
    `Matplotlib <https://matplotlib.org/stable/gallery/color/named_colors.html>`, with a few extra.
    You can find more info about the accepted colors in :doc:`../appendix/color_schemes`.

    .. versionadded:: 2.0

    Returns
    -------
    dict(str: tuple(int))
        A dictionary of the color names and their RGBA values.
    """
    colors_dict = {}

    with open(op.join(MODULE_DIR, "res/color_codes.txt"), "r", encoding="utf-8") as f:
        content = f.read().split("\n")

    for line in content:
        els = line.split("\t")
        colors_dict[els[0]] = (int(els[1]), int(els[2]), int(els[3]), int(els[4]))

    return colors_dict


def load_color_schemes():
    """Returns a dictionary containing the color gradients saved in ``res/color_gradients.txt``.

    .. versionadded:: 2.0

    Returns
    -------
    dict(int: list(tuple(int, int, int, int)))
        A dictionary with the name of color gradients as keys, and a list of their corresponding RGBA colors as values.
    """
    color_schemes = {}
    colors_dict = load_color_names()
    with open(op.join(MODULE_DIR, "res/color_schemes.txt"), "r", encoding="utf-8") as f:
        content = f.read().split("\n")

    for line in content:
        elements = line.split("\t")
        colors = []
        for e in range(1, len(elements)):
            if elements[e][0] == "(":
                colors_string = elements[e][1:-1].split(",")
                color = []
                for c in colors_string:
                    color.append(int(c))
                colors.append(tuple(color))
            elif elements[e][0] == "#":
                if len(elements[e]) == 7:
                    color = list(int(elements[e][i:i + 2], 16) for i in (1, 3, 5))
                    color.append(255)
                else:
                    color = list(int(elements[e][i:i + 2], 16) for i in (1, 3, 5, 7))
                colors.append(tuple(color))
            else:
                colors.append(colors_dict[elements[e]])

        color_schemes[elements[0]] = colors

    return color_schemes


def hex_color_to_rgb(color, include_alpha=None):
    """Converts a color from its hexadecimal value to its RGB or RGBA value.

    Parameters
    ----------
    color: str
        The hexadecimal value of a color, with or without a leading number sign (``"#"``).
    include_alpha: bool|None, optional
        If ``True``, returns the RGB color with an alpha value. If an alpha value is not present in the hexadecimal
        value, the alpha channel will be set to 255. If set on ``None`` (default), the returned value will
        contain an alpha value, only if the input color contains one. If ``False``, the returned value will never
        contain the alpha value.

    Returns
    -------
    tuple(int, int, int) or tuple(int, int, int, int)
        A RGB or RGBA value.

    Examples
    --------
    >>> hex_color_to_rgb("#c0ffee")
    (192, 255, 238)
    >>> hex_color_to_rgb("#f0ccac1a")
    (240, 204, 172, 26)
    """
    if color[0] != "#":
        color = "#" + color

    if not set(color[1:].lower()).issubset(set("0123456789abcdef")):
        raise Exception("Invalid color hexadecimal value: " + str(color))

    if len(color) == 7:
        rgba_color = list(int(color[i:i + 2], 16) for i in (1, 3, 5))
        if include_alpha:
            rgba_color.append(255)
        rgba_color = tuple(rgba_color)
    elif len(color) == 9:
        rgba_color = tuple(int(color[i:i + 2], 16) for i in (1, 3, 5, 7))
        if include_alpha is not None and not include_alpha:
            rgba_color = rgba_color[0:3]
    else:
        raise Exception(f"Invalid color hexadecimal value: {str(color)}")

    return rgba_color


def rgb_color_to_hex(color, include_alpha=None):
    """Converts a color from its RGB or RGBA value to its hexadecimal value.

    Parameters
    ----------
    color: tuple(int, int, int) or tuple(int, int, int, int)
        The RGB or RGBA value of a color.
    include_alpha: bool|None, optional
        If ``True``, returns the hexadecimal color with an alpha value. If an alpha value is not present in the
        RGB value, the alpha channel will be set to ff. If ``False``, returns the hexadecimal color without an alpha
        value. If ``None`` (default), returns the hexadecimal color if the RGB color contains an alpha value.

    Returns
    -------
    str
        A hexadecimal value, with a leading number sign (``"#"``).

    Examples
    --------
    >>> rgb_color_to_hex((192, 255, 238))
    "#c0ffee"
    >>> rgb_color_to_hex((240, 204, 172, 26))
    "#f0ccac1a"
    """
    for i in range(len(color)):
        if color[i] < 0 or color[i] > 255:
            raise Exception(f"Invalid color index: {color[i]}")

    if len(color) == 3:
        if include_alpha:
            return '#%02x%02x%02xff' % color
        else:
            return '#%02x%02x%02x' % color

    elif len(color) == 4:
        if include_alpha is None or include_alpha:
            return '#%02x%02x%02x%02x' % color
        else:
            color = color[0:3]
            return '#%02x%02x%02x' % color

    else:
        raise Exception("Invalid number of arguments in the color.")


def convert_color(color, color_format="RGB", include_alpha=True):
    """Converts a color to the desired format.

    .. versionadded:: 2.0

    Parameters
    ----------
    color: str or tuple(int, int, int) or tuple(int, int, int, int)
        This parameter can take a number of forms:

            • The `HTML/CSS name of a color <https://en.wikipedia.org/wiki/X11_color_names>`_ (e.g. ``"red"`` or
              ``"blanched almond"``),
            • The hexadecimal code of a color, starting with a number sign (``#``, e.g. ``"#ffcc00"`` or ``"#c0ffee"``).
            • The RGB or RGBA tuple of a color (e.g. ``(153, 204, 0)`` or ``(77, 77, 77, 255)``).

    color_format: str, optional
        The format in which you want to convert the color. This parameter can be ``"RGB"`` (default), ``"RGBA"`` or
        ``"HEX"``.

    include_alpha: bool, optional
        If ``True``, returns the color with an alpha value. If an alpha value is not present in the original value,
        the alpha channel will be set to 255 (or ff).

    Returns
    -------
    list(tuple(int, int, int)) or list(tuple(int, int, int, int)) or list(string)
        A list of RGB or RGBA values, or hexadecimal strings with a leading number sign.

    Examples
    --------
    >>> convert_color("red", "RGB")
    (255, 0, 0, 255)
    >>>  convert_color("red", "HEX", False)
    "#ff0000"
    """

    # Color: tuple
    if type(color) is tuple:
        if len(color) == 3:
            converted_color = list(color)
            if include_alpha:
                converted_color.append(255)
            converted_color = tuple(converted_color)
        elif len(color) == 4:
            converted_color = list(color[:3])
            if include_alpha:
                converted_color.append(color[3])
            converted_color = tuple(converted_color)
        else:
            raise Exception("Invalid color tuple: " + str(color) + ". Its length should be 3 or 4 elements.")
        for i in range(len(converted_color)):
            if converted_color[i] < 0 or converted_color[i] > 255:
                raise Exception("Invalid color value: " + str(color) + ". The values should be comprised between 0 " +
                                "and 255.")

    # Color: string
    elif type(color) is str:
        color = color.replace(" ", "")
        if color[0] == "#":
            converted_color = hex_color_to_rgb(color, include_alpha)

        else:
            colors_dict = load_color_names()

            if color not in colors_dict.keys():
                raise Exception("Invalid color name: " + str(color) + ".")

            else:
                converted_color = colors_dict[color]
                if not include_alpha:
                    converted_color = converted_color[:3]

    else:
        raise Exception("Invalid color type: " + str(type(color)) + ". Should be tuple or string.")

    if color_format.lower() in ["rgb", "rgba"]:
        return converted_color
    elif color_format.lower() == "hex":
        return rgb_color_to_hex(converted_color, include_alpha)
    else:
        raise Exception('Invalid parameter color_format: ' + str(color_format) + '. Should be "rgb", "rgba" or "hex".')


def convert_colors(color_scheme_or_colors, color_format="RGB", include_alpha=True):
    """Converts a list of colors to the desired format.

    .. versionadded:: 2.0

    Parameters
    ----------
    color_scheme_or_colors: str or list(str)
        This parameter can take a number of forms:

        • **The name of a color scheme:** a string matching one of the color gradients available in
          :doc:`../appendix/color_schemes`.
        • **A list of colors:** a list containing colors, either using:

            • Their `HTML/CSS names <https://en.wikipedia.org/wiki/X11_color_names>`_ (e.g. ``"red"`` or
              ``"blanched almond"``),
            • Their hexadecimal code, starting with a number sign (``#``, e.g. ``"#ffcc00"`` or ``"#c0ffee"``).
            • Their RGB or RGBA tuples (e.g. ``(153, 204, 0)`` or ``(77, 77, 77, 255)``).

          These different codes can be used concurrently, e.g. ``["red", (14, 18, 32), "#a1b2c3"]``.

    color_format: str, optional
        The format in which you want to convert the colors. This parameter can be ``"RGB"`` (default), ``"RGBA"`` or
        ``"HEX"``.

    include_alpha: bool, optional
        If ``True``, returns the colors with an alpha value. If an alpha value is not present in the
        original value, the alpha channel will be set to 255 (or ff).

    Returns
    -------
    tuple(int, int, int) or tuple(int, int, int, int) or string
        A RGB or RGBA value, or a hexadecimal string with a leading number sign.

    Examples
    --------
    >>> colors = ["#ff0000", (255, 0, 0), "red", (255, 0, 0, 255)]
    >>> convert_colors(colors, "RGB")
    [(255, 0, 0, 255), (255, 0, 0, 255), (255, 0, 0, 255), (255, 0, 0, 255)]
    >>>  convert_colors("default", "HEX", False)
    ['#9acd32', '#ff0000']
    """

    if type(color_scheme_or_colors) is str:

        color_schemes = load_color_schemes()

        if color_scheme_or_colors in color_schemes.keys():
            color_scheme_or_colors = color_schemes[color_scheme_or_colors]

        else:
            colors_dict = load_color_names()

            if color_scheme_or_colors not in colors_dict.keys():
                raise Exception("Invalid color scheme: " + str(color_scheme_or_colors) + ".")
            else:
                warnings.warn("The string entered is not a valid color scheme, but is a valid color name. The value of "
                              "the color will be returned, but convert_color should be used instead of convert_colors.")

            return colors_dict[color_scheme_or_colors]

    converted_colors = []
    for color in color_scheme_or_colors:
        converted_colors.append(convert_color(color, color_format, include_alpha))

    return converted_colors

def calculate_color_points_on_gradient(color_scheme, no_points):
    """Given a color scheme and a number of points, creates a list of equidistant colors to create a gradient.

    Parameters
    ----------
    color_scheme: str|list
        The name of a color scheme (see :doc:`../appendix/color_schemes`. for the available schemes), or a list of
        colors names, RGB or hexadecimal values.
    no_points: int
        The number of colors to get on the gradient.

    Returns
    -------
    list(tuple(int, int, int, int))
        A list of colors in a progressive gradient following the colors of the color scheme.

    Examples
    --------
    >>> color_scheme = ["red", "orange", "yellow"]
    >>> calculate_color_points_on_gradient(color_scheme, 9)
    [(255, 0, 0, 255), (255, 41, 0, 255), (255, 82, 0, 255), (255, 123, 0, 255), (255, 165, 0, 255),
     (255, 187, 0, 255), (255, 210, 0, 255), (255, 232, 0, 255), (255, 255, 0, 255)]
    >>> calculate_color_points_on_gradient("viridis", 7)
    [(66, 1, 86, 255), (60, 58, 122, 255), (48, 106, 140, 255), (32, 146, 140, 255), (74, 183, 110, 255),
     (148, 212, 83, 255), (252, 233, 59, 255)]
    """
    color_scheme = convert_colors(color_scheme, "rgb", True)

    colors = []
    for i in range(no_points):
        colors.append(calculate_color_ratio(color_scheme, i / (no_points - 1)))

    return colors


def calculate_color_ratio(colors, ratio, type_return="RGB", include_alpha=True):
    """Given a gradient of colors, returns a color located at a specific ratio on the gradient.

    .. versionadded:: 2.0

    Parameters
    ----------
    colors: list(tuple(int, int, int)) or list(tuple(int, int, int, int))
        A list of RGB or RGBA color tuples.
    ratio: float
        The ratio of the color, between 0 and 1, on the gradient. If this parameter is equal to or inferior to 0, the
        function will return the first color of the ``colors`` parameter. If the ratio is equal or superior to 1, the
        function will return the last color of the ``colors`` parameter.
    type_return: str, optional
        The way the color should be returned. If set on ``"rgb"`` (default) or ``"rgba"``, the color will be returned
        as a tuple of integers. If set on ``"hex"``, the color will be returned as a string containing hexadecimal
        values, with a number sign (``#``) as a leading character.
    include_alpha: bool, optional
        If ``True``, returns the hexadecimal color with an alpha value. If an alpha value is not present in the
        RGB value, the alpha channel will be set to 255.

    Returns
    -------
    list(tuple(int, int, int)) or list(tuple(int, int, int, int)) or str
        A color in RGB, RGBA or hex form.

    Example
    -------
    >>> calculate_color_ratio((0, 0, 0), (255, 255, 255), 0.5, "RGB", False)
    (127, 127, 127)
    """

    if 0 < ratio < 1:
        index_start = int((ratio / (1 / (len(colors) - 1))))
        index_end = int((ratio / (1 / (len(colors) - 1)))) + 1
    elif ratio >= 1:
        index_start = len(colors) - 2
        index_end = len(colors) - 1
        ratio = 1
    else:
        index_start = 0
        index_end = 1
        ratio = 0

    color_start = colors[index_start]
    color_end = colors[index_end]

    new_ratio = (ratio - index_start / (len(colors) - 1)) * (len(colors) - 1)

    diff_r = color_end[0] - color_start[0]
    diff_g = color_end[1] - color_start[1]
    diff_b = color_end[2] - color_start[2]
    if len(color_start) == 4:
        diff_a = color_end[3] - color_start[3]
    else:
        diff_a = 0

    r = color_start[0] + diff_r * new_ratio
    g = color_start[1] + diff_g * new_ratio
    b = color_start[2] + diff_b * new_ratio

    if len(color_start) == 4 and include_alpha:
        a = color_start[3] + diff_a * new_ratio
        color = (int(r), int(g), int(b), int(a))

    else:
        color = (int(r), int(g), int(b))

    if type_return.lower() in ["rgb", "rgba"]:
        return color
    elif type_return.lower() == "hex":
        return rgb_color_to_hex(color, include_alpha)


def calculate_colors_by_values(dict_values, color_scheme="default", min_value=None, max_value=None,
                               type_return="RGB", include_alpha=True):
    """Given a dictionary of values and a color scheme, returns a dictionary where each key is attributed to a color
    from the scheme. The extreme colors of the
    color scheme will be attributed to the extreme values of the dictionary; all the intermediate values are
    attributed their corresponding intermediate color on the gradient.

    .. versionadded:: 2.0

    Parameters
    ----------
    dict_values: dict(str: float or int)
        A dictionary where keys (e.g., joint labels) are attributed to values (floats or integers).
    color_scheme: str|list
        A color scheme or a list of colors that can be parsed by the function
        :func:`tool_functions.convert_colors_rgba`.
    min_value: int|float, optional
        The value that will be attributed to the first color of the scheme. If not specified (default), the minimum
        value will be set on the minimum value of the dictionary ``dict_values``.
    max_value: int|float, optional
        The value that will be attributed to the last color of the scheme. If not specified (default), the maximum
        value will be set on the maximum value of the dictionary ``dict_values``.
    type_return: str, optional
        The way the color should be returned. If set on ``"rgb"`` (default) or ``"rgba"``, the color will be returned
        as a tuple of integers. If set on ``"hex"``, the color will be returned as a string containing hexadecimal
        values, with a number sign (``#``) as a leading character.
    include_alpha: bool, optional
        If ``True``, returns the hexadecimal color with an alpha value. If an alpha value is not present in the
        RGB value, the alpha channel will be set to 255.

    Returns
    -------
    dict(str: tuple(int, int, int, int))
        A dictionary where each key is matched to an RGBA color.

    Example
    -------
    >>> values = {"Head": 0.2, "Hand": 0.59}
    >>> colors = calculate_colors_by_values(values, "celsius")
    >>> colors["Head"]
    (64, 224, 208, 255)
    >>> colors["Hand"]
    (255, 170, 0, 255)
    """
    keys_colors = {}

    # Get the min and max quantities in the dictionary
    if min_value is None:
        min_value = min(list(dict_values.values()))
    if max_value is None:
        max_value = max(list(dict_values.values()))

    colors = convert_colors(color_scheme, "rgb", True)

    # Apply the colors to all joints
    for j in dict_values.keys():
        ratio = (dict_values[j] - min_value) / (max_value - min_value)
        keys_colors[j] = calculate_color_ratio(colors, ratio, type_return, include_alpha)

    return keys_colors


def generate_random_color(random_alpha=False, color_format="RGB", include_alpha=True):
    """Generates a random color, and returns its RGB or HEX value.

    .. versionadded:: 2.0

    Parameters
    ----------
    random_alpha: bool
        If set on `False` (default), returns a color with an alpha value set on 255. If set on `True`, randomizes the
        alpha value in the color.
    color_format: str
        The format of the color to be returned: either ``"RGBA"`` (default), ``"RGB"``, or ``"HEX"``.
    include_alpha: bool
        If `True` (default), includes the alpha value in the returned color.

    Returns
    -------
    tuple(int, int, int, int)|tuple(int, int, int)|str
        A tuple containing the RGB or RGBA values of a random color, or a string containing its hexadecimal value with
        a leading ``#``.

    Example
    -------
    >>> generate_random_color()
    (255, 128, 64, 32)
    """
    r = random.randrange(0, 256)
    g = random.randrange(0, 256)
    b = random.randrange(0, 256)
    if random_alpha:
        a = random.randrange(0, 256)
    else:
        a = 255

    if color_format.upper() == "RGB":
        if include_alpha:
            return r, g, b, a
        else:
            return r, g, b
    else:
        return convert_color((r, g, b, a), color_format, include_alpha)


# === Audio functions ===
def scale_audio(audio_array, max_value, use_abs_values=False, set_lowest_at_zero=False, verbosity=1):
    """Scale an array of audio samples according to a maximum value. This function is used by the toolbox for display
    purposes.

    .. versionadded::2.0

    Parameters
    ----------
    audio_array: list(int or float) or numpy.array(int or float)
        An array of audio samples.
    max_value: int or float
        The value that will become the maximum value of the scaled array.
    use_abs_values: bool, optional
        If set on ``True``, the values of the samples will be converted to absolute values. If set on ``False``
        (default), the sign of the samples will be preserved.
    set_lowest_at_zero: bool, optional
        If set on ``True``, if the minimum sample value of the audio is over zero, the minimal value will be subtracted
        from all the samples. This parameter is set on ``False`` by default.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    list(float)
        An array of scaled audio samples.
    """
    audio_array = np.array(audio_array)

    if verbosity > 0:
        print("Scaling audio...")

    if use_abs_values:
        audio_array = np.abs(audio_array)

    if set_lowest_at_zero:
        audio_array -= np.min(audio_array)

    audio_array = audio_array / np.max(np.abs(audio_array)) * max_value

    if verbosity > 0:
        print("100% - Done.")

    return audio_array

def stereo_to_mono(audio_data, mono_channel=0, verbosity=1):
    """Converts an audio array to mono.

    .. versionadded:: 2.0

    Parameters
    ----------
    audio_data: `np.ndarray <https://numpy.org/doc/stable/reference/generated/numpy.ndarray.html>`_ (1D or 2D)
        The parameter data resulting from reading a WAV file with
        `scipy.io.wavfile.read <https://docs.scipy.org/doc/scipy/reference/generated/scipy.io.wavfile.read.html>`_.

    mono_channel: int|str, optional
        Defines the method to use to convert multiple-channel WAV files to mono, By default, this parameter value is
        ``0``: the channel with index 0 in the WAV file is used as the array, while all the other channels are
        discarded. This value can be any of the channels indices (using ``1`` will preserve the channel with index
        1, etc.). This parameter can also take the value ``"average"``: in that case, a new channel is created by
        averaging the values of all the channels of the WAV file.

    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    np.array
        A 1D numpy array containing the audio converted to mono.
    """

    if verbosity > 0:
        print("\n\tConverting audio samples from stereo to mono...", end=" ")

    # If mono_channel is a channel number
    if isinstance(mono_channel, int):

        # If the channel number is negative or nonexistent
        if mono_channel >= np.size(audio_data[1]) or mono_channel < 0:
            if np.size(audio_data[1] == 0):
                raise Exception(f"""The channel chosen for the parameter "mono_channel" ({mono_channel}) is not valid.
                As the audio data is mono, the channel chosen should be 0.""")
            else:
                raise Exception(f"""The channel chosen for the parameter "mono_channel" ({mono_channel}) is not valid.
                Please choose a channel between 0 and {np.size(audio_data[1]) - 1}.""")

        # If the audio data is already mono
        elif np.size(audio_data[1]) == 1:
            mono_array = audio_data
            if verbosity > 0:
                print(f"\tThe audio data is already in mono, no conversion necessary.")

        # If the audio data is not mono
        else:
            mono_array = audio_data[:, mono_channel]  # Turn to mono
            if verbosity > 0:
                print(f"\tFile converted to mono by keeping channel with index {mono_channel}. The original file"
                      f" contains {np.size(audio_data[1])} channels.")

    # If mono_channel is "average"
    elif mono_channel == "average":
        mono_array = np.mean(audio_data, 1)

    # Any other case
    else:
        raise Exception(f"""The parameter "mono_channel" should be an integer (the channel index) or "average", not
                        {mono_channel}.""")

    return mono_array

def remove_average(array):
    """Removes the average of an array of values.

    .. versionadded:: 2.0

    Parameters
    ----------
    array: list|`np.ndarray <https://numpy.org/doc/stable/reference/generated/numpy.ndarray.html>`_ (1D or 2D)
        An array of values, for example from an audio file.
    """

    if type(array) is list:
        array = np.array(array)

    return array - np.mean(array)

# === Joint labels functions ===
def load_joint_labels(system="kinect", part="all", label_style="original"):
    """Loads and returns a list of joint labels from a file.

    .. versionadded:: 2.0

    Parameters
    ---------
    system: str, optional
        The system from which to load the joints. This parameter can be set on ``"kinect"`` (default) or
        ``"qualisys"``/``"kualisys"`` to load a preset list of joints. It is also possible to pass a path to a file
        containing joints on different lines.
    part: str, optional
        The part of the joints to load. This parameter can be set on ``"all"`` (default), ``"top"`` or ``"bottom"``.
        If the parameter ``system`` is set on a path, this parameter is ignored.
    label_style: str, optional
        When used for loading Qualisys joints:
            · If set on ``"original"`` (default), the returned joint labels are the original ones from the Qualisys
              system.
            · If set on ``"new"``, ``"replaced"``, ``"krajjat"`` or ``"kualisys"``, the returned joint labels are the
              renamed Kualisys joint labels from the Krajjat toolbox. For more information, see
              :doc:`../appendix/joint_labels`.

        When used for paths, the function checks if each line contains two strings separated by a tab, and returns
        the list of the first strings if set on ``"original"``, and the list of the second strings if set on ``"new"``
        or ``"replaced"``.

        This parameter is ignored if ``system`` is set on ``"krajjat"``.

    Returns
    -------
    list(str)
        The list of the Kinect joint labels.
    """
    if system.lower() == "kinect":
        with open(op.join(MODULE_DIR, "res/kinect_joint_labels.txt"), "r", encoding="utf-8") as f:
            joint_labels = f.read().split("\n")

        if part.lower() == "top":
            joint_labels = joint_labels[:13]
        elif part.lower() == "bottom":
            joint_labels = joint_labels[13:]
        elif part.lower() != "all":
            raise InvalidParameterValueException("part", part, ["top", "bottom", "all"])

        return joint_labels

    elif system.lower() in ["qualisys", "kualisys"]:
        with open(op.join(MODULE_DIR, "res/kualisys_joint_labels.txt"), "r", encoding="utf-8") as f:
            joint_labels = f.read().split("\n")

        if part.lower() == "top":
            joint_labels = joint_labels[:28]
        elif part.lower() == "bottom":
            joint_labels = joint_labels[28:]
        elif part.lower() != "all":
            raise InvalidParameterValueException("part", part, ["top", "bottom", "all"])

        joint_labels_original = []
        joint_labels_krajjat = []

        for line in joint_labels:
            elements = line.split("\t")
            joint_labels_original.append(elements[0])
            joint_labels_krajjat.append(elements[1])

        if label_style == "original":
            return joint_labels_original
        elif label_style in ["new", "replaced", "krajjat", "kualisys"]:
            return joint_labels_krajjat

    else:
        if os.path.exists(system):
            # noinspection PyArgumentList
            with open(system, "r", encoding="utf-8") as f:
                joint_labels = f.read().split("\n")
        else:
            raise InvalidPathException("system", "text file", "the file does not exist.")

        if "\t" in joint_labels[0]:
            joint_labels_original = []
            joint_labels_new = []

            for line in joint_labels:
                elements = line.split("\t")
                joint_labels_original.append(elements[0])
                joint_labels_new.append(elements[1])

            if label_style == "original":
                return joint_labels_original
            elif label_style in ["new", "replaced", "krajjat", "kualisys"]:
                return joint_labels_new

        else:
            return joint_labels

def load_qualisys_joint_label_conversion():
    """Returns a dictionary containing the original Qualisys joint labels as keys, and the renamed Kualisys joints
    labels as values. The dictionary is loaded from ``res/kualisys_joint_labels.txt``. For more information, see
    :doc:`../appendix/joint_labels`.

    .. versionadded:: 2.0

    Returns
    -------
    dict(str: str)
        A dictionary with the original Qualisys joint labels as keys, and the Kualisys renamed joint labels as values.
    """

    with open(op.join(MODULE_DIR, "res/kualisys_joint_labels.txt"), "r", encoding="utf-8") as f:
        content = f.read().split("\n")

    joints_conversions = {}

    for line in content:
        elements = line.split("\t")
        joints_conversions[elements[0]] = elements[1]

    return joints_conversions


def load_joint_connections(system="kinect", part="all"):
    """Returns a list of joint pairs between which a line can be traced to form a skeleton.

    .. versionadded:: 2.0

    Parameters
    ----------
    system: str, optional
        The system from which to load the joints. This parameter can be set on ``"kinect"`` (default) or
        ``"qualisys"``/``"kualisys"`` to load a preset list of joints. It is also possible to pass a path to a file
        containing joints on different lines.
    part: str, optional
        The part of the joints to load. This parameter can be set on ``"all"`` (default), ``"top"`` or ``"bottom"``.
        If the parameter ``system`` is set on a path, this parameter is ignored.

    Returns
    -------
    list(list(str))
        A list of sub-lists, each containing two elements (two joint labels).
    """
    if system.lower() == "kinect":
        with open(op.join(MODULE_DIR, "res/kinect_skeleton_connections.txt"), "r", encoding="utf-8") as f:
            joint_connections = f.read().split("\n")

        if part.lower() == "top":
            joint_connections = joint_connections[:12]
        elif part.lower() == "bottom":
            joint_connections = joint_connections[12:]
        elif part.lower() != "all":
            raise InvalidParameterValueException("part", part, ["top", "bottom", "all"])

    elif system.lower() in ["qualisys", "kualisys"]:
        with open(op.join(MODULE_DIR, "res/kualisys_skeleton_connections.txt"), "r", encoding="utf-8") as f:
            joint_connections = f.read().split("\n")

        if part.lower() == "top":
            joint_connections = joint_connections[:43]
        elif part.lower() == "bottom":
            joint_connections = joint_connections[43:]
        elif part.lower() != "all":
            raise InvalidParameterValueException("part", part, ["top", "bottom", "all"])

    else:
        if os.path.exists(system):
            # noinspection PyArgumentList
            with open(system, "r", encoding="utf-8") as f:
                joint_connections = f.read().split("\n")
        else:
            raise InvalidPathException("system", "text file", "the file does not exist.")

    joint_connections = [tuple(joints.split("\t")) for joints in joint_connections]
    return joint_connections


def load_qualisys_to_kinect():
    """Loads a dictionary containing Kinect joint labels as keys, and a series of Kualisys joint labels as values.
    The Kinect joints can be averaged from the joints in values. The dictionary is loaded from
    ``res/kualisys_to_kinect.txt``. For more information, see :doc:`../appendix/joint_labels`.

    .. versionadded:: 2.0

    Returns
    -------
    dict(str: list(str))
        A dictionary of Kinect joint labels as keys, and a series of Kualisys joint labels as values.
    """
    with open(op.join(MODULE_DIR, "res", "kualisys_to_kinect.txt"), "r", encoding="utf-8") as f:
        content = f.read().split("\n")
    connections = {}

    for line in content:
        elements = line.split("\t")
        connections[elements[0]] = tuple(elements[1:])

    return connections


def load_joints_subplot_layout(joint_layout):
    """Returns a dictionary of the subplot positions of the joints on a skeleton graph. Loads the data from
    ``"res/kinect_joints_subplot_layout.txt"`` or ``"res/kualisys_joints_subplot_layout.txt"``.

    .. versionadded:: 2.0

    Parameters
    ----------
    joint_layout: str
        The layout to load, either ``"kinect"``, ``"kualisys"``/``"qualisys"``, or the path to a text file containing
        a custom layout.

    Returns
    -------
    dict(str: int)
        A dictionary containing joint labels as keys, and subplot positions as values.
    """

    if joint_layout.lower() == "kinect":
        with open(op.join(MODULE_DIR, "res", "kinect_joints_subplot_layout.txt")) as f:
            content = f.read().split("\n")
    elif joint_layout.lower() in ["qualisys", "kualisys"]:
        with open(op.join(MODULE_DIR, "res", "kualisys_joints_subplot_layout.txt")) as f:
            content = f.read().split("\n")
    else:
        try:
            with open(op.join(joint_layout)) as f:
                content = f.read().split("\n")
        except FileNotFoundError:
            raise Exception("The file " + str(joint_layout) + " is not a valid file.")

    joints_positions = {}

    for line in content:
        elements = line.split("\t")
        joints_positions[elements[0]] = int(elements[1])

    return joints_positions


def load_joints_silhouette_layout(joint_layout):
    """Returns a dictionary of the positions and radii of the joints, and a list of the order in which to plot the
    joints for the function :func:`plot_functions.plot_silhouette` Loads the data from
    ``"res/kinect_joints_silhouette_layout.txt"``, ``"res/kualisys_joints_silhouette_layout.txt"``, or a custom file.

    .. versionadded:: 2.0

    Parameters
    ----------
    joint_layout: str
        The layout to load, either ``"kinect"`` or ``"qualisys"``/``"kualisys"``. This parameter can also be the path
        to a custom layout (in .txt). In that case, each line must contain a joint label, the horizontal position,
        the vertical position and the radius of the circle to plot, all separated by a tabulation. The positions and
        radius are expected to be specified in pixels. The output list of the function, which defines the order in
        which to plot the joints, is defined by the order of the joints in the file.

    Returns
    -------
    dict(str: tuple)
        A dictionary containing joint labels as keys, and tuples containing three elements as values: the horizontal
        position, the vertical position, and the radius of the circle.

    list(str)
        The list of the joints, in the order they should be plotted.
    """

    if joint_layout.lower() == "kinect":
        with open(op.join(MODULE_DIR, "res", "kinect_joints_silhouette_layout.txt")) as f:
            content = f.read().split("\n")
    elif joint_layout.lower() in ["qualisys", "kualisys"]:
        with open(op.join(MODULE_DIR, "res", "kualisys_joints_silhouette_layout.txt")) as f:
            content = f.read().split("\n")
    else:
        try:
            with open(joint_layout) as f:
                content = f.read().split("\n")
        except FileNotFoundError:
            raise Exception("The file " + str(joint_layout) + " is not a valid file.")

    joints_positions = {}
    joint_layout = []

    for line in content:
        elements = line.split("\t")
        joints_positions[elements[0]] = (int(elements[1]), int(elements[2]), int(elements[3]))
        joint_layout.append(elements[0])

    return joints_positions, joint_layout


def load_default_steps_gui():
    """Loads the steps for the modification of the parameters in the GUI of the graphic functions, from the file
    ``res/default_steps_gui.txt``.

    .. versionadded:: 2.0
    """
    with open(op.join(MODULE_DIR, "res/default_steps_gui.txt"), "r", encoding="utf-8") as f:
        lines = f.read().split("\n")

    steps = {}
    for line in lines:
        elements = line.split("\t")
        steps[elements[0]] = float(elements[1])

    return steps


# === Time functions ===
def convert_timestamp_to_seconds(timestamp, time_unit="s"):
    """Converts a timestamp in the specified time unit to a timestamp in seconds.

    .. versionadded:: 2.0

    Parameters
    ----------
    timestamp: int|float
        A timestamp, in the unit specified by `time_unit`.
    time_unit: str, optional
        The time unit of the timestamp. This parameter can take the following values: "ns", "1ns", "10ns", "100ns",
        "µs", "1µs", "10µs", "100µs", "ms", "1ms", "10ms", "100ms", "s", "sec", "1s", "min", "mn", "h", "hr", "d",
        "day".

    Returns
    -------
    float
        The converted timestamp, in seconds.

    Examples
    --------
    >>> convert_timestamp_to_seconds(5, "ms")
    0.005
    >>> convert_timestamp_to_seconds(1, "h")
    3600
    """
    return timestamp / UNITS[time_unit.lower()]


def format_time(time, time_unit="s", time_format="hh:mm:ss"):
    """Formats a given time in a given unit according to a time format.

    .. versionadded:: 2.0

    Parameters
    ----------
    time: int or float
        A value of time.
    time_unit: str, optional
        The unit of the ``time`` parameter. This parameter can take the following values: "ns", "1ns", "10ns", "100ns",
        "µs", "1µs", "10µs", "100µs", "ms", "1ms", "10ms", "100ms", "s", "sec", "1s", "min", "mn", "h", "hr", "d",
        "day".
    time_format: str, optional
        The format in which to return the time. Can be either "hh:mm:ss", "hh:mm:ss.ms", "hh:mm", "mm:ss" or "mm:ss.ms".

    Examples
    --------
    >>> format_time(1000, "s", "hh:mm:ss")
    "00:16:40"
    >>> format_time(1000, "s", "hh:mm:ss.ms")
    "00:16:40.000"
    >>> format_time(1000, "s", "mm:ss")
    "16:40"
    >>> format_time(1000, "s", "mm:ss.ms")
    "16:40.000"
    """
    time = convert_timestamp_to_seconds(time, time_unit)
    hour = int(time // 3600)
    minute = int((time // 60) % 60)
    second = int((time % 60) // 1)
    millisecond = int((time % 1) * 1000)

    if time_format.startswith("hh:mm:ss"):
        if time_format.endswith(".ms"):
            return str(hour).zfill(2) + ":" + str(minute).zfill(2) + ":" + str(second).zfill(2) + "." + \
                   str(millisecond).zfill(3)
        else:
            return str(hour).zfill(2) + ":" + str(minute).zfill(2) + ":" + str(second).zfill(2)

    elif time_format == "hh:mm":
        return str(hour).zfill(2) + ":" + str(minute).zfill(2)

    elif time_format.startswith("mm:ss"):
        minute = int(time // 60)
        if time_format.endswith(".ms"):
            return str(minute).zfill(2) + ":" + str(second).zfill(2) + "." + str(millisecond).zfill(3)
        else:
            return str(minute).zfill(2) + ":" + str(second).zfill(2)

    else:
        raise Exception('Wrong value for the time_format parameter. It should be "hh:mm:ss" or "mm:ss".')


def time_unit_to_datetime(time, time_unit="s"):
    """Turns any time unit into a ``datetime`` format.

    .. versionadded:: 2.0

    Parameters
    ----------
    time: int or float
        A value of time.
    time_unit: str, optional
        The unit of the ``time`` parameter. This parameter can take the following values: "ns", "1ns", "10ns", "100ns",
        "µs", "1µs", "10µs", "100µs", "ms", "1ms", "10ms", "100ms", "s", "sec", "1s", "min", "mn", "h", "hr", "d",
        "day".

    Example
    -------
    >>> time_unit_to_datetime(1000000, "ms")
    datetime.datetime(1, 1, 1, 0, 16, 40)
    """

    time = convert_timestamp_to_seconds(time, time_unit)
    day = int(time // 86400)
    hour = int(time // 3600) % 24
    minute = int((time // 60) % 60)
    second = int((time % 60) // 1)
    microsecond = int((time % 1) * 1000000)

    return datetime.datetime(1, 1, 1+day, hour, minute, second, microsecond)


def time_unit_to_timedelta(time, time_unit="s"):
    """Turns any time unit into a ``timedelta`` format.

    .. versionadded:: 2.0

    Parameters
    ----------
    time: int or float
        A value of time.
    time_unit: str, optional
        The unit of the ``time`` parameter. This parameter can take the following values: "ns", "1ns", "10ns", "100ns",
        "µs", "1µs", "10µs", "100µs", "ms", "1ms", "10ms", "100ms", "s", "sec", "1s", "min", "mn", "h", "hr", "d",
        "day".

    Example
    -------
    >>> time_unit_to_timedelta(1000000, "ms")
    datetime.timedelta(0, 1000, 0)
    """

    time = convert_timestamp_to_seconds(time, time_unit)
    day = int(time // 86400)
    second = int((time % 86400) // 1)
    microsecond = int((time % 1) * 1000000)

    return datetime.timedelta(day, second, microsecond)


# === Miscellaneous functions ===
def show_progression(verbosity, current_iteration, goal, next_percentage, step=10):
    """Shows a percentage of progression if verbosity is equal to 1.

    .. versionadded:: 2.0

    Parameters
    ----------
    verbosity: int
        This parameter must be equal to 1 for the function to print information. If the value is set on 2 (chatty mode),
        the percentage will not be displayed, in favour of more detailed information related to the ongoing operations.
    current_iteration: int or float
        The current iteration in a loop.
    goal: int or float
        The stopping value of the loop.
    next_percentage: int or float
        The next percentage at which to print the progression.
    step: int or float, optional
        The step between each print of the progression.

    Returns
    -------
    int or float
        The next percentage at which to print the progression.

    Example
    -------
    >>> next_percentage = 10
    >>> goal = 100
    >>> for i in range(goal):
    ...       next_percentage = show_progression(1, i, goal, next_percentage)
    10% 20̀% 30% 40% 50% 60% 70% 80M 90%
    """

    if verbosity == 1:
        while current_iteration / goal > next_percentage / 100:
            print(str(next_percentage) + "%", end=" ")
            next_percentage += step

    return next_percentage


def get_min_max_values_from_plot_dictionary(plot_dictionary, keys_to_exclude=None, xlim=None):
    """Returns the minimum and maximum values of all the graphs contained in a plot dictionary.

    .. versionadded:: 2.0

    Parameters
    ----------
    plot_dictionary: dict
        A dictionary with labels as keys and Graph elements, float, int or arrays as values.
    keys_to_exclude: list(string)
        A list of keys of the dictionary to exclude from the search of the minimum and maximum values.
    xlim: list(int|float, int|float), optional
        The limits of the horizontal axis of the graphs.

    Returns
    -------
    float
        The minimum value detected in all the series.
    float
        The maximum value detected in all the series.
    """
    min_value = None
    max_value = None

    if keys_to_exclude is None:
        keys_to_exclude = []

    for key in plot_dictionary.keys():
        if key not in keys_to_exclude:
            if isinstance(plot_dictionary[key], (collections.abc.Sequence, np.ndarray)):
                if len(plot_dictionary[key]) > 0:
                    local_min = np.min(np.array(plot_dictionary[key]).flatten())
                    local_max = np.max(np.array(plot_dictionary[key]).flatten())
                    if min_value is None or local_min < min_value:
                        min_value = local_min
                    if max_value is None or local_max > max_value:
                        max_value = local_max
            elif isinstance(plot_dictionary[key], Graph):
                for plot in plot_dictionary[key].plots:
                    if xlim is None:
                        local_min, local_max = plot.get_extrema()
                    else:
                        local_min, local_max = plot.get_extrema(xlim[0], xlim[1])
                    if min_value is None or local_min < min_value:
                        min_value = local_min
                    if max_value is None or local_max > max_value:
                        max_value = local_max
            elif isinstance(plot_dictionary[key], numbers.Number):
                if min_value is None or plot_dictionary[key] < min_value:
                    min_value = plot_dictionary[key]
                if max_value is None or plot_dictionary[key] > max_value:
                    max_value = plot_dictionary[key]
            else:
                raise Exception('The plot dictionary values must be Graph, float or int.')

    return min_value, max_value


def kwargs_parser(dictionary, suffix):
    """Given a dictionary of keyword arguments and a suffix, returns a dictionary that removes the suffix from the given
    keywords. If a key without the suffix already exists in the dictionary, it is left untouched and the key with the
    suffix is removed.

    .. versionadded:: 2.0

    Parameters
    ----------
    dictionary: dict
        A dictionary of keyword arguments.
    suffix: str
        The suffix to look for in the keys of the dictionary.
    """
    new_dictionary = {}

    for key in dictionary.keys():
        if key.endswith(suffix):
            if len(key) >= len(suffix):
                key_without_suffix = key[:len(key)-len(suffix)]
                if key_without_suffix not in dictionary:
                    new_dictionary[key_without_suffix] = dictionary[key]
        else:
            new_dictionary[key] = dictionary[key]

    return new_dictionary
