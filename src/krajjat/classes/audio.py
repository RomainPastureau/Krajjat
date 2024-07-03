"""Default class for audio recordings matching a Sequence, typically the voice of the subject of the motion capture.
This class allows to perform a variety of transformations of the audio stream, such as getting the envelope, pitch and
formants of the speech.
"""
from collections import OrderedDict
import datetime as dt

from scipy import signal

from krajjat.plot_functions import _plot_figure_find_excerpt
from krajjat.tool_functions import *
from krajjat.classes.exceptions import *
from krajjat.classes.audio_derivatives import *
from scipy.signal import butter, lfilter


class Audio(object):
    """Default class for audio clips matching a Sequence, typically the voice of the subject of the motion
    capture. This class allows to perform a variety of transformations of the audio stream, such as getting the
    envelope, pitch and formants of the speech.

    .. versionadded:: 2.0

    Parameters
    ----------
    path_or_samples: str or list(int) or numpy.ndarray(int)
        The path to the audio file, or a list containing the samples of an audio file. If the file is a path, it should
        either point to a `.wav` file, or to a file containing the timestamps and samples in a text form (`.json`,
        `.csv`, `.tsv`, `.txt` or `.mat`). It is also possible to point to a folder containing one file per sample.
        See :ref:`Audio formats <wav_example>` for the acceptable file types.

    frequency: int or float, optional
        The frequency, in Hz (or samples per sec) at which the parameter `path_or_samples` is set. This parameter
        will be ignored if ``path_or_samples`` is a path, but will be used to define the :attr:`timestamps` of the
        Audio object if ``path_or_samples`` is a list of samples.

    name: str, optional
        Defines a name for the Audio instance. If a string is provided, the attribute :attr:`name` will take
        its value. If not, see :meth:`Audio._define_name_init()`.

    condition: str or None, optional
        Optional field to represent in which experimental condition the audio was recorded.

    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Attributes
    ----------
    samples: np.ndarray(int)
        A list containing the audio samples, in chronological order.
    timestamps: list(float)
        A list containing the timestamps matching each audio sample. Consequently, :attr:`samples` and
        :attr:`timestamps` should have the same length.
    frequency: int or float
        The amount of samples per second.
    name: str
        Custom name given to the audio. If no name has been provided upon initialisation, it will be defined by
        :meth:`Audio._define_name_init()`.
    condition: str
        Defines in which experimental condition the audio clip was recorded.
    metadata: dict
        A dictionary containing metadata about the recording, extracted from the file.
    path: str
        Path to the audio file passed as a parameter upon creation; if samples were provided, this attribute will be
        `None`.
    files: list(str)
        List of files contained in the path. The list will be of size 1 if the path points to a single file.
    kind: str
        A parameter that is set on ``"Audio"``, to differentiate it from the different types of AudioDerivative.
    """

    def __init__(self, path_or_samples, frequency=None, name=None, condition=None, verbosity=1):
        self.samples = np.ndarray([])
        self.timestamps = []
        self.frequency = None
        self.name = None  # Placeholder for the name of the sequence
        self.condition = condition
        self.metadata = {}
        self.path = None
        self.files = None
        self.kind = "Audio"

        if type(path_or_samples) is str:
            self._load_from_path(path_or_samples, verbosity)
        elif isinstance(path_or_samples, (list, np.ndarray)):
            self._load_from_samples(path_or_samples, frequency, verbosity)
        else:
            raise Exception("Invalid type for the argument path_or_samples: should be str or list.")

        self._define_name_init(name, verbosity)

        # self.max_sample = max(self.samples)
        # self.min_sample = min(self.samples)
        #
        # self.envelope = None
        # self.max_sample_envelope = None
        # self.min_sample_envelope = None
        # self.get_envelope()

    # === Setter functions ===
    def set_name(self, name):
        """Sets the :attr:`name` attribute of the Audio instance. This name can be used as display functions or as
        a means to identify the audio.

        .. versionadded:: 2.0

        Parameters
        ----------
        name : str
            A name to describe the audio clip.

        Example
        -------
        >>> aud = Audio("C:/Users/Walter/Sequences/audio.wav")
        >>> aud.set_name("Audio 28980")
        """

        self.name = name

    def set_condition(self, condition):
        """Sets the :py:attr:`condition` attribute of the Audio instance. This attribute can be used to save the
        experimental condition in which the Audio instance was recorded.

        .. versionadded:: 2.0

        Parameters
        ----------
        condition: str
            The experimental condition in which the audio clip was recorded.

        Example
        -------
        >>> aud1 = Audio("C:/Users/Dwight/Sequences/English/seq.wav")
        >>> aud1.set_condition("English")
        >>> aud2 = Audio("C:/Users/Dwight/Sequences/Spanish/seq.wav")
        >>> aud2.set_condition("Spanish")
        """
        self.condition = condition

    # === Loading functions ===
    def _define_name_init(self, name, verbosity=1):
        """Sets the name attribute for an instance of the Audio class, using the name provided during the
        initialization, or the path. If no ``name`` is provided, the function will create the name based on the
        :attr:`path` attribute, by defining the name as the last element of the path hierarchy (last subfolder, or file
        name). For example, if ``path`` is ``"C:/Users/Bender/Documents/Recording001/"``, the function will define the
        name on ``"Recording001"``. If both ``name`` and ``path`` are set on ``None``, the sequence name will be defined
        as ``"Unnamed audio"``.

        .. versionadded:: 2.0

        Parameters
        ----------
        name: str
            The name passed as parameter in :meth:`Audio.__init__()`
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        if name is not None:
            self.name = name
            if verbosity > 1:
                print("The provided name " + str(name) + " will be attributed to the audio.")

        elif self.path is not None:
            if len(self.path.split("/")) >= 1:
                self.name = self.path.split("/")[-1]
            else:
                self.name = str(self.path)
            if verbosity > 1:
                print(
                    "No name was provided. Instead, the name " + str(self.name) + " was attributed by extracting it " +
                    "from the provided path.")

        else:
            self.name = "Unnamed audio"
            if verbosity > 1:
                print("No name nor path was provided. The placeholder name " + str(
                    self.name) + " was attributed to the " +
                      "audio.")

    def _load_from_path(self, path, verbosity=1):
        """Loads the audio data from the :attr:`path` provided during the initialization.

        .. versionadded:: 2.0

        Parameters
        ----------
        path: str
            Path to the audio file passed as a parameter upon creation.
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        if verbosity > 1:
            print("Parameter path_or_samples detected as being a path.")

        self.path = path

        # If it's a folder, we fetch all the files
        if not os.path.exists(self.path):
            if "." in self.path.split("/")[-1]:
                raise InvalidPathException(self.path, "audio clip", "The file doesn't exist.")
            else:
                raise InvalidPathException(self.path, "audio clip", "The folder doesn't exist.")

        # If it's a folder, we fetch all the files
        if os.path.isdir(self.path):
            self._fetch_files_from_folder(verbosity)  # Fetches all the files
        self._load_samples(verbosity)  # Loads the files into poses

        if len(self.samples) == 0:
            raise EmptyAudioException()

    def _load_from_samples(self, samples, frequency, verbosity=1):
        """Loads the audio data when samples and frequency have been provided upon initialisation.

        .. versionadded:: 2.0

        Parameters
        ----------
        samples: list(int) or numpy.ndarray(int)
            A list containing the audio samples, in chronological order.
        frequency: int or float
            The amount of samples per second.
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        if verbosity > 1:
            print("Parameter path_or_samples detected as being samples.")

        if not isinstance(samples, np.ndarray):
            self.samples = np.array(samples)
        else:
            self.samples = samples
        if frequency is not None:
            self.frequency = frequency
        else:
            raise Exception("If samples are provided, frequency cannot be None. Please provide a value for " +
                            "frequency.")

        self._calculate_timestamps()

    def _fetch_files_from_folder(self, verbosity=1):
        """Finds all the files ending with the accepted extensions (``.csv``, ``.json``, ``.tsv``, ``.txt``, or
        ``.xlsx``) in the folder defined by path, and orders the files according to their name.

        .. versionadded:: 2.0

        Note
        ----
        This functions ignores the elements of the directory defined by :attr:`path` if:
            •	They don’t have an extension
            •	They are a folder
            •	Their extension is not one of the accepted ones (``.csv``, ``.json``, ``.tsv``, ``.txt``, or
                ``.xlsx``)
            •	The file name does not contain an underscore (``_``)

        If a file has a valid extension, the function tries to detect an underscore (``_``) in the name. The file
        names should be ``xxxxxx_0.ext``, where ``xxxxxx`` can be any series of characters, ``0`` must be the index of
        the sample (with or without leading zeros), and ``ext`` must be an accepted extension (``.csv``, ``.json``,
        ``.tsv``, ``.txt``, or ``.xlsx``). The first pose of the sequence must have the index 0. If the file does not
        have an underscore in the name, it is ignored. The indices must be coherent with the chronological order of the
        timestamps.

        The function uses the number after the underscore to order the samples. This is due to differences in how file
        systems handle numbers without leading zeros: some place ``sample_11.json`` alphabetically before
        ``sample_2.json`` (1 comes before 2), while some other systems place it after as 11 is greater than 2. In order
        to avoid these, the function converts the number after the underscore into an integer to place it properly
        according to its index.

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """
        if verbosity > 0:
            print("Fetching sample files...", end=" ")

        file_list = os.listdir(self.path)  # List all the files in the folder
        self.files = ["" for _ in range(len(file_list))]  # Create an empty list the length of the files

        extensions_found = []  # Save the valid extensions found in the folder

        # Here, we find all the files that are either .json or .meta in the folder.
        for f in file_list:

            # If a file has an accepted extension, we get its index from its name to order it correctly in the list.
            # This is necessary as some file systems will order sample_2.json after sample_11.json because,
            # alphabetically, "1" is before "2".

            # We check if the element has a dot, i.e. if it is a file. If not, it is a folder, and we ignore it.
            if "." in f:

                extension = f.split(".")[-1]  # We get the file extension
                if extension in ["json", "csv", "tsv", "txt", "xlsx"]:

                    if verbosity > 1:
                        print("\tAdding file: " + str(f))

                    if "_" in f:
                        self.files[int(f.split(".")[0].split("_")[1])] = f
                        if extension not in extensions_found:
                            extensions_found.append(extension)
                    else:
                        if verbosity > 1:
                            print("""\tIgnoring file (no "_" detected in the name): """ + str(f))

                    if len(extensions_found) > 1:
                        raise InvalidPathException(self.path, "audio clip",
                                                   "More than one of the accepted extensions has been found in the " +
                                                   "provided folder: " + str(extensions_found[0]) + " and " +
                                                   str(extensions_found[1]))

                else:
                    if verbosity > 1:
                        print("\tIgnoring file (not an accepted filetype): " + str(f))

            else:
                if verbosity > 1:
                    print("\tIgnoring folder: " + str(f))

        # If files that weren't of the accepted extension are in the folder, then "self.files" is larger than
        # the number of samples. The list is thus ending by a series of empty strings that we trim.
        if "" in self.files:
            limit = self.files.index("")
            self.files = self.files[0:limit]

        for i in range(len(self.files)):
            if self.files[i] == "":
                raise InvalidPathException(self.path, "audio clip", "At least one of the files is missing (index " +
                                           str(i) + ").")

        print(str(len(self.files)) + " sample file(s) found.")

    def _load_samples(self, verbosity=1):
        """Loads the single sample files or the global file containing all the samples. Depending on the input, this
        function calls either :meth:`Audio._load_single_sample_file` or :meth:`Audio._load_audio_file`.

        .. versionadded:: 2.0

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        if verbosity > 0:
            print("Opening audio from " + self.path + "...", end=" ")

        perc = 10  # Used for the progression percentage

        # If the path is a folder, we load every single file
        if os.path.isdir(self.path):

            self.samples = np.zeros((len(self.files),))

            for i in range(len(self.files)):

                if verbosity > 1:
                    print("Loading file " + str(i) + " of " + str(len(self.files)) + ":" + self.path + "...", end=" ")

                # Show percentage if verbosity
                perc = show_progression(verbosity, i, len(self.files), perc)

                # Loads a file containing one timestamp and one sample
                self._load_single_sample_file(self.path + "/" + self.files[i], i)

                if verbosity > 1:
                    print("OK.")

            self._calculate_frequency()

            if verbosity > 0:
                print("100% - Done.")

        # Otherwise, we load the one file
        else:
            self._load_audio_file(verbosity)

    def _load_single_sample_file(self, path, i):
        """Loads the content of a single sample file into the Audio object. Depending on the file type, this function
        handles the content differently (see :ref:`Audio formats <wav_example>`).

        .. versionadded:: 2.0

        Parameters
        ----------
        path: str
            The path of a file containing a single sample and timestamp.
        i: int
            The index of the sample to load.
        """

        # JSON file
        if path.split(".")[-1] == "json":
            data = read_json(path)
            self.samples[i] = data["Sample"]
            self.timestamps.append(data["Timestamp"])

        # Excel file
        elif path.split(".")[-1] == "xlsx":
            import openpyxl as op
            workbook = op.load_workbook(path)
            sheet = workbook[workbook.sheetnames[0]]

            self.timestamps.append(float(sheet.cell(2, 1).value))
            self.samples[i] = float(sheet.cell(2, 2).value)

        # Text file
        elif path.split(".")[-1] in ["txt", "csv", "tsv"]:

            separator = get_filetype_separator(path.split(".")[-1])

            # Open the file and read the data
            data, self.metadata = read_text_table(path)

            for s in range(1, len(data)):
                d = data[s].split(separator)
                self.timestamps.append(float(d[0][s]))
                self.samples[i] = int(d[1][s])

        else:
            raise InvalidPathException(self.path, "audio clip", "Invalid file extension: " + path.split(".")[-1] + ".")

    def _load_audio_file(self, verbosity=1):
        """Loads the content of a file containing all the samples of the audio stream.
        Depending on the file type, this function handles the content differently (see
        :ref:`Audio formats <wav_example>`).

        .. versionadded:: 2.0

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        self.name = self.path

        if self.path.split(".")[-1] == "wav":
            self._read_wav(verbosity)
        elif self.path.split(".")[-1] in ["xlsx", "json", "csv", "txt"]:
            self._read_text_file(verbosity)
        else:
            raise InvalidPathException(self.path, "audio clip", "Invalid file extension: " + self.path.split(".")[-1] +
                                       ".")

    def _read_wav(self, verbosity=1):
        """Opens a ``.wav`` file using `scipy.io.wavfile.read
        <https://docs.scipy.org/doc/scipy/reference/generated/scipy.io.wavfile.read.html>`_, and loads the attributes
        :attr:`samples` and :attr:`frequency`. If the wav file has more than one channel, it is converted to mono by
        averaging the values from all the samples via :func:`tool_functions.stereo_to_mono`.

        .. versionadded:: 2.0

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        if verbosity > 0:
            print("\n\tOpening the audio...", end=" ")

        try:
            from scipy.io import wavfile
        except ImportError:
            raise ModuleNotFoundException("scipy", "open a file in .wav format.")

        try:
            import numpy as np
        except ImportError:
            raise ModuleNotFoundException("numpy", "open a file in .wav format.")

        audio_data = wavfile.read(self.path)
        self.frequency = audio_data[0]

        # Turn stereo to mono if the file is stereo
        if len(np.shape(audio_data[1])) != 1:
            self.samples = stereo_to_mono(audio_data[1], verbosity)
        else:
            self.samples = audio_data[1]

        self._calculate_timestamps()

        if verbosity > 0:
            print("Audio loaded.\n")

    def _read_text_file(self, verbosity=1):
        """Opens a file containing the samples in ``.json``, ``.xlsx``, ``.csv``, ``.tsv`` or ``.txt`` file containing
        the timestamps and samples of the audio and loads the attributes :attr:`samples`, :attr:`timestamps`, and
        :attr:`frequency`. Depending on the file type, this function handles the content differently (see
        :ref:`Audio formats <wav_example>`).

        .. versionadded:: 2.0

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        if verbosity > 0:
            print("Opening the audio...", end=" ")

        if self.path.split(".")[-1] == "json":
            data = read_json(self.path)
            self.samples = data["Sample"]
            self.timestamp = data["Timestamp"]
            self._calculate_frequency()

        elif self.path.split(".")[-1] == "xlsx":
            import openpyxl as op
            workbook = op.load_workbook(self.path)
            sheet = workbook[workbook.sheetnames[0]]

            joints_labels = []
            for cell in sheet["1"]:
                joints_labels.append(str(cell.value))

            # For each pose
            for s in range(2, len(sheet["A"]) + 1):

                if verbosity > 1:
                    print("Loading sample " + str(s) + " of " + str(len(sheet["A"])) + "...", end=" ")

                self.timestamps.append(float(sheet.cell(s, 1).value))
                self.samples.append(int(sheet.cell(s, 2).value))

                if verbosity > 1:
                    print("OK.")

            self._calculate_frequency()

        elif self.path.split(".")[-1] in ["csv", "txt", "tsv"]:

            separator = get_filetype_separator(self.path.split(".")[-1])

            # Open the file and read the data
            data, self.metadata = read_text_table(self.path)

            for s in range(1, len(data)):

                d = data[s].split(separator)

                if verbosity > 1:
                    print("Loading sample " + str(s) + " of " + str(len(data) - 1) + "...", end=" ")

                self.timestamps.append(float(d[0][s]))
                self.samples.append(int(d[1][s]))

                self._calculate_frequency()

                if verbosity > 1:
                    print("OK.")

        if verbosity:
            print("100% - Done.")

    def _calculate_frequency(self):
        """Determines the frequency (number of samples per second) of the audio by calculating the time elapsed between
        the two first timestamps. This function is automatically called when reading the timestamps and samples from
        a text file.

        .. versionadded:: 2.0
        """
        self.frequency = 1 / (self.timestamps[1] - self.timestamps[0])

    def _calculate_timestamps(self):
        """Calculates the timestamps of the audio samples from the frequency and the number of samples. This function
        is automatically called when reading the frequency and samples from an audio file, or when a list of samples
        and a frequency are passed as a parameters during the initialisation.

        .. versionadded:: 2.0
        """
        self.timestamps = [i / self.frequency for i in range(len(self.samples))]

    # === Getter functions ===
    def get_path(self):
        """Returns the attribute :attr:`path` of the Audio instance.

        . versionadded:: 2.0

        Returns
        -------
        str
            The path of the Audio instance.
        """
        return self.path

    def get_name(self):
        """Returns the attribute :attr:`name` of the Audio instance.

        . versionadded:: 2.0

        Returns
        -------
        str
            The name of the Audio instance.
        """
        return self.name

    def get_condition(self):
        """Returns the attribute :attr:`condition` of the Audio instance.

        .. versionadded:: 2.0

        Returns
        -------
        str
            The experimental condition in which the recording of the audio clip was performed.
        """
        return self.condition

    def get_samples(self):
        """Returns the attribute :attr:`samples` of the Audio instance.

        . versionadded:: 2.0

        Returns
        -------
        list(int)
            The samples of the Audio instance.
        """
        return self.samples

    def get_sample(self, sample_index):
        """Returns the sample corresponding to the index passed as parameter.

        .. versionadded:: 2.0

        Parameters
        ----------
        sample_index: int
            The index of the sample.

        Returns
        -------
        int
            A sample from the sequence.
        """
        if len(self.samples) == 0:
            raise Exception("The Audio does not have any sample.")
        elif not 0 <= sample_index < len(self.samples):
            raise Exception("The pose index must be between 0 and " + str(len(self.samples) - 1) + ". ")

        return self.samples[sample_index]

    def get_number_of_samples(self):
        """Returns the number of samples in the audio clip.

        .. versionadded:: 2.0

        Returns
        -------
        int
            The amount of samples in the audio clip.
        """
        return len(self.samples)

    def get_timestamps(self):
        """Returns a list of the timestamps for every sample, in seconds.

        .. versionadded:: 2.0

        Returns
        -------
        list(float)
            List of the timestamps of all the samples of the audio clip, in seconds.
        """
        return self.timestamps

    def get_duration(self):
        """Returns the duration of the audio clip, in seconds.

        .. versionadded:: 2.0

        Returns
        -------
        float
            The duration of the audio clip, in seconds.
        """
        return self.timestamps[-1]

    def get_frequency(self):
        """Returns the frequency of the audio clip, in hertz.

        .. versionadded:: 2.0

        Returns
        -------
        int or float
            The frequency of the audio clip, in hertz.
        """
        return self.frequency

    def get_info(self):
        """Returns a dictionary containing data from the Audio clip.

        .. versionadded:: 2.0

        Returns
        -------
        OrderedDict
            An ordered dictionary where each descriptor is associated to its value. The included information fields are:

                • ``"Name"``: The :attr:`name` attribute of the audio clip.
                • ``"Path"``: The :attr:`path` attribute of the audio clip.
                • ``"Condition"``: The :attr:`condition` attribute of the audio clip.
                • ``"Frequency"``: Output of :meth:`Audio.get_frequency`.
                • ``"Number of samples"``: Output of :meth:`Audio.get_number_of_samples`.
                • ``"Duration"``: Output of :meth:`Audio.get_duration`.
        """

        stats = OrderedDict()
        stats["Name"] = self.name
        stats["Path"] = self.path
        stats["Condition"] = self.condition
        stats["Frequency"] = self.frequency
        stats["Number of samples"] = self.get_number_of_samples()
        stats["Duration"] = self.get_duration()

        return stats

    # === Transformation functions ===
    def get_envelope(self, window_size=1e6, overlap_ratio=0.5, filter_below=None, filter_over=None, name=None,
                     verbosity=1):
        """Calculates the envelope of an array, and returns it. The function can also optionally perform a band-pass
        filtering, if the corresponding parameters are provided.

        Parameters
        ----------
        window_size: int or None, optional
            The size of the windows (in samples) in which to cut the audio clip to calculate the envelope. Cutting the
            audio clips in windows allows, in the case where they are long, to speed up the computation. If this
            parameter is set on `None`, the window size will be set on the number of samples. A good value for this
            parameter is generally 1 million. If this parameter is set on 0, on None or on a number of samples bigger
            than the amount of samples in the Audio instance, the window size is set on the length of the samples.

        overlap_ratio: float or None, optional
            The ratio of samples overlapping between each window. If this parameter is not `None`, each window will
            overlap with the previous (and, logically, the next) for an amount of samples equal to the number of samples
            in a window times the overlap ratio. Then, only the central values of each window will be preserved and
            concatenated; this allows to discard any "edge" effect due to the windowing. If the parameter is set on
            `None` or 0, the windows will not overlap.

        filter_below: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the lowest frequency of the band-pass filter.

        filter_over: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the highest frequency of the band-pass filter.

        name: str or None, optional
            Defines the name of the envelope. If set on ``None``, the name will be the same as the original Audio
            instance, with the suffix ``"(ENV)"``.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        np.array
            The envelope of the original array.
        """

        if verbosity > 0:
            print("Creating an Envelope object...")

        try:
            from scipy.signal import hilbert
        except ImportError:
            raise ModuleNotFoundException("scipy", "extracting the envelope of an audio file.")

        try:
            import numpy as np
        except ImportError:
            raise ModuleNotFoundException("numpy", "extracting the envelope of an audio file.")

        if window_size == 0 or window_size > len(self.samples) or window_size is None:
            window_size = len(self.samples)

        if overlap_ratio is None:
            overlap_ratio = 0

        if name is None:
            name = self.name + " (ENV)"

        window_size = int(window_size)
        overlap = int(np.ceil(overlap_ratio * window_size))
        number_of_windows = get_number_of_windows(len(self.samples), window_size, overlap_ratio, True)

        # Hilbert transform
        if verbosity == 1:
            print("\tGetting the Hilbert transform...", end=" ")
        elif verbosity > 1:
            print("\tGetting the Hilbert transform...")
            print("\t\tDividing the samples in " + str(number_of_windows) + " window(s) of " + str(window_size) +
                  " samples, with an overlap of " + str(overlap) + " samples.")

        envelope_samples = np.zeros(len(self.samples))
        j = 0
        next_percentage = 10

        for i in range(number_of_windows):

            if verbosity == 1:
                while i / number_of_windows > next_percentage / 100:
                    print(str(next_percentage) + "%", end=" ")
                    next_percentage += 10

            # Get the Hilbert transform of the window
            array_start = i * (window_size - overlap)
            array_end = np.min([(i + 1) * window_size - i * overlap, len(self.samples)])
            if verbosity > 1:
                print("\t\t\tGetting samples from window " + str(i + 1) + "/" + str(number_of_windows) + ": samples " +
                      str(array_start) + " to " + str(array_end) + "... ", end=" ")
            hilbert_window = np.abs(hilbert(self.samples[array_start:array_end]))

            # Keep only the center values
            if i == 0:
                slice_start = 0
            else:
                slice_start = overlap // 2  # We stop one before if the overlap is odd

            if i == number_of_windows - 1:
                slice_end = len(hilbert_window)
            else:
                slice_end = window_size - int(np.ceil(overlap / 2))

            if verbosity > 1:
                print("\n\t\t\tKeeping the samples from " + str(slice_start) + " to " + str(slice_end) + " in the " +
                      "window: samples " + str(array_start + slice_start) + " to " + str(
                    array_start + slice_end) + "...",
                      end=" ")

            preserved_samples = hilbert_window[slice_start:slice_end]
            envelope_samples[j:j + len(preserved_samples)] = preserved_samples
            j += len(preserved_samples)

            if verbosity > 1:
                print("Done.")

        if verbosity == 1:
            print("100% - Done.")
        elif verbosity > 1:
            print("Done.")

        envelope = Envelope(envelope_samples, self.timestamps, self.frequency, name, self.condition)

        # Filtering
        if filter_below is not None or filter_over is not None:
            envelope = envelope.filter_frequencies(filter_below, filter_over, name, verbosity)

        return envelope

    # noinspection PyArgumentList
    def get_pitch(self, filter_below=None, filter_over=None, name=None, method="parselmouth", zeros_as_nan=False,
                  verbosity=1):
        """Calculates the pitch of the voice in the audio clip, and returns a Pitch object.

        .. versionadded:: 2.0

        Parameters
        ----------
        filter_below: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the lowest frequency of the band-pass filter.

        filter_over: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the highest frequency of the band-pass filter.

        name: str or None, optional
            Defines the name of the pitch. If set on ``None``, the name will be the same as the original Audio
            instance, with the suffix ``"(PIT)"``.

        zeros_as_nan: bool, optional
            If set on True, the values where the pitch is equal to 0 will be replaced by
            `numpy.nan <https://numpy.org/doc/stable/reference/constants.html#numpy.nan>`_ objects.

        method: str, optional
            Defines the pitch tracking method used. If set on `"parselmouth` (default), the to_pitch method from
            Parselmouth will be used to get the pitch. If set on `"crepe"`, the CREPE Python module will be used.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        Pitch
            The pitch of the voice in the audio clip.
        """
        if verbosity > 0:
            print("Creating a Pitch object...")

        if method == "parselmouth":
            try:
                from parselmouth import Sound
            except ImportError:
                raise ModuleNotFoundException("parselmouth", "get the pitch of an audio clip.")
        elif method == "crepe":
            try:
                import crepe
            except ImportError:
                raise ModuleNotFoundException("crepe", "get the pitch of an audio clip.")
        else:
            raise Exception("""The parameter "method" can only be set on "parselmouth" or "crepe".""")

        try:
            import numpy as np
        except ImportError:
            raise ModuleNotFoundException("numpy", "get the pitch of an audio clip.")

        samples = np.array(self.samples, dtype=np.float64)

        if verbosity > 0:
            if method == "parselmouth":
                print("\tTurning the audio into a parselmouth object...", end=" ")
            else:
                print("\tGetting the pitch from crepe...", end=" ")

        if method == "parselmouth":
            parselmouth_sound = Sound(np.ndarray(np.shape(samples), dtype=np.float64, buffer=samples), self.frequency)

            if name is None:
                name = self.name + " (PIT)"

            if verbosity > 0:
                print("Done.")
                print("\tGetting the pitch...", end=" ")

            pitch = parselmouth_sound.to_pitch(time_step=1 / self.frequency)

            if verbosity > 0:
                print("Done.")
                print("\tPadding the data...", end=" ")

            pitch, timestamps = pad(pitch.selected_array["frequency"], pitch.xs(), self.timestamps)

            if zeros_as_nan:
                pitch[pitch == 0] = np.nan

            if verbosity > 0:
                print("Done.")

            pitch = Pitch(pitch, self.timestamps, self.frequency, self.name, self.condition)

        else:
            time, frequency, confidence, activation = crepe.predict(samples, self.frequency, viterbi=True)

            pitch = Pitch(frequency, time, self.frequency, self.name, self.condition)

        if filter_below is not None or filter_over is not None:
            pitch = pitch.filter_frequencies(filter_below, filter_over, name, verbosity)

        return pitch

    # noinspection PyArgumentList
    def get_intensity(self, filter_below=None, filter_over=None, name=None, verbosity=1):
        """Calculates the intensity of the voice in the audio clip, and returns an Intensity object. The function can
        also optionally perform a band-pass filtering and a resampling, if the corresponding parameters are provided.

        .. versionadded:: 2.0

        Parameters
        ----------
        filter_below: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the lowest frequency of the band-pass filter.

        filter_over: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the highest frequency of the band-pass filter.

        name: str or None, optional
            Defines the name of the intensity. If set on ``None``, the name will be the same as the original Audio
            instance, with the suffix ``"(INT)"``.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        Intensity
            The intensity of the voice in the audio clip.
        """
        if verbosity > 0:
            print("Creating an Intensity object...")

        try:
            from parselmouth import Sound
        except ImportError:
            raise ModuleNotFoundException("parselmouth", "get the intensity of an audio clip")

        try:
            import numpy as np
        except ImportError:
            raise ModuleNotFoundException("numpy", "get the intensity of an audio clip.")

        samples = np.array(self.samples, dtype=np.float64)

        if verbosity > 0:
            print("\tTurning the audio into a parselmouth object...", end=" ")

        parselmouth_sound = Sound(np.ndarray(np.shape(samples), dtype=np.float64, buffer=samples), self.frequency)

        if name is None:
            name = self.name + " (INT)"

        if verbosity > 0:
            print("Done.")
            print("\tGetting the intensity...", end=" ")

        intensity = parselmouth_sound.to_intensity(time_step=1 / self.frequency)
        intensity_timestamps = add_delay(intensity.xs(), -intensity.xs()[0] % (1 / self.frequency))

        if verbosity > 0:
            print("Done.")
            print("\tPadding the data...", end=" ")

        samples, timestamps = pad(intensity.values.T, intensity_timestamps, self.timestamps)

        if verbosity > 0:
            print("Done.")

        intensity = Intensity(samples, timestamps, self.frequency, name, self.condition)

        if filter_below is not None or filter_over is not None:
            intensity = intensity.filter_frequencies(filter_below, filter_over, name, verbosity)

        return intensity

    # noinspection PyArgumentList
    def get_formant(self, filter_below=None, filter_over=None, name=None, formant_number=1, verbosity=1):
        """Calculates the formants of the voice in the audio clip, and returns a Formant object.

        .. versionadded:: 2.0

        Parameters
        ----------
        filter_below: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the lowest frequency of the band-pass filter.

        filter_over: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the highest frequency of the band-pass filter.

        name: str or None, optional
            Defines the name of the intensity. If set on ``None``, the name will be the same as the original Audio
            instance, with the suffix ``"(Fn)"``, with n being the formant number.

        formant_number: int, optional.
            One of the formants of the voice in the audio clip (1 (default), 2, 3, 4 or 5).

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        Formant
            The value of a formant of the voice in the audio clip.
        """
        if verbosity > 0:
            print("Creating a Formant object...")

        try:
            from parselmouth import Sound
        except ImportError:
            raise ModuleNotFoundException("parselmouth", "get one of the formants of an audio clip.")

        if verbosity > 0:
            print("\tTurning the audio into a parselmouth object...", end=" ")

        try:
            import numpy as np
        except ImportError:
            raise ModuleNotFoundException("numpy", "get the formants of an audio clip.")

        samples = np.array(self.samples, dtype=np.float64)
        parselmouth_sound = Sound(np.ndarray(np.shape(samples), dtype=np.float64, buffer=samples), self.frequency)

        if verbosity > 0:
            print("Done.")
            print("\tGetting the formant...", end=" ")

        if name is None:
            name = self.name + " (" + str(formant_number) + ")"

        formant = parselmouth_sound.to_formant_burg(time_step=1 / self.frequency)
        formant_timestamps = add_delay(formant.xs(), -1 / (2 * self.frequency))

        if verbosity > 0:
            print("Done.")
            print("\tPadding the data...", end=" ")

        number_of_points = formant.get_number_of_frames()
        f = np.zeros(number_of_points)
        for i in range(1, number_of_points + 1):
            t = formant.get_time_from_frame_number(i)
            f[i-1] = formant.get_value_at_time(formant_number=formant_number, time=t)
        f, timestamps = pad(f, formant_timestamps, self.timestamps)

        if verbosity > 0:
            print("Done.")

        formant = Formant(f, timestamps, self.frequency, formant_number, name, self.condition)

        if filter_below is not None or filter_over is not None:
            formant = formant.filter_frequencies(filter_below, filter_over, name, verbosity)

        return formant

    def filter_frequencies(self, filter_below=None, filter_over=None, name=None, verbosity=1):
        """Applies a low-pass, high-pass or band-pass filter to the data in the attribute :attr:`samples`.

        .. versionadded: 2.0

        Parameters
        ----------
        filter_below: float or None, optional
            The value below which you want to filter the data. If set on None or 0, this parameter will be ignored.
            If this parameter is the only one provided, a high-pass filter will be applied to the samples; if
            ``filter_over`` is also provided, a band-pass filter will be applied to the samples.

        filter_over: float or None, optional
            The value over which you want to filter the data. If set on None or 0, this parameter will be ignored.
            If this parameter is the only one provided, a low-pass filter will be applied to the samples; if
            ``filter_below`` is also provided, a band-pass filter will be applied to the samples.

        name: str or None, optional
            Defines the name of the output audio. If set on ``None``, the name will be the same as the
            original audio derivative, with the suffix ``"+FF"``.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        Audio
            The Audio instance, with filtered values.
        """

        # Band-pass filter
        if filter_below not in [None, 0] and filter_over not in [None, 0]:
            if verbosity > 0:
                print("Applying a band-pass filter for frequencies between " + str(filter_below) + " and " +
                      str(filter_over) + " Hz...", end=" ")
            b, a = butter(2, [filter_below, filter_over], "band", fs=self.frequency)
            new_samples = lfilter(b, a, self.samples)

        # High-pass filter
        elif filter_below not in [None, 0]:
            if verbosity > 0:
                print("Applying a high-pass filter for frequencies over " + str(filter_below) + " Hz...", end=" ")
            b, a = butter(2, filter_below, "high", fs=self.frequency)
            new_samples = lfilter(b, a, self.samples)

        # Low-pass filter
        elif filter_over not in [None, 0]:
            if verbosity > 0:
                print("Applying a low-pass filter for frequencies below " + str(filter_over) + " Hz...", end=" ")
            b, a = butter(2, filter_over, "low", fs=self.frequency)
            new_samples = lfilter(b, a, self.samples)

        else:
            new_samples = self.samples

        if verbosity > 0:
            print("Done.")

        if name is None:
            name = self.name

        new_audio = Audio(new_samples, self.timestamps, self.frequency, name, verbosity=0)
        return new_audio

    def resample(self, frequency, mode="cubic", window_size=1e7, overlap_ratio=0.5, name=None, verbosity=1):
        """Resamples an audio clip to the `frequency` parameter. It first creates a new set of timestamps at the
        desired frequency, and then interpolates the original data to the new timestamps.

        .. versionadded:: 2.0

        Parameters
        ----------
        frequency: float
            The frequency, in hertz, at which you want to resample the audio clip. A frequency of 4 will return samples
            at 0.25 s intervals.

        mode: str, optional
            This parameter allows for all the values accepted for the ``kind`` parameter in the function
            :func:`scipy.interpolate.interp1d`: ``"linear"``, ``"nearest"``, ``"nearest-up"``, ``"zero"``,
            ``"slinear"``, ``"quadratic"``, ``"cubic"``”, ``"previous"``, and ``"next"``. See the `documentation
            for this Python module
            <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`_ for more.
            This parameter also allows another special value, ``"take"``, which keeps one out of :math:`n` samples,
            where :math:`n` is equal to the original frequency divided by the resampling frequency. This allows for
            faster computation. Note that this function will return a warning if the resampling frequency is not an
            integer divider of the original frequency.

        window_size: int, optional
            The size of the windows in which to cut the audio samples to perform the resampling. Cutting long arrays
            in windows allows to speed up the computation. If this parameter is set on `None`, the window size will be
            set on the number of samples. A good value for this parameter is generally 10 million (1e7). If this
            parameter is set on 0, on None or on a number of samples bigger than the amount of samples in the Audio
            instance, the window size is set on the length of the samples.

        overlap_ratio: float, optional
            The ratio of samples overlapping between each window. If this parameter is not `None`, each window will
            overlap with the previous (and, logically, the next) for an amount of samples equal to the number of samples
            in a window times the overlap ratio. Then, only the central values of each window will be preserved and
            concatenated; this allows to discard any "edge" effect due to the windowing. If the parameter is set on
            `None` or 0, the windows will not overlap. By default, this parameter is set on 0.5, meaning that each
            window will overlap for half of their values with the previous, and half of their values with the next.

        name: str or None, optional
            Defines the name of the output audio clip. If set on ``None``, the name will be the same as the input
            audio clip, with the suffix ``"+RS"``.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        Audio
            A new audio clip containing resampled timestamps and samples.

        Warning
        -------
        This function allows both the **upsampling** and the **downsampling** of audio clips. However, during any of
        these operations, the algorithm only **estimates** the real values of the samples. You should then consider
        the upsampling (and the downsampling, to a lesser extent) with care.
        You can control the frequency of the original audio clip with :meth:`Audio.get_frequency()`.
        """

        if verbosity > 0:
            print("Resampling the audio clip at " + str(frequency) + " Hz (mode: " + str(mode) + ")...")
            print("\tOriginal frequency: " + str(round(self.get_frequency(), 2)))

        if verbosity > 0:
            print("\tPerforming the resampling...", end=" ")

        if window_size == 0 or window_size > len(self.samples) or window_size is None:
            window_size = len(self.samples)

        if overlap_ratio is None:
            overlap_ratio = 0

        if mode == "take":
            if frequency > self.frequency:
                raise Exception("""The mode "take" does not allow for upsampling of the data. Please input a resampling
                frequency inferior to the original (""" + str(self.frequency) + ").")
            factor_resampling = self.frequency / frequency
            if factor_resampling != int(factor_resampling):
                raise Warning("""The downsampling factor is not an integer, meaning that the downsampling may not be
                accurate. To ensure an accurate resampling with the "take" mode, use a resampling frequency that is
                an integer divider of the original frequency.""")
            indices = np.arange(0, len(self.samples), factor_resampling, dtype=int)
            resampled_audio_array = np.take(self.samples, indices)

        else:
            resampled_audio_array, resampled_audio_times = resample_data(self.samples, self.timestamps, frequency,
                                                                         window_size, overlap_ratio, mode,
                                                                         verbosity=verbosity)
            resampled_audio_array = list(resampled_audio_array)

        new_audio = Audio(resampled_audio_array, frequency, name, verbosity=verbosity)
        new_audio.path = self.path

        if verbosity > 0:
            print("100% - Done.")
            print("\tOriginal audio had " + str(len(self.samples)) + " samples.")
            print("\tNew audio has " + str(len(new_audio.samples)) + " samples.\n")

        return new_audio

    def filter_and_resample(self, filter_below, filter_over, resampling_frequency, resampling_mode="cubic",
                            verbosity=1):
        """Returns the samples contained in :attr:`samples`, after applying a band-pass filter and a resampling if
        parameters are provided.

        .. versionadded:: 2.0

        Parameters
        ----------
        filter_below: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the lowest frequency of the band-pass filter.

        filter_over: int, float or None, optional
            If not ``None`` nor 0, this value will be provided as the highest frequency of the band-pass filter.

        resampling_frequency: int, float or None, optional
            If not ``None``, the pitch will be resampled at the provided frequency before being returned.

        resampling_mode: str, optional
            This parameter also allows for all the values accepted for the ``kind`` parameter in the function
            :func:`scipy.interpolate.interp1d`: ``"linear"``, ``"nearest"``, ``"nearest-up"``, ``"zero"``,
            ``"slinear"``, ``"quadratic"``, ``"cubic"``”, ``"previous"``, and ``"next"``. See the `documentation
            for this Python module
            <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`_ for more.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        Audio
            The filtered and resampled Audio instance.
        """
        audio = self
        if filter_below is not None or filter_over is not None:
            audio = audio.filter_frequencies(filter_below, filter_over, verbosity=verbosity)
        if resampling_frequency is not None:
            audio = audio.resample(resampling_frequency, resampling_mode, verbosity=verbosity)

        return audio

    def trim(self, start=None, end=None, name=None, verbosity=1, add_tabs=0):
        """Trims an audio clip according to a starting and an ending timestamps. Timestamps must be provided in seconds.

        .. versionadded:: 2.0

        Parameters
        ----------
        start: int or None, optional
            The timestamp after which the samples will be preserved. If set on ``None``, the beginning of the audio clip
            will be set as the timestamp of the first sample.

        end: int or None, optional
            The timestamp before which the samples will be preserved. If set on ``None``, the end of the audio clip will
            be set as the timestamp of the last sample.

        name: str or None, optional
            Defines the name of the output Audio instance. If set on ``None``, the name will be the same as the original
            Audio instance, with the suffix ``"+TR"``.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        add_tabs: int, optional
            Adds the specified amount of tabulations to the verbosity outputs. This parameter may be used by other
            functions to encapsulate the verbosity outputs by indenting them. In a normal use, it shouldn't be set
            by the user.

        Returns
        -------
        Audio
            A new Audio instance containing a subset of the samples of the original.
        """

        start_index = None
        end_index = None

        if start is None:
            start = self.timestamps[0]
            start_index = 0
        if end is None:
            end = self.timestamps[-1]
            end_index = -1

        if end < start:
            raise Exception("End timestamp should be inferior to beginning timestamp.")
        elif start < self.timestamps[0]:
            raise Exception("Starting timestamp (" + str(start) + ") lower than the first timestamp of the Audio " +
                            "instance (" + str(self.timestamps[0]) + ").")
        elif start < 0:
            raise Exception("Starting timestamp (" + str(start) + ") must be equal to or higher than 0.")
        elif end > self.timestamps[-1]:
            raise Exception("Ending timestamp (" + str(end) + ") exceeds last timestamp of the Audio instance (" +
                            str(self.timestamps[-1]) + ").")

        if name is None:
            name = self.name + " +TR"

        if start_index is None:
            start_index = int(np.floor(start * self.frequency))
            if self.timestamps[start_index] <= start:
                start_index += 1

        if verbosity > 1:
            print("Closest (above) starting timestamp from " + str(start) + ": " + str(self.timestamps[start_index]))

        if end_index is None:
            end_index = int(np.ceil(end * self.frequency))
        if self.timestamps[end_index] >= end:
            end_index -= 1

        if verbosity > 1:
            print("Closest (below) starting timestamp from " + str(end) + ": " + str(self.timestamps[end_index]))

        new_audio = Audio(self.samples[start_index:end_index + 1], self.frequency, name, self.condition, verbosity)

        return new_audio

    def find_excerpt(self, other, window_size_env=1e6, overlap_ratio_env=0.5, filter_below=None, filter_over=50,
                     resampling_rate=1000, window_size_res=1e7, overlap_ratio_res=0.5, resampling_mode="cubic",
                     return_delay_format="s", return_correlation_value=False, threshold=0.9, plot_figure=False,
                     plot_intermediate_steps=False, path_figure=None, verbosity=1):
        """This function tries to find the timestamp at which an excerpt of the current Audio instance begins.
        The computation is performed through cross-correlation, by first turning the audio clips into filtered envelopes
        and downsampling them to accelerate the processing. The function returns the timestamp or the index of the
        maximal correlation value, or `None` if this value is below threshold.

        .. versionadded:: 2.0

        Parameters
        ----------
        other: Audio
            An Audio instance, smaller than or of equal size to the current object, that is allegedly an excerpt from
            the current Audio instance. The amplitude, frequency or values do not have to match exactly the ones from
            the current Audio instance.

        window_size_env: int or None, optional
            The size of the windows in which to cut the audio clips to calculate the envelope. Cutting the audio clips
            in windows allows, in the case where they are long, to speed up the computation. If this parameter is set on
            `None`, the window size will be set on the number of samples. A good value for this parameter is generally
            1 million.

        overlap_ratio_env: float or None, optional
            The ratio of samples overlapping between each window. If this parameter is not `None`, each window will
            overlap with the previous (and, logically, the next) for an amount of samples equal to the number of samples
            in a window times the overlap ratio. Then, only the central values of each window will be preserved and
            concatenated; this allows to discard any "edge" effect due to the windowing. If the parameter is set on
            `None` or 0, the windows will not overlap. By default, this parameter is set on 0.5, meaning that each
            window will overlap for half of their values with the previous, and half of their values with the next.

        filter_below: int or None, optional
            If set, a high-pass filter will be applied on the envelopes before performing the cross-correlation
            (default: 0 Hz).

        filter_over: int or None, optional
            If set, a low-pass filter will be applied on the envelopes before performing the cross-correlation (default:
            50 Hz).

        resampling_rate: int or None, optional
            The sampling rate at which to downsample the envelopes for the cross-correlation. A larger value will
            result in longer computation times. Setting the parameter on `None` will not downsample the envelopes, which
            will result in an error if the two audio clips do not have the same sampling rate. If this parameter is
            `None`, the next parameters related to resampling will be ignored. A recommended value for this parameter is
            1000, as it will speed up the computation of the cross-correlation while still giving a millisecond-
            precision delay.

        window_size_res: int or None, optional
            The size of the windows in which to cut the envelopes. Cutting the envelope in windows allows, in the case
            where audio clips are long, to speed up the computation. If this parameter is set on `None`, the window size
            will be set on the number of samples. A good value for this parameter is generally 10 million (1e7).

        overlap_ratio_res: float or None, optional
            The ratio of samples overlapping between each window. If this parameter is not `None`, each window will
            overlap with the previous (and, logically, the next) for an amount of samples equal to the number of samples
            in a window times the overlap ratio. Then, only the central values of each window will be preserved and
            concatenated; this allows to discard any "edge" effect due to the windowing. If the parameter is set on
            `None` or 0, the windows will not overlap. By default, this parameter is set on 0.5, meaning that each
            window will overlap for half of their values with the previous, and half of their values with the next.

        resampling_mode: str, optional
            This parameter allows for various values:

            • ``"linear"`` performs a linear
              `numpy.interp <https://numpy.org/devdocs/reference/generated/numpy.interp.html>`_ interpolation. This
              method, though simple, may not be very precise for upsampling naturalistic stimuli.
            • ``"cubic"`` performs a cubic interpolation via `scipy.interpolate.CubicSpline
              <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.CubicSpline.html>`_. This method,
              while smoother than the linear interpolation, may lead to unwanted oscillations nearby strong variations
              in the data.
            • ``"pchip"`` performs a monotonic cubic spline interpolation (Piecewise Cubic Hermite Interpolating
              Polynomial) via `scipy.interpolate.PchipInterpolator
              <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.PchipInterpolator.html>`_.
            • ``"akima"`` performs another type of monotonic cubic spline interpolation, using
              `scipy.interpolate.Akima1DInterpolator
              <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.Akima1DInterpolator.html>`_.
            • ``"take"`` keeps one out of n samples from the original array. While being the fastest computation, it
              will be prone to imprecision if the downsampling factor is not an integer divider of the original
              frequency.
            • ``"interp1d_XXX"`` uses the function `scipy.interpolate.interp1d
              <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`. The XXX part of
              the parameter can be replaced by ``"linear"``, ``"nearest"``, ``"nearest-up"``, ``"zero"``, "slinear"``,
              ``"quadratic"``, ``"cubic"``”, ``"previous"``, and ``"next"`` (see the documentation of this function for
              specifics).

        threshold: float, optional
            The threshold of the minimum correlation value between the two audio clips to accept a delay
            as a solution. If multiple delays are over threshold, the delay with the maximum correlation
            value will be returned. This value should be between 0 and 1; if the maximum found value is below the
            threshold, the function will return `None` instead of a timestamp.

        return_delay_format: str, optional
            This parameter can be either ``"index"``, ``"ms"``, ``"s"``, or ``"timedelta"``, and returns a marker of
            where the excerpt begins in the current Audio instance according to the highest cross-correlation value:

                • If ``"index"`` (default), the function will return the sample index from the current Audio instance.
                • If ``"ms"``, the function will return the timestamp from the current Audio instance, in milliseconds.
                • If ``"s"``, the function will return the timestamp from the current Audio instance, in seconds.
                • If ``"timedelta"``, the function will return the timestamp from the current Audio instance, as a
                  `datetime.timedelta <https://docs.python.org/3/library/datetime.html#timedelta-objects>`_ object.

        return_correlation_value: bool, optional
            If `True`, the function returns a second value: the correlation value at the returned delay.
            This value will be None if it is below the specified threshold. By default, this parameter is set on
            `False`.

        plot_figure: bool, optional
            If set on `True`, plots a graph showing the result of the cross-correlation using Matplotlib. Note that
            plotting the figure causes an interruption of the code execution.

        plot_intermediate_steps: bool, optional
            If set on `True`, plots the original audio clips, the envelopes and the resampled arrays (if
            calculated) besides the cross-correlation.

        path_figure: str or None, optional
            If set, saves the figure at the given path.

        threshold: float, optional
            The threshold of the maximum correlation value between the two audio clips, relative to the maximum
            correlation value between the excerpt and itself. This value should be between 0 and 1; if the maximum
            found value is below the threshold, the function will return `None` instead of a timestamp.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        int, float, timedelta or None
            The sample index, timestamp or timedelta of the current Audio instance at which the excerpt can be found,
            or `None` if the excerpt is not contained in the current Audio instance.
        float or None, optional
            Optionally, if ``return_correlation_value`` is `True`, the correlation value at the corresponding
            index/timestamp.
        """

        # Introduction
        if verbosity > 0:
            print("Trying to find when " + str(other.name) + " starts in " + str(self.name) + ".")
            print("\t" + str(self.name) + " contains " + str(len(self.samples)) + " samples, at a rate of " +
                  str(self.frequency) + " Hz.")
            print("\t" + str(other.name) + " contains " + str(len(other.samples)) + " samples, at a rate of " +
                  str(other.frequency) + " Hz.")

        # Envelopes
        if verbosity > 0:
            print("Getting the envelope from the first Audio instance...")
        envelope1 = self.get_envelope(window_size_env, overlap_ratio_env, filter_below, filter_over, verbosity)
        if verbosity > 0:
            print("Getting the envelope from the second Audio instance...")
        envelope2 = other.get_envelope(window_size_env, overlap_ratio_env, filter_below, filter_over, verbosity)
        if verbosity > 0:
            print("Envelopes calculated.\n")

        # Resampling
        if resampling_rate is not None:
            rate = resampling_rate
            if verbosity > 0:
                print("\tResampling the first envelope...")
            y1 = envelope1.resample(resampling_rate, window_size_res, overlap_ratio_res, resampling_mode, verbosity)
            if verbosity > 0:
                print("\tResampling the second envelope...")
            y2 = envelope2.resample(resampling_rate, window_size_res, overlap_ratio_res, resampling_mode, verbosity)
            if verbosity > 0:
                print("\tResampling done.")
        else:
            rate = self.frequency
            if self.frequency != other.frequency:
                raise Exception("The rate of the two Audio instances you are trying to correlate are different (" +
                                str(self.frequency) + " and " + str(other.frequency) + "). You must indicate a " +
                                "resampling rate to perform the cross-correlation.")
            y1 = envelope1
            y2 = envelope2

        # Cross-correlation
        if verbosity > 0:
            print("\tComputing the correlation...")
        y2_normalized = (y2.samples - y2.samples.mean()) / y2.samples.std() / np.sqrt(y2.samples.size)
        y1_m = signal.correlate(y1.samples, np.ones(y2.samples.size), 'valid') ** 2 / y2_normalized.size
        y1_m2 = signal.correlate(y1.samples ** 2, np.ones(y2.samples.size), "valid")
        cross_correlation = signal.correlate(y1.samples, y2_normalized, "valid") / np.sqrt(y1_m2 - y1_m)
        max_correlation_value = np.max(cross_correlation)

        index_max_correlation = np.argmax(cross_correlation)
        index = int(round(index_max_correlation * self.frequency / rate, 0))
        delay_in_seconds = index_max_correlation / rate
        t = dt.timedelta(days=delay_in_seconds // 86400, seconds=int(delay_in_seconds % 86400),
                         microseconds=(delay_in_seconds % 1) * 1000000)

        if verbosity > 0:
            print("Done.")

            if max_correlation_value >= threshold:
                print("\tMaximum correlation (" + str(round(max_correlation_value, 3)) + ") found at sample " +
                      str(index) + " (timestamp " + str(t) + ").")

            else:
                print("\tNo correlation over threshold found (max correlation: " +
                      str(round(max_correlation_value, 3)) + ") found at sample " + str(index) +
                      " (timestamp " + str(t) + ").")

        # Return values: None if below threshold
        if return_delay_format == "index":
            return_value = index
        elif return_delay_format == "ms":
            return_value = delay_in_seconds * 1000
        elif return_delay_format == "s":
            return_value = delay_in_seconds
        elif return_delay_format == "timedelta":
            return_value = t
        else:
            raise Exception(
                "Wrong value for the parameter return_delay_format: " + str(return_delay_format) + ". The " +
                'value should be either "index", "ms", "s" or "timedelta".')

        # Plot and/or save the figure
        if plot_figure or path_figure is not None:
            _plot_figure_find_excerpt(self, other, envelope1, envelope2, y1, y2, window_size_env,
                                      overlap_ratio_env, filter_below, filter_over, resampling_rate,
                                      window_size_res, overlap_ratio_res, cross_correlation, threshold,
                                      return_delay_format, return_value, plot_figure, path_figure, None,
                                      plot_intermediate_steps, verbosity)

        if max_correlation_value >= threshold:
            if return_correlation_value:
                return return_value, max_correlation_value
            else:
                return return_value
        else:
            if return_correlation_value:
                return None, None
            else:
                return None

    def find_excerpts(self, excerpts, window_size_env=1e6, overlap_ratio_env=0.5, filter_below=None, filter_over=50,
                     resampling_rate=1000, window_size_res=1e7, overlap_ratio_res=0.5, resampling_mode="cubic",
                     return_delay_format="s", return_correlation_values=False, threshold=0.9, plot_figure=False,
                     plot_intermediate_steps=False, path_figures=None, name_figures="figure", verbosity=1):
        """This function tries to find the timestamp at which multiple excerpts of the current Audio instance begin.
        The computation is performed through cross-correlation, by first turning the audio clips into downsampled and
        filtered envelopes to accelerate the processing. For each excerpt, the function returns the timestamp of the
        maximal correlation value, or `None` if this value is below threshold.

        .. versionadded:: 2.0

        Parameters
        ----------
        excerpts: list(Audio)
            A list of Audio instances, smaller than or of equal size to the current object, that are all allegedly
            excerpts from current Audio instance. The amplitude, frequency or values do not have to match exactly the
            ones from the current Audio instance.

        window_size_env: int or None, optional
            The size of the windows in which to cut the audio clips to calculate the envelope. Cutting the audio clips
            in windows allows, in the case where they are long, to speed up the computation. If this parameter is set on
            `None`, the window size will be set on the number of samples. A good value for this parameter is generally 1
            million.

        overlap_ratio_env: float or None, optional
            The ratio of samples overlapping between each window. If this parameter is not `None`, each window will
            overlap with the previous (and, logically, the next) for an amount of samples equal to the number of samples
            in a window times the overlap ratio. Then, only the central values of each window will be preserved and
            concatenated; this allows to discard any "edge" effect due to the windowing. If the parameter is set on
            `None` or 0, the windows will not overlap. By default, this parameter is set on 0.5, meaning that each
            window will overlap for half of their values with the previous, and half of their values with the next.

        filter_below: int or None, optional
            If set, a high-pass filter will be applied on the envelopes before performing the cross-correlation
            (default: 0 Hz).

        filter_over: int or None, optional
            If set, a low-pass filter will be applied on the envelopes before performing the cross-correlation (default:
            50 Hz).

        resampling_rate: int or None, optional
            The sampling rate at which to downsample the envelopes for the cross-correlation. A larger value will
            result in longer computation times. Setting the parameter on `None` will not downsample the envelopes, which
            will result in an error if the two audio clips do not have the same sampling rate. If this parameter is
            `None`, the next parameters related to resampling will be ignored. A recommended value for this parameter is
            1000, as it will speed up the computation of the cross-correlation while still giving a millisecond-
            precision delay.

        window_size_res: int or None, optional
            The size of the windows in which to cut the envelopes. Cutting the envelope in windows allows, in the case
            where audio clips are long, to speed up the computation. If this parameter is set on `None`, the window size
            will be set on the number of samples. A good value for this parameter is generally 10 million (1e7).

        overlap_ratio_res: float or None, optional
            The ratio of samples overlapping between each window. If this parameter is not `None`, each window will
            overlap with the previous (and, logically, the next) for an amount of samples equal to the number of samples
            in a window times the overlap ratio. Then, only the central values of each window will be preserved and
            concatenated; this allows to discard any "edge" effect due to the windowing. If the parameter is set on
            `None` or 0, the windows will not overlap. By default, this parameter is set on 0.5, meaning that each
            window will overlap for half of their values with the previous, and half of their values with the next.

        resampling_mode: str, optional
            This parameter allows for various values:

            • ``"linear"`` performs a linear
              `numpy.interp <https://numpy.org/devdocs/reference/generated/numpy.interp.html>`_ interpolation. This
              method, though simple, may not be very precise for upsampling naturalistic stimuli.
            • ``"cubic"`` performs a cubic interpolation via `scipy.interpolate.CubicSpline
              <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.CubicSpline.html>`_. This method,
              while smoother than the linear interpolation, may lead to unwanted oscillations nearby strong variations
              in the data.
            • ``"pchip"`` performs a monotonic cubic spline interpolation (Piecewise Cubic Hermite Interpolating
              Polynomial) via `scipy.interpolate.PchipInterpolator
              <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.PchipInterpolator.html>`_.
            • ``"akima"`` performs another type of monotonic cubic spline interpolation, using
              `scipy.interpolate.Akima1DInterpolator
              <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.Akima1DInterpolator.html>`_.
            • ``"take"`` keeps one out of n samples from the original array. While being the fastest computation, it
              will be prone to imprecision if the downsampling factor is not an integer divider of the original
              frequency.
            • ``"interp1d_XXX"`` uses the function `scipy.interpolate.interp1d
              <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`. The XXX part of
              the parameter can be replaced by ``"linear"``, ``"nearest"``, ``"nearest-up"``, ``"zero"``, "slinear"``,
              ``"quadratic"``, ``"cubic"``”, ``"previous"``, and ``"next"`` (see the documentation of this function for
              specifics).

        threshold: float, optional
            The threshold of the minimum correlation value between the audio clips to accept a delay as
            a solution. If multiple delays are over threshold, the delay with the maximum correlation
            value will be returned. This value should be between 0 and 1; if the maximum found value is below the
            threshold, the function will return `None` instead of a timestamp.

        return_delay_format: str, optional
            This parameter can be either ``"index"``, ``"ms"``, ``"s"``, or ``"timedelta"``, and returns a marker of
            where the excerpt begins in the current Audio instance according to the highest cross-correlation value:

                • If ``"index"`` (default), the function will return the sample index from the current Audio instance.
                • If ``"ms"``, the function will return the timestamp from the current Audio instance, in milliseconds.
                • If ``"s"``, the function will return the timestamp from the current Audio instance, in seconds.
                • If ``"timedelta"``, the function will return the timestamp from the current Audio instance, as a
                  `datetime.timedelta <https://docs.python.org/3/library/datetime.html#timedelta-objects>`_ object.

        return_correlation_values: bool, optional
            If `True`, the function returns a second value: the correlation value at the returned delay.
            This value will be None if it is below the specified threshold. By default, this parameter is set on
            `False`.

        plot_figure: bool, optional
            If set on `True`, plots a graph showing the result of the cross-correlation using Matplotlib. Note that
            plotting the figure causes an interruption of the code execution.

        plot_intermediate_steps: bool, optional
            If set on `True`, plots the original audio clips, the envelopes and the resampled arrays (if
            calculated) besides the cross-correlation.

        path_figures: str or None, optional
            If set, saves the figure at the given directory.

        name_figures: str, optional
            The name to give to each figure in the directory set by `path_figures`. The figures will be found in
            `path_figures/name_figures_n.png`, where n is the index of the excerpt in `excerpts`, starting at 0.

        threshold: float, optional
            The threshold of the maximum correlation value between the audio clips, relative to the maximum
            correlation value between the excerpt and itself. This value should be between 0 and 1; if the maximum
            found value is below the threshold, the function will return `None` instead of a timestamp.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        list(float or None)
            The timestamps of the current Audio instance at which each excerpt can be found, or `None` if the excerpt is
            not contained in the current Audio instance.
        """

        delays = []
        correlation_values = []

        # Introduction
        if verbosity > 0:
            print("Trying to find when the excerpts starts in the array.")
            print("\t" + str(self.name) + " contains " + str(len(self.samples)) + " samples, at a rate of " +
                  str(self.frequency) + " Hz.")
            print("\t" + str(len(excerpts)) + " excerpts to find.")

        # Envelope
        if verbosity > 0:
            print("Getting the envelope from the first Audio instance...")
        envelope1 = self.get_envelope(window_size_env, overlap_ratio_env, filter_below, filter_over, verbosity)
        if verbosity > 0:
            print("Envelope calculated.\n")

        # Resampling
        if resampling_rate is not None:
            rate = resampling_rate
            if verbosity > 0:
                print("\tResampling the first envelope...")
            y1 = envelope1.resample(resampling_rate, window_size_res, overlap_ratio_res, resampling_mode, verbosity)
            if verbosity > 0:
                print("\tResampling done.")
        else:
            rate = self.frequency
            y1 = envelope1

        # Main loop
        for i in range(len(excerpts)):

            # Introduction
            if verbosity > 0:
                print("\nExcerpt " + str(i + 1) + "/" + str(len(excerpts)))

            excerpt = excerpts[i]
            if verbosity > 0:
                print("\t" + str(excerpt.name) + " contains " + str(len(excerpt.samples)) + " samples, at a rate of " +
                      str(excerpt.frequency) + " Hz.")

            # Envelope
            if verbosity > 0:
                print("Getting the envelope from the second Audio instance...")
            envelope2 = excerpt.get_envelope(window_size_env, overlap_ratio_env, filter_below, filter_over, verbosity)
            if verbosity > 0:
                print("Envelopes calculated.\n")

            # Resampling
            if resampling_rate is not None:
                if verbosity > 0:
                    print("\tResampling the second envelope...")
                y2 = envelope2.resample(resampling_rate, window_size_res, overlap_ratio_res, resampling_mode, verbosity)
                if verbosity > 0:
                    print("\tResampling done.")
            else:
                rate = self.frequency
                if self.frequency != excerpt.frequency:
                    raise Exception("The rate of the two Audio instances you are trying to correlate are different (" +
                                    str(self.frequency) + " and " + str(excerpt.frequency) + "). You must indicate a " +
                                    "resampling rate to perform the cross-correlation.")
                y2 = envelope2

            # Cross-correlation
            if verbosity > 0:
                print("\tComputing the correlation...")
            y2_normalized = (y2.samples - y2.samples.mean()) / y2.samples.std() / np.sqrt(y2.samples.size)
            y1_m = signal.correlate(y1.samples, np.ones(y2.samples.size), 'valid') ** 2 / y2_normalized.size
            y1_m2 = signal.correlate(y1.samples ** 2, np.ones(y2.samples.size), "valid")
            cross_correlation = signal.correlate(y1.samples, y2_normalized, "valid") / np.sqrt(y1_m2 - y1_m)
            max_correlation_value = np.max(cross_correlation)

            index_max_correlation = np.argmax(cross_correlation)
            index = int(round(index_max_correlation * self.frequency / rate, 0))
            delay_in_seconds = index_max_correlation / rate
            t = dt.timedelta(days=delay_in_seconds // 86400, seconds=int(delay_in_seconds % 86400),
                             microseconds=(delay_in_seconds % 1) * 1000000)

            if verbosity > 0:
                print("Done.")

                if max_correlation_value >= threshold:
                    print("\tMaximum correlation (" + str(round(max_correlation_value, 3)) + ") found at sample " +
                          str(index) + " (timestamp " + str(t) + ").")

                else:
                    print("\tNo correlation over threshold found (max correlation: " +
                          str(round(max_correlation_value, 3)) + ") found at sample " + str(index) + " (timestamp " +
                          str(t) + ").")

            # Return values: None if below threshold
            if return_delay_format == "index":
                return_value = index
            elif return_delay_format == "ms":
                return_value = delay_in_seconds * 1000
            elif return_delay_format == "s":
                return_value = delay_in_seconds
            elif return_delay_format == "timedelta":
                return_value = t
            else:
                raise Exception("Wrong value for the parameter return_delay_format: " + str(return_delay_format) +
                                '. The value should be either "index", "ms", "s" or "timedelta".')

            # Plot and/or save the figure
            if plot_figure or path_figures is not None:
                _plot_figure_find_excerpt(self, excerpt, envelope1, envelope2, y1, y2, window_size_env,
                                          overlap_ratio_env, filter_below, filter_over, resampling_rate,
                                          window_size_res, overlap_ratio_res, cross_correlation, threshold,
                                          return_delay_format, return_value, plot_figure, path_figures, name_figures,
                                          plot_intermediate_steps, verbosity)

            if max_correlation_value >= threshold:
                if return_correlation_values:
                    delays.append(return_value)
                    correlation_values.append(max_correlation_value)
                else:
                    delays.append(return_value)

            else:

                if return_correlation_values:
                    delays.append(None)
                    correlation_values.append(None)
                else:
                    delays.append(None)

        if return_correlation_values:
            return delays, correlation_values
        else:
            return delays

    # === Conversion functions ===

    def convert_to_table(self):
        """Returns a list of lists where each sublist contains a timestamp and a sample. The first sublist contains the
        headers of the table. The output then resembles the table found in :ref:`Tabled formats <table_example>`.

        .. versionadded:: 2.0

        Returns
        -------
        list(list)
            A list of lists that can be interpreted as a table, containing headers, and with the timestamps and the
            coordinates of the joints from the sequence on each row.
        """

        table = [["Timestamp", "Sample"]]

        # For each pose
        for s in range(len(self.samples)):
            table.append([self.timestamps[s], self.samples[s]])

        return table

    def convert_to_json(self):
        """Returns a list ready to be exported in JSON. The returned JSON data is a dictionary with two keys:
        "Sample" is the key to the list of samples, while "Frequency" is the key to the sampling frequency of
        the audio clip.

        .. versionadded:: 2.0

        Returns
        -------
        dict
            A dictionary containing the data of the audio clip, ready to be exported in JSON.
        """

        data = {"Sample": self.samples, "Frequency": self.frequency}
        return data

    # === Print functions ===
    def print_details(self, include_name=True, include_condition=True, include_frequency=True,
                      include_number_of_samples=True, include_duration=True, add_tabs=0):
        """Prints a series of details about the audio clip.

        .. versionadded:: 2.0

        Parameters
        ----------
        include_name: bool, optional
            If set on ``True`` (default), adds the attribute :attr:`name` to the printed string.
        include_condition: bool, optional
            If set on ``True`` (default), adds the attribute :attr:`condition` to the printed string.
        include_frequency: bool, optional
            If set on ``True`` (default), adds the attribute :attr:`frequency` to the printed string.
        include_number_of_samples: bool, optional
            If set on ``True`` (default), adds the length of the attribute :attr:`samples` to the printed string.
        include_duration: bool, optional
            If set on ``True`` (default), adds the duration of the Sequence to the printed string.
        add_tabs: int, optional
            Adds the specified amount of tabulations to the verbosity outputs. This parameter may be used by other
            functions to encapsulate the verbosity outputs by indenting them. In a normal use, it shouldn't be set
            by the user.
        """
        t = add_tabs * "\t"

        string = t + "Audio clip · "
        if include_name:
            string += "Name: " + str(self.name) + " · "
        if include_condition:
            string += "Condition: " + str(self.condition) + " · "
        if include_frequency:
            string += "Frequency: " + str(self.frequency) + " Hz · "
        if include_number_of_samples:
            string += "Number of samples: " + str(self.get_number_of_samples()) + " · "
        if include_duration:
            t = self.get_duration()
            string += "Duration: " + str(time_unit_to_timedelta(t)) + " · "
        if len(string) > 3:
            string = string[:-3]

        print(string)

    # === Saving functions ===

    def save(self, folder_out, name=None, file_format="xlsx", individual=False, verbosity=1):
        """Saves an audio clip in a file or a folder. The function saves the sequence under
        ``folder_out/name.file_format``. All the non-existent subfolders present in the ``folder_out`` path will be
        created by the function. The function also updates the :attr:`path` attribute of the Audio clip.

        .. versionadded:: 2.0

        Parameters
        ----------
        folder_out: str, optional
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them. If the string provided is empty (by default), the audio clip will be saved in
            the current working directory. If the string provided contains a file with an extension, the fields ``name``
            and ``file_format`` will be ignored.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on the attribute :attr:`name` of the audio clip; if that attribute is also set on ``None``, the name will be
            set on ``"out"``. If ``individual`` is set on ``True``, each sample will be saved as a different file,
            having the index of the pose as a suffix after the name (e.g. if the name is ``"sample"`` and the file
            format is ``"txt"``, the poses will be saved as ``sample_0.txt``, ``sample_1.txt``, ``sample_2.txt``, etc.).

        file_format: str or None, optional
            The file format in which to save the audio clip. The file format must be ``"json"`` (default), ``"xlsx"``,
            ``"txt"``, ``"csv"``, ``"tsv"``, ``"wav"``, or, if you are a masochist, ``"mat"``. Notes:

                • ``"xls"`` will save the file with an ``.xlsx`` extension.
                • Any string starting with a dot will be accepted (e.g. ``".csv"`` instead of ``"csv"``).
                • ``"csv;"`` will force the value separator on ``;``, while ``"csv,"`` will force the separator
                  on ``,``. By default, the function will detect which separator the system uses.
                • ``"txt"`` and ``"tsv"`` both separate the values by a tabulation.
                • Any other string will not return an error, but rather be used as a custom extension. The data will
                  be saved as in a text file (using tabulations as values separators).

            .. warning::
                While it is possible to save audio clips as ``.mat`` or custom extensions, the toolbox will not
                recognize these files upon opening. The support for ``.mat`` and custom extensions as input may come in
                a future release, but for now these are just offered as output options.

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.
            This option is not available and will be ignored if ``file_format`` is set on ``"wav"``.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        if folder_out == "":
            folder_out = os.getcwd()

        if not individual:
            subfolders = folder_out.split("/")
            if len(subfolders) != 0:
                if "." in subfolders[-1]:
                    folder_out = "/".join(subfolders[:-1])
                    name = ".".join(subfolders[-1].split(".")[:-1])
                    file_format = subfolders[-1].split(".")[-1]

        # Automatic creation of all the folders of the path if they don't exist
        os.makedirs(folder_out, exist_ok=True)

        if name is None and self.name is not None:
            name = self.name
        elif name is None:
            name = "out"

        file_format = file_format.strip(".")  # We remove the dot in the format
        if file_format == "xls":
            file_format = "xlsx"

        if verbosity > 0:
            if individual:
                print("Saving " + file_format.upper() + " individual files...")
            else:
                print("Saving " + file_format.upper() + " global file: " + folder_out.strip("/") + "/" + name + "." +
                      file_format + "...")

        if file_format == "json":
            self._save_json(folder_out, name, individual, verbosity)

        elif file_format == "mat":
            self._save_mat(folder_out, name, individual, verbosity)

        elif file_format == "xlsx":
            self._save_xlsx(folder_out, name, individual, verbosity)

        elif file_format == "wav":
            self._save_wav(folder_out, name)

        else:
            self._save_txt(folder_out, name, file_format, individual, verbosity)

        if individual:
            self.path = folder_out + name
        else:
            self.path = folder_out + name + file_format

        if verbosity > 0:
            print("100% - Done.")

    def _save_json(self, folder_out, name=None, individual=False, verbosity=1):
        """Saves an audio clip as a json file or files. This function is called by the :meth:`Audio.save`
        method, and saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"`` if individual is ``False``, or on ``"sample"`` if individual is ``True``.

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        perc = 10  # Used for the progression percentage

        # Save the data
        if not individual:
            data = self.convert_to_json()
            if name is None:
                name = "out"
            with open(folder_out + "/" + name + ".json", 'w', encoding="utf-16-le") as f:
                json.dump(data, f)
        else:
            if name is None:
                name = "pose"
            for s in range(len(self.samples)):
                perc = show_progression(verbosity, s, len(self.samples), perc)
                with open(folder_out + "/" + name + "_" + str(s) + ".json", 'w', encoding="utf-16-le") as f:
                    d = {"Sample": self.samples[s], "Timestamp": self.timestamps[s]}
                    json.dump(d, f)

    def _save_mat(self, folder_out, name=None, individual=False, verbosity=1):
        """Saves an audio clip as a Matlab .mat file or files. This function is called by the :meth:`Audio.save`
        method, and saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Important
        ---------
            This function is dependent of the module `scipy <https://scipy.org/>`_.

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"`` if individual is ``False``, or on ``"sample"`` if individual is ``True``.

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        perc = 10  # Used for the progression percentage

        try:
            from scipy.io import savemat
        except ImportError:
            raise ModuleNotFoundException("scipy", "save a file in .mat format.")

        # Save the data
        if not individual:
            if name is None:
                name = "out"
            data = self.convert_to_json()
            savemat(folder_out + "/" + name + ".mat", {"data": data})
        else:
            for s in range(len(self.samples)):
                if name is None:
                    name = "sample"
                perc = show_progression(verbosity, s, len(self.samples), perc)
                savemat(folder_out + "/" + name + "_" + str(s) + ".mat",
                        {"data": [["Timestamp", "Sample"], [self.timestamps[s], self.samples[s]]]})

    def _save_xlsx(self, folder_out, name=None, individual=False, verbosity=1):
        """Saves an audio clip as an Excel .xlsx file or files. This function is called by the :meth:`Audio.save`
        method, and saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Important
        ---------
            This function is dependent of the module `openpyxl <https://pypi.org/project/openpyxl/>`_.

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"`` if individual is ``False``, or on ``"sample"`` if individual is ``True``.

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        perc = 10  # Used for the progression percentage

        data = self.convert_to_table()

        # Save the data
        if not individual:
            if name is None:
                name = "out"
            write_xlsx(data, folder_out + "/" + name + ".xlsx", verbosity)

        else:
            for s in range(len(self.samples)):
                if name is None:
                    name = "sample"
                d = [data[0], data[s + 1]]
                perc = show_progression(verbosity, s, len(self.samples), perc)
                write_xlsx(d, folder_out + "/" + name + "_" + str(s) + ".xlsx", False)

    def _save_wav(self, folder_out, name=None):
        """Saves an audio clip as a .wav file or files. This function is called by the :meth:`Audio.save` method, and
        saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"``.
        """

        try:
            from scipy.io import wavfile
        except ImportError:
            raise ModuleNotFoundException("scipy", "save a file in .wav format.")

        try:
            import numpy as np
        except ImportError:
            raise ModuleNotFoundException("numpy", "save a file in .wav format.")

        if name is None:
            name = "out"
        wavfile.write(folder_out + "/" + name + ".wav", self.frequency, self.samples)

    def _save_txt(self, folder_out, name=None, file_format="csv", individual=False, verbosity=1):
        """Saves an audio clip as a .txt, .csv, .tsv, or custom extension file or files. This function is called by the
        :meth:`Audio.save` method, and saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"`` if individual is ``False``, or on ``"sample"`` if individual is ``True``.

        file_format: str, optional
            The file format in which to save the audio clip. The file format can be ``"txt"``, ``"csv"`` (default) or
            ``"tsv"``. ``"csv;"`` will force the value separator on ``";"``, while ``"csv,"`` will force the separator
            on ``","``. By default, the function will detect which separator the system uses. ``"txt"`` and ``"tsv"``
            both separate the values by a tabulation. Any other string will not return an error, but rather be used as
            a custom extension. The data will be saved as in a text file (using tabulations as values separators).

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        perc = 10  # Used for the progression percentage

        # Force comma or semicolon separator
        if file_format == "csv,":
            separator = ","
            file_format = "csv"
        elif file_format == "csv;":
            separator = ";"
            file_format = "csv"
        elif file_format[0:3] == "csv":  # Get the separator from local user (to get , or ;)
            separator = get_system_csv_separator()
        elif file_format == "txt":  # For text files, tab separator
            separator = "\t"
        else:
            separator = "\t"

        # Save the data
        if not individual:
            if name is None:
                name = "out"
            write_text_table(self.convert_to_table(), separator, folder_out + "/" + name + "." + file_format, verbosity)
        else:
            for s in range(len(self.samples)):
                if name is None:
                    name = "pose"
                write_text_table(self.samples[s].convert_to_table(), separator,
                                 folder_out + "/" + name + "_" + str(s) + "." + file_format, 0)
                perc = show_progression(verbosity, s, len(self.samples), perc)

    def __len__(self):
        """Returns the number of samples in the audio clip (i.e., the length of the attribute :attr:`samples`).

        .. versionadded:: 2.0

        Returns
        -------
        int
            The number of samples in the audio clip.
        """
        return len(self.samples)

    def __getitem__(self, index):
        """Returns the sample of index specified by the parameter ``index``.

        Parameters
        ----------
        index: int
            The index of the sample to return.

        Returns
        -------
        float
            A sample from the attribute :attr:`samples`.
        """
        return self.samples[index]
