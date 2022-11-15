#!/usr/bin/env python
"""Krajjat 1.10
Kinect Realignment Algorithm for Joint Jumps And Twitches
Author: Romain Pastureau
This file contains all the classes used by the other files,
along with the realignment algorithm.
"""

import math
import numpy as np
import os
from pose import *
from joint import *
import random
from scipy import interpolate
import statistics
import wave
import tool_functions

__author__ = "Romain Pastureau"
__version__ = "1.10"
__email__ = "r.pastureau@bcbl.eu"
__license__ = "GPL"


class Sequence(object):
    """Defines a motion sequence by opening an existing one or creating a new."""

    def __init__(self, path=None, path_audio=None, name=None, convert_to_seconds=True):
        self.files = None  # List of the files in the target folder
        self.poses = []  # List of the poses in the sequence, ordered
        self.randomized = False  # True if the joints positions are randomized
        self.path_audio = path_audio  # Path to the audio file matched with this series of gestures

        self.name = None  # Placeholder for the name of the sequence
        self.define_name_init(name, path)  # Defines the name of the sequence

        # In the case where a file or folder is passed as argument,
        # we load the sequence
        self.path = None  # Placeholder for the path

        if path is not None:
            self.load_from_path(path, convert_to_seconds)

    # === Name and setter functions ===
    def define_name_init(self, name, path):
        """Defines the name of the sequence from a given name as a parameter of the declaration of the sequence,
        or takes the last folder (or filename) of the path."""
        if name is not None:
            self.name = name
        elif path is not None:
            if len(path.split("/")) >= 1:
                self.name = path.split("/")[-1]
        else:
            self.name = "Unnamed sequence"

    def set_name(self, name):
        """Sets a name or an identifier for the sequence."""
        self.name = name

    def set_audio_path(self, path_audio):
        """Sets the path to the audio file matching the gesture sequence."""
        self.path_audio = path_audio

    # === Loading functions ===
    def load_from_path(self, path, convert_to_seconds):
        self.path = path

        # If it's a folder, we fetch all the files
        if os.path.isdir(self.path):
            self.get_files()  # Fetches all the files
        self.load_poses()  # Loads the files into poses

        if len(self.poses) == 0:
            raise Exception("The path " + path + " is not a valid sequence.")

        self.load_relative_timestamps(convert_to_seconds)  # Sets the relative time from the first pose for each pose
        self.calculate_velocity()

    def get_files(self):
        """Opens a folder and fetches all the individual files (.json, .csv, .txt or .xlsx)."""

        file_list = tool_functions.os.listdir(self.path)  # List all the files in the folder
        self.files = ["" for _ in range(len(file_list))]  # Create an empty list the length of the files

        # Here, we find all the files that are either .json or .meta in the folder.
        for f in file_list:

            # If a file is ".json", then we get its index from its name to order it correctly in the list.
            # This is necessary as some file systems will order frame_2.json after frame_11.json because,
            # alphabetically, "1" is before "2".
            if f.split(".")[-1] in ["json", "csv", "txt", "xlsx"]:
                self.files[int(f.split(".")[0].split("_")[1])] = f

        # If there were files that weren't ".json" in the folder, then "self.files" is larger than the number of
        # frames. The list is thus ending by a series of empty strings that we trim.
        if "" in self.files:
            limit = self.files.index("")
            self.files = self.files[0:limit]

        print(str(len(self.files)) + " pose file(s) found.\n")

    def load_poses(self, verbose=1):
        """Opens successively all the .json files, read them and creates a Pose object for each of them.
        If verbose is True, returns a progression percentage every 10%."""

        if verbose > 0:
            print("Opening poses...")

        perc = 10  # Used for the progression percentage

        # If the path is a folder, we load every single file
        if tool_functions.os.path.isdir(self.path):

            for i in range(len(self.files)):

                if verbose > 1:
                    print("Loading file " + str(i) + " of " + str(len(self.files)) + ":" + self.path + "...", end=" ")

                # Show percentage if verbose
                perc = tool_functions.show_percentage(verbose, i, len(self.files), perc)

                # Create the Pose object, passes as parameter the index and the file path
                self.load_single_file(i, self.path + "/" + self.files[i])

                if verbose > 1:
                    print("OK.")

            if verbose > 0:
                print("100% - Done.\n")

        # Otherwise, we load the one file
        else:
            self.load_global_file(verbose)

    def load_single_file(self, pose_number, path):

        values = None
        timestamp = None

        # JSON file
        if path.split(".")[-1] == "json":
            data = tool_functions.open_json(path)
            values = data["Bodies"][0]["Joints"]
            timestamp = data["Timestamp"]

        # Excel file
        elif path.split(".")[-1] == "xlsx":
            data, joints_labels = tool_functions.open_xlsx(path)

            values = OrderedDict()
            for i in range(len(joints_labels)):
                values[joints_labels[i]] = float(data.cell(1, i + 1).value)

            values, timestamp = tool_functions.table_to_json_joints(values)

        elif path.split(".")[-1] in ["txt", "csv"]:

            # Open the file and read the data
            separator = tool_functions.get_filetype_separator(path)
            data = tool_functions.open_txt(path)

            # Get the joints labels and values
            joints_labels = data[0].split(separator)
            elements = data[1].split(separator)

            values = OrderedDict()
            for i in range(len(joints_labels)):
                values[joints_labels[i]] = float(elements[i])

            values, timestamp = tool_functions.table_to_json_joints(values)

        # Create the Pose object, passes as parameter the index and the data
        self.read_pose(pose_number, values, timestamp)

    def load_global_file(self, verbose):
        """Loads a global file containing the whole sequence instead of single files for each pose."""

        perc = 10  # Used for the progression percentage

        # JSON file
        if self.path.split(".")[-1] == "json":

            # Open the file and read the data
            data = tool_functions.open_json(self.path)

            for p in range(len(data)):
                if verbose > 1:
                    print("Loading pose " + str(p) + " of " + str(len(data)) + "...", end=" ")

                # Show percentage if verbose
                perc = tool_functions.show_percentage(verbose, p, len(data), perc)

                # Create the Pose object, passes as parameter the index and the data
                self.read_pose(p, data[p]["Bodies"][0]["Joints"], data[p]["Timestamp"])

                if verbose > 1:
                    print("OK.")

        # Excel file
        elif self.path.split(".")[-1] == "xlsx":
            import openpyxl as op
            workbook = op.load_workbook(self.path)
            sheet = workbook[workbook.sheetnames[0]]

            # Get the labels (timestamp and joint labels) from the first row
            joints_labels = []
            for cell in sheet["1"]:
                joints_labels.append(str(cell.value))

            # For each pose
            for p in range(len(sheet["A"]) - 1):

                if verbose > 1:
                    print("Loading pose " + str(p) + " of " + str(len(sheet["A"])) + "...", end=" ")

                # Get the values (coordinates)
                values = OrderedDict()
                for i in range(len(joints_labels)):
                    values[joints_labels[i]] = float(sheet.cell(p + 2, i + 1).value)
                values, timestamp = tool_functions.table_to_json_joints(values)

                # Create the Pose object, passes as parameter the index and the data
                self.read_pose(p, values, timestamp)

                if verbose > 1:
                    print("OK.")

        # Text file (csv or txt)
        elif self.path.split(".")[-1] in ["csv", "txt"]:

            separator = tool_functions.get_filetype_separator(self.path)

            # Open the file and read the data
            data = tool_functions.open_txt(self.path)

            # Get the joints labels
            joints_labels = data[0].split(separator)

            # For each pose
            for line in range(len(data) - 1):

                if verbose > 1:
                    print("Loading pose " + str(line) + " of " + str(len(data)-1) + "...", end=" ")

                elements = data[line + 1].split(separator)
                values = OrderedDict()
                for i in range(len(joints_labels)):
                    values[joints_labels[i]] = float(elements[i])

                values, timestamp = tool_functions.table_to_json_joints(values)

                # Create the Pose object, passes as parameter the index and the data
                self.read_pose(line, values, timestamp)

                if verbose > 1:
                    print("OK.")

        if verbose:
            print("100% - Done.\n")

    def read_pose(self, pose_number, data, timestamp):
        """Reads a pose from a file, creates a pose object and adds it to the sequence"""
        pose = Pose(pose_number)
        for j in data:
            joint = Joint(j["JointType"], j["Position"]["X"], j["Position"]["Y"], j["Position"]["Z"])
            pose.add_joint(j["JointType"], joint)
        pose.set_timestamp(timestamp)
        self.poses.append(pose)

    # === Calculation functions ===

    def load_relative_timestamps(self, convert_to_seconds=True):
        """For all the poses, sets the relative time from the first pose, by subtracting the absolute UNIX time
        of the first pose from the absolute UNIX time of every pose."""
        t = self.poses[0].get_timestamp()
        for f in self.poses:
            f.calculate_relative_timestamp(t, convert_to_seconds)

    def calculate_velocity(self):
        """Calculates the distance travelled by each joint between two poses, and divides it by the time elapsed
        between the poses."""

        for p in range(1, len(self.poses)):  # For each pose

            for j in self.poses[0].joints.keys():  # For each joint
                velocity = self.get_velocity(self.poses[p - 1], self.poses[p], j)  # Get the velocity
                self.poses[p].joints[j].set_velocity(velocity)  # Set the velocity

    def get_velocity(self, pose1, pose2, joint):
        """Given two poses and a joint, returns the velocity of the joint between the two poses. To do so, this
        function calculates the distance travelled by the joint between the two poses and divides it by the
        time elapsed between the two poses. The returned value is in cm/s."""

        # Get the distance travelled by a joint between two poses (meters)
        dist = self.get_distance(pose1.joints[joint], pose2.joints[joint])

        # Get the time elapsed between two poses (hundreds of seconds)
        delay = self.get_delay(pose1, pose2)

        # Calculate the velocity (meters per hundreds of seconds, or centimeters per second)
        velocity = dist / delay

        return velocity

    def get_total_velocity(self, joint):
        """Returns the sum of the velocity of a joint for the whole duration of the sequence."""
        total_velocity = 0

        for p in range(1, len(self.poses)):
            total_velocity += self.get_velocity(self.poses[p - 1], self.poses[p], joint)

        return total_velocity

    @staticmethod
    def get_distance(joint_a, joint_b):
        """Uses the Euclidian formula to calculate the distance between two joints. This can be used to calculate
        the distance travelled by one joint between two poses, or the distance between two joints on the same pose."""
        x = (joint_b.x - joint_a.x) ** 2
        y = (joint_b.y - joint_a.y) ** 2
        z = (joint_b.z - joint_a.z) ** 2

        return math.sqrt(x + y + z)

    @staticmethod
    def get_delay(frame_a, frame_b):
        """Returns the delay between two poses, in tenths of microsecond (10^-7)."""
        time_a = frame_a.get_relative_timestamp()
        time_b = frame_b.get_relative_timestamp()
        return time_b - time_a

    # === Correction functions ===

    def realign(self, velocity_threshold, window, verbose=1, name=None):
        """Realignment function: detects twitches and jumps in a sequence, corrects them linearly and outputs
        a new realigned sequence. For more information on how the realignment is performed, see the documentation."""

        # Create an empty sequence, the same length in poses as the original, just keeping the timestamp information
        # for each pose.
        new_sequence = self.create_new_sequence_with_original_timestamps()
        if name is None:
            new_sequence.name = self.name + " +RA"
        else:
            new_sequence.name = name

        # If verbose, show all the information. Else, show only the progress in percentage.
        if verbose > 0:
            print("Starting realignment...")

        # Define the counters
        realigned_points = 0
        perc = 10

        # For every pose starting on the second one
        for p in range(1, len(self.poses)):

            if verbose > 1:
                print("\nNew sequence:\n" + str(new_sequence.poses[p - 1]))
                print("\n== POSE NUMBER " + str(p + 1) + " ==")

            perc = tool_functions.show_percentage(verbose, p, len(self.poses), perc)

            # For every joint
            for j in self.poses[0].joints.keys():

                # We get the cm/s travelled by the joint between this pose and the previous
                velocity_before = self.get_velocity(new_sequence.poses[p - 1], self.poses[p], j)

                if verbose > 1:
                    print(j + ": " + str(velocity_before))

                # If we already corrected this joint, we ignore it to avoid overcorrection
                if j in new_sequence.poses[p].joints:

                    if verbose > 1:
                        print("\tAlready corrected.")

                # If the velocity is over threshold, we check if it is a jump or a twitch
                elif velocity_before > velocity_threshold and p != len(self.poses) - 1:

                    if verbose > 1:
                        print("\tVelocity over threshold. Check in subsequent poses...")

                    self.poses[p].joints[j].set_movement_over_threshold(True)

                    # We check the next poses (defined by the window) if the position of the joint comes back below
                    # threshold
                    for k in range(p, min(p + window, len(self.poses))):

                        if verbose > 1:
                            print("\t\tPose " + str(k + 1) + ":", end=" ")

                        # Method 1 (original): get the subsequent velocities
                        velocity = self.get_velocity(new_sequence.poses[p - 1], self.poses[k], j)

                        # Method 2 (to test): get the distance over the time between the two first joints
                        # dist = self.get_distance(new_sequence.poses[p - 1].joints[joint], self.poses[k].joints[joint])
                        # delay = self.get_delay(new_sequence.poses[p - 1], new_sequence.poses[p]) / 1000000000
                        # velocity = dist / delay

                        # Twitch case: One of the poses of the window is below threshold compared to previous pose.
                        if velocity < velocity_threshold:

                            if verbose > 1:
                                print("no threshold deviation compared to pose " + str(p + 1) + ".")
                                print("\t\t\tCorrecting for twitch...")

                            new_sequence, realigned_points = self.realign_window(new_sequence, p - 1, k,
                                                                                 j, realigned_points, verbose)

                            break

                        # Jump case: No pose of the window is below threshold.
                        elif k == p + window - 1 or k == len(self.poses) - 1:

                            if verbose > 1:
                                print("original deviation not found.")
                                print("\t\t\tCorrecting for jump...")

                            new_sequence, realigned_points = self.realign_window(new_sequence, p - 1, k,
                                                                                 j, realigned_points, verbose)

                            break

                        # Wait: if there are still poses in the window, we continue.
                        else:

                            if verbose > 1:
                                print("still over threshold.")

                    if verbose > 1:
                        print("")

                # If we're still in threshold, there is no correction, we copy the joint
                else:

                    joint = self.copy_joint(p, j)
                    new_sequence.poses[p].add_joint(j, joint)
                    new_sequence.poses[p].joints[j].corrected = False

        new_sequence.load_relative_timestamps(False)  # Sets the relative time from the first pose for each pose
        new_sequence.calculate_velocity()  # Sets the velocity for each joint of each pose

        if verbose > 0:
            print("100% - Done.\n")
            print("Realignment over. " + str(realigned_points) + " point(s) realigned over " + str(
                len(self.poses) * len(list(self.poses[0].joints.keys()))) + ".\n")

        return new_sequence

    def realign_window(self, new_sequence, start_pose_number, end_pose_number, joint_name, realigned_points, verbose):
        """Performs a jump or twitch correction"""

        # We extract the data from the joints at the beginning and at the end of the partial sequence to correct
        joint_before = new_sequence.poses[start_pose_number].joints[joint_name]
        joint_after = self.poses[end_pose_number].joints[joint_name]

        # If the starting and ending joint are the same, we don't correct, it means it is the last pose of the sequence
        if start_pose_number == end_pose_number:
            joint = self.copy_joint(start_pose_number, joint_name)
            new_sequence.poses[start_pose_number].add_joint(joint_name, joint)

            if verbose > 1:
                print("\t\t\t\tDid not corrected joint " + str(start_pose_number) +
                      " as it is the last pose of the sequence.")

        # Otherwise we correct for all the intermediate joints
        for pose_number in range(start_pose_number + 1, end_pose_number):

            # If a joint was already corrected we don't correct it to avoid overcorrection
            if self.poses[pose_number].joints[joint_name].corrected:

                if verbose > 1:
                    print("\t\t\t\tDid not corrected joint " + str(pose_number + 1) + " as it was already corrected.")

            # Otherwise we correct it
            else:

                # We get the new coordinates
                x, y, z = self.realign_joint(joint_before, joint_after, start_pose_number, pose_number, end_pose_number,
                                             verbose)

                # We copy the original joint, apply the new coordinates and add it to the new sequence
                joint = self.copy_joint(pose_number, joint_name)
                joint.realign_joint(x, y, z)
                new_sequence.poses[pose_number].add_joint(joint_name, joint)

                if verbose > 1:
                    print("\t\t\t\tCorrecting joint: " + str(pose_number + 1) + ". Original coordinates: (" + str(
                        self.poses[pose_number].joints[joint_name].x) +
                          ", " + str(self.poses[pose_number].joints[joint_name].y) + ", " + str(
                        self.poses[pose_number].joints[joint_name].z) + ")")
                    print("\t\t\t\tPrevious joint: " + str(start_pose_number + 1) + ". (" + str(
                        new_sequence.poses[start_pose_number].joints[joint_name].x) +
                          ", " + str(new_sequence.poses[start_pose_number].joints[joint_name].y) + ", " + str(
                        new_sequence.poses[start_pose_number].joints[joint_name].z) + ")")
                    print("\t\t\t\tNext joint: " + str(end_pose_number + 1) + ". (" + str(
                        self.poses[end_pose_number].joints[joint_name].x) +
                          ", " + str(self.poses[end_pose_number].joints[joint_name].y) + ", " + str(
                        self.poses[end_pose_number].joints[joint_name].z) + ")")
                    print("\t\t\t\tCorrected joint " + str(pose_number + 1) + ". New coordinates: (" + str(
                        x) + ", " + str(y) + ", " + str(z) + ")\n")

                # To count the number of joints that were realigned overall
                realigned_points += 1

        return new_sequence, realigned_points

    def realign_joint(self, joint_before, joint_after, pose_before, pose_current, pose_after, verbose):
        """Corrects one single joint (depending on the time between frames)"""

        # We calculate the percentage of time elapsed between the pose at the beginning of the partial sequence to
        # correct and the pose at the end
        percentage_time = (self.poses[pose_current].timestamp - self.poses[pose_before].timestamp) / (
                self.poses[pose_after].timestamp - self.poses[pose_before].timestamp)

        if verbose > 1:
            print("\t\t\t\tJoint " + str(pose_current + 1) + " positioned at " + str(
                percentage_time * 100) + " % between poses " + str(pose_before + 1) + " and " + str(
                pose_after + 1) + ".")

        # We correct linearly every joint. See the documentation for more precision.
        x = joint_before.x - percentage_time * (joint_before.x - joint_after.x)
        y = joint_before.y - percentage_time * (joint_before.y - joint_after.y)
        z = joint_before.z - percentage_time * (joint_before.z - joint_after.z)

        return x, y, z

    def re_reference(self, reference_joint="SpineMid", place_at_zero=True, verbose=1, name=None):
        """Subtracts the movement of one joint to the rest of the joints, to measure the movements from the
        joints independently of the global movements of the body in the room."""

        # Create an empty sequence, the same length in poses as the original, just keeping the timestamp information
        # for each pose.
        new_sequence = self.create_new_sequence_with_original_timestamps()
        if name is None:
            new_sequence.name = self.name + " +RF"
        else:
            new_sequence.name = name

        # If verbose, show all the information. Else, show only the progress in percentage.
        if verbose > 0:
            print("Starting normalization...")

        # Get reference : position of the reference joint at time 0
        if place_at_zero:
            start_ref_x = 0  # self.poses[0].joints[reference_joint].x
            start_ref_y = 0  # self.poses[0].joints[reference_joint].y
            start_ref_z = 0  # self.poses[0].joints[reference_joint].z
        else:
            start_ref_x = self.poses[0].joints[reference_joint].x
            start_ref_y = self.poses[0].joints[reference_joint].y
            start_ref_z = self.poses[0].joints[reference_joint].z

        # Define the percentage counter
        perc = 10

        # For every pose starting on the second one
        for p in range(len(self.poses)):

            if verbose > 1:
                print("\n== POSE NUMBER " + str(p + 1) + " ==")

            perc = tool_functions.show_percentage(verbose, p, len(self.poses), perc)

            # Get movement from the reference point
            curr_ref_x = self.poses[p].joints[reference_joint].x
            curr_ref_y = self.poses[p].joints[reference_joint].y
            curr_ref_z = self.poses[p].joints[reference_joint].z

            # Compute the differences
            diff_ref_x = curr_ref_x - start_ref_x
            diff_ref_y = curr_ref_y - start_ref_y
            diff_ref_z = curr_ref_z - start_ref_z

            if verbose > 1:
                print("Removing " + str(diff_ref_x) + " to every joint in x;")
                print("Removing " + str(diff_ref_y) + " to every joint in y;")
                print("Removing " + str(diff_ref_z) + " to every joint in z.")

            # For every joint
            for j in self.poses[0].joints.keys():

                # Compute new joint position
                if j == reference_joint:
                    new_x = 0
                    new_y = 0
                    new_z = 0
                else:
                    new_x = self.poses[p].joints[j].x - diff_ref_x
                    new_y = self.poses[p].joints[j].y - diff_ref_y
                    new_z = self.poses[p].joints[j].z - diff_ref_z

                # Add to the sequence
                joint = self.copy_joint(p, j)
                joint.realign_joint(new_x, new_y, new_z)
                new_sequence.poses[p].add_joint(j, joint)

                if verbose > 1:
                    print(j + ": ")
                    print("X: " + str(self.poses[p].joints[j].x) + " -> " + str(new_x))
                    print("Y: " + str(self.poses[p].joints[j].y) + " -> " + str(new_y))
                    print("Z: " + str(self.poses[p].joints[j].z) + " -> " + str(new_z))

        new_sequence.load_relative_timestamps(False)  # Sets the relative time from the first pose for each pose
        new_sequence.calculate_velocity()  # Sets the velocity for each joint of each pose

        if verbose > 0:
            print("100% - Done.\n")
            print("Normalization over.")
        return new_sequence

    def synchronize(self, delay_beginning, path_audio=None, trim=False, verbose=1, name=None):
        """Synchronizes the timestamps with an audio file"""

        if verbose > 0:
            if trim:
                print("Synchronizing the sequence with an audio file, and trimming...")
            else:
                print("Synchronizing the sequence with an audio file...")

        # Create an empty sequence
        new_sequence = Sequence(None)
        if name is None:
            if trim:
                new_sequence.name = self.name + " +TR"
            else:
                new_sequence.name = self.name + " +SY"
        else:
            new_sequence.name = name

        # Define the percentage counter
        perc = 10

        # Get the duration of the audio file
        if path_audio is not None:
            self.path_audio = path_audio
        sound = wave.open(self.path_audio, "r")
        sampling = sound.getframerate()
        audio_duration = sound.getnframes() / sampling

        if verbose > 0:
            print("Moving the beginning by a delay of "+str(delay_beginning)+".")
            print("Old duration: "+str(self.get_duration()))
            print("New duration: "+str(audio_duration))

        for p in range(len(self.poses)):
            perc = tool_functions.show_percentage(verbose, p, len(self.poses), perc)

            pose = self.copy_pose(p)
            pose.set_timestamp(pose.relative_timestamp + delay_beginning)
            if trim and 0 <= pose.timestamp <= audio_duration:
                new_sequence.poses.append(pose)
            elif not trim:
                if not 0 <= pose.timestamp <= audio_duration:
                    pose.to_trim = True
                new_sequence.poses.append(pose)

        new_sequence.load_relative_timestamps(False)  # Sets the relative time from the first pose for each pose
        new_sequence.calculate_velocity()  # Sets the velocity for each joint of each pose

        if verbose > 0:
            print("100% - Done.")

        return new_sequence

    def resample(self, frequency, mode="cubic", verbose=1, name=None):
        """Resamples the sequence to a fixed frequency defined by the user.
        Module to perform the resampling can be Scipy or Resampy."""

        if verbose > 0:
            print("Original average framerate: "+str(self.get_average_framerate()))
            print("Original min framerate: " + str(self.get_min_framerate()))
            print("Original max framerate: " + str(self.get_max_framerate()))
            print("New framerate: "+str(frequency))
            print("\nResampling the sequence: creating vectors...")

        # Create an empty sequence
        new_sequence = Sequence(None)
        if name is None:
            new_sequence.name = self.name + " +RS" + str(frequency)
        else:
            new_sequence.name = name

        # Define the percentage counter
        perc = 10

        # Define positions lists
        x_points = OrderedDict()
        y_points = OrderedDict()
        z_points = OrderedDict()
        time_points = []

        # Create vectors of position and time
        for p in range(len(self.poses)):

            perc = tool_functions.show_percentage(verbose, p, len(self.poses), perc)

            for j in self.poses[p].joints.keys():
                if p == 0:
                    x_points[j] = []
                    y_points[j] = []
                    z_points[j] = []
                x_points[j].append(self.poses[p].joints[j].x)
                y_points[j].append(self.poses[p].joints[j].y)
                z_points[j].append(self.poses[p].joints[j].z)
            time_points.append(self.poses[p].relative_timestamp)

        if verbose > 0:
            print("100% - Done.")
            print("Resampling the sequence: resampling...")

        # Define the percentage counter
        perc = 10

        # Resample
        new_x_points = OrderedDict()
        new_y_points = OrderedDict()
        new_z_points = OrderedDict()
        new_time_points = []

        no_joints = len(x_points.keys())
        i = 0

        for joint in x_points.keys():
            perc = tool_functions.show_percentage(verbose, i, no_joints, perc)
            new_x_points[joint], new_time_points = self.resample_data(x_points[joint], time_points, frequency, mode)
            new_y_points[joint], new_time_points = self.resample_data(y_points[joint], time_points, frequency, mode)
            new_z_points[joint], new_time_points = self.resample_data(z_points[joint], time_points, frequency, mode)
            i += 1

        # Define the percentage counter
        if verbose > 0:
            print("100% - Done.")
            print("Resampling the sequence: saving the new sequence...")
        perc = 10

        # Save data
        for p in range(len(new_time_points)):
            perc = tool_functions.show_percentage(verbose, i, no_joints, perc)
            pose = Pose(p)
            for j in new_x_points.keys():
                x = new_x_points[j][p]
                y = new_y_points[j][p]
                z = new_z_points[j][p]
                joint = Joint(j, x, y, z)
                joint.joint_type = j
                pose.add_joint(j, joint)
            pose.timestamp = new_time_points[p]
            new_sequence.poses.append(pose)

        if verbose > 0:
            print("100% - Done.")
            print("Original sequence had "+str(len(self.poses))+" poses.")
            print("New sequence has " + str(len(new_sequence.poses)) + " poses.")

        new_sequence.load_relative_timestamps(False)  # Sets the relative time from the first pose for each pose
        new_sequence.calculate_velocity()  # Sets the velocity for each joint of each pose

        return new_sequence

    @staticmethod
    def resample_data(data, time_points, frequency, mode="linear"):
        """Resamples non-uniform data to a uniform time series according to a specific frequency."""
        np_data = np.array(data)
        np_time_points = np.array(time_points)
        interp = interpolate.interp1d(np_time_points, np_data, kind=mode)

        resampled_time_points = np.arange(0, max(time_points), 1/frequency)
        resampled_data = interp(resampled_time_points)
        return resampled_data, resampled_time_points

    # === Copy and blank sequence creation functions ===

    def create_new_sequence_with_original_timestamps(self, verbose=1):
        """Creates an empty sequence with empty poses and empty joints, with the same number of poses as
        the sequence in which it is created. The only data kept for the poses is the timestamp and relative
        timestamp."""

        if verbose > 0:
            print("Creating an empty sequence...")
        new_sequence = Sequence(None)

        # Used for the progression percentage
        perc = 10
        for p in range(len(self.poses)):

            # Show percentage if verbose
            perc = tool_functions.show_percentage(verbose, p, len(self.poses), perc)

            pose = self.poses[p].get_clear_copy()  # Creates an empty pose with the same timestamp as the original
            new_sequence.poses.append(pose)  # Add to the sequence

        new_sequence = self.add_first_pose(new_sequence)

        if verbose > 0:
            print("100% - Done.\n")

        return new_sequence

    def add_first_pose(self, sequence):
        """Copies the joints from the first pose that doesn't already exist in the target sequence"""

        for i in self.poses[0].joints.keys():  # For every joint

            # If this joint doesn't already exist in the new sequence, we copy it
            if i not in sequence.poses[0].joints.keys():
                sequence.poses[0].joints[i] = self.copy_joint(0, i)

        return sequence

    def copy_pose(self, pose_number):
        """Copies the pose"""

        pose = Pose(pose_number)

        for i in self.poses[pose_number].joints.keys():  # For every joint

            self.poses[pose_number].joints[i] = self.copy_joint(pose_number, i)

        pose.set_timestamp(self.poses[pose_number].get_timestamp())
        pose.relative_timestamp = self.poses[pose_number].get_relative_timestamp()

        return pose

    def copy_joint(self, pose_number, joint_name):
        """Creates a copy of a single joint and returns it"""

        joint = self.poses[pose_number].get_joint(joint_name).copy()

        return joint

    # === Getter functions ===

    def get_timestamps(self):
        """Returns a list of all the timestamps (relative) from the poses"""

        timestamps = []
        for pose in self.poses:
            timestamps.append(pose.get_relative_timestamp())

        return timestamps

    def get_time_between_two_poses(self, pose1, pose2):
        """Returns the time delay between two poses."""
        timestamp1 = self.poses[pose1].relative_timestamp
        timestamp2 = self.poses[pose2].relative_timestamp
        if timestamp1 > timestamp2:
            return timestamp1 - timestamp2
        else:
            return timestamp2 - timestamp1

    def get_duration(self):
        """Returns the duration of the sequence, in seconds."""
        return self.poses[-1].relative_timestamp

    def get_table(self, include_relative_timestamp=False):
        """Returns a list of lists containing a header and the x, y and z coordinates of the joints for all the
        poses of the sequence"""

        table = []

        # For each pose
        for p in range(len(self.poses)):
            if p == 0:
                table = self.poses[0].get_table(include_relative_timestamp)
            else:
                table.append(self.poses[p].get_table(include_relative_timestamp)[1])

        return table

    def get_json_joint_list(self, include_relative_timestamp=False):
        """Returns a json format containing the timestamps and the x, y ans z coordinates of the joints for all
        the poses of the sequence"""

        data = []

        # For each pose
        for p in range(len(self.poses)):
            data.append(self.poses[p].get_json_joint_list(include_relative_timestamp))

        return data

    def get_framerate(self):
        """Returns the framerate across time for the whole sequence."""
        time_points = []
        framerates = []
        for p in range(1, len(self.poses) - 1):
            time_points.append(self.poses[p].relative_timestamp)
            framerate = 1 / (self.poses[p].relative_timestamp - self.poses[p - 1].relative_timestamp)
            framerates.append(framerate)

        return framerates, time_points

    def get_average_framerate(self):
        """Returns the average framerate of the sequence."""
        framerates, time_points = self.get_framerate()
        average = sum(framerates)/len(framerates)
        return average

    def get_min_framerate(self):
        """Returns the min framerate of the sequence."""
        framerates, time_points = self.get_framerate()
        return min(framerates)

    def get_max_framerate(self):
        """Returns the min framerate of the sequence."""
        framerates, time_points = self.get_framerate()
        return max(framerates)

    def get_stats(self, tabled=False):
        """Returns some statistics and information over the sequence."""
        stats = OrderedDict()
        stats["Path"] = self.path
        stats["Duration"] = self.poses[-1].relative_timestamp
        stats["Number of poses"] = len(self.poses)

        # Framerate stats
        framerates, time_points = self.get_framerate()
        stats["Average framerate"] = sum(framerates) / len(framerates)
        stats["SD framerate"] = statistics.stdev(framerates)
        stats["Max framerate"] = max(framerates)
        stats["Min framerate"] = min(framerates)

        # Movement stats
        for joint in self.poses[0].joints.keys():
            stats["Total velocity " + joint] = self.get_total_velocity(joint)

        if tabled:
            table = [[], []]
            for key in stats.keys():
                table[0].append(key)
                table[1].append(stats[key])
            return table

        return stats

    # === Print functions ===

    def print_json(self, pose):
        """Print the original json data from a pose"""
        data = self.poses[pose].get_json_data()
        print(tool_functions.json.dumps(data, indent=4))

    def print_pose(self, pose):
        """Prints all the information relative to one pose"""
        txt = "Pose " + str(pose + 1) + " of " + str(len(self.poses)) + "\n"
        txt += str(self.poses[pose])
        print(txt)

    def print_stats(self):
        """Prints the statistics and information over the sequence."""
        stats = self.get_stats()
        for key in stats.keys():
            print(str(key) + ": " + str(stats[key]))

    # === Randomization functions ===

    @staticmethod
    def generate_random_joints():
        """Creates uniform, random positions for the joints"""
        random_joints = []
        for i in range(21):
            x = random.uniform(-0.2, 0.2)
            y = random.uniform(-0.5, 0.5)
            z = random.uniform(-0.5, 0.5)
            j = Joint(None, [x, y, z])
            random_joints.append(j)
        return random_joints

    def randomize(self):
        """Randomizes the joints"""

        # Get the starting positions of all joints
        starting_positions = []
        for joint in self.poses[0].joints.keys():
            starting_positions.append(copy.copy(self.poses[0].joints[joint]))

        # Randomize new starting positions for joints
        randomized_positions = self.generate_random_joints()
        random.shuffle(randomized_positions)

        # Assign these new starting positions to the joints and moves the joints in every subsequent pose
        # in relation to the new starting positions
        joints_list = list(self.poses[0].joints.keys())
        for p in self.poses:
            for j in range(len(joints_list)):
                p.joints[joints_list[j]].move_joint_randomized(starting_positions[j], randomized_positions[j])
        self.randomized = True

    # === Magic methods ===
    def __len__(self):
        """Return the amount of poses in the sequence"""
        return len(self.poses)

    def __repr__(self):
        return self.name
