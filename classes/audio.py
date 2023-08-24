"""Default class for audio recordings matching a Sequence, typically the voice of the subject of the motion capture.
This class allows to perform a variety of transformations of the audio stream, such as getting the envelope, pitch and
formants of the speech.
"""

from classes.audio_derivatives import *
from classes.exceptions import *
from tool_functions import *


class Audio(object):
    """Default class for audio clips matching a Sequence, typically the voice of the subject of the motion
    capture. This class allows to perform a variety of transformations of the audio stream, such as getting the
    envelope, pitch and formants of the speech.

    .. versionadded:: 2.0

    Parameters
    ----------
    path_or_samples: str or list(int)
        The path to the audio file, or a list containing the samples of an audio file. If the file is a path, it should
        either point to a `.wav` file, or to a file containing the timestamps and samples in a text form (`.json`,
        `.csv`, `.tsv`, `.txt` or `.mat`). It is also possible to point to a folder containing one file per sample.
        See :ref:`Audio formats <wav_example>` for the acceptable file types.

    frequency: int or float, optional
        The frequency, in Hz (or samples per sec) at which the parameter `path_or_samples` is set. This parameter
        will be ignored if ``path_or_samples`` is a path, but will be used to define the :attr:`timestamps` of the
        Audio object if ``path_or_samples`` is a list of samples.

    name: str, optional
        Defines a name for the Audio instance. If a string is provided, the attribute :attr:`name` will take
        its value. If not, see :meth:`Audio._define_name_init()`.

    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Attributes
    ----------
    samples: list(int)
        A list containing the audio samples, in chronological order.
    timestamps: list(float)
        A list containing the timestamps matching each audio sample. Consequently, :attr:`samples` and
        :attr:`timestamps` should have the same length.
    frequency: int or float
        The amount of samples per second.
    name: str
        Custom name given to the audio. If no name has been provided upon initialisation, it will be defined by
        :meth:`Audio._define_name_init()`.
    path: str
        Path to the audio file passed as a parameter upon creation; if samples were provided, this attribute will be
        `None`.
    files: list(str)
        List of files contained in the path. The list will be of size 1 if the path points to a single file.
    """

    def __init__(self, path_or_samples, frequency=None, name=None, verbosity=1):
        self.samples = []
        self.timestamps = []
        self.frequency = None
        self.name = None  # Placeholder for the name of the sequence
        self.path = None
        self.files = None

        if type(path_or_samples) is str:
            self._load_from_path(path_or_samples, verbosity)
        elif type(path_or_samples) is list:
            self._load_from_samples(path_or_samples, frequency, verbosity)
        else:
            raise Exception("Invalid type for the argument path_or_samples: should be str or list.")

        self._define_name_init(name, verbosity)

        # self.max_sample = max(self.samples)
        # self.min_sample = min(self.samples)
        #
        # self.envelope = None
        # self.max_sample_envelope = None
        # self.min_sample_envelope = None
        # self.get_envelope()

    # === Setter functions ===
    def set_name(self, name):
        """Sets the :attr:`name` attribute of the Audio instance. This name can be used as display functions or as
        a means to identify the audio.

        .. versionadded:: 2.0

        Parameters
        ----------
        name : str
            A name to describe the audio clip.

        Example
        -------
        >>> aud = Audio("C:/Users/Walter/Sequences/audio.wav")
        >>> aud.set_name("Audio 28980")
        """

        self.name = name

    # === Loading functions ===
    def _define_name_init(self, name, verbosity=1):
        """Sets the name attribute for an instance of the Audio class, using the name provided during the
        initialization, or the path. If no ``name`` is provided, the function will create the name based on the
        :attr:`path` attribute, by defining the name as the last element of the path hierarchy (last subfolder, or file
        name). For example, if ``path`` is ``"C:/Users/Bender/Documents/Recording001/"``, the function will define the
        name on ``"Recording001"``. If both ``name`` and ``path`` are set on ``None``, the sequence name will be defined
        as ``"Unnamed audio"``.

        .. versionadded:: 2.0

        Parameters
        ----------
        name: str
            The name passed as parameter in :meth:`Audio.__init__()`
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        if name is not None:
            self.name = name
            if verbosity > 1:
                print("The provided name " + str(name) + " will be attributed to the audio.")

        elif self.path is not None:
            if len(self.path.split("/")) >= 1:
                self.name = self.path.split("/")[-1]
            else:
                self.name = str(self.path)
            if verbosity > 1:
                print(
                    "No name was provided. Instead, the name " + str(self.name) + " was attributed by extracting it " +
                    "from the provided path.")

        else:
            self.name = "Unnamed audio"
            if verbosity > 1:
                print("No name nor path was provided. The placeholder name " + str(
                    self.name) + " was attributed to the " +
                      "audio.")

    def _load_from_path(self, path, verbosity=1):
        """Loads the audio data from the :attr:`path` provided during the initialization.

        .. versionadded:: 2.0

        Parameters
        ----------
        path: str
            Path to the audio file passed as a parameter upon creation.
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        if verbosity > 0:
            print("Parameter path_or_samples detected as being a path.")

        self.path = path

        # If it's a folder, we fetch all the files
        if not os.path.exists(self.path):
            if "." in self.path.split("/")[-1]:
                raise InvalidPathException(self.path, "audio clip", "The file doesn't exist.")
            else:
                raise InvalidPathException(self.path, "audio clip", "The folder doesn't exist.")

        # If it's a folder, we fetch all the files
        if os.path.isdir(self.path):
            self._fetch_files_from_folder(verbosity)  # Fetches all the files
        self._load_samples()  # Loads the files into poses

        if len(self.samples) == 0:
            raise EmptyAudioException()

    def _load_from_samples(self, samples, frequency, verbosity=1):
        """Loads the audio data when samples and frequency have been provided upon initialisation.

        .. versionadded:: 2.0

        Parameters
        ----------
        samples: list(int)
            A list containing the audio samples, in chronological order.
        frequency: int or float
            The amount of samples per second.
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        if verbosity > 0:
            print("Parameter path_or_samples detected as being samples.")

        self.samples = samples
        if frequency is not None:
            self.frequency = frequency
        else:
            raise Exception("If samples are provided, frequency cannot be None. Please provide a value for " +
                            "frequency.")

        self._calculate_timestamps()

    def _fetch_files_from_folder(self, verbosity=1):
        """Finds all the files ending with the accepted extensions (``.csv``, ``.json``, ``.tsv``, ``.txt``, or
        ``.xlsx``) in the folder defined by path, and orders the files according to their name.

        .. versionadded:: 2.0

        Note
        ----
        This functions ignores the elements of the directory defined by :attr:`path` if:
            •	They don’t have an extension
            •	They are a folder
            •	Their extension is not one of the accepted ones (``.csv``, ``.json``, ``.tsv``, ``.txt``, or
                ``.xlsx``)
            •	The file name does not contain an underscore (``_``)

        If a file has a valid extension, the function tries to detect an underscore (``_``) in the name. The file
        names should be ``xxxxxx_0.ext``, where ``xxxxxx`` can be any series of characters, ``0`` must be the index of
        the sample (with or without leading zeros), and ``ext`` must be an accepted extension (``.csv``, ``.json``,
        ``.tsv``, ``.txt``, or ``.xlsx``). The first pose of the sequence must have the index 0. If the file does not
        have an underscore in the name, it is ignored. The indices must be coherent with the chronological order of the
        timestamps.

        The function uses the number after the underscore to order the samples. This is due to differences in how file
        systems handle numbers without leading zeros: some place ``sample_11.json`` alphabetically before
        ``sample_2.json`` (1 comes before 2), while some other systems place it after as 11 is greater than 2. In order
        to avoid these, the function converts the number after the underscore into an integer to place it properly
        according to its index.

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """
        if verbosity > 0:
            print("Fetching sample files...", end=" ")

        file_list = os.listdir(self.path)  # List all the files in the folder
        self.files = ["" for _ in range(len(file_list))]  # Create an empty list the length of the files

        extensions_found = []  # Save the valid extensions found in the folder

        # Here, we find all the files that are either .json or .meta in the folder.
        for f in file_list:

            # If a file has an accepted extension, we get its index from its name to order it correctly in the list.
            # This is necessary as some file systems will order sample_2.json after sample_11.json because,
            # alphabetically, "1" is before "2".

            # We check if the element has a dot, i.e. if it is a file. If not, it is a folder, and we ignore it.
            if "." in f:

                extension = f.split(".")[-1]  # We get the file extension
                if extension in ["json", "csv", "tsv", "txt", "xlsx"]:

                    if verbosity > 1:
                        print("\tAdding file: " + str(f))

                    if "_" in f:
                        self.files[int(f.split(".")[0].split("_")[1])] = f
                        if extension not in extensions_found:
                            extensions_found.append(extension)
                    else:
                        if verbosity > 1:
                            print("""\tIgnoring file (no "_" detected in the name): """ + str(f))

                    if len(extensions_found) > 1:
                        raise InvalidPathException(self.path, "audio clip",
                                                   "More than one of the accepted extensions has been found in the " +
                                                   "provided folder: " + str(extensions_found[0]) + " and " +
                                                   str(extensions_found[1]))

                else:
                    if verbosity > 1:
                        print("\tIgnoring file (not an accepted filetype): " + str(f))

            else:
                if verbosity > 1:
                    print("\tIgnoring folder: " + str(f))

        # If files that weren't of the accepted extension are in the folder, then "self.files" is larger than
        # the number of samples. The list is thus ending by a series of empty strings that we trim.
        if "" in self.files:
            limit = self.files.index("")
            self.files = self.files[0:limit]

        for i in range(len(self.files)):
            if self.files[i] == "":
                raise InvalidPathException(self.path, "audio clip", "At least one of the files is missing (index " +
                                           str(i) + ").")

        print(str(len(self.files)) + " sample file(s) found.")

    def _load_samples(self, verbosity=1):
        """Loads the single sample files or the global file containing all the samples. Depending on the input, this
        function calls either :meth:`Audio._load_single_sample_file` or :meth:`Audio._load_audio_file`.

        .. versionadded:: 2.0

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        if verbosity > 0:
            print("Opening audio from " + self.path + "...", end=" ")

        perc = 10  # Used for the progression percentage

        # If the path is a folder, we load every single file
        if os.path.isdir(self.path):

            for i in range(len(self.files)):

                if verbosity > 1:
                    print("Loading file " + str(i) + " of " + str(len(self.files)) + ":" + self.path + "...", end=" ")

                # Show percentage if verbosity
                perc = show_progression(verbosity, i, len(self.files), perc)

                # Loads a file containing one timestamp and one sample
                self._load_single_sample_file(self.path + "/" + self.files[i])

                if verbosity > 1:
                    print("OK.")

            self._calculate_frequency()

            if verbosity > 0:
                print("100% - Done.")

        # Otherwise, we load the one file
        else:
            self._load_audio_file(verbosity)

    def _load_single_sample_file(self, path):
        """Loads the content of a single sample file into the Audio object. Depending on the file type, this function
        handles the content differently (see :ref:`Audio formats <wav_example>`).

        .. versionadded:: 2.0

        Parameters
        ----------
        path: str
            The path of a file containing a single sample and timestamp.
        """

        # JSON file
        if path.split(".")[-1] == "json":
            data = read_json(path)
            self.samples.append(data["Sample"])
            self.timestamps.append(data["Timestamp"])

        # Excel file
        elif path.split(".")[-1] == "xlsx":
            import openpyxl as op
            workbook = op.load_workbook(path)
            sheet = workbook[workbook.sheetnames[0]]

            self.timestamps.append(float(sheet.cell(2, 1).value))
            self.samples.append(float(sheet.cell(2, 2).value))

        # Text file
        elif path.split(".")[-1] in ["txt", "csv", "tsv"]:

            separator = get_filetype_separator(path.split(".")[-1])

            # Open the file and read the data
            data = read_text_table(path)

            for s in range(1, len(data)):
                d = data[s].split(separator)
                self.timestamps.append(float(d[0][s]))
                self.samples.append(int(d[1][s]))

        else:
            raise InvalidPathException(self.path, "audio clip", "Invalid file extension: " + path.split(".")[-1] + ".")

    def _load_audio_file(self, verbosity=1):
        """Loads the content of a file containing all the samples of the audio stream.
        Depending on the file type, this function handles the content differently (see
        :ref:`Audio formats <wav_example>`).

        .. versionadded:: 2.0

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        self.name = self.path

        if self.path.split(".")[-1] == "wav":
            self._read_wav(verbosity)
        elif self.path.split(".")[-1] in ["xlsx", "json", "csv", "txt"]:
            self._read_text_file(verbosity)
        else:
            raise InvalidPathException(self.path, "audio clip", "Invalid file extension: " + self.path.split(".")[-1] +
                                       ".")

    def _read_wav(self, verbosity=1):
        """Opens a ``.wav`` file using `scipy.io.wavfile.read
        <https://docs.scipy.org/doc/scipy/reference/generated/scipy.io.wavfile.read.html>`_, and loads the attributes
        :attr:`samples` and :attr:`frequency`. If the wav file has more than one channel, it is converted to mono by
        averaging the values from all the samples via :func:`tool_functions.stereo_to_mono`.

        .. versionadded:: 2.0

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        if verbosity > 0:
            print("\n\tOpening the audio...", end=" ")

        try:
            from scipy.io import wavfile
        except ImportError:
            raise ModuleNotFoundException("scipy", "open a file in .wav format.")

        audio_data = wavfile.read(self.path)
        self.frequency = audio_data[0]

        # Turn stereo to mono if the file is stereo
        if audio_data[1].shape[1] == 2:
            self.samples = stereo_to_mono(audio_data[1])
        else:
            self.samples = audio_data[1]

        self._calculate_timestamps()

        if verbosity > 0:
            print("Audio loaded.\n")

    def _read_text_file(self, verbosity=1):
        """Opens a file containing the samples in ``.json``, ``.xlsx``, ``.csv``, ``.tsv`` or ``.txt`` file containing
        the timestamps and samples of the audio and loads the attributes :attr:`samples`, :attr:`timestamps`, and
        :attr:`frequency`. Depending on the file type, this function handles the content differently (see
        :ref:`Audio formats <wav_example>`).

        .. versionadded:: 2.0

        Parameters
        ----------
        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        """

        if verbosity > 0:
            print("Opening the audio...", end=" ")

        if self.path.split(".")[-1] == "json":
            data = read_json(self.path)
            self.samples = data["Sample"]
            self.timestamp = data["Timestamp"]
            self._calculate_frequency()

        elif self.path.split(".")[-1] == "xlsx":
            import openpyxl as op
            workbook = op.load_workbook(self.path)
            sheet = workbook[workbook.sheetnames[0]]

            joints_labels = []
            for cell in sheet["1"]:
                joints_labels.append(str(cell.value))

            # For each pose
            for s in range(2, len(sheet["A"]) + 1):

                if verbosity > 1:
                    print("Loading sample " + str(s) + " of " + str(len(sheet["A"])) + "...", end=" ")

                self.timestamps.append(float(sheet.cell(s, 1).value))
                self.samples.append(int(sheet.cell(s, 2).value))

                if verbosity > 1:
                    print("OK.")

            self._calculate_frequency()

        elif self.path.split(".")[-1] in ["csv", "txt", "tsv"]:

            separator = get_filetype_separator(self.path.split(".")[-1])

            # Open the file and read the data
            data = read_text_table(self.path)

            for s in range(1, len(data)):

                d = data[s].split(separator)

                if verbosity > 1:
                    print("Loading sample " + str(s) + " of " + str(len(data) - 1) + "...", end=" ")

                self.timestamps.append(float(d[0][s]))
                self.samples.append(int(d[1][s]))

                self._calculate_frequency()

                if verbosity > 1:
                    print("OK.")

        if verbosity:
            print("100% - Done.")

    def _calculate_frequency(self):
        """Determines the frequency (number of samples per second) of the audio by calculating the time elapsed between
        the two first timestamps. This function is automatically called when reading the timestamps and samples from
        a text file.

        .. versionadded:: 2.0
        """
        self.frequency = 1 / (self.timestamps[1] - self.timestamps[0])

    def _calculate_timestamps(self):
        """Calculates the timestamps of the audio samples from the frequency and the number of samples. This function
        is automatically called when reading the frequency and samples from an audio file, or when a list of samples
        and a frequency are passed as a parameters during the initialisation.

        .. versionadded:: 2.0
        """
        self.timestamps = [i / self.frequency for i in range(len(self.samples))]

    # === Getter functions ===
    def get_path(self):
        """Returns the attribute :attr:`path` of the Audio instance.

        . versionadded:: 2.0

        Returns
        -------
        str
            The path of the Audio instance.
        """
        return self.path

    def get_name(self):
        """Returns the attribute :attr:`name` of the Audio instance.

        . versionadded:: 2.0

        Returns
        -------
        str
            The name of the Audio instance.
        """
        return self.name

    def get_samples(self):
        """Returns the attribute :attr:`samples` of the Audio instance.

        . versionadded:: 2.0

        Returns
        -------
        list(int)
            The samples of the Audio instance.
        """
        return self.samples

    def get_sample(self, sample_index):
        """Returns the sample corresponding to the index passed as parameter.

        .. versionadded:: 2.0

        Parameters
        ----------
        sample_index: int
            The index of the sample.

        Returns
        -------
        int
            A sample from the sequence.
        """
        if len(self.samples) == 0:
            raise Exception("The Audio does not have any sample.")
        elif not 0 <= sample_index < len(self.samples):
            raise Exception("The pose index must be between 0 and " + str(len(self.samples) - 1) + ". ")

        return self.samples[sample_index]

    def get_number_of_samples(self):
        """Returns the number of samples in the audio clip.

        .. versionadded:: 2.0

        Returns
        -------
        int
            The amount of samples in the audio clip.
        """
        return len(self.samples)

    def get_timestamps(self):
        """Returns a list of the timestamps for every sample, in seconds.

        .. versionadded:: 2.0

        Returns
        -------
        list(float)
            List of the timestamps of all the samples of the audio clip, in seconds.
        """
        return self.timestamps

    def get_duration(self):
        """Returns the duration of the audio clip, in seconds.

        .. versionadded:: 2.0

        Returns
        -------
        float
            The duration of the audio clip, in seconds.
        """
        return self.timestamps[-1]

    def get_frequency(self):
        """Returns the frequency of the audio clip, in hertz.

        .. versionadded:: 2.0

        Returns
        -------
        int or float
            The frequency of the audio clip, in hertz.
        """
        return self.frequency

    # === Transformation functions ===
    def get_envelope(self, filter_below=0, filter_over=50):
        """Calculates the envelope of the audio clip, applies a band-pass filter if values are provided,
        and returns an Envelope object.

        .. versionadded:: 2.0

        Parameters
        ----------
        filter_below: int or None, optional
            If not None nor 0, this value will be provided as the lowest frequency of the band-pass filter.
        filter_over: int or None, optional
            If not None nor 0, this value will be provided as the highest frequency of the high-pass filter.

        Returns
        -------
        Envelope
            The filtered envelope of the audio clip.
        """
        envelope = Envelope(self)
        envelope = envelope.filter_frequencies(filter_below, filter_over)
        return envelope

    def get_pitch(self, zeros_as_nan=False):
        """Calculates the pitch of the voice in the audio clip, and returns a Pitch object.

        .. versionadded:: 2.0

        Parameters
        ----------
        zeros_as_nan: bool, optional
            If set on True, the values where the pitch is equal to 0 will be replaced by
            `numpy.nan <https://numpy.org/doc/stable/reference/constants.html#numpy.nan>`_ objects.

        Returns
        -------
        Pitch
            The pitch of the voice in the audio clip.
        """
        return Pitch(self, zeros_as_nan=zeros_as_nan)

    def get_intensity(self):
        """Calculates the intensity of the voice in the audio clip, and returns an Intensity object.

        .. versionadded:: 2.0

        Returns
        -------
        Intensity
            The intensity of the voice in the audio clip.
        """
        return Intensity(self)

    def get_formant(self, formant=1):
        """Calculates the formants of the voice in the audio clip, and returns a Formant object.

        .. versionadded:: 2.0

        Parameters
        ----------
        formant: int, optional.
            One of the formants of the voice in the audio clip (1 (default), 2, 3, 4 or 5).

        Returns
        -------
        Formant
            The value of a formant of the voice in the audio clip.
        """
        return Formant(self, formant_number=formant)

    def resample(self, frequency, mode="cubic", name=None, verbosity=1):
        """Resamples an audio clip to the `frequency` parameter. It first creates a new set of timestamps at the
        desired frequency, and then interpolates the original data to the new timestamps.

        .. versionadded:: 2.0

        Parameters
        ----------
        frequency: float
            The frequency, in hertz, at which you want to resample the audio clip. A frequency of 4 will return samples
            at 0.25 s intervals.

        mode: str, optional
            This parameter also allows for all the values accepted for the ``kind`` parameter in the function
            :func:`scipy.interpolate.interp1d`: ``"linear"``, ``"nearest"``, ``"nearest-up"``, ``"zero"``,
            ``"slinear"``, ``"quadratic"``, ``"cubic"``”, ``"previous"``, and ``"next"``. See the `documentation
            for this Python module
            <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`_ for more.

        name: str or None, optional
            Defines the name of the output audio clip. If set on ``None``, the name will be the same as the input
            audio clip, with the suffix ``"+RS"``.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.

        Returns
        -------
        Audio
            A new audio clip containing resampled timestamps and samples.

        Warning
        -------
        This function allows both the **upsampling** and the **downsampling** of audio clips. However, during any of
        these operations, the algorithm only **estimates** the real values of the samples. You should then consider
        the upsampling (and the downsampling, to a lesser extent) with care.
        You can control the frequency of the original audio clip with :meth:`Audio.get_frequency()`.
        """

        if verbosity > 0:
            print("Resampling the audio clip at " + str(frequency) + " Hz (mode: " + str(mode) + ")...")
            print("\tOriginal frequency: " + str(round(self.get_frequency(), 2)))

        if verbosity > 0:
            print("\tPerforming the resampling...", end=" ")

        resampled_audio_array, resampled_audio_times = resample_data(self.samples, self.timestamps, frequency, mode)
        resampled_audio_array = list(resampled_audio_array)

        new_audio = Audio(resampled_audio_array, frequency, name)
        new_audio.path = self.path

        if verbosity > 0:
            print("100% - Done.")
            print("\tOriginal audio had " + str(len(self.samples)) + " samples.")
            print("\tNew audio has " + str(len(new_audio.samples)) + " samples.\n")

        return new_audio

    # === Conversion functions ===

    def convert_to_table(self):
        """Returns a list of lists where each sublist contains a timestamp and a sample. The first sublist contains the
        headers of the table. The output then resembles the table found in :ref:`Tabled formats <table_example>`.

        .. versionadded:: 2.0

        Returns
        -------
        list(list)
            A list of lists that can be interpreted as a table, containing headers, and with the timestamps and the
            coordinates of the joints from the sequence on each row.
        """

        table = [["Timestamp", "Sample"]]

        # For each pose
        for s in range(len(self.samples)):
            table.append([self.timestamps[s], self.samples[s]])

        return table

    def convert_to_json(self):
        """Returns a list ready to be exported in JSON. The returned JSON data is a dictionary with two keys:
        "Sample" is the key to the list of samples, while "Frequency" is the key to the sampling frequency of
        the audio clip.

        .. versionadded:: 2.0

        Returns
        -------
        dict
            A dictionary containing the data of the audio clip, ready to be exported in JSON.
        """

        data = {"Sample": self.samples, "Frequency": self.frequency}
        return data

    # === Saving functions ===

    def save(self, folder_out, name=None, file_format="xlsx", individual=False, verbosity=1):
        """Saves an audio clip in a file or a folder. The function saves the sequence under
        ``folder_out/name.file_format``. All the non-existent subfolders present in the ``folder_out`` path will be
        created by the function. The function also updates the :attr:`path` attribute of the Audio clip.

        .. versionadded:: 2.0

        Parameters
        ----------
        folder_out: str, optional
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them. If the string provided is empty (by default), the audio clip will be saved in
            the current working directory. If the string provided contains a file with an extension, the fields ``name``
            and ``file_format`` will be ignored.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on the attribute :attr:`name` of the audio clip; if that attribute is also set on ``None``, the name will be
            set on ``"out"``. If ``individual`` is set on ``True``, each sample will be saved as a different file,
            having the index of the pose as a suffix after the name (e.g. if the name is ``"sample"`` and the file
            format is ``"txt"``, the poses will be saved as ``sample_0.txt``, ``sample_1.txt``, ``sample_2.txt``, etc.).

        file_format: str or None, optional
            The file format in which to save the audio clip. The file format must be ``"json"`` (default), ``"xlsx"``,
            ``"txt"``, ``"csv"``, ``"tsv"``, ``"wav"``, or, if you are a masochist, ``"mat"``. Notes:

                • ``"xls"`` will save the file with an ``.xlsx`` extension.
                • Any string starting with a dot will be accepted (e.g. ``".csv"`` instead of ``"csv"``).
                • ``"csv;"`` will force the value separator on ``;``, while ``"csv,"`` will force the separator
                  on ``,``. By default, the function will detect which separator the system uses.
                • ``"txt"`` and ``"tsv"`` both separate the values by a tabulation.
                • Any other string will not return an error, but rather be used as a custom extension. The data will
                  be saved as in a text file (using tabulations as values separators).

            .. warning::
                While it is possible to save audio clips as ``.mat`` or custom extensions, the toolbox will not
                recognize these files upon opening. The support for ``.mat`` and custom extensions as input may come in
                a future release, but for now these are just offered as output options.

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.
            This option is not available and will be ignored if ``file_format`` is set on ``"wav"``.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        if folder_out == "":
            folder_out = os.getcwd()

        if not individual:
            subfolders = folder_out.split("/")
            if len(subfolders) != 0:
                if "." in subfolders[-1]:
                    folder_out = "/".join(subfolders[:-1])
                    name = ".".join(subfolders[-1].split(".")[:-1])
                    file_format = subfolders[-1].split(".")[-1]

        # Automatic creation of all the folders of the path if they don't exist
        create_subfolders(folder_out)

        if name is None and self.name is not None:
            name = self.name
        elif name is None:
            name = "out"

        file_format = file_format.strip(".")  # We remove the dot in the format
        if file_format == "xls":
            file_format = "xlsx"

        if verbosity > 0:
            if individual:
                print("Saving " + file_format.upper() + " individual files...")
            else:
                print("Saving " + file_format.upper() + " global file: " + folder_out.strip("/") + "/" + name + "." +
                      file_format + "...")

        if file_format == "json":
            self._save_json(folder_out, name, individual, verbosity)

        elif file_format == "mat":
            self._save_mat(folder_out, name, individual, verbosity)

        elif file_format == "xlsx":
            self._save_xlsx(folder_out, name, individual, verbosity)

        elif file_format == "wav":
            self._save_wav(folder_out, name)

        else:
            self._save_txt(folder_out, name, file_format, individual, verbosity)

        if individual:
            self.path = folder_out + name
        else:
            self.path = folder_out + name + file_format

        if verbosity > 0:
            print("100% - Done.")

    def _save_json(self, folder_out, name=None, individual=False, verbosity=1):
        """Saves an audio clip as a json file or files. This function is called by the :meth:`Audio.save`
        method, and saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"`` if individual is ``False``, or on ``"sample"`` if individual is ``True``.

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        perc = 10  # Used for the progression percentage

        # Save the data
        if not individual:
            data = self.convert_to_json()
            if name is None:
                name = "out"
            with open(folder_out + "/" + name + ".json", 'w', encoding="utf-16-le") as f:
                json.dump(data, f)
        else:
            if name is None:
                name = "pose"
            for s in range(len(self.samples)):
                perc = show_progression(verbosity, s, len(self.samples), perc)
                with open(folder_out + "/" + name + "_" + str(s) + ".json", 'w', encoding="utf-16-le") as f:
                    d = {"Sample": self.samples[s], "Timestamp": self.timestamps[s]}
                    json.dump(d, f)

    def _save_mat(self, folder_out, name=None, individual=False, verbosity=1):
        """Saves an audio clip as a Matlab .mat file or files. This function is called by the :meth:`Audio.save`
        method, and saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Important
        ---------
            This function is dependent of the module `scipy <https://scipy.org/>`_.

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"`` if individual is ``False``, or on ``"sample"`` if individual is ``True``.

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        perc = 10  # Used for the progression percentage

        try:
            from scipy.io import savemat
        except ImportError:
            raise ModuleNotFoundException("scipy", "save a file in .mat format.")

        # Save the data
        if not individual:
            if name is None:
                name = "out"
            data = self.convert_to_json()
            savemat(folder_out + "/" + name + ".mat", {"data": data})
        else:
            for s in range(len(self.samples)):
                if name is None:
                    name = "sample"
                perc = show_progression(verbosity, s, len(self.samples), perc)
                savemat(folder_out + "/" + name + "_" + str(s) + ".mat",
                        {"data": [["Timestamp", "Sample"], [self.timestamps[s], self.samples[s]]]})

    def _save_xlsx(self, folder_out, name=None, individual=False, verbosity=1):
        """Saves an audio clip as an Excel .xlsx file or files. This function is called by the :meth:`Audio.save`
        method, and saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Important
        ---------
            This function is dependent of the module `openpyxl <https://pypi.org/project/openpyxl/>`_.

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"`` if individual is ``False``, or on ``"sample"`` if individual is ``True``.

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        perc = 10  # Used for the progression percentage

        data = self.convert_to_table()

        # Save the data
        if not individual:
            if name is None:
                name = "out"
            write_xlsx(data, folder_out + "/" + name + ".xlsx", verbosity)

        else:
            for s in range(len(self.samples)):
                if name is None:
                    name = "sample"
                d = [data[0], data[s + 1]]
                perc = show_progression(verbosity, s, len(self.samples), perc)
                write_xlsx(d, folder_out + "/" + name + "_" + str(s) + ".xlsx", False)

    def _save_wav(self, folder_out, name=None):
        """Saves an audio clip as a .wav file or files. This function is called by the :meth:`Audio.save` method, and
        saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"``.
        """

        try:
            from scipy.io import wavfile
        except ImportError:
            raise ModuleNotFoundException("scipy", "save a file in .wav format.")

        try:
            import numpy as np
        except ImportError:
            raise ModuleNotFoundException("numpy", "save a file in .wav format.")

        if name is None:
            name = "out"
        wavfile.write(folder_out + "/" + name + ".wav", self.frequency, np.ndarray(self.samples))

    def _save_txt(self, folder_out, name=None, file_format="csv", individual=False, verbosity=1):
        """Saves an audio clip as a .txt, .csv, .tsv, or custom extension file or files. This function is called by the
        :meth:`Audio.save` method, and saves the Audio instance as ``folder_out/name.file_format``.

        .. versionadded:: 2.0

        Parameters
        ----------
        folder_out: str
            The path to the folder where to save the file or files. If one or more subfolders of the path do not exist,
            the function will create them.

        name: str or None, optional
            Defines the name of the file or files where to save the audio clip. If set on ``None``, the name will be set
            on ``"out"`` if individual is ``False``, or on ``"sample"`` if individual is ``True``.

        file_format: str, optional
            The file format in which to save the audio clip. The file format can be ``"txt"``, ``"csv"`` (default) or
            ``"tsv"``. ``"csv;"`` will force the value separator on ``";"``, while ``"csv,"`` will force the separator
            on ``","``. By default, the function will detect which separator the system uses. ``"txt"`` and ``"tsv"``
            both separate the values by a tabulation. Any other string will not return an error, but rather be used as
            a custom extension. The data will be saved as in a text file (using tabulations as values separators).

        individual: bool, optional
            If set on ``False`` (default), the function will save the audio clip in a unique file.
            If set on ``True``, the function will save each sample of the audio clip in an individual file, appending an
            underscore and the index of the sample (starting at 0) after the name.

            .. warning::
                It is not recommended to save each sample in a different file. This incredibly tedious way of handling
                audio files has only been implemented to follow the same logic as for the Sequence files, and should be
                avoided.

        verbosity: int, optional
            Sets how much feedback the code will provide in the console output:

            • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
            • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
              current steps.
            • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
              may clutter the output and slow down the execution.
        """

        perc = 10  # Used for the progression percentage

        # Force comma or semicolon separator
        if file_format == "csv,":
            separator = ","
            file_format = "csv"
        elif file_format == "csv;":
            separator = ";"
            file_format = "csv"
        elif file_format[0:3] == "csv":  # Get the separator from local user (to get , or ;)
            separator = get_system_csv_separator()
        elif file_format == "txt":  # For text files, tab separator
            separator = "\t"
        else:
            separator = "\t"

        # Save the data
        if not individual:
            if name is None:
                name = "out"
            write_text_table(self.convert_to_table(), separator, folder_out + "/" + name + "." + file_format, verbosity)
        else:
            for s in range(len(self.samples)):
                if name is None:
                    name = "pose"
                write_text_table(self.samples[s].convert_to_table(), separator,
                                 folder_out + "/" + name + "_" + str(s) + "." + file_format, 0)
                perc = show_progression(verbosity, s, len(self.samples), perc)

    def __len__(self):
        """Returns the number of samples in the audio clip (i.e., the length of the attribute :attr:`samples`).

        .. versionadded:: 2.0

        Returns
        -------
        int
            The number of samples in the audio clip.
        """
        return len(self.samples)

    def __getitem__(self, index):
        """Returns the sample of index specified by the parameter ``index``.

        Parameters
        ----------
        index: int
            The index of the sample to return.

        Returns
        -------
        float
            A sample from the attribute :attr:`samples`.
        """
        return self.samples[index]
