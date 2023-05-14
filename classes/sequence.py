#!/usr/bin/env python
"""Krajjat 2.0
Kinect Realignment Algorithm for Joint Jumps And Twitches
Author: Romain Pastureau
This file contains all the classes used by the other files,
along with the realignment algorithm.
"""

from classes.audio import *
from classes.pose import *
from classes.joint import *
from tool_functions import *

import statistics
from datetime import datetime, timedelta


class Sequence(object):
    """Default class for motion sequences in the toolbox. An instance of this class will be one motion sequence. The
    class contains several methods to perform pre-processing or displaying summarized data in text form (see public
    methods)."""

    def __init__(self, path=None, path_audio=None, name=None, time_unit="auto", start_timestamps_at_zero=False,
                 verbose=1):

        self.path_audio = path_audio  # Path to the audio file matched with this series of gestures

        self.name = None  # Placeholder for the name of the sequence
        self._define_name_init(name, path, verbose)  # Defines the name of the sequence

        self.files = None  # List of the files in the target folder
        self.poses = []  # List of the poses in the sequence, ordered
        self.randomized = False  # True if the joints positions are randomized
        self.date_recording = None  # Placeholder for the date of the recording.

        # In the case where a file or folder is passed as argument,
        # we load the sequence
        self.path = path  # Placeholder for the path

        if path is not None:
            self._load_from_path(time_unit)

        if start_timestamps_at_zero:
            self.set_first_timestamp(0)

    # === Name and setter functions ===
    def set_name(self, name):
        """Sets the name attribute of the Sequence instance."""
        self.name = name

    def set_path_audio(self, path_audio):
        """Sets the path_audio attribute of the Sequence instance."""
        self.path_audio = path_audio

    def set_first_timestamp(self, first_timestamp):
        """Attributes a new timestamp to the first pose of the sequence and delays the timestamps of the other poses
        accordingly."""
        timestamps = self.get_timestamps(relative=False)

        for p in range(len(self.poses)):
            self.poses[p].set_timestamp(timestamps[p] + first_timestamp)

    # === Loading functions ===
    def _define_name_init(self, name, path, verbose=1):
        """Sets the name attribute for an instance of the Sequence class, using the name provided during the
        initialization, or the path."""

        if name is not None:
            self.name = name
            if verbose > 1:
                print("The provided name " + str(name) + " will be attributed to the sequence.")

        elif path is not None:
            if len(path.split("/")) >= 1:
                self.name = path.split("/")[-1]
            else:
                self.name = str(path)
            if verbose > 1:
                print(
                    "No name was provided. Instead, the name " + str(self.name) + " was attributed by extracting it " +
                    "from the provided path.")

        else:
            self.name = "Unnamed sequence"
            if verbose > 1:
                print("No name nor path was provided. The placeholder name " + str(
                    self.name) + " was attributed to the " +
                      "sequence.")

    def _load_from_path(self, time_unit, verbose=1):
        """Loads the sequence data from the path provided during the initialization, and calculates the relative
        timestamps from the first pose for each pose, and velocities of each joint between each pose."""

        # If it's a folder, we fetch all the files
        if os.path.isdir(self.path):
            self._fetch_files_from_folder(verbose)  # Fetches all the files
        else:
            self.files = self.path
        self._load_poses(verbose)  # Loads the files into poses

        if len(self.poses) == 0:
            raise Exception("The path " + self.path + " is not a valid sequence.")

        self._calculate_relative_timestamps(time_unit)  # Sets the relative time from the first pose for each pose

    def _fetch_files_from_folder(self, verbose=1):
        """Finds all the files ending with the accepted extensions (.csv, .json, .tsv, .txt or .xlsx) in the folder
        defined by path, and orders the files according to their name."""

        if verbose > 0:
            print("Fetching sequence files...", end=" ")

        file_list = os.listdir(self.path)  # List all the files in the folder
        self.files = ["" for _ in range(len(file_list))]  # Create an empty list the length of the files

        extensions_found = []  # Save the valid extensions found in the folder

        # Here, we find all the files that are either .json or .meta in the folder.
        for f in file_list:

            # If a file has an accepted extension, we get its index from its name to order it correctly in the list.
            # This is necessary as some file systems will order frame_2.json after frame_11.json because,
            # alphabetically, "1" is before "2".

            # We check if the element has a dot, i.e. if it is a file. If not, it is a folder, and we ignore it.
            if "." in f:

                extension = f.split(".")[-1]  # We get the file extension
                if extension in ["json", "csv", "tsv", "txt", "xlsx"]:

                    if verbose > 1:
                        print("\tAdding file: " + str(f))

                    if "_" in f:
                        self.files[int(f.split(".")[0].split("_")[1])] = f
                        if extension not in extensions_found:
                            extensions_found.append(extension)
                    else:
                        if verbose > 1:
                            print("""\tIgnoring file (no "_" detected in the name): """ + str(f))

                    if len(extensions_found) > 1:
                        raise Exception("More than one of the accepted extensions has been found in the provided " +
                                        "folder: " + str(extensions_found[0]) + " and " + str(extensions_found[1]))

                else:
                    if verbose > 1:
                        print("\tIgnoring file (not an accepted filetype): " + str(f))

            else:
                if verbose > 1:
                    print("\tIgnoring folder: " + str(f))

        # If files that weren't of the accepted extension are in the folder, then "self.files" is larger than
        # the number of frames. The list is thus ending by a series of empty strings that we trim.
        if "" in self.files:
            limit = self.files.index("")
            self.files = self.files[0:limit]

        if verbose > 0:
            print(str(len(self.files)) + " pose file(s) found.")

    def _load_poses(self, verbose=1):
        """Loads all the single pose files or the global file containing all the poses."""

        if verbose > 0:
            print("Opening sequence from " + self.path + "...", end=" ")

        perc = 10  # Used for the progression percentage

        # If the path is a folder, we load every single file
        if os.path.isdir(self.path):

            for i in range(len(self.files)):

                if verbose > 1:
                    print("Loading file " + str(i) + " of " + str(len(self.files)) + ":" + self.path + "...", end=" ")

                # Show percentage if verbose
                perc = show_percentage(verbose, i, len(self.files), perc)

                # Create the Pose object, passes as parameter the index and the file path
                self._load_single_pose_file(i, self.path + "/" + self.files[i], verbose)

                if verbose > 1:
                    print("OK.")

            if verbose > 0:
                print("100% - Done.\n")

        # Otherwise, we load the one file
        else:
            self._load_sequence_file(verbose)

    def _load_single_pose_file(self, pose_number, path, verbose=1):
        """Loads the content of a single pose file into a Pose object."""

        file_extension = path.split(".")[-1]

        # JSON file
        if file_extension == "json":
            data = open_json(path)
            values = data["Bodies"][0]["Joints"]
            timestamp = data["Timestamp"]
            if pose_number == 0:
                self._load_date_recording(data, verbose)

        # Excel file
        elif file_extension == "xlsx":
            data, joints_labels = open_xlsx(path)

            values = OrderedDict()
            for i in range(len(joints_labels)):
                values[joints_labels[i]] = float(data.cell(1, i + 1).value)

            values, timestamp = table_to_dict_joints(values)

        elif file_extension in ["txt", "csv", "tsv"]:

            # Open the file and read the data
            separator = get_filetype_separator(path)
            data = open_txt(path)

            # If the data is exported from QTM (Qualisys), convert the tsv to manageable data
            if file_extension == "tsv":
                if data[0].split("\t")[0] == "NO_OF_FRAMES":
                    data = convert_data_from_qtm(data)

            # Get the joints labels and values
            joints_labels = data[0].split(separator)
            elements = data[1].split(separator)

            values = OrderedDict()
            for i in range(len(joints_labels)):
                values[joints_labels[i]] = float(elements[i])

            values, timestamp = table_to_dict_joints(values)

        else:
            raise Exception("Invalid file extension: " + file_extension + ".")

        # Create the Pose object, passes as parameter the index and the data
        self._create_pose(pose_number, values, timestamp)

    def _load_sequence_file(self, verbose=1):
        """Loads the content of a global sequence file containing individual poses into Pose objects."""

        perc = 10  # Used for the progression percentage

        # JSON file
        if self.path.split(".")[-1] == "json":

            # Open the file and read the data
            data = open_json(self.path)

            self._load_date_recording(data, verbose)

            for p in range(len(data)):
                if verbose > 1:
                    print("Loading pose " + str(p) + " of " + str(len(data)) + "...", end=" ")

                # Show percentage if verbose
                perc = show_percentage(verbose, p, len(data), perc)

                # Create the Pose object, passes as parameter the index and the data
                self._create_pose(p, data[p]["Bodies"][0]["Joints"], data[p]["Timestamp"])

                if verbose > 1:
                    print("OK.")

        # Excel file
        elif self.path.split(".")[-1] == "xlsx":
            try:
                import openpyxl as op
            except ImportError:
                raise Exception("Module openpyxl not found; please install it to be able to open xlsx files.")
            workbook = op.load_workbook(self.path)
            sheet = workbook[workbook.sheetnames[0]]

            # Get the labels (timestamp and joint labels) from the first row
            joints_labels = []
            for cell in sheet["1"]:
                joints_labels.append(str(cell.value))

            # For each pose
            for p in range(len(sheet["A"]) - 1):

                if verbose > 1:
                    print("Loading pose " + str(p) + " of " + str(len(sheet["A"])) + "...", end=" ")

                # Get the values (coordinates)
                values = OrderedDict()
                for i in range(len(joints_labels)):
                    values[joints_labels[i]] = float(sheet.cell(p + 2, i + 1).value)
                values, timestamp = table_to_dict_joints(values)

                # Create the Pose object, passes as parameter the index and the data
                self._create_pose(p, values, timestamp)

                if verbose > 1:
                    print("OK.")

        # Text file (csv or txt)
        elif self.path.split(".")[-1] in ["csv", "tsv", "txt"]:

            separator = get_filetype_separator(self.path)

            # Open the file and read the data
            data = open_txt(self.path)

            # If the data is exported from QTM (Qualisys), convert the tsv to manageable data
            if self.path.split(".")[-1] == "tsv":
                if data[0].split("\t")[0] == "NO_OF_FRAMES":
                    self._load_date_recording(data, verbose)
                    data = convert_data_from_qtm(data)

            # Get the joints labels
            joints_labels = data[0].split(separator)

            # For each pose
            for line in range(len(data) - 1):

                if verbose > 1:
                    print("Loading pose " + str(line) + " of " + str(len(data) - 1) + "...", end=" ")

                elements = data[line + 1].split(separator)
                values = OrderedDict()
                for i in range(len(joints_labels)):
                    values[joints_labels[i]] = float(elements[i])

                values, timestamp = table_to_dict_joints(values)

                # Create the Pose object, passes as parameter the index and the data
                self._create_pose(line, values, timestamp)

                if verbose > 1:
                    print("OK.")

        if verbose:
            print("100% - Done.\n")

    def _create_pose(self, pose_number, data, timestamp):
        """Creates a pose based on its index, timestamp and joint information, and adds it to the Sequence instance."""
        pose = Pose(pose_number, timestamp)
        for j in data:
            joint = Joint(j["JointType"], j["Position"]["X"], j["Position"]["Y"], j["Position"]["Z"])
            pose.add_joint(j["JointType"], joint)
        self.poses.append(pose)

    def _load_date_recording(self, data, verbose=1):
        """Loads the date of a recording from the information contained in the file(s)."""

        if verbose > 1:
            print("Trying to find recording date...", end=" ")

        if type(data) is dict:
            if "Timestamp" in data.keys():
                self.date_recording = datetime(1, 1, 1, 0, 0, 0) + timedelta(microseconds=data["Timestamp"] / 10)

        elif type(data) is list:
            if len(data) > 7:
                splitted = data[7].split("\t")
                if splitted[0] == "TIME_STAMP":
                    self.date_recording = datetime.fromisoformat(splitted[1].replace(",", ""))

        if verbose > 1:
            if self.date_recording is not None:
                print("found: " + self.get_printable_date_recording())
            else:
                print("not found.")

    # === Calculation functions ===

    def _calculate_relative_timestamps(self, time_unit="auto"):
        """For all the poses of the sequence, sets and converts the relative_timestamp attribute taking the first pose
        of the sequence as reference."""
        if time_unit == "auto":
            if len(self.poses) > 1:
                if self.poses[1].get_timestamp() - self.poses[0].get_timestamp() >= 1000:
                    time_unit = "100ns"
                elif self.poses[1].get_timestamp() - self.poses[0].get_timestamp() >= 1:
                    time_unit = "ms"
                else:
                    time_unit = "s"
            else:
                time_unit = "s"

        # Validity check
        time_unit = time_unit.lower().replace(" ", "")  # Removes spaces
        units = ["ns", "1ns", "10ns", "100ns", "µs", "1µs", "10µs", "100µs", "ms", "1ms", "10ms", "100ms",
                 "s", "sec", "1s", "min", "mn", "h", "hr", "d", "day"]
        if time_unit not in units:
            raise Exception("Invalid time unit. Should be ns, µs, ms or s.")

        t = self.poses[0].get_timestamp()
        for p in range(len(self.poses)):
            if self.poses[p].timestamp - t < 0:
                raise Exception("The timestamps aren't in chronological order: you can't travel back " +
                                "in time (yet). Timestamp for pose " + str(p + 1) + "/" + str(len(self.poses)) +
                                " is " + str(self.poses[p].timestamp) + " while the timestamp for the first pose is " +
                                str(t) + " (difference of " + str(self.poses[p].timestamp - t) + ").")
            elif p > 0:
                if self.poses[p].timestamp - self.poses[p - 1].timestamp < 0:
                    raise Exception("The timestamps aren't in chronological order: you can't travel back " +
                                    "in time (yet). Timestamp for pose " + str(p + 1) + "/" + str(len(self.poses)) +
                                    " is " + str(self.poses[p].timestamp) + " while the timestamp for pose " + str(p) +
                                    "/" + str(len(self.poses)) + " is " + str(t) + " (difference of " +
                                    str(self.poses[p].timestamp - t) + ").")
            self.poses[p]._calculate_relative_timestamp(t, time_unit)

    # === Getter functions ===

    def get_pose(self, pose_index):
        """Returns the pose instance corresponding to the index passed as parameter."""
        if len(self.poses) == 0:
            raise Exception("The Sequence does not have any pose.")
        elif not 0 <= pose_index < len(self.poses):
            raise Exception("The pose index must be between 0 and " + str(len(self.poses) - 1) + ". ")

        return self.poses[pose_index]

    def get_number_of_poses(self):
        """Returns the number of poses in the sequence."""
        return len(self.poses)

    def get_date_recording(self):
        """Returns the date and time of the recording as a datetime object, if it exists."""
        return self.date_recording

    def get_printable_date_recording(self):
        """Returns the date and time of the recording as a string, if it exists."""
        days = {0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday", 4: "Friday", 5: "Saturday", 6: "Sunday"}
        months = {1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
                  7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December"}

        if self.date_recording is not None:
            day = days[self.date_recording.weekday()] + " "
            month = months[self.date_recording.month] + " "
            date = day + str(self.date_recording.day) + " " + month + str(self.date_recording.year) + ", " + \
                   str(self.date_recording.hour).zfill(2) + ":" + str(self.date_recording.minute).zfill(2) + ":" + \
                   str(self.date_recording.second).zfill(2)
        else:
            date = "No date found"
        return date

    def get_timestamps(self, relative=False):
        """Returns a list of the timestamps for every pose, in seconds."""

        timestamps = []
        for pose in self.poses:
            if relative:
                timestamps.append(pose.get_relative_timestamp())
            else:
                timestamps.append(pose.get_timestamp())

        return timestamps

    def get_time_between_two_poses(self, pose1, pose2):
        """Returns the difference between the timestamps of two poses, in seconds."""
        timestamp1 = self.poses[pose1].relative_timestamp
        timestamp2 = self.poses[pose2].relative_timestamp
        return timestamp2 - timestamp1

    def get_duration(self):
        """Returns the duration of the sequence, in seconds."""
        return self.poses[-1].relative_timestamp

    def get_frequencies(self):
        """Returns a list of the frequencies of poses per second and a list of the matching timestamps."""
        time_points = []
        frequencies = []
        for p in range(1, len(self.poses) - 1):
            time_points.append(self.poses[p].relative_timestamp)
            framerate = 1 / (self.poses[p].relative_timestamp - self.poses[p - 1].relative_timestamp)
            frequencies.append(framerate)

        return frequencies, time_points

    def get_average_frequency(self):
        """Returns the average number of poses per second of the sequence."""
        frequencies, time_points = self.get_frequencies()
        average_frequency = len(frequencies) / self.get_duration()
        return average_frequency

    def get_min_frequency(self):
        """Returns the minimum frequency of poses per second of the sequence."""
        frequencies, time_points = self.get_frequencies()
        return min(frequencies)

    def get_max_frequency(self):
        """Returns the maximum frequency of poses per second of the sequence."""
        frequencies, time_points = self.get_frequencies()
        return max(frequencies)

    def get_joint_coordinate_as_list(self, joint_label, axis):
        """Returns a list of all the values for one specified coordinate of a specified joint label."""
        if joint_label not in self.poses[0].keys():
            raise Exception("Invalid joint name: " + str(joint_label) + ".")
        elif axis not in ["x", "y", "z"]:
            raise Exception("Invalid axis: " + str(axis) + ". The axis must be x, y or z.")

        values = []
        for p in range(len(self.poses)):
            if axis == "x":
                values.append(self.poses[p].joints[joint_label].x)
            elif axis == "y":
                values.append(self.poses[p].joints[joint_label].y)
            elif axis == "z":
                values.append(self.poses[p].joints[joint_label].z)

        return values

    def get_joint_velocity_as_list(self, joint_label):
        """Returns a list of all the velocities (distance travelled over time) for a specified joint."""
        if joint_label not in self.poses[0].keys():
            raise Exception("Invalid joint name: " + str(joint_label) + ".")

        values = []
        for p in range(1, len(self.poses)):
            values.append(calculate_velocity(self.poses[p - 1], self.poses[p], joint_label))
        return values

    def get_velocities(self):
        """Returns a dictionary containing a list of velocities (distance travelled over time between two poses)
        for each joint."""

        dict_velocities = OrderedDict()

        for p in range(1, len(self.poses)):
            for joint_label in self.poses[p].joints:
                dict_velocities[joint_label] = self.get_joint_velocity_as_list(joint_label)

        return dict_velocities

    def get_max_velocity_whole_sequence(self):
        """Returns the single maximum value of the velocity across every joint of the sequence."""
        dict_velocities = self.get_velocities()
        max_velocity_whole_sequence = 0

        for joint_label in dict_velocities.keys():
            if max(dict_velocities[joint_label]) > max_velocity_whole_sequence:
                max_velocity_whole_sequence = max(dict_velocities[joint_label])

        return max_velocity_whole_sequence

    def get_max_velocity_single_joint(self, joint_label):
        """Returns the maximum value of the velocity for a given joint."""
        dict_velocities = self.get_velocities()

        if joint_label not in dict_velocities.keys():
            raise Exception("Invalid joint name.")

        return max(dict_velocities[joint_label])

    def get_max_velocity_per_joint(self):
        """Returns the maximum value of the velocity for each joint of the sequence."""
        dict_velocities = self.get_velocities()
        max_velocity_per_joint = OrderedDict()

        for joint_label in dict_velocities.keys():
            max_velocity_per_joint[joint_label] = max(dict_velocities[joint_label])

        return max_velocity_per_joint

    def get_total_velocity_whole_sequence(self):
        """Returns the sum of the velocities of every joint across all the poses of the sequence."""
        total_velocity_whole_sequence = 0
        dict_velocities = self.get_velocities()

        for joint_label in dict_velocities.keys():
            total_velocity_whole_sequence += sum(dict_velocities[joint_label])

        return total_velocity_whole_sequence

    def get_total_velocity_single_joint(self, joint):
        """Returns the sum of the velocities for a given joint."""
        total_velocity_single_joint = 0

        for p in range(1, len(self.poses)):
            total_velocity_single_joint += calculate_velocity(self.poses[p - 1], self.poses[p], joint)

        return total_velocity_single_joint

    def get_total_velocity_per_joint(self):
        """Returns the sum of the velocities for each individual joint of the sequence."""
        total_velocity_per_joint = OrderedDict()

        for joint_label in self.poses[0].joints.keys():
            total_velocity_per_joint[joint_label] = self.get_total_velocity_single_joint(joint_label)

        return total_velocity_per_joint

    def get_subject_height(self, verbose=1):
        """Returns an estimation of the height of the subject, in meters."""
        kinect_joints = ["Head", "Neck", "SpineShoulder", "SpineMid", "SpineBase", ["KneeRight", "KneeLeft"],
                         ["AnkleRight", "AnkleLeft"], ["FootRight", "FootLeft"]]
        qualisys_joints = ["HeadTop", ["ShoulderTopRight", "ShoulderTopLeft"], "Chest", ["WaistBackRight",
                                                                                         "WaistBackLeft",
                                                                                         "WaistFrontRight",
                                                                                         "WaistFrontLeft"],
                           ["KneeRight", "KneeLeft"],
                           ["AnkleRight", "AnkleLeft"], ["ForefootOutRight", "ForefootOutLeft"]]

        if "Head" in self.poses[0].joints.keys():
            list_joints = kinect_joints
        else:
            list_joints = qualisys_joints

        height_sum = 0  # Summing the heights through all poses to average

        if verbose > 1:
            print("Calculating the height of the subject based on the poses...")

        for p in range(len(self.poses)):

            pose = self.poses[p]  # Current pose
            dist = 0  # Height for the current pose
            joints = []  # Joints and average joints

            # Getting the average joints
            for i in range(len(list_joints)):

                if type(list_joints[i]) is list:
                    joint = pose.generate_average_joint(list_joints[i], "Average", False)
                else:
                    joint = pose.joints[list_joints[i]]

                joints.append(joint)

            # Calculating the distance
            for i in range(len(joints) - 1):
                dist += calculate_distance(joints[i], joints[i + 1])

            if verbose > 1:
                print(
                    "\tHeight on frame " + str(p + 1) + "/" + str(len(self.poses)) + ": " + str(round(dist, 3)) + " m.")
            height_sum += dist

        height = height_sum / len(self.poses)

        if verbose == 1:
            print("Average height estimated at " + str(round(height, 3)) + " m.")

        return height

    def get_subject_arm_length(self, side="left", verbose=1):
        """Returns an estimation of the length of the left or right arm of the subject, in meters."""
        kinect_joints = ["Shoulder", "Elbow", "Wrist", "Hand"]
        qualisys_joints = ["ShoulderTop", "Arm", "Elbow", "WristOut", "HandOut"]

        if "Head" in self.poses[0].joints.keys():
            list_joints = kinect_joints
        else:
            list_joints = qualisys_joints

        if side.lower() == "left":
            side = "Left"
        elif side.lower() == "right":
            side = "Right"
        else:
            raise Exception('Wrong side parameter: should be "left" or "right"')

        arm_length_sum = 0  # Summing the arm_lengths through all poses to average

        if verbose > 1:
            print("Calculating the " + side.lower() + " arm length of the subject based on the poses...")

        for p in range(len(self.poses)):

            pose = self.poses[p]  # Current pose
            dist = 0  # Arm length for the current pose

            # Calculating the distance
            for i in range(len(list_joints) - 1):
                dist += calculate_distance(pose.joints[list_joints[i] + side], pose.joints[list_joints[i + 1] + side])

            if verbose > 1:
                print("\t" + side + " arm length on frame " + str(p + 1) + "/" + str(len(self.poses)) + ": " + str(
                    round(dist, 3)) + " m.")
            arm_length_sum += dist

        arm_length = arm_length_sum / len(self.poses)

        if verbose == 1:
            print("Average " + side.lower() + " arm length estimated at " + str(round(arm_length, 3)) + " m.")

        return arm_length

    def get_stats(self, tabled=False):
        """Returns a dictionary containing a series of statistics regarding the sequence."""
        stats = OrderedDict()
        stats["Path"] = self.path
        stats["Date of recording"] = self.get_printable_date_recording()
        stats["Duration"] = self.get_duration()
        stats["Number of poses"] = self.get_number_of_poses()

        # Height and distances stats
        stats["Subject height"] = self.get_subject_height(0)
        stats["Left arm length"] = self.get_subject_arm_length("left", 0)
        stats["Right arm length"] = self.get_subject_arm_length("right", 0)

        # Framerate stats
        frequencies, time_points = self.get_frequencies()
        stats["Average frequency"] = self.get_average_frequency()
        stats["SD frequency"] = statistics.stdev(frequencies)
        stats["Min frequency"] = self.get_min_frequency()
        stats["Max frequency"] = self.get_max_frequency()

        # Movement stats
        for joint_label in self.poses[0].joints.keys():
            stats["Average velocity " + joint_label] = \
                self.get_total_velocity_single_joint(joint_label) / len(self.poses)

        if tabled:
            table = [[], []]
            for key in stats.keys():
                table[0].append(key)
                table[1].append(stats[key])
            return table

        return stats

    # === Correction functions ===

    def correct_jitter(self, velocity_threshold, window, window_unit="poses", method="default", name=None, verbose=1):
        """Detects and corrects rapid twitches and jumps in a motion sequence."""

        dict_joints_to_correct = {}
        for key in self.poses[0].joints.keys():
            dict_joints_to_correct[key] = []

        if verbose > 0:
            print("Correcting jumps and twitches...", end=" ")

        # Create an empty sequence, the same length in poses as the original, just keeping the timestamp information
        # for each pose.
        new_sequence = self._create_new_sequence_with_timestamps(verbose)
        if name is None:
            new_sequence.name = self.name + " +CJ"
        else:
            new_sequence.name = name

        # If verbose, show all the information. Else, show only the progress in percentage.
        if verbose > 0:
            if method in ["old", "default"]:
                print("\tPerforming realignment...", end=" ")
            else:
                print("\tDetecting jumps and twitches...", end=" ")

        # If the window unit is defined in ms, we convert it to s.
        if window_unit == "ms":
            window /= 1000
            window_unit = "s"

        # Define the counters
        realigned_points = 0
        jumps = 0
        twitches = 0
        perc = 10

        # For every pose starting on the second one
        for p in range(1, len(self.poses)):

            if verbose > 1:
                print("\nNew sequence:\n" + str(new_sequence.poses[p - 1]))
                print("\n== POSE NUMBER " + str(p + 1) + " ==")

            perc = show_percentage(verbose, p, len(self.poses), perc)

            # If the window unit is defined in seconds, we calculate how many poses that is.
            if window_unit == "s":
                if verbose > 1:
                    print("Calculating number of poses in the window...", end=" ")

                window_effective = 0
                time_diff = 0
                previous_time_diff = 0

                # Get window length in poses
                for i in range(p, len(self.poses)):
                    window_effective = i - p
                    time_diff = self.get_time_between_two_poses(p, i)
                    if time_diff >= window:
                        break
                    previous_time_diff = time_diff

                # We get the closest window from the target
                if p + window_effective != len(self.poses) - 1:
                    if abs(window - previous_time_diff) < abs(window - time_diff):
                        window_effective -= 1

                if verbose > 1:
                    print("Window size: " + str(window_effective) + " (" + str(time_diff) + " s).")

            else:
                window_effective = window

            # For every joint
            for joint_label in self.poses[0].joints.keys():

                # We get the m/s travelled by the joint between this pose and the previous
                velocity_before = calculate_velocity(new_sequence.poses[p - 1], self.poses[p], joint_label)

                if verbose > 1:
                    print(joint_label + ": " + str(velocity_before))

                # If we already corrected this joint, we ignore it to avoid overcorrection
                if joint_label in new_sequence.poses[p].joints:

                    if verbose > 1:
                        print("\tAlready corrected.")

                # If the velocity is over threshold, we check if it is a jump or a twitch
                elif velocity_before > velocity_threshold and p != len(self.poses) - 1:

                    if verbose > 1:
                        print("\tVelocity over threshold between poses " + str(p) + " and " + str(p + 1) + ".", end=" ")

                    self.poses[p].joints[joint_label].set_movement_over_threshold(True)

                    if window_effective < 2:
                        if verbose > 1:
                            print("Window size inferior to 2, copying joint...")
                        new_sequence, realigned_points = \
                            self._correct_jitter_window(new_sequence, p - 1, p, joint_label, realigned_points, verbose)

                    else:
                        if verbose > 1:
                            print("Checking in subsequent poses...")

                        # We check the next poses (defined by the window) if the position of the joint comes back below
                        # threshold
                        for k in range(p, min(p + window_effective, len(self.poses))):

                            if verbose > 1:
                                print("\t\tPose " + str(k + 1) + ":", end=" ")

                            # Method 1 (original): get the subsequent velocities
                            if method == "old":
                                velocity = calculate_velocity(new_sequence.poses[p - 1], self.poses[k], joint_label)

                            # Method 2 (to test): get the distance over the time between the two first joints
                            else:
                                dist = calculate_distance(new_sequence.poses[p - 1].joints[joint_label],
                                                          self.poses[k].joints[joint_label])
                                delay = calculate_delay(new_sequence.poses[p - 1], new_sequence.poses[p])
                                velocity = dist / delay

                            # Twitch case: One of the poses of the window is below threshold compared to previous pose.
                            if velocity < velocity_threshold:

                                if verbose > 1:
                                    print("velocity back under threshold defined by pose " + str(p) + ".")
                                    print("\t\t\tCorrecting for twitch...")

                                new_sequence, realigned_points = \
                                    self._correct_jitter_window(new_sequence, p - 1, k, joint_label, realigned_points,
                                                                verbose)

                                if method not in ["old", "default"]:
                                    for i in range(p, k):
                                        dict_joints_to_correct[joint_label].append(i)

                                twitches += 1

                                break

                            # Jump case: No pose of the window is below threshold.
                            if k == p + window_effective - 1 or k == len(self.poses) - 1:

                                if verbose > 1:
                                    print("velocity still over threshold at the end of the window.")
                                    print("\t\t\tCorrecting for jump...")

                                new_sequence, realigned_points = \
                                    self._correct_jitter_window(new_sequence, p - 1, k, joint_label, realigned_points,
                                                                verbose)

                                if method not in ["old", "default"]:
                                    for i in range(p, k):
                                        dict_joints_to_correct[joint_label].append(i)

                                jumps += 1

                                break

                            # Wait: if there are still poses in the window, we continue.
                            else:

                                if verbose > 1:
                                    print("still over threshold.")

                    if verbose > 1:
                        print("")

                # If we're still in threshold, there is no correction, we copy the joint
                else:

                    joint = self.copy_joint(p, joint_label)
                    new_sequence.poses[p].add_joint(joint_label, joint)
                    new_sequence.poses[p].joints[joint_label].corrected = False

        new_sequence._calculate_relative_timestamps()  # Sets the relative time from the first pose for each pose

        if verbose > 0:
            print("100% - Done.")

        # If we use one of the interpolation methods, we call the correct_zeros function
        if method not in ["old", "default"]:

            if verbose > 0:
                print("\t" + str(realigned_points) + " point(s) with twitches and jumps detected.")
                print("\tSetting faulty joints to (0, 0, 0)...", end=" ")

            for joint_label in dict_joints_to_correct:
                for p in dict_joints_to_correct[joint_label]:
                    new_sequence.poses[p].joints[joint_label].set_to_zero()

            if verbose > 0:
                print("100% - Done.")
                print("\tInterpolation of the data of the joints set to (0, 0, 0)...")

            new_sequence = new_sequence.correct_zeros(mode=method, name=name, verbose=verbose, add_tabs=2)

        if verbose > 0:
            percentage = round((realigned_points / (len(self.poses) * len(list(self.poses[0].joints.keys())))) * 100, 1)
            print("De-jittering over. " + str(realigned_points) + " point(s) corrected over " + str(
                len(self.poses) * len(list(self.poses[0].joints.keys()))) + " (" + str(percentage) + "%).")
            print(str(jumps) + " jump(s) and " + str(twitches) + " twitch(es) corrected.\n")

        return new_sequence

    def _correct_jitter_window(self, new_sequence, start_pose_number, end_pose_number, joint_label, realigned_points,
                               verbose):
        """Corrects linearly the jumps and twitches of the positions of a joint, between two given time points."""

        # We extract the data from the joints at the beginning and at the end of the partial sequence to correct
        joint_before = new_sequence.poses[start_pose_number].joints[joint_label]
        joint_after = self.poses[end_pose_number].joints[joint_label]

        # If the starting and ending joint are the same, we don't correct, it means it is the last pose of the sequence
        if start_pose_number == end_pose_number:
            joint = self.copy_joint(start_pose_number, joint_label)
            new_sequence.poses[start_pose_number + 1].add_joint(joint_label, joint)

            if verbose > 1:
                print("\t\t\t\tDid not correct joint " + str(start_pose_number + 1) +
                      " as the window is of size 0 or the joint is the last pose of the sequence.")

        elif start_pose_number == end_pose_number - 1:
            joint = self.copy_joint(start_pose_number, joint_label)
            new_sequence.poses[start_pose_number + 1].add_joint(joint_label, joint)

            if verbose > 1:
                print("\t\t\t\tDid not correct joint " + str(start_pose_number + 1) +
                      " as the window is of size 1.")

        # Otherwise we correct for all the intermediate joints
        for pose_number in range(start_pose_number + 1, end_pose_number):

            # If a joint was already corrected we don't correct it to avoid overcorrection
            if self.poses[pose_number].joints[joint_label].corrected:

                if verbose > 1:
                    print("\t\t\t\tDid not correct joint " + str(pose_number + 1) + " as it was already corrected.")

            # Otherwise we correct it
            else:

                x, y, z = self._correct_jitter_single_joint(joint_before, joint_after, start_pose_number, pose_number,
                                                            end_pose_number, verbose)

                # We copy the original joint, apply the new coordinates and add it to the new sequence
                joint = self.copy_joint(pose_number, joint_label)
                joint.correct_joint(x, y, z)
                new_sequence.poses[pose_number].add_joint(joint_label, joint)

                if verbose > 1:
                    print("\t\t\t\tCorrecting joint: " + str(pose_number + 1) + ". Original coordinates: (" + str(
                        self.poses[pose_number].joints[joint_label].x) +
                          ", " + str(self.poses[pose_number].joints[joint_label].y) + ", " + str(
                        self.poses[pose_number].joints[joint_label].z) + ")")
                    print("\t\t\t\tPrevious joint: " + str(start_pose_number + 1) + ". (" + str(
                        new_sequence.poses[start_pose_number].joints[joint_label].x) +
                          ", " + str(new_sequence.poses[start_pose_number].joints[joint_label].y) + ", " + str(
                        new_sequence.poses[start_pose_number].joints[joint_label].z) + ")")
                    print("\t\t\t\tNext joint: " + str(end_pose_number + 1) + ". (" + str(
                        self.poses[end_pose_number].joints[joint_label].x) +
                          ", " + str(self.poses[end_pose_number].joints[joint_label].y) + ", " + str(
                        self.poses[end_pose_number].joints[joint_label].z) + ")")
                    print("\t\t\t\tCorrected joint " + str(pose_number + 1) + ". New coordinates: (" + str(
                        x) + ", " + str(y) + ", " + str(z) + ")\n")

                # To count the number of joints that were realigned overall
                realigned_points += 1

        return new_sequence, realigned_points

    def _correct_jitter_single_joint(self, joint_before, joint_after, pose_before, pose_current, pose_after, verbose):
        """Corrects linearly a Joint object following the algorithm in Sequence.correct_jitter()."""

        # We calculate the percentage of time elapsed between the pose at the beginning of the partial sequence to
        # correct and the pose at the end
        percentage_time = (self.poses[pose_current].timestamp - self.poses[pose_before].timestamp) / (
                self.poses[pose_after].timestamp - self.poses[pose_before].timestamp)

        if verbose > 1:
            print("\t\t\t\tJoint " + str(pose_current + 1) + " positioned at " + str(
                percentage_time * 100) + " % between poses " + str(pose_before + 1) + " and " + str(
                pose_after + 1) + ".")

        # We correct linearly every joint. See the documentation for precisions.
        x = joint_before.x - percentage_time * (joint_before.x - joint_after.x)
        y = joint_before.y - percentage_time * (joint_before.y - joint_after.y)
        z = joint_before.z - percentage_time * (joint_before.z - joint_after.z)

        return x, y, z

    def re_reference(self, reference_joint_label="auto", place_at_zero=True, name=None, verbose=1):
        """Changes the position of all the joints of a sequence to be relative to the position of a reference joint."""

        if reference_joint_label == "auto":
            joints_labels = self.poses[0].joints.keys()
            if "SpineMid" in joints_labels:
                reference_joint_label = "SpineMid"
            elif "Chest" in joints_labels:
                reference_joint_label = "Chest"
            else:
                raise Exception("Automatic reference joint finder failed. Please enter a valid joint label as a " +
                                "reference joint.")

        if verbose > 0:
            print("Re-referencing to " + str(reference_joint_label) + "...", end=" ")

        # Create an empty sequence, the same length in poses as the original, just keeping the timestamp information
        # for each pose.
        new_sequence = Sequence(None)
        if name is None:
            new_sequence.name = self.name + " +RF"
        else:
            new_sequence.name = name

        # If verbose, show all the information. Else, show only the progress in percentage.
        if verbose > 0:
            print("\tPerforming re-referencing...", end=" ")

        # Get reference : position of the reference joint at time 0
        if place_at_zero:
            start_ref_x = 0
            start_ref_y = 0
            start_ref_z = 0
        else:
            start_ref_x = self.poses[0].joints[reference_joint_label].x
            start_ref_y = self.poses[0].joints[reference_joint_label].y
            start_ref_z = self.poses[0].joints[reference_joint_label].z

        # Define the percentage counter
        perc = 10

        # For every pose starting on the second one
        for p in range(len(self.poses)):

            if verbose > 1:
                print("\n== POSE NUMBER " + str(p + 1) + " ==")

            perc = show_percentage(verbose, p, len(self.poses), perc)

            # Get movement from the reference point
            curr_ref_x = self.poses[p].joints[reference_joint_label].x
            curr_ref_y = self.poses[p].joints[reference_joint_label].y
            curr_ref_z = self.poses[p].joints[reference_joint_label].z

            # Compute the differences
            diff_ref_x = curr_ref_x - start_ref_x
            diff_ref_y = curr_ref_y - start_ref_y
            diff_ref_z = curr_ref_z - start_ref_z

            if verbose > 1:
                print("Removing " + str(diff_ref_x) + " to every joint in x;")
                print("Removing " + str(diff_ref_y) + " to every joint in y;")
                print("Removing " + str(diff_ref_z) + " to every joint in z.")

            # For every joint
            for joint_label in self.poses[0].joints.keys():

                # Compute new joint position
                if joint_label == reference_joint_label:
                    new_x = 0
                    new_y = 0
                    new_z = 0
                else:
                    new_x = self.poses[p].joints[joint_label].x - diff_ref_x
                    new_y = self.poses[p].joints[joint_label].y - diff_ref_y
                    new_z = self.poses[p].joints[joint_label].z - diff_ref_z

                # Add to the sequence
                joint = self.copy_joint(p, joint_label)
                joint.correct_joint(new_x, new_y, new_z)
                new_sequence.poses[p].add_joint(joint_label, joint)

                if verbose > 1:
                    print(joint_label + ": ")
                    print("X: " + str(self.poses[p].joints[joint_label].x) + " -> " + str(new_x))
                    print("Y: " + str(self.poses[p].joints[joint_label].y) + " -> " + str(new_y))
                    print("Z: " + str(self.poses[p].joints[joint_label].z) + " -> " + str(new_z))

        new_sequence._calculate_relative_timestamps()  # Sets the relative time from the first pose for each pose

        if verbose > 0:
            print("100% - Done.")
            print("Re-referencing over.\n")
        return new_sequence

    def trim(self, start=None, end=None, use_relative_timestamps=False, name=None, verbose=1, add_tabs=0):
        """Trims a sequence according to a starting timestamp (by default the beginning of the original sequence) and
        an ending timestamp (by default the end of the original sequence). Timestamps must be provided in seconds."""

        tabs = add_tabs * "\t"

        if start is None:
            if use_relative_timestamps:
                start = 0
            else:
                start = self.poses[0].timestamp

        if end is None:
            if use_relative_timestamps:
                end = self.get_duration()
            else:
                end = self.poses[-1].timestamp

        if end < start:
            raise Exception("End timestamp should be inferior to beginning timestamp.")
        elif start < self.poses[0].timestamp and not use_relative_timestamps:
            raise Exception("Starting timestamp (" + str(start) + ") lower than the first timestamp of the sequence (" +
                            str(self.poses[0].timestamp) + ").")
        elif start < 0 and use_relative_timestamps:
            raise Exception("Starting timestamp (" + str(start) + ") must be equal to or higher than 0.")
        elif end > self.poses[-1].timestamp and not use_relative_timestamps:
            raise Exception("Ending timestamp (" + str(end) + ") exceeds last timestamp of the sequence (" +
                            str(self.poses[-1].timestamp) + ").")
        elif end > self.get_duration() and use_relative_timestamps:
            raise Exception("Ending timestamp (" + str(end) + ") exceeds the duration of the sequence (" +
                            str(self.get_duration()) + ").")

        if verbose > 0:
            print(tabs + "Trimming the sequence:")
            print(tabs + "\tStarting timestamp: " + str(start) + " seconds")
            print(tabs + "\tEnding timestamp: " + str(end) + " seconds")
            print(tabs + "\tOriginal duration: " + str(self.get_duration()) + " seconds")
            print(tabs + "\tDuration after trimming: " + str(end - start) + " seconds")
        if verbose == 1:
            print(tabs + "Starting the trimming...", end=" ")

        # Create an empty sequence
        new_sequence = Sequence(None)
        if name is None:
            new_sequence.name = self.name + " +TR"
        else:
            new_sequence.name = name

        # Define the percentage counter
        perc = 10

        for p in range(len(self.poses)):

            if verbose == 1:
                perc = show_percentage(verbose, p, len(self.poses), perc)

            if use_relative_timestamps:
                t = self.poses[p].relative_timestamp
            else:
                t = self.poses[p].timestamp

            if verbose > 1:
                print(tabs + "\t\tPose " + str(p + 1) + " of " + str(len(self.poses)) + ": " + str(t), end=" ")

            if start <= t <= end:
                pose = self.copy_pose(p)
                new_sequence.poses.append(pose)
                if verbose > 1:
                    print("Preserved.")
            elif verbose > 1:
                print("Trimmed.")

        new_sequence._calculate_relative_timestamps()  # Sets the relative time from the first pose for each pose

        if verbose > 0:
            if verbose == 1:
                print("100% - Done.")
            print(tabs + "\tNew sequence duration: " + str(new_sequence.get_duration()) + " s.")
            print(tabs + "\tOriginal number of poses: " + str(len(self.poses)), end=" · ")
            print("New number of poses: " + str(len(new_sequence.poses)))
            print(tabs + "Trimming over.\n")

        return new_sequence

    def trim_to_audio(self, delay=0, audio=None, name=None, verbose=1):
        """Synchronizes the timestamps to the duration of an audio file."""

        if audio is None:
            if self.path_audio is None:
                raise Exception("No audio path defined in the sequence, please specify one.")
            audio = Audio(self.path_audio, name=self.path_audio)

        if type(audio) is str:
            audio = Audio(audio, name=audio)

        if verbose > 0:
            print("Trimming the sequence according to the duration of an audio file...")
            print("\tParameters:")
            print("\t\tSequence duration: " + str(self.get_duration()) + " s.")
            print("\t\tAudio duration: " + str(audio.duration) + " s.")

        if audio.duration > self.get_duration():
            print("\n\tSequence duration: " + str(self.get_duration()))
            print("\tAudio duration: " + str(audio.duration))
            raise Exception("Error: The duration of the audio exceeds the duration of the sequence.")
        elif verbose > 0:
            print("\t\tMoving the beginning by a delay of " + str(delay) + " s.")
            print("\t\tIgnoring " + str(round(self.get_duration() - audio.duration - delay, 2)) +
                  " s at the end of the sequence.")

        new_sequence = self.trim(delay, audio.duration + delay, True, name, verbose, 1)

        if abs(audio.duration - new_sequence.get_duration()) > 1:
            raise Exception("The duration of the audio is different of more than one second with the duration of the " +
                            "new sequence.")

        return new_sequence

    def resample(self, frequency, mode="cubic", name=None, verbose=1):
        """Resamples a sequence to a fixed frequency."""

        if verbose > 0:
            print("Resampling the sequence at " + str(frequency) + " Hz (mode: " + str(mode) + ")...")
            print("\tOriginal framerate: ")
            print("\t\tAverage: " + str(round(self.get_average_frequency(), 2)), end=" · ")
            print("Min: " + str(round(self.get_min_frequency(), 2)), end=" · ")
            print("Max: " + str(round(self.get_max_frequency(), 2)))
            print("\tCreating vectors...", end=" ")

        # Create an empty sequence
        new_sequence = Sequence(None)
        if name is None:
            new_sequence.name = self.name + " +RS" + str(frequency)
        else:
            new_sequence.name = name

        # Define the percentage counter
        perc = 10

        # Define positions lists
        x_points = OrderedDict()
        y_points = OrderedDict()
        z_points = OrderedDict()
        time_points = []

        # Create vectors of position and time
        for p in range(len(self.poses)):

            perc = show_percentage(verbose, p, len(self.poses), perc)

            for j in self.poses[p].joints.keys():
                if p == 0:
                    x_points[j] = []
                    y_points[j] = []
                    z_points[j] = []
                x_points[j].append(self.poses[p].joints[j].x)
                y_points[j].append(self.poses[p].joints[j].y)
                z_points[j].append(self.poses[p].joints[j].z)
            time_points.append(self.poses[p].timestamp)

        if verbose > 0:
            print("100% - Done.")
            print("\tPerforming the resampling...", end=" ")

        # Define the percentage counter
        perc = 10

        # Resample
        new_x_points = OrderedDict()
        new_y_points = OrderedDict()
        new_z_points = OrderedDict()
        new_time_points = []

        no_joints = len(x_points.keys())
        i = 0

        for joint_label in x_points.keys():
            perc = show_percentage(verbose, i, no_joints, perc)
            new_x_points[joint_label], new_time_points = resample_data(x_points[joint_label],
                                                                       time_points, frequency, mode)
            new_y_points[joint_label], new_time_points = resample_data(y_points[joint_label],
                                                                       time_points, frequency, mode)
            new_z_points[joint_label], new_time_points = resample_data(z_points[joint_label],
                                                                       time_points, frequency, mode)
            i += 1

        # Define the percentage counter
        if verbose > 0:
            print("100% - Done.")
            print("\tSaving the new sequence...", end=" ")
        perc = 10

        # Save data
        for p in range(len(new_time_points)):
            perc = show_percentage(verbose, i, no_joints, perc)
            pose = Pose(p, new_time_points[p])
            for j in new_x_points.keys():
                x = new_x_points[j][p]
                y = new_y_points[j][p]
                z = new_z_points[j][p]
                joint_label = Joint(j, x, y, z)
                joint_label.joint_type = j
                pose.add_joint(j, joint_label)
            new_sequence.poses.append(pose)

        if verbose > 0:
            print("100% - Done.")
            print("\tOriginal sequence had " + str(len(self.poses)) + " poses.")
            print("\tNew sequence has " + str(len(new_sequence.poses)) + " poses.\n")

        new_sequence._calculate_relative_timestamps()  # Sets the relative time from the first pose for each pose

        return new_sequence

    def correct_zeros(self, mode="cubic", name=None, verbose=1, add_tabs=0):
        """Detects the joints set at (0, 0, 0) and correct their coordinates by interpolating the data."""

        t = add_tabs * "\t"

        if verbose > 0:
            print(t + "Correcting the zeros (mode: " + str(mode) + ")...")
            print(t + "\tCreating vectors and finding zeros...", end=" ")

        # Create an empty sequence
        new_sequence = Sequence(None)
        if name is None:
            new_sequence.name = self.name + " +CZ"
        else:
            new_sequence.name = name

        # Define the percentage counter
        perc = 10

        # Define positions lists
        x_points = OrderedDict()
        y_points = OrderedDict()
        z_points = OrderedDict()
        time_points = OrderedDict()
        zero_points_detected = 0
        total_joints = 0
        longest_zero = 0

        # Create vectors of position and time
        for p in range(len(self.poses)):

            perc = show_percentage(verbose, p, len(self.poses), perc)
            if verbose > 1:
                if p == 0:
                    print("")
                print(t + "\t\tChecking pose " + str(p + 1) + " out of " + str(len(self.poses)))

            faulty_joints = []

            for joint_label in self.poses[p].joints.keys():

                # First iteration, we create the vectors for this joint
                if p == 0:
                    x_points[joint_label] = []
                    y_points[joint_label] = []
                    z_points[joint_label] = []
                    time_points[joint_label] = []

                # If coordinates are non-zero
                if (self.poses[p].joints[joint_label].x != 0 and self.poses[p].joints[joint_label].y != 0 and
                        self.poses[p].joints[joint_label].z != 0):

                    # If these are the first non-zero coordinates, we create a non-zero coordinate
                    if p != 0 and len(x_points) == 0:
                        if verbose > 1:
                            print(t + "\t\t\tThis is the first pose for the joint " + str(joint_label) + " having " +
                                  "non-zero coordinates. Creating a pose at timestamp 0 with these coordinates.")
                        x_points[joint_label].append(self.poses[p].joints[joint_label].x)
                        y_points[joint_label].append(self.poses[p].joints[joint_label].y)
                        z_points[joint_label].append(self.poses[p].joints[joint_label].z)
                        time_points[joint_label].append(self.poses[0].timestamp)

                    # We add the coordinates and timestamps
                    x_points[joint_label].append(self.poses[p].joints[joint_label].x)
                    y_points[joint_label].append(self.poses[p].joints[joint_label].y)
                    z_points[joint_label].append(self.poses[p].joints[joint_label].z)
                    time_points[joint_label].append(self.poses[p].timestamp)

                # If coordinate is zero
                else:

                    # If it's the last pose of the sequence
                    if p == len(self.poses) - 1:

                        # If the vector is still empty, we add a first coordinate
                        if len(x_points) == 0:
                            if verbose > 1:
                                print(t + "\t\t\tAll the coordinates for the joint " + str(joint_label) + " were " +
                                      "(0, 0, 0). Creating a pose at timestamp 0 with (0, 0, 0) coordinates.")
                            x_points[joint_label].append(self.poses[p].joints[joint_label].x)
                            y_points[joint_label].append(self.poses[p].joints[joint_label].y)
                            z_points[joint_label].append(self.poses[p].joints[joint_label].z)
                            time_points[joint_label].append(self.poses[0].timestamp)

                        # We copy the last coordinate we added to the vector as an ending point
                        if verbose > 1:
                            print(
                                t + "\t\t\tThis is the last pose for the joint " + str(joint_label) + ", and it has " +
                                "(0, 0, 0) coordinates. Copying the last pose having non-zero coordinates to put at " +
                                "the end of the series.")
                        x_points[joint_label].append(x_points[joint_label][-1])
                        y_points[joint_label].append(y_points[joint_label][-1])
                        z_points[joint_label].append(z_points[joint_label][-1])
                        time_points[joint_label].append(self.poses[p].timestamp)

                    zero_points_detected += 1

                    if joint_label not in faulty_joints:
                        faulty_joints.append(joint_label)

                    if len(time_points[joint_label]) > 1:
                        t_diff = time_points[joint_label][-1] - time_points[joint_label][-2]
                        if t_diff > longest_zero:
                            longest_zero = t_diff

                total_joints += 1

            if verbose > 1:
                if len(faulty_joints) == 0:
                    print(t + "\t\t\tNo joints found with coordinate (0, 0, 0).")
                else:
                    to_print = ""
                    if len(faulty_joints) == 1:
                        print(t + "\t\t\t 1 joint with coordinate (0, 0, 0) detected: ", end="")
                    else:
                        print(t + "\t\t\t" + str(len(faulty_joints)) + " joints with coordinate (0, 0, 0) detected: ",
                              end="")
                    for joint_label in faulty_joints:
                        to_print += joint_label + ", "
                    print(to_print[:-2])

        if verbose == 1:
            print("100% - Done.")
        if verbose > 0:
            print(t + "\t\t" + str(zero_points_detected) + " joints at coordinate (0, 0, 0) detected over " +
                  str(total_joints) + " (" + str(round(zero_points_detected / total_joints * 100, 2)) + "%).")
            print(t + "\t\t" + "Longest chain of zeros detected: " + str(longest_zero) + "s. ")

        if zero_points_detected != 0:
            if verbose > 0:
                print(t + "\tPerforming the correction..", end=" ")

            # Define the percentage counter
            perc = 10

            # Resample
            new_x_points = OrderedDict()
            new_y_points = OrderedDict()
            new_z_points = OrderedDict()
            new_time_points = []

            no_joints = len(x_points.keys())
            i = 0

            for joint_label in x_points.keys():
                perc = show_percentage(verbose, i, no_joints, perc)
                if verbose > 1:
                    if i == 0:
                        print("")
                    print(t + "\t\tCorrecting the time series for the joint " + joint_label + "...", end=" ")
                new_x_points[joint_label], new_time_points = interpolate_data(x_points[joint_label],
                                                                              time_points[joint_label],
                                                                              self.get_timestamps(relative=False), mode)
                new_y_points[joint_label], new_time_points = interpolate_data(y_points[joint_label],
                                                                              time_points[joint_label],
                                                                              self.get_timestamps(relative=False), mode)
                new_z_points[joint_label], new_time_points = interpolate_data(z_points[joint_label],
                                                                              time_points[joint_label],
                                                                              self.get_timestamps(relative=False), mode)
                if verbose > 1:
                    print("OK")
                i += 1

            # Define the percentage counter
            if verbose == 1:
                print("100% - Done.")
            if verbose > 0:
                print(t + "\tSaving the new sequence...", end=" ")
            perc = 10

            # Save data
            for p in range(len(new_time_points)):
                perc = show_percentage(verbose, i, no_joints, perc)
                if verbose > 1:
                    if p == 0:
                        print("")
                    print(t + "\t\tCreating pose " + str(p + 1) + " out of " + str(len(self.poses)) + "... ", end="")
                pose = Pose(p, new_time_points[p])
                for joint_label in new_x_points.keys():
                    x = new_x_points[joint_label][p]
                    y = new_y_points[joint_label][p]
                    z = new_z_points[joint_label][p]
                    joint_label = Joint(joint_label, x, y, z)
                    joint_label.joint_type = joint_label
                    pose.add_joint(joint_label, joint_label)
                new_sequence.poses.append(pose)
                if verbose > 1:
                    print("OK")

            if verbose == 1:
                print("100% - Done.")
            if verbose > 1:
                print(t + "\tOriginal sequence had " + str(len(self.poses)) + " poses.")
                print(t + "\tNew sequence has " + str(len(new_sequence.poses)) + " poses.")
                print(t + "\t\t" + str(zero_points_detected) + " out of " + str(total_joints) + " joints have been " +
                      "corrected (" + str(round(zero_points_detected / total_joints * 100, 2)) + "%).")

            new_sequence._calculate_relative_timestamps()  # Sets the relative time from the first pose for each pose

            return new_sequence

        else:
            if verbose != 0:
                print(t + "No zero coordinate found, returning the original sequence.")
            return self

    def randomize(self, verbose=1):
        """Returns a sequence that randomizes the starting position of all the joints of the original sequence."""

        if verbose > 0:
            print("Randomizing the starting points of the joints...")
            print("\tCreating a new sequence...", end=" ")

        new_sequence = Sequence(None)
        for p in range(len(self.poses)):
            new_sequence.poses.append(self.copy_pose(p))

        if verbose > 0:
            print("OK. \n\tRandomizing starting positions...", end=" ")

        # Get the starting positions of all joints
        starting_positions = []
        for joint_label in new_sequence.poses[0].joints.keys():
            starting_positions.append(copy.deepcopy(new_sequence.poses[0].joints[joint_label]))

        # Randomize new starting positions for joints
        randomized_positions = generate_random_joints(len(new_sequence.poses[0].joints))
        # random.shuffle(randomized_positions)

        if verbose > 0:
            print("OK. \n\tMoving the joints...", end=" ")

        # Assign these new starting positions to the joints and moves the joints in every subsequent pose
        # in relation to the new starting positions
        joints_labels = list(new_sequence.poses[0].joints.keys())
        for p in new_sequence.poses:
            for j in range(len(joints_labels)):
                p.joints[joints_labels[j]].move_joint_randomized(starting_positions[j], randomized_positions[j])
        new_sequence.randomized = True
        new_sequence._calculate_relative_timestamps()

        if verbose > 0:
            print("OK.")

        return new_sequence

    # === Copy and blank sequence creation functions ===

    def _create_new_sequence_with_timestamps(self, verbose=1):
        """Creates a sequence with the same number of poses as in the original, but only containing timestamps."""

        if verbose > 0:
            print("\n\tCreating an empty sequence...", end=" ")
        new_sequence = Sequence(None)

        # Used for the progression percentage
        perc = 10
        for p in range(len(self.poses)):
            # Show percentage if verbose
            perc = show_percentage(verbose, p, len(self.poses), perc)
            # Creates an empty pose with the same timestamp as the original
            pose = self.poses[p].get_copy_with_empty_joints(False)
            new_sequence.poses.append(pose)  # Add to the sequence

        # Copy the joints from the first pose
        for joint_label in self.poses[0].joints.keys():  # For every joint

            # If this joint doesn't already exist in the new sequence, we copy it
            if joint_label not in new_sequence.poses[0].joints.keys():
                new_sequence.poses[0].joints[joint_label] = self.copy_joint(0, joint_label)

        if verbose > 0:
            print("100% - Done.")

        return new_sequence

    def copy_pose(self, pose_index):
        """Returns a deep copy of a specified pose."""

        if len(self.poses) == 0:
            raise Exception("The Sequence does not have any pose.")
        elif not 0 <= pose_index < len(self.poses):
            raise Exception("The pose index must be between 0 and " + str(len(self.poses) - 1) + ". ")

        pose = Pose(pose_index, self.poses[pose_index].get_timestamp())

        for i in self.poses[pose_index].joints.keys():  # For every joint

            pose.joints[i] = self.copy_joint(pose_index, i)

        pose.relative_timestamp = self.poses[pose_index].get_relative_timestamp()

        return pose

    def copy_joint(self, pose_index, joint_label):
        """Returns a deep copy of a specified joint from a specified pose."""

        if len(self.poses) == 0:
            raise Exception("The Sequence does not have any pose.")
        elif not 0 <= pose_index < len(self.poses):
            raise Exception("The pose index must be between 0 and " + str(len(self.poses) - 1) + ". ")

        joint = self.poses[pose_index].get_joint(joint_label).get_copy()

        return joint

    # === Miscellaneous functions ===

    def average_qualisys_joints(self, joints_labels_to_exclude=None, remove_averaged_joints=False,
                                remove_non_kinect_joints=False):
        """Create missing Kinect joints from the Qualisys labelling system by averaging the distance between Qualisys
        joints."""

        joints_to_average = load_qualisys_to_kinect()
        joints_to_remove = ["ArmRight", "ArmLeft", "ThighRight", "ThighLeft", "ShinRight", "ShinLeft"]

        for p in range(len(self.poses)):

            for joint_label in joints_to_average.keys():

                if joint_label not in joints_labels_to_exclude:
                    self.poses[p].generate_average_joint(joints_to_average[joint_label], joint_label)
                    if remove_averaged_joints:
                        self.poses[p].remove_joints(joints_to_average[joint_label])

            if remove_non_kinect_joints:

                if p == 0:
                    list_kinect_joints = load_kinect_joints()
                    for joint_label in self.poses[p].joints.keys():
                        if joint_label not in list_kinect_joints:
                            joints_to_remove.append(joint_label)

                self.poses[p].remove_joints(joints_to_remove)

    def average_joints(self, joints_labels_to_average, new_joint_label, remove_averaged_joints=False):
        """Create a joint located at the average arithmetic distance of specified joint labels."""

        for pose in self.poses:

            pose.generate_average_joint(joints_labels_to_average, new_joint_label)

            if remove_averaged_joints:
                pose.remove_joints(joints_labels_to_average)

    def concatenate(self, other, delay):
        """Adds all the poses of another sequence at the end of the sequence."""

        last_pose_timestamp = self.poses[-1].timestamp

        for p in range(len(other.poses)):
            self.poses.append(other.poses[p])
            self.poses[p].timestamp = last_pose_timestamp + delay + self.poses[p].relative_timestamp

        self._calculate_relative_timestamps()

    # === Print functions ===

    def print_pose(self, pose_index):
        """Prints the information related to one specific pose of the sequence. """
        if len(self.poses) == 0:
            raise Exception("The Sequence does not have any pose.")
        elif not 0 <= pose_index < len(self.poses):
            raise Exception("The pose index must be between 0 and " + str(len(self.poses) - 1) + ". ")
        txt = "Pose " + str(pose_index + 1) + " of " + str(len(self.poses)) + "\n"
        txt += str(self.poses[pose_index])
        print(txt)

    def print_stats(self):
        """Prints a series of statistics related to the sequence. """
        stats = self.get_stats()
        for key in stats.keys():
            if key == "Duration":
                print(str(key) + ": " + str(stats[key]) + " s")
            elif key in ["Subject height", "Left arm length", "Right arm length"]:
                print(str(key) + ": " + str(round(stats[key], 3)) + " m")
            elif "Average velocity" in key:
                print(str(key) + ": " + str(round(stats[key] * 1000, 1)) + " mm/s")
            else:
                print(str(key) + ": " + str(stats[key]))

    # === Conversion functions ===
    def convert_to_table(self, use_relative_timestamp=False):
        """Returns a list of lists with headers, and containing the timestamps and coordinates for each joint."""

        table = []

        # For each pose
        for p in range(len(self.poses)):
            if p == 0:
                table = self.poses[0].convert_to_table(use_relative_timestamp)
            else:
                table.append(self.poses[p].convert_to_table(use_relative_timestamp)[1])

        return table

    def convert_to_json(self, use_relative_timestamp=False):
        """Returns a list ready to be exported in json."""

        data = []

        # For each pose
        for p in range(len(self.poses)):
            data.append(self.poses[p].convert_to_json(use_relative_timestamp))

        return data

    # === Saving functions ===

    def save(self, folder_out, name=None, file_format="json", individual=False, correct_timestamp=True, verbose=1):
        """Saves a sequence in a file or a folder."""

        # Automatic creation of all the folders of the path if they don't exist
        subfolder_creation(folder_out)

        if name is None and self.name is not None:
            name = self.name
        elif name is None:
            name = "out"

        file_format = file_format.strip(".")  # We remove the dot in the format
        if file_format == "xls":
            file_format = "xlsx"

        if verbose > 0:
            if individual:
                print("Saving " + file_format.upper() + " individual files...")
            else:
                print("Saving " + file_format.upper() + " global file: " + folder_out.strip("/") + "/" + name + "." +
                      file_format + "...")

        # If file format is JSON
        if file_format == "json":
            self._save_json(folder_out, name, individual, correct_timestamp, verbose)

        elif file_format == "mat":
            self._save_mat(folder_out, name, individual, correct_timestamp, verbose)

        # If file format is XLS or XLSX
        elif file_format == "xlsx":
            self._save_xlsx(folder_out, name, individual, correct_timestamp, verbose)

        # If file format is CSV or TXT
        else:
            self._save_txt(folder_out, name, file_format, individual, correct_timestamp, verbose)

        if verbose > 0:
            print("100% - Done.\n")

    def _save_json(self, folder_out, name=None, individual=False, use_relative_timestamp=False, verbose=1):
        """Saves a sequence as a json or mat file or files."""

        perc = 10  # Used for the progression percentage

        # Get the data
        data = self.convert_to_json(use_relative_timestamp)

        # Save the data
        if not individual:
            with open(folder_out + "/" + name + ".json", 'w', encoding="utf-16-le") as f:
                json.dump(data, f)
        else:
            for p in range(len(self.poses)):
                perc = show_percentage(verbose, p, len(self.poses), perc)
                with open(folder_out + "/frame_" + str(p) + ".json", 'w', encoding="utf-16-le") as f:
                    json.dump(data[p], f)

    def _save_mat(self, folder_out, name=None, individual=False, use_relative_timestamp=True, verbose=1):
        """Saves the sequence as a Matlab .mat file or files."""
        perc = 10  # Used for the progression percentage

        try:
            import scipy
        except ImportError:
            raise Exception("Module scipy not found; please install it to be able to save a file in mat format.")

        # Save the data
        if not individual:
            data = self.convert_to_json(use_relative_timestamp)
            scipy.io.savemat(folder_out + "/" + name + ".mat", {"data": data})
        else:
            for p in range(len(self.poses)):
                data = self.poses[p].convert_to_table(use_relative_timestamp)
                perc = show_percentage(verbose, p, len(self.poses), perc)
                scipy.io.savemat(folder_out + "/" + name + ".mat", {"data": data})

    def _save_xlsx(self, folder_out, name=None, individual=False, use_relative_timestamp=True, verbose=1):
        """Saves the sequence as an Excel file or Excel files."""
        try:
            import openpyxl as op
        except ImportError:
            raise Exception("Module openpyxl not found; please install it to be able to save a file in xlsx format.")

        workbook_out = op.Workbook()
        sheet_out = workbook_out.active

        perc = 10  # Used for the progression percentage

        # Save the data
        if not individual:
            data = self.convert_to_table(use_relative_timestamp)
            excel_write(workbook_out, sheet_out, data, folder_out + "/" + name + ".xlsx", perc, verbose)

        else:
            for p in range(len(self.poses)):
                data = self.poses[p].convert_to_table(use_relative_timestamp)
                perc = show_percentage(verbose, p, len(self.poses), perc)
                excel_write(workbook_out, sheet_out, data, folder_out + "/frame_" + str(p) + ".xlsx", perc, False)

    def _save_txt(self, folder_out, name=None, file_format="csv", individual=False, use_relative_timestamp=True,
                  verbose=1):
        """Saves the sequence as a csv, txt, tsv or custom extension file or files."""
        perc = 10  # Used for the progression percentage

        # Force comma or semicolon separator
        if file_format == "csv,":
            separator = ","
            file_format = "csv"
        elif file_format == "csv;":
            separator = ";"
            file_format = "csv"
        elif file_format[0:3] == "csv":  # Get the separator from local user (to get , or ;)
            separator = get_system_separator()
        elif file_format == "txt":  # For text files, tab separator
            separator = "\t"
        else:
            separator = "\t"

        # Save the data
        if not individual:
            data = table_to_string(self.convert_to_table(use_relative_timestamp), separator, perc, verbose)
            with open(folder_out + "/" + name + "." + file_format, 'w', encoding="utf-16-le") as f:
                f.write(data)
        else:
            for p in range(len(self.poses)):
                data = table_to_string(self.poses[p].convert_to_table(use_relative_timestamp), separator, perc, False)
                perc = show_percentage(verbose, p, len(self.poses), perc)
                with open(folder_out + "/frame_" + str(p) + "." + file_format, 'w', encoding="utf-16-le") as f:
                    f.write(data)

    # === Magic methods ===
    def __len__(self):
        """Returns the number of poses in the sequence (i.e., the length of the attribute poses).  """
        return len(self.poses)

    def __getitem__(self, index):
        """Returns the pose or poses of index specified by the parameter key."""
        return self.poses[index]

    def __repr__(self):
        """Returns the name attribute of the sequence."""
        return self.name
