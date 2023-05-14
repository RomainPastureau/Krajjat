#!/usr/bin/env python
"""Krajjat 1.10
Kinect Realignment Algorithm for Joint Jumps And Twitches
Author: Romain Pastureau
This file contains the main core functions. This is the file
to run to realign one or more recordings.
"""
from classes.audio import Audio
from classes.sequence import *
from tool_functions import *

__author__ = "Romain Pastureau"
__version__ = "1.10"
__email__ = "r.pastureau@bcbl.eu"
__license__ = "GPL"


def save_sequence(sequence, folder_out, original_sequence=None, name_out="global", file_format="json", individual=True,
                  correct_timestamp=False, verbose=True):
    """Saves a realigned sequence in json, txt, csv or xlsx. In the first case, the original sequence is necessary
    in order to only modify the timestamp and xyz position from the original file; all the other formats lose the
    information from the original sequence. It is possible to save the data with one file per pose or a meta file."""

    # Automatic creation of all the folders of the path if they don't exist
    subfolder_creation(folder_out)

    file_format = file_format.strip(".")  # We remove the dot in the format
    if file_format == "xls":
        file_format = "xlsx"

    if verbose:
        if individual:
            print("Saving " + file_format.upper() + " individual files...")
        else:
            print("Saving " + file_format.upper() + " global file" + folder_out + name_out + "." + file_format + "...")

    perc = 10  # Used for the progression percentage

    # If file format is JSON
    if file_format == "json":

        # Get the data
        if original_sequence is not None:
            data = merge_original_and_new_sequences(original_sequence, sequence, correct_timestamp, perc, verbose)
        else:
            data = sequence.convert_to_json(correct_timestamp)

        # Save the data
        if not individual:
            with open(folder_out + "/" + name_out + ".json", 'w', encoding="utf-16-le") as f:
                json.dump(data, f)
        else:
            for p in range(len(sequence.poses)):
                perc = show_percentage(verbose, p, len(sequence.poses), perc)
                with open(folder_out + "/frame_" + str(p) + ".json", 'w', encoding="utf-16-le") as f:
                    json.dump(data[p], f)

    elif file_format == "xlsx":
        import openpyxl as op
        workbook_out = op.Workbook()
        sheet_out = workbook_out.active

        # Save the data
        if not individual:
            data = sequence.convert_to_table(correct_timestamp)
            excel_write(workbook_out, sheet_out, data, folder_out + "/" + name_out + ".xlsx", perc, verbose)

        else:
            for p in range(len(sequence.poses)):
                data = sequence.poses[p].convert_to_table(correct_timestamp)
                perc = show_percentage(verbose, p, len(sequence.poses), perc)
                excel_write(workbook_out, sheet_out, data, folder_out + "/frame_" + str(p) + ".xlsx", perc, False)

    # If the file format is CSV or TXT
    else:

        # Force comma or semicolon separator
        if file_format == "csv,":
            separator = ","
            file_format = "csv"
        elif file_format == "csv;":
            separator = ";"
            file_format = "csv"
        elif file_format[0:3] == "csv":  # Get the separator from local user (to get , or ;)
            separator = get_system_separator()
        elif file_format == "txt":  # For text files, tab separator
            separator = "\t"
        else:
            separator = "\t"

        # Save the data
        if not individual:
            data = table_to_string(sequence.convert_to_table(correct_timestamp), separator, perc, verbose)
            with open(folder_out + "/" + name_out + "." + file_format, 'w', encoding="utf-16-le") as f:
                f.write(data)
        else:
            for p in range(len(sequence.poses)):
                data = table_to_string(sequence.poses[p].convert_to_table(correct_timestamp), separator, perc, False)
                perc = show_percentage(verbose, p, len(sequence.poses), perc)
                with open(folder_out + "/frame_" + str(p) + "." + file_format, 'w', encoding="utf-16-le") as f:
                    f.write(data)

    print("100% - Done.\n")


def save_audio(audio, folder_out, name_out="audio", file_format="xlsx", individual=False, verbose=1):

    # Automatic creation of all the folders of the path if they don't exist
    subfolder_creation(folder_out)

    file_format = file_format.strip(".")  # We remove the dot in the format
    if file_format == "xls":
        file_format = "xlsx"

    if verbose:
        if individual:
            print("Saving " + file_format.upper() + " individual files...")
        else:
            print("Saving " + file_format.upper() + " global file " + folder_out + "/" + name_out
                  + "." + file_format + "...")

    perc = 10  # Used for the progression percentage

    # If file format is JSON
    if file_format == "json":

        # Get the data
        data = audio.get_json_samples()

        # Save the data
        if not individual:
            with open(folder_out + "/" + name_out + ".json", 'w', encoding="utf-16-le") as f:
                json.dump(data, f)
        else:
            for s in range(len(audio.samples)):
                perc = show_percentage(verbose, s, len(audio.samples), perc)
                with open(folder_out + "/sample_" + str(s) + ".json", 'w', encoding="utf-16-le") as f:
                    json.dump(data[s], f)

    elif file_format == "xlsx":
        import openpyxl as op
        workbook_out = op.Workbook()
        sheet_out = workbook_out.active

        # Save the data
        if not individual:
            data = audio.convert_to_table()
            excel_write(workbook_out, sheet_out, data, folder_out + "/" + name_out + ".xlsx", perc, verbose)

        else:
            for s in range(len(audio.samples)):
                data = audio.samples[s].convert_to_table()
                perc = show_percentage(verbose, s, len(audio.samples), perc)
                excel_write(workbook_out, sheet_out, data, folder_out + "/sample_" + str(s) + ".xlsx", perc, False)

    # If the file format is CSV or TXT
    else:

        # Force comma or semicolon separator
        if file_format == "csv,":
            separator = ","
            file_format = "csv"
        elif file_format == "csv;":
            separator = ";"
            file_format = "csv"
        elif file_format[0:3] == "csv":  # Get the separator from local user (to get , or ;)
            separator = get_system_separator()
        elif file_format == "txt":  # For text files, tab separator
            separator = "\t"
        else:
            separator = "\t"

        # Save the data
        if not individual:
            data = table_to_string(audio.convert_to_table(), separator, perc, verbose)
            with open(folder_out + "/" + name_out + "." + file_format, 'w', encoding="utf-16-le") as f:
                f.write(data)
        else:
            for s in range(len(audio.samples)):
                data = table_to_string(audio.samples[s].convert_to_table(), separator, perc, False)
                perc = show_percentage(verbose, s, len(audio.samples), perc)
                with open(folder_out + "/sample_" + str(s) + "." + file_format, 'w', encoding="utf-16-le") as f:
                    f.write(data)

    print("100% - Done.\n")


def save_stats(sequenceOrSequences, output_file, verbose=1):
    """Saves the sequence or sequences statistics and information in a table file (xlsx, csv or txt)."""

    # Automatic creation of all the folders of the path if they don't exist
    subfolder_creation("/".join(output_file.split("/")[:-1]))

    # Open the sequence or sequences
    if type(sequenceOrSequences) is Sequence:
        sequenceOrSequences = [sequenceOrSequences]
    elif type(sequenceOrSequences) is str:
        sequenceOrSequences = load_sequences_recursive(sequenceOrSequences)

    file_format = output_file.split(".")[-1]
    if file_format == "xls":
        file_format = "xlsx"
        output_file += "x"

    perc = 10  # Used for the progression percentage

    if verbose:
        print("Saving " + file_format.upper() + " stat file...")

    # Get the table
    global_stats = []
    for sequence in sequenceOrSequences:
        stats = sequence.get_stats(tabled=True)
        if global_stats == []:
            global_stats.append(stats[0].copy())
        global_stats.append(stats[1].copy())

    if file_format == "xlsx":
        import openpyxl as op
        workbook_out = op.Workbook()
        sheet_out = workbook_out.active
        excel_write(workbook_out, sheet_out, global_stats, output_file, perc, verbose)

    else:
        # Force comma or semicolon separator
        if file_format == "csv,":
            separator = ","
            output_file = output_file[:-1]
        elif file_format == "csv;":
            separator = ";"
            output_file = output_file[:-1]
        elif file_format[0:3] == "csv":  # Get the separator from local user (to get , or ;)
            separator = get_system_separator()
        elif file_format == "txt":  # For text files, tab separator
            separator = "\t"
        else:
            separator = "\t"

        data = table_to_string(global_stats, separator, perc, False)
        with open(output_file, 'w', encoding="utf-16-le") as f:
            f.write(data)

    print("100% - Done.")


def realign_single(input_folder, output_folder, velocity_threshold, window, verbose=1):
    """Realigns one sequence and saves it in an output folder."""

    sequence = Sequence(input_folder)
    new_sequence = sequence.correct_jitter(velocity_threshold, window, verbose)
    save_sequence(sequence, new_sequence, output_folder)


def realign_folder(input_folder, output_folder, velocity_threshold, window, verbose=1):
    """Realigns all sequences in a folder and saves them in an output folder."""

    contents = os.listdir(input_folder)

    for c in contents:
        if os.path.isdir(input_folder + "/" + c):
            print("======= " + c + " =======")
            if not os.path.exists(output_folder + "/" + c):
                os.mkdir(output_folder + "/" + c)
            realign_single(input_folder + "/" + c, output_folder + "/" + c, velocity_threshold, window, verbose)


def realign_recursive(input_folder, output_folder, velocity_threshold, window, verbose=1):
    """Realigns all sequences in a recursive fashion, save them using the same folder structure."""

    contents = os.listdir(input_folder)

    for c in contents:
        if os.path.isdir(input_folder + "/" + c):
            subcontents = os.listdir(input_folder + "/" + c)
            for file_name in subcontents:
                if file_name.endswith('.json'):
                    print("======= " + input_folder + "/" + c + " =======")
                    realign_single(input_folder + "/" + c, output_folder + "/" + c, velocity_threshold, window, verbose)
                    break
            else:
                realign_recursive(input_folder + "/" + c, output_folder + "/" + c, velocity_threshold, window, verbose)


def realign_sequences(sequence_or_sequences, velocity_threshold, window, verbose=1):
    """Realigns sequences and returns them."""

    if sequence_or_sequences is Sequence:
        sequence_or_sequences = [sequence_or_sequences]

    output_sequences = []
    for sequence in sequence_or_sequences:
        output_sequences.append(sequence.realign(velocity_threshold, window, verbose))

    return(output_sequences)


def re_reference_single(input_folder, output_folder, reference_joint, place_at_zero, verbose=1):
    """Realigns one sequence and saves it in an output folder."""

    sequence = Sequence(input_folder)
    new_sequence = sequence.re_reference(reference_joint, place_at_zero, verbose)
    save_sequence(sequence, new_sequence, output_folder)


def re_reference_folder(input_folder, output_folder, reference_joint, place_at_zero, verbose=1):
    """Realigns all sequences in a folder and saves them in an output folder."""

    contents = os.listdir(input_folder)

    for c in contents:
        if os.path.isdir(input_folder + "/" + c):
            print("======= " + c + " =======")
            if not os.path.exists(output_folder + "/" + c):
                os.mkdir(output_folder + "/" + c)
            re_reference_single(input_folder + "/" + c, output_folder + "/" + c, reference_joint, place_at_zero,
                                verbose)


def re_reference_recursive(input_folder, output_folder, reference_joint, place_at_zero, verbose=1):
    """Realigns all sequences in a recursive fashion, save them using the same folder structure."""

    contents = os.listdir(input_folder)

    for c in contents:
        if os.path.isdir(input_folder + "/" + c):
            subcontents = os.listdir(input_folder + "/" + c)
            for file_name in subcontents:
                if file_name.endswith('.json'):
                    print("======= " + input_folder + "/" + c + " =======")
                    re_reference_single(input_folder + "/" + c, output_folder + "/" + c, reference_joint, place_at_zero,
                                        verbose)
                    break
            else:
                re_reference_recursive(input_folder + "/" + c, output_folder + "/" + c, reference_joint, place_at_zero,
                                       verbose)


def re_reference_sequences(sequence_or_sequences, reference_joint, place_at_zero, verbose=1):
    """Realigns sequences and returns them."""

    if sequence_or_sequences is Sequence:
        sequence_or_sequences = [sequence_or_sequences]

    output_sequences = []
    for sequence in sequence_or_sequences:
        output_sequences.append(sequence.re_reference(reference_joint, place_at_zero, verbose))

    return (output_sequences)


def trim_single(input_folder, output_folder, delay_beginning, path_audio=None, verbose=1):
    """Trims a sequence by deleting the poses at the beginning in the duration of delay_beginning, and the poses
    at the end after the duration of the audio file."""

    sequence = Sequence(input_folder)
    new_sequence = sequence.synchronize_with_audio(delay_beginning, path_audio, True, verbose)
    save_sequence(sequence, new_sequence, output_folder)


def trim_folder(input_folder, output_folder, delays_beginning, paths_audio=None, verbose=1):
    """Trims all sequences in a folder and saves them in an output folder."""

    contents = os.listdir(input_folder)

    # Iterative value
    i = 0

    for c in contents:
        if os.path.isdir(input_folder + "/" + c):
            print("======= " + c + " =======")
            if not os.path.exists(output_folder + "/" + c):
                os.mkdir(output_folder + "/" + c)

            if type(delays_beginning) is dict:
                db = delays_beginning[input_folder + "/" + c]
            elif type(delays_beginning) is list:
                db = delays_beginning[i]
            if type(paths_audio) is dict:
                pa = paths_audio[input_folder + "/" + c]
            elif type(paths_audio) is list:
                pa = paths_audio[i]
            elif type(paths_audio) is None:
                pa = None

            trim_single(input_folder + "/" + c, output_folder + "/" + c, db, pa, verbose)
            i += 1


def trim_recursive(input_folder, output_folder, delays_beginning, paths_audio=None, verbose=1):
    """Trims all sequences in a recursive fashion, save them using the same folder structure."""

    contents = os.listdir(input_folder)

    # Iterative value
    i = 0

    for c in contents:
        if os.path.isdir(input_folder + "/" + c):
                subcontents = os.listdir(input_folder + "/" + c)
                for file_name in subcontents:
                    if file_name.endswith('.json'):
                        print("======= " + input_folder + "/" + c + " =======")

                        if type(delays_beginning) is dict:
                            db = delays_beginning[input_folder + "/" + c]
                        elif type(delays_beginning) is list:
                            db = delays_beginning[i]
                        if type(paths_audio) is dict:
                            pa = paths_audio[input_folder + "/" + c]
                        elif type(paths_audio) is list:
                            pa = paths_audio[i]
                        elif type(paths_audio) is None:
                            pa = None

                        trim_single(input_folder + "/" + c, output_folder + "/" + c, db, pa, verbose)
                        i += 1
                        break
                else:
                    if type(delays_beginning) is dict:
                        db = delays_beginning
                    elif type(delays_beginning) is list:
                        db = delays_beginning[i:]
                    if type(paths_audio) is dict:
                        pa = paths_audio
                    elif type(paths_audio) is list:
                        pa = paths_audio[i:]
                    elif type(paths_audio) is None:
                        pa = None

                    trim_recursive(input_folder + "/" + c, output_folder + "/" + c, db, pa, verbose)


def trim_sequences(sequence_or_sequences, delays_beginning, paths_audio=None, verbose=1):
    """Realigns sequences and returns them."""

    if sequence_or_sequences is Sequence:
        sequence_or_sequences = [sequence_or_sequences]

    # Iteration value
    i = 0

    output_sequences = []
    for sequence in sequence_or_sequences:
        print("\n=== "+str(sequence.path)+" ===")
        if type(delays_beginning) is dict:
            db = delays_beginning[sequence.path]
        elif type(delays_beginning) is list:
            db = delays_beginning[i]
        else:
            return Exception("Delays_beginning has to be a list or a dict.")
        if type(paths_audio) is dict:
            pa = paths_audio[sequence.path]
        elif type(paths_audio) is list:
            pa = paths_audio[i]
        else:
            print("Warning: no path_audio defined.")
            pa = None

        output_sequences.append(sequence.synchronize(db, pa, True, verbose))
        i += 1

    return (output_sequences)


def resample_single(input_folder, output_folder, frequency, mode="cubic", verbose=1):
    """Resamples one sequence and saves it in an output folder."""

    sequence = Sequence(input_folder)
    new_sequence = sequence.resample(frequency, mode, verbose)
    save_sequence(sequence, new_sequence, output_folder)


def resample_folder(input_folder, output_folder, frequency, mode="cubic", verbose=1):
    """Resamples all sequences in a folder and saves them in an output folder."""

    contents = os.listdir(input_folder)

    for c in contents:
        if os.path.isdir(input_folder + "/" + c):
            print("======= " + c + " =======")
            if not os.path.exists(output_folder + "/" + c):
                os.mkdir(output_folder + "/" + c)
            re_reference_single(input_folder + "/" + c, output_folder + "/" + c, frequency, mode, verbose)


def resample_recursive(input_folder, output_folder, frequency, mode, verbose=1):
    """Resamples all sequences in a recursive fashion, save them using the same folder structure."""

    contents = os.listdir(input_folder)

    for c in contents:
        if os.path.isdir(input_folder + "/" + c):
            subcontents = os.listdir(input_folder + "/" + c)
            for file_name in subcontents:
                if file_name.endswith('.json'):
                    print("======= " + input_folder + "/" + c + " =======")
                    re_reference_single(input_folder + "/" + c, output_folder + "/" + c, frequency, mode, verbose)
                    break
            else:
                re_reference_recursive(input_folder + "/" + c, output_folder + "/" + c, frequency, mode, verbose)


def resample_sequences(sequence_or_sequences, frequency, mode, verbose=1):
    """Resamples sequences and returns them."""

    if sequence_or_sequences is Sequence:
        sequence_or_sequences = [sequence_or_sequences]

    output_sequences = []
    for sequence in sequence_or_sequences:
        output_sequences.append(sequence.re_reference(frequency, mode, verbose))

    return (output_sequences)

def load_sequences_folder(input_folder):
    """Opens sequences contained in a folder."""
    print("Loading sequences...")
    return (load_sequences(input_folder, False))


def load_sequences_recursive(input_folder):
    """Finds and opens sequences recursively from a folder."""
    print("Loading sequences recursively...")
    return (load_sequences(input_folder, True))


def load_sequences(input_folder, recursive=False):

    content = os.listdir(input_folder)
    sequences = []

    for c in content:
        try:
            print("======= " + input_folder + "/" + c + " =======")
            sequence = Sequence(input_folder + "/" + c, name=c)
            sequences.append(sequence)
        except Exception as ex:
            print("The path " + input_folder + "/" + c + " is not a valid sequence.")
            if hasattr(ex, 'message'):
                print(ex.message)
            else:
                print(ex)
            if recursive:
                if os.path.isdir(input_folder + "/" + c):
                    sequences += load_sequences(input_folder + "/" + c, True)

    print("All the sequences have been loaded.")

    return(sequences)

def load_audios_folder(input_folder):
    """Opens sequences contained in a folder."""
    print("Loading audios...")
    return (load_audios(input_folder, False))


def load_audios_recursive(input_folder):
    """Finds and opens sequences recursively from a folder."""
    print("Loading audios recursively...")
    return (load_audios(input_folder, True))


def load_audios(input_folder, recursive=False):

    content = os.listdir(input_folder)
    audios = []

    for c in content:
        try:
            print("======= " + input_folder + "/" + c + " =======")
            audio = Audio(input_folder + "/" + c, name=c)
            audios.append(audio)
        except Exception as ex:
            print("The path " + input_folder + "/" + c + " is not a valid sequence.")
            if hasattr(ex, 'message'):
                print(ex.message)
            else:
                print(ex)
            if recursive:
                if os.path.isdir(input_folder + "/" + c):
                    audios += load_sequences(input_folder + "/" + c, True)

    print("All the sequences have been loaded.")

    return(audios)