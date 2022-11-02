#!/usr/bin/env python
"""Krajjat 1.8
Kinect Realignment Algorithm for Joint Jumps And Twitches
Author: Romain Pastureau
This file contains all the classes used by the other files,
along with the realignment algorithm.
"""

import copy
import math
import os
import random
import statistics
import tool_functions
from collections import OrderedDict

__author__ = "Romain Pastureau"
__version__ = "1.8"
__email__ = "r.pastureau@bcbl.eu"
__license__ = "GPL"


class Sequence(object):
    """Defines a motion sequence by opening an existing one or creating a new."""

    def __init__(self, path):
        self.files = None  # List of the files in the target folder
        self.poses = []  # List of the poses in the sequence, ordered
        self.randomized = False  # True if the joints positions are randomized

        # In the case where a file or folder is passed as argument,
        # we load the sequence
        if path is not None:
            self.path = path

            # If it's a folder, we fetch all the files

            if os.path.isdir(self.path):
                self.get_files()  # Fetches all the files
            self.load_poses()  # Loads the files into poses
            self.load_relative_timestamps()  # Sets the relative time from the first pose for each pose

        if len(self.poses) == 0:
            raise Exception("The path "+path+" is not a valid sequence.")

        self.calculate_velocity()

    def get_files(self):
        """Opens a folder and fetches all the individual files (.json, .csv, .txt or .xlsx)."""

        file_list: list[str] | list[bytes] = tool_functions.os.listdir(self.path)  # List all the files in the folder
        self.files = ["" for _ in range(len(file_list))]  # Create an empty list the length of the files

        # Here, we find all the files that are either .json or .meta in the folder.
        for f in file_list:

            # If a file is ".json", then we get its index from its name to order it correctly in the list.
            # This is necessary as some file systems will order frame_2.json after frame_11.json because,
            # alphabetically, "1" is before "2".
            if f.split(".")[-1] in ["json", "csv", "txt", "xlsx"]:
                self.files[int(f.split(".")[0].split("_")[1])] = f

        # If there were files that weren't ".json" in the folder, then "self.files" is larger than the number of
        # frames. The list is thus ending by a series of empty strings that we trim.
        if "" in self.files:
            limit = self.files.index("")
            self.files = self.files[0:limit]

        print(str(len(self.files)) + " pose file(s) found.\n")

    def load_poses(self, verbose=True):
        """Opens successively all the .json files, read them and creates a Pose object for each of them.
        If verbose is True, returns a progression percentage every 10%."""

        if verbose:
            print("Opening poses...")

        perc = 10  # Used for the progression percentage

        # If the path is a folder, we load every single file
        if tool_functions.os.path.isdir(self.path):

            for i in range(len(self.files)):

                # Show percentage if verbose
                perc = tool_functions.show_percentage(verbose, i, len(self.files), perc)

                # Create the Pose object, passes as parameter the index and the file path
                self.load_single_file(i, self.path + "/" + self.files[i])

            if verbose:
                print("100% - Done.\n")

        # Otherwise, we load the one file
        else:
            self.load_global_file(verbose)

    def load_single_file(self, pose_number, path):

        file = None
        values = None
        timestamp = None

        # JSON file
        if path.split(".")[-1] == "json":
            data = tool_functions.open_json(path)
            values = data["Bodies"][0]["Joints"]
            timestamp = data["Timestamp"]
            file = self.path.split("/")[-1]

        # Excel file
        elif path.split(".")[-1] == "xlsx":
            data, joints_labels = tool_functions.open_xlsx(path)

            values = OrderedDict()
            for i in range(len(joints_labels)):
                values[joints_labels[i]] = float(data.cell(1, i+1).value)

            values, timestamp = tool_functions.table_to_json_joints(values)

        elif path.split(".")[-1] in ["txt", "csv"]:

            # Open the file and read the data
            separator = tool_functions.get_filetype_separator(path)
            data = tool_functions.open_txt(path)

            # Get the joints labels and values
            joints_labels = data[0].split(separator)
            elements = data[1].split(separator)

            values = OrderedDict()
            for i in range(len(joints_labels)):
                values[joints_labels[i]] = float(elements[i])

            values, timestamp = tool_functions.table_to_json_joints(values)

        # Create the Pose object, passes as parameter the index and the data
        pose = Pose(pose_number, values, file)
        pose.set_timestamp(timestamp)
        self.poses.append(pose)

    def load_global_file(self, verbose):
        """Loads a global file containing the whole sequence instead of single files for each pose."""

        perc = 10  # Used for the progression percentage

        # JSON file
        if self.path.split(".")[-1] == "json":

            # Open the file and read the data
            data = tool_functions.open_json(self.path)

            for p in range(len(data)):

                # Show percentage if verbose
                perc = tool_functions.show_percentage(verbose, p, len(data), perc)

                # Create the Pose object, passes as parameter the index and the data
                pose = Pose(p, data[p]["Bodies"][0]["Joints"])
                pose.set_timestamp(data[p]["Timestamp"])
                self.poses.append(pose)

        # Excel file
        elif self.path.split(".")[-1] == "xlsx":
            import openpyxl as op
            workbook = op.load_workbook(self.path)
            sheet = workbook[workbook.sheetnames[0]]

            # Get the labels (timestamp and joint labels) from the first row
            joints_labels = []
            for cell in sheet["1"]:
                joints_labels.append(str(cell.value))

            # For each pose
            for p in range(len(sheet["A"])-1):

                # Get the values (coordinates)
                values = OrderedDict()
                for i in range(len(joints_labels)):
                    values[joints_labels[i]] = float(sheet.cell(p+2, i+1).value)
                values, timestamp = tool_functions.table_to_json_joints(values)

                # Create the Pose object, passes as parameter the index and the data
                pose = Pose(p, values)
                pose.set_timestamp(timestamp)
                self.poses.append(pose)

        # Text file (csv or txt)
        elif self.path.split(".")[-1] in ["csv", "txt"]:

            separator = tool_functions.get_filetype_separator(self.path)

            # Open the file and read the data
            data = tool_functions.open_txt(self.path)

            # Get the joints labels
            joints_labels = data[0].split(separator)

            # For each pose
            for line in range(len(data)-1):
                elements = data[line+1].split(separator)
                values = OrderedDict()
                for i in range(len(joints_labels)):
                    values[joints_labels[i]] = float(elements[i])

                values, timestamp = tool_functions.table_to_json_joints(values)

                # Create the Pose object, passes as parameter the index and the data
                pose = Pose(line, values)
                pose.set_timestamp(timestamp)
                self.poses.append(pose)

        if verbose:
            print("100% - Done.\n")

    def load_relative_timestamps(self):
        """For all the poses, sets the relative time from the first pose, by subtracting the absolute UNIX time
        of the first pose from the absolute UNIX time of every pose."""
        t = self.poses[0].get_timestamp()
        for f in self.poses:
            f.set_relative_timestamp(t)
            
    def calculate_velocity(self):
        """Calculates the distance travelled by each joint between two poses, and divides it by the time elapsed
        between the poses."""

        for p in range(1, len(self.poses)):  # For each pose

            for j in self.poses[0].joints.keys():  # For each joint
                velocity = self.get_velocity(self.poses[p - 1], self.poses[p], j)  # Get the velocity
                self.poses[p].joints[j].set_velocity(velocity)  # Set the velocity

    def get_velocity(self, pose1, pose2, joint):
        """Given two poses and a joint, returns the velocity of the joint between the two poses. To do so, this
        function calculates the distance travelled by the joint between the two poses and divides it by the
        time elapsed between the two poses. The returned value is in cm/s."""

        # Get the distance travelled by a joint between two poses (meters)
        dist = self.get_distance(pose1.joints[joint], pose2.joints[joint])

        # Get the time elapsed between two poses (hundreds of seconds)
        delay = self.get_delay(pose1, pose2) / 1000000000

        # Calculate the velocity (meters per hundreds of seconds, or centimeters per second)
        velocity = dist / delay

        return velocity

    def get_total_velocity(self, joint):
        """Returns the sum of the velocity of a joint for the whole duration of the sequence."""
        total_velocity = 0

        for p in range(1, len(self.poses)):
            total_velocity += self.get_velocity(self.poses[p-1], self.poses[p], joint)

        return total_velocity

    @staticmethod
    def get_distance(joint_a, joint_b):
        """Uses the Euclidian formula to calculate the distance between two joints. This can be used to calculate
        the distance travelled by one joint between two poses, or the distance between two joints on the same pose."""
        x = (joint_b.x - joint_a.x) ** 2
        y = (joint_b.y - joint_a.y) ** 2
        z = (joint_b.z - joint_a.z) ** 2

        return math.sqrt(x + y + z)

    @staticmethod
    def get_delay(frame_a, frame_b):
        """Returns the delay between two poses, in tenths of microsecond (10^-7)."""
        time_a = frame_a.get_timestamp()
        time_b = frame_b.get_timestamp()
        return time_b - time_a

    @staticmethod
    def get_x(joint):
        """Returns the x value of a joint."""
        return joint.x

    @staticmethod
    def get_y(joint):
        """Returns the y value of a joint."""
        return joint.y

    @staticmethod
    def get_z(joint):
        """Returns the z value of a joint."""
        return joint.z

    def realign(self, velocity_threshold, window, verbose=False):
        """Realignment function: detects twitches and jumps in a sequence, corrects them linearly and outputs
        a new realigned sequence. For more information on how the realignment is performed, see the documentation."""

        # Create an empty sequence, the same length in poses as the original, just keeping the timestamp information
        # for each pose.
        new_sequence = self.create_empty_sequence()

        # If verbose, show all the information. Else, show only the progress in percentage.
        print("Starting realignment...")

        # Define the counters
        realigned_points = 0
        perc = 10

        # For every pose starting on the second one
        for p in range(1, len(self.poses)):

            if verbose:
                print("\nNew sequence:\n" + str(new_sequence.poses[p - 1]))
                print("\n== POSE NUMBER " + str(p + 1) + " ==")

            if not verbose and p / len(self.poses) > perc / 100:
                print(str(perc) + "%", end=" ")
                perc += 10

            # For every joint
            for j in self.poses[0].joints.keys():

                # We get the cm/s travelled by the joint between this pose and the previous
                velocity_before = self.get_velocity(new_sequence.poses[p - 1], self.poses[p], j)

                if verbose:
                    print(j + ": " + str(velocity_before))

                # If we already corrected this joint, we ignore it to avoid overcorrection
                if j in new_sequence.poses[p].joints:

                    if verbose:
                        print("\tAlready corrected.")

                # If the velocity is over threshold, we check if it is a jump or a twitch
                elif velocity_before > velocity_threshold and p != len(self.poses) - 1:

                    if verbose:
                        print("\tVelocity over threshold. Check in subsequent poses...")

                    self.poses[p].joints[j].set_move_much(True)

                    # We check the next poses (defined by the window) if the position of the joint comes back below
                    # threshold
                    for k in range(p, min(p + window, len(self.poses))):

                        if verbose:
                            print("\t\tPose " + str(k + 1) + ":", end=" ")

                        # Method 1 (original): get the subsequent velocities
                        velocity = self.get_velocity(new_sequence.poses[p - 1], self.poses[k], j)

                        # Method 2 (to test): get the distance over the time between the two first joints
                        # dist = self.get_distance(new_sequence.poses[p - 1].joints[joint], self.poses[k].joints[joint])
                        # delay = self.get_delay(new_sequence.poses[p - 1], new_sequence.poses[p]) / 1000000000
                        # velocity = dist / delay

                        # Twitch case: One of the poses of the window is below threshold compared to previous pose.
                        if velocity < velocity_threshold:

                            if verbose:
                                print("no threshold deviation compared to pose " + str(p + 1) + ".")
                                print("\t\t\tCorrecting for twitch...")

                            new_sequence, realigned_points = self.correction(new_sequence, p - 1, k,
                                                                             j, realigned_points, verbose)

                            break

                        # Jump case: No pose of the window is below threshold.
                        elif k == p + window - 1 or k == len(self.poses) - 1:

                            if verbose:
                                print("original deviation not found.")
                                print("\t\t\tCorrecting for jump...")

                            new_sequence, realigned_points = self.correction(new_sequence, p - 1, k,
                                                                             j, realigned_points, verbose)

                            break

                        # Wait: if there are still poses in the window, we continue.
                        else:

                            if verbose:
                                print("still over threshold.")

                    if verbose:
                        print("")

                # If we're still in threshold, there is no correction, we copy the joint
                else:

                    joint = self.copy_joint(p, j)
                    new_sequence.poses[p].add_joint(j, joint)
                    new_sequence.poses[p].joints[j].corrected = False

        print("100% - Done.\n")
        print("Realignment over. " + str(realigned_points) + " point(s) realigned over " + str(
            len(self.poses) * len(list(self.poses[0].joints.keys()))) + ".\n")

        return new_sequence

    def correction(self, new_sequence, start_pose_number, end_pose_number, joint_name, realigned_points, verbose):
        """Performs a jump or twitch correction"""

        # We extract the data from the joints at the beginning and at the end of the partial sequence to correct
        joint_before = new_sequence.poses[start_pose_number].joints[joint_name]
        joint_after = self.poses[end_pose_number].joints[joint_name]

        # If the starting and ending joint are the same, we don't correct, it means it is the last pose of the sequence
        if start_pose_number == end_pose_number:
            joint = self.copy_joint(start_pose_number, joint_name)
            new_sequence.poses[start_pose_number].add_joint(joint_name, joint)

            if verbose:
                print("\t\t\t\tDid not corrected joint " + str(start_pose_number) +
                      " as it is the last pose of the sequence.")

        # Otherwise we correct for all the intermediate joints
        for pose_number in range(start_pose_number + 1, end_pose_number):

            # If a joint was already corrected we don't correct it to avoid overcorrection
            if self.poses[pose_number].joints[joint_name].corrected:

                if verbose:
                    print("\t\t\t\tDid not corrected joint " + str(pose_number + 1) + " as it was already corrected.")

            # Otherwise we correct it
            else:

                # We get the new coordinates
                x, y, z = self.correct_joint(joint_before, joint_after, start_pose_number, pose_number, end_pose_number,
                                             verbose)

                # We copy the original joint, apply the new coordinates and add it to the new sequence
                joint = self.copy_joint(pose_number, joint_name)
                joint.correct_joint(x, y, z)
                new_sequence.poses[pose_number].add_joint(joint_name, joint)

                if verbose:
                    print("\t\t\t\tCorrecting joint: " + str(pose_number + 1) + ". Original coordinates: (" + str(
                        self.poses[pose_number].joints[joint_name].x) +
                          ", " + str(self.poses[pose_number].joints[joint_name].y) + ", " + str(
                        self.poses[pose_number].joints[joint_name].z) + ")")
                    print("\t\t\t\tPrevious joint: " + str(start_pose_number + 1) + ". (" + str(
                        new_sequence.poses[start_pose_number].joints[joint_name].x) +
                          ", " + str(new_sequence.poses[start_pose_number].joints[joint_name].y) + ", " + str(
                        new_sequence.poses[start_pose_number].joints[joint_name].z) + ")")
                    print("\t\t\t\tNext joint: " + str(end_pose_number + 1) + ". (" + str(
                        self.poses[end_pose_number].joints[joint_name].x) +
                          ", " + str(self.poses[end_pose_number].joints[joint_name].y) + ", " + str(
                        self.poses[end_pose_number].joints[joint_name].z) + ")")
                    print("\t\t\t\tCorrected joint " + str(pose_number + 1) + ". New coordinates: (" + str(
                        x) + ", " + str(y) + ", " + str(z) + ")\n")

                # To count the number of joints that were realigned overall
                realigned_points += 1

        return new_sequence, realigned_points

    def correct_joint(self, joint_before, joint_after, pose_before, pose_current, pose_after, verbose):
        """Corrects one single joint (depending on the time between frames)"""

        # We calculate the percentage of time elapsed between the pose at the beginning of the partial sequence to
        # correct and the pose at the end
        percentage_time = (self.poses[pose_current].timestamp - self.poses[pose_before].timestamp) / (
                self.poses[pose_after].timestamp - self.poses[pose_before].timestamp)

        if verbose:
            print("\t\t\t\tJoint " + str(pose_current + 1) + " positioned at " + str(
                percentage_time * 100) + " % between poses " + str(pose_before + 1) + " and " + str(
                pose_after + 1) + ".")

        # We correct linearly every joint. See the documentation for more precision.
        x = joint_before.x - percentage_time * (joint_before.x - joint_after.x)
        y = joint_before.y - percentage_time * (joint_before.y - joint_after.y)
        z = joint_before.z - percentage_time * (joint_before.z - joint_after.z)

        return x, y, z

    def re_reference(self, reference_joint="SpineMid", place_at_zero=True, verbose=False):
        """Subtracts the movement of one joint to the rest of the joints, to measure the movements from the
        joints independently of the global movements of the body in the room."""

        # Create an empty sequence, the same length in poses as the original, just keeping the timestamp information
        # for each pose.
        new_sequence = self.create_empty_sequence()

        # If verbose, show all the information. Else, show only the progress in percentage.
        print("Starting normalization...")

        # Get reference : position of the reference joint at time 0
        if place_at_zero:
            start_ref_x = 0  # self.poses[0].joints[reference_joint].x
            start_ref_y = 0  # self.poses[0].joints[reference_joint].y
            start_ref_z = 0  # self.poses[0].joints[reference_joint].z
        else:
            start_ref_x = self.poses[0].joints[reference_joint].x
            start_ref_y = self.poses[0].joints[reference_joint].y
            start_ref_z = self.poses[0].joints[reference_joint].z

        # Define the percentage counter
        perc = 10

        # For every pose starting on the second one
        for p in range(len(self.poses)):

            if verbose:
                print("\n== POSE NUMBER " + str(p + 1) + " ==")

            if not verbose and p / len(self.poses) > perc / 100:
                print(str(perc) + "%", end=" ")
                perc += 10

            # Get movement from the reference point
            curr_ref_x = self.poses[p].joints[reference_joint].x
            curr_ref_y = self.poses[p].joints[reference_joint].y
            curr_ref_z = self.poses[p].joints[reference_joint].z

            # Compute the differences
            diff_ref_x = curr_ref_x - start_ref_x
            diff_ref_y = curr_ref_y - start_ref_y
            diff_ref_z = curr_ref_z - start_ref_z

            if verbose:
                print("Removing " + str(diff_ref_x) + " to every joint in x;")
                print("Removing " + str(diff_ref_y) + " to every joint in y;")
                print("Removing " + str(diff_ref_z) + " to every joint in z.")

            # For every joint
            for j in self.poses[0].joints.keys():

                # Compute new joint position
                if j == reference_joint:
                    new_x = 0
                    new_y = 0
                    new_z = 0
                else:
                    new_x = self.poses[p].joints[j].x - diff_ref_x
                    new_y = self.poses[p].joints[j].y - diff_ref_y
                    new_z = self.poses[p].joints[j].z - diff_ref_z

                # Add to the sequence
                joint = self.copy_joint(p, j)
                joint.correct_joint(new_x, new_y, new_z)
                new_sequence.poses[p].add_joint(j, joint)

                if verbose:
                    print(j + ": ")
                    print("X: " + str(self.poses[p].joints[j].x) + " -> " + str(new_x))
                    print("Y: " + str(self.poses[p].joints[j].y) + " -> " + str(new_y))
                    print("Z: " + str(self.poses[p].joints[j].z) + " -> " + str(new_z))

        print("100% - Done.\n")
        print("Normalization over.")
        return new_sequence

    def create_empty_sequence(self, verbose=True):
        """Creates an empty sequence with empty poses and empty joints, with the same number of poses as
        the sequence in which it is created. The only data kept for the poses is the timestamp and relative
        timestamp."""

        if verbose:
            print("Creating an empty sequence...")
        new_sequence: Sequence = Sequence(None)

        # Used for the progression percentage
        perc = 10
        for p in range(len(self.poses)):

            # Show percentage if verbose
            if verbose and p / len(self.poses) > perc / 100:
                print(str(perc) + "%", end=" ")
                perc += 10

            pose = self.poses[p].get_clear_copy()  # Creates an empty pose with the same timestamp as the original
            new_sequence.poses.append(pose)  # Add to the sequence

        new_sequence = self.copy_pose(new_sequence, 0)

        if verbose:
            print("100% - Done.\n")

        return new_sequence

    def copy_pose(self, sequence, pose_number):
        """Copies the joints from a pose that doesn't already exist in the target sequence"""

        for i in self.poses[pose_number].joints.keys():  # For every joint

            # If this joint doesn't already exist in the new sequence, we copy it
            if i not in sequence.poses[pose_number].joints.keys():
                sequence.poses[pose_number].joints[i] = self.copy_joint(pose_number, i)

        return sequence

    def copy_joint(self, pose_number, joint_name):
        """Creates a copy of a single joint and returns it"""

        joint = copy.copy(self.poses[pose_number].get_joint(joint_name))

        return joint

    def get_timestamps(self):
        """Returns a list of all the timestamps (relative) from the poses"""

        timestamps = []
        for pose in self.poses:
            timestamps.append(pose.get_relative_timestamp())

        return timestamps

    def get_table(self, include_relative_timestamp=False):
        """Returns a list of lists containing a header and the x, y and z coordinates of the joints for all the
        poses of the sequence"""

        table = []

        # For each pose
        for p in range(len(self.poses)):
            if p == 0:
                table = self.poses[0].get_table(include_relative_timestamp)
            else:
                table.append(self.poses[p].get_table(include_relative_timestamp)[1])

        return table

    def get_json_joint_list(self, include_relative_timestamp=False):
        """Returns a json format containing the timestamps and the x, y ans z coordinates of the joints for all
        the poses of the sequence"""

        data = []

        # For each pose
        for p in range(len(self.poses)):
            data.append(self.poses[p].get_json_joint_list(include_relative_timestamp))

        return data
    
    def print_json(self, pose):
        """Print the original json data from a pose"""
        data = self.poses[pose].get_json_data()
        print(tool_functions.json.dumps(data, indent=4))

    def print_pose(self, pose):
        """Prints all the information relative to one pose"""
        txt = "Pose " + str(pose + 1) + " of " + str(len(self.poses)) + "\n"
        txt += str(self.poses[pose])
        print(txt)

    def randomize(self):
        """Randomizes the joints"""

        # Get the starting positions of all joints
        starting_positions = []
        for joint in self.poses[0].joints.keys():
            starting_positions.append(copy.copy(self.poses[0].joints[joint]))

        # Randomize new starting positions for joints
        randomized_positions = self.generate_random_joints()
        random.shuffle(randomized_positions)

        # Assign these new starting positions to the joints and moves the joints in every subsequent pose
        # in relation to the new starting positions
        joints_list = list(self.poses[0].joints.keys())
        for p in self.poses:
            for j in range(len(joints_list)):
                p.joints[joints_list[j]].move_joint_randomized(starting_positions[j], randomized_positions[j])
        self.randomized = True

    def get_framerate(self):
        """Returns the framerate across time for the whole sequence."""
        time_points = []
        framerates = []
        for p in range(1, len(self.poses)-1):
            time_points.append(self.poses[p].relative_timestamp)
            framerate = 1/(self.poses[p].relative_timestamp - self.poses[p-1].relative_timestamp)
            framerates.append(framerate)

        return framerates, time_points

    def get_stats(self, tabled=False):
        """Returns some statistics and information over the sequence."""
        stats = OrderedDict()
        stats["Path"] = self.path
        stats["Duration"] = self.poses[-1].relative_timestamp
        stats["Number of poses"] = len(self.poses)

        # Framerate stats
        framerates, time_points = self.get_framerate()
        stats["Average framerate"] = sum(framerates)/len(framerates)
        stats["SD framerate"] = statistics.stdev(framerates)
        stats["Max framerate"] = max(framerates)
        stats["Min framerate"] = min(framerates)

        # Movement stats
        for joint in self.poses[0].joints.keys():
            stats["Total velocity "+joint] = self.get_total_velocity(joint)

        if tabled:
            table = [[],[]]
            for key in stats.keys():
                table[0].append(key)
                table[1].append(stats[key])
            return table

        return stats

    def print_stats(self):
        """Prints the statistics and information over the sequence."""
        stats = self.get_stats()
        for key in stats.keys():
            print(str(key)+": "+str(stats[key]))

    def __len__(self):
        """Return the amount of poses in the sequence"""
        return len(self.poses)

    @staticmethod
    def generate_random_joints():
        """Creates uniform, random positions for the joints"""
        random_joints = []
        for i in range(21):
            x = random.uniform(-0.2, 0.2)
            y = random.uniform(-0.5, 0.5)
            z = random.uniform(-0.5, 0.5)
            j = Joint(None, [x, y, z])
            random_joints.append(j)
        return random_joints




class Pose(object):
    """Defines a pose (frame) pertaining to a motion sequence."""

    def __init__(self, no, data, file=None):
        self.no = no  # Index of the pose in the sequence
        self.joints = OrderedDict()  # Dictionary of joint objects
        self.timestamp = None  # Original timestamp of the pose
        self.relative_timestamp = None  # Timestamp relative to the first pose
        self.file = file

        if data is not None:
            self.read(data)

    def read(self, data):
        """Opens the file, reads the content and create the joints"""
        for joint in data:
            self.joints[joint["JointType"]] = Joint(joint)

    def get_json_data(self):
        """Reads and returns the contents of one json file"""
        return tool_functions.open_json(self.file)

    def add_joint(self, name, joint):
        """Adds a joint to the pose"""
        self.joints[name] = joint

    def get_joint(self, name):
        """Returns a joint object from the name of the joint"""
        return self.joints[name]

    def get_timestamp(self):
        """Return the original json timestamp"""
        return self.timestamp

    def set_timestamp(self, timestamp):
        """Sets the original json timestamp"""
        self.timestamp = float(timestamp)

    def set_relative_timestamp(self, t):
        """Sets the relative time compared to the time from the first pose (in sec)"""
        self.relative_timestamp = (self.timestamp - t) / 10000000

    def get_relative_timestamp(self):
        """Gets the relative time of this pose compared to the first one"""
        return self.relative_timestamp

    def get_table(self, include_relative_timestamp=False):
        """Returns the table version of the pose"""
        labels = ["Timestamp"]
        if include_relative_timestamp:
            values = [self.relative_timestamp]
        else:
            values = [self.timestamp]

        for joint in self.joints.keys():
            labels.append(self.joints[joint].joint_type + "_X")
            labels.append(self.joints[joint].joint_type + "_Y")
            labels.append(self.joints[joint].joint_type + "_Z")
            values.append(self.joints[joint].x)
            values.append(self.joints[joint].y)
            values.append(self.joints[joint].z)
        return [labels, values]

    def get_json_joint_list(self, include_relative_timestamp=False):

        data = {"Bodies": [{"Joints": []}]}

        if include_relative_timestamp:
            data["Timestamp"] = self.relative_timestamp
        else:
            data["Timestamp"] = self.timestamp

        joints = []
        for j in self.joints.keys():
            joint = {"JointType": j.joint_type, "Position": {"X": j.x, "Y": j.y, "Z": j.z}}
            joints.append(joint)

        data["Bodies"][0]["Joints"] = joints

        return data

    def __repr__(self):
        """Prints all the joints"""
        txt = ""
        if len(list(self.joints.keys())) == 0:
            txt = "Empty list of joints."
        for j in self.joints.keys():
            txt += str(self.joints[j]) + "\n"
        return txt

    def get_clear_copy(self):
        """Creates a copy of the pose and returns it"""
        p = Pose(self.no, self.file)
        p.joints = {}
        p.set_timestamp(self.timestamp)
        p.relative_timestamp = self.relative_timestamp
        return p


class Joint(object):
    """Defines a joint pertaining to a pose."""
    
    def __init__(self, data, coord=None):
        
        self.joint_type = None  # Label of the joint (e.g. "Head")
        
        # If we don't define a coordinate, then we read the data
        if coord is None:
            self.read_data(data)
            self.color = (255, 255, 255)  # Color white for the unchanged joints
            
        # Otherwise, we create a joint with new coordinates
        else:
            self.x = float(coord[0])
            self.y = float(coord[1])
            self.z = float(coord[2])
            self.color = (0, 255, 0)  # Color green for the corrected joints

        self.position = [self.x, self.y, self.z]  # List of the x, y and z coordinates of the joint
        self.velocity = None  # Velocity of the joint compared to the previous pose
        self.move_much = False  # True if this joint has moved above threshold during correction
        self.corrected = False  # True if this joint has been corrected by the algorithm
        self.randomized = False  # True if this joint coordinates have been randomly generated

    def read_data(self, data):
        """Reads the data from the json file and creates the corresponding attributes"""
        self.joint_type = data["JointType"]
        self.x = float(data["Position"]["X"])
        self.y = float(data["Position"]["Y"])
        self.z = float(data["Position"]["Z"])
        self.position = [self.x, self.y, self.z]

    def correct_joint(self, x, y, z):
        """Assigns new x, y and z coordinates to the joint and marks it as corrected"""
        self.x = float(x)
        self.y = float(y)
        self.z = float(z)
        self.position = [self.x, self.y, self.z]
        self.corrected = True

    def set_move_much(self, move):
        """Defines that the displacement of a joint compared to the previous pose is above or below threshold"""
        self.move_much = move
        if move:
            self.color = (255, 50, 50)  # Color red for the joints above threshold
        else:
            self.color = (255, 255, 255)  # Color white for the joints below threshold

    def set_velocity(self, velocity):
        """Sets the velocity of the joint compared to the previous pose"""
        self.velocity = velocity

    def get_velocity(self):
        """Returns the velocity of the joint compared to the previous pose"""
        return self.velocity

    def get_copy(self):
        """Returns a copy of itself"""
        return copy.copy(self)

    def move_joint_randomized(self, origin, joint):
        """Changes the coordinates of a joint after randomization"""
        # print("Before: "+str(self.x)+", "+str(self.y)+", "+str(self.z))
        self.x = joint.x + (self.x - origin.x)
        self.y = joint.y + (self.y - origin.y)
        self.z = joint.z + (self.z - origin.z)
        self.randomized = True
        # print("After: "+str(self.x)+", "+str(self.y)+", "+str(self.z))

    def __repr__(self):
        """Returns a textual representation of the joint data"""
        txt = self.joint_type + ": (" + str(self.x) + ", " + str(self.y) + ", " + str(self.z) + ")"
        txt += " - Shift: " + str(self.velocity)
        if self.move_much:
            txt += " OVER THRESHOLD"
        if self.corrected:
            txt += " CORRECTED"
        return txt
