"""Functions to perform simple tasks that can be used throughout the toolbox and beyond."""
import datetime
import random

import chardet
import math
import os
import json
import warnings
import numpy as np

from classes.exceptions import ModuleNotFoundException, InvalidParameterValueException
from classes.graph_element import *
from classes.joint import Joint

# Constants
UNITS = {"ns": 1000000000, "1ns": 1000000000, "10ns": 100000000, "100ns": 10000000,
         "µs": 1000000, "1µs": 1000000, "10µs": 100000, "100µs": 10000,
         "ms": 1000, "1ms": 1000, "10ms": 100, "100ms": 10,
         "s": 1, "sec": 1, "1s": 1, "min": 1 / 60, "mn": 1 / 60, "h": 1 / 3600, "hr": 1 / 3600,
         "d": 1 / 86400, "day": 1 / 86400}


# === Folder and path functions ===
def find_common_parent_path(paths):
    """Finds the common root to a series of paths. If the root of the paths is different, the function returns an
    empty string.

    .. versionadded:: 2.0

    Parameters
    ----------
    paths: list(str)
        The list of paths. If the list contains only one element, this element will be returned by the function.

    Returns
    -------
    str
        The common parent path between the paths passed as parameters.

    Example
    -------
    >>> path1 = "C:/Users/Olivia/Documents/Word/Novels/Celsius 233.docx"
    >>> path2 = "C:/Users/Olivia/Documents/Word/Novels/Arrogance and Preconception.docx"
    >>> path3 = "C:/Users/Olivia/Documents/Word/Documentation/Krajjat documentation.docx"
    >>> path4 = "C:/Users/Olivia/Documents/Excel/Results/Results_1.xlsx"
    >>> path5 = "C:/Users/Olivia/Documents/Excel/Results/Results_2.xlsx"
    >>> path6 = "C:/Users/Olivia/Documents/Contract.pdf"
    >>> print(find_common_parent_path([path1, path2, path3, path4, path5, path6]))
    C:/Users/Olivia/Documents/
    """

    if len(paths) == 1:
        return paths[0]

    paths_separated = []
    min_length = None

    for path in paths:
        sub_paths = path.split("/")
        paths_separated.append(sub_paths)
        if min_length is None:
            min_length = len(sub_paths)
        elif len(sub_paths) < min_length:
            min_length = len(sub_paths)

    common_parent_path = ""
    add_path = True
    for i in range(min_length):
        for j in range(1, len(paths_separated)):
            if not paths_separated[j - 1][i] == paths_separated[j][i]:
                add_path = False
                break

        if add_path:
            common_parent_path += paths_separated[0][i] + "/"

    return common_parent_path


def compute_subpath(path1, path2):
    """Considering two absolute paths path1 and path2, where one is a parent path of the other, this function
    returns the subdirectories absent from the other.

    .. versionadded:: 2.0

    Parameters
    ----------
    path1: str
        An absolute path.
    path2: str
        An absolute path.

    Returns
    -------
    str
        The subdirectories of one of the two paths that is absent from the other.

    Example
    -------
    >>> path1 = "C:/Users/Shawn/"
    >>> path2 = "C:/Users/Shawn/Music/Coldplay/A Rush of Blood to the Head/"
    >>> print(compute_subpath(path1, path2))
    Music/Coldplay/A Rush of Blood to the Head/
    """

    path1 = path1.split("/")
    path2 = path2.split("/")

    i = 0

    for i in range(min(len(path1), len(path2))):
        if path1[i] != path2[i]:
            break

    if len(path1) > len(path2):
        sub_path = "/".join(path1[i:])
    else:
        sub_path = "/".join(path2[i:])

    return sub_path


def get_difference_paths(paths):
    """Returns, for each path, a list of the subfolders that are not common with all the other paths.

    .. versionadded:: 2.0

    Parameters
    ----------
    paths: list(str)
        The list of paths.

    Returns
    -------
    list(list(str))
        A list containing a list for each path, which itself contains the subfolders that are not common with the other
        paths.

    Example
    -------
    >>> path1 = "C:/Sequences/Raw/Subject1/Feb1/Sequence1"
    >>> path2 = "C:/Sequences/Resampled/Subject1/Mar1/Sequence1"
    >>> get_difference_paths([path1, path2])
    [['Raw', 'Feb1'], ['Original', 'Mar1']]
    """
    paths_splitted = []
    min_len = None

    for path in paths:
        path_splitted = path.split("/")
        paths_splitted.append(path_splitted)
        length_path = len(path_splitted)
        if min_len is None:
            min_len = length_path
        elif min_len > length_path:
            min_len = length_path

    new_paths = [[] for _ in range(len(paths))]

    if min_len > 1:
        for i in range(min_len):
            keep = False
            for p in range(len(paths_splitted)):
                if paths_splitted[p][i] != paths_splitted[0][i]:
                    keep = True
            if keep:
                for p in range(len(paths_splitted)):
                    new_paths[p].append(paths_splitted[p][i])

        for p in range(len(paths_splitted)):
            if len(paths_splitted[p]) > min_len:
                for i in range(min_len, len(paths_splitted[p])):
                    new_paths[p].append(paths_splitted[p][i])

    else:
        new_paths = paths_splitted

    return new_paths


def get_objects_paths(list_of_objects):
    """Returns a list of the attributes ``path`` of the objects passed as parameter (Sequence or Audio instances).

    .. versionadded:: 2.0

    Parameters
    ----------
    list_of_objects: list(Sequence) or list(Audio)
        A list of Sequence or Audio instances.

    Returns
    -------
    list(str)
        A list of paths.
    """
    paths = []
    for element in list_of_objects:
        paths.append(element.get_path())
    return paths


def create_subfolders(path, verbosity=1):
    """Creates all the subfolders that do not exist in a path. For example, if the folder ``"C:/Recordings/"`` is empty,
    and the parameter ``"path"`` is set on ``"C:/Recordings/Subject_01/Session_01/Video_01/"``, the function will
    successively create the folders ``"C:/Recordings/Subject_01/"``,  ``"C:/Recordings/Subject_01/Session_01/"``, and
    ``"C:/Recordings/Subject_01/Session_01/Video_01/"``.

    .. versionadded:: 2.0

    Parameters
    ----------
    path: str
        The absolute path to a folder.

    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.
    """

    folders = path.split("/")  # We create a list of all the sub-folders of the path
    for folder in range(len(folders)):
        partial_path = ""
        for i in range(folder + 1):
            partial_path += folders[i] + "/"
        if "." in folders[-1]:
            if verbosity > 0:
                print("Not creating " + path + " as it is not a valid folder.")
            break
        if not os.path.exists(partial_path):
            os.mkdir(partial_path)
            if verbosity > 0:
                print("Creating folder: " + partial_path)


# === Name functions ===
def get_objects_names(list_of_objects):
    """Returns a list of the attributes ``name`` of the objects passed as parameter (Sequence or Audio instances).

    .. versionadded:: 2.0

    Parameters
    ----------
    list_of_objects: list(Sequence) or list(Audio)
        A list of Sequence or Audio instances.

    Returns
    -------
    list(str)
        A list of names.
    """
    names = []
    for element in list_of_objects:
        names.append(element.get_name())
    return names


# === Sequences functions ===
def compute_different_joints(sequence1, sequence2, verbosity=False):
    """Returns the number of joints having different coordinates between two sequences having the same amount of poses.
    This function can be used to check how many joints have been modified by a processing function.

    .. versionadded:: 2.0

    Note
    ----
    To discard the rounding errors due to other functions, this function rounds the coordinates to the fifth decimal.

    Parameters
    ----------
    sequence1: Sequence
        A Sequence instance.
    sequence2: Sequence
        A Sequence instance.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    int
        The number of different joints between the two sequences.
    int
        The total number of joints in each of the sequences.
    """
    if len(sequence1) != len(sequence2):
        raise Exception("The two sequences do not have the same size (" + str(len(sequence1)) + " and " +
                        str(len(sequence2)) + ").")

    different_joints = 0
    total_joints = 0

    for p in range(len(sequence1.poses)):  # For all the poses
        for j in sequence1.poses[p].joints.keys():  # For all the joints

            # We round the coordinates to 5 significant digits in order to account for rounding errors
            if round(sequence1.poses[p].joints[j].x, 5) == round(sequence2.poses[p].joints[j].x, 5) and \
                    round(sequence1.poses[p].joints[j].y, 5) == round(sequence2.poses[p].joints[j].y, 5) and \
                    round(sequence1.poses[p].joints[j].z, 5) == round(sequence2.poses[p].joints[j].z, 5):
                total_joints += 1

            else:
                # If verbosity, we show the number of the pose, the name of the joint, and the coordinates of the
                # two sequences
                if verbosity:
                    print("Pose n°" + str(p) + ", " + str(j))
                    print("X: " + str(sequence1.poses[p].joints[j].x) +
                          "; Y: " + str(sequence1.poses[p].joints[j].y) +
                          "; Z: " + str(sequence1.poses[p].joints[j].z))
                    print("X: " + str(sequence2.poses[p].joints[j].x) +
                          "; Y: " + str(sequence2.poses[p].joints[j].y) +
                          "; Z: " + str(sequence2.poses[p].joints[j].z))
                different_joints += 1
                total_joints += 1

    # We print the result
    print("Different joints: " + str(different_joints) + "/" + str(total_joints) + " (" + str(
        different_joints / total_joints) + "%).")

    return different_joints, total_joints


def align_two_sequences(sequence1, sequence2):
    """Checks if one sequence is a subset of another; if true, the function returns the indices at which the two
    sequences start synchronizing. Otherwise, it returns ``False``.

    .. versionadded:: 2.0

    Parameters
    ----------
    sequence1: Sequence
        A Sequence instance.
    sequence2: Sequence
        A Sequence instance.

    Returns
    -------
    list or False
        If one of the sequences is a subsequence of the other, the function returns a list of two integers,
        corresponding to the indices of the first and the second sequences where the two sequences synchronize,
        respectively. Otherwise, this function returns ``False``.
    """

    indices = None

    # Define longest sequence
    if len(sequence1) > len(sequence2):
        longest_sequence = 1
    else:
        longest_sequence = 2

    joint_labels_1 = sequence1.get_joint_labels()
    joint_labels_2 = sequence2.get_joint_labels()

    # If the joints are not the same between the two sequences, we return False.
    if joint_labels_1 != joint_labels_2:
        return False

    # We compare for the x coordinate of the first joint
    subsequence1 = []
    for p in range(0, len(sequence1.poses)):
        subsequence1.append(sequence1.poses[p].joints[joint_labels_1[0]].x)

    subsequence2 = []
    for p in range(0, len(sequence2.poses)):
        subsequence2.append(sequence2.poses[p].joints[joint_labels_2[0]].x)

    if longest_sequence == 2:
        for i in range(0, len(subsequence2) - len(subsequence1)):
            if subsequence1 == subsequence2[i:i + len(subsequence1)]:
                indices = [0, i]
    else:
        for i in range(0, len(subsequence1) - len(subsequence2)):
            if subsequence2 == subsequence1[i:i + len(subsequence2)]:
                indices = [i, 0]

    # If we didn't find any subsequence, we return False
    if indices is None:
        return False

    # Otherwise, we check for every other coordinate
    for joint_label in joint_labels_1:

        for coordinate in range(0, 3):

            subsequence1 = []
            for p in range(0, len(sequence1.poses)):
                subsequence1.append(sequence1.poses[p].joints[joint_label].position[coordinate])
            subsequence2 = []
            for p in range(0, len(sequence2.poses)):
                subsequence2.append(sequence2.poses[p].joints[joint_label].position[coordinate])

            if longest_sequence == 2:
                if subsequence1 != subsequence2[indices[1]:indices[1] + len(subsequence1)]:
                    return False
            else:
                if subsequence2 != subsequence1[indices[0]:indices[0] + len(subsequence2)]:
                    return False

    return indices


def align_multiple_sequences(sequences, verbosity=1):
    """Given a list of sequences, returns """

    # Get the time array
    timestamps = []

    for sequence_index in range(len(sequences)):
        sequence = sequences[sequence_index]
        aligned = False

        # We compare the current sequence to every previous sequence
        for prior_sequence_index in range(sequence_index):
            prior_sequence = sequences[prior_sequence_index]

            # Returns False if the two sequence don't align, returns two indices otherwise
            alignment_indices = align_two_sequences(sequence, prior_sequence)

            # If the sequences do align
            if alignment_indices:

                # Case where the current sequence is a subsequence of the prior
                if alignment_indices[0] == 0:
                    if verbosity > 0:
                        print("Aligning sequence " + str(sequence_index + 1) + " to sequence " +
                              str(prior_sequence_index + 1) + ".")
                    timestamps.append(timestamps[prior_sequence_index][alignment_indices[1]:min(alignment_indices[1] +
                                                                                                len(sequence),
                                                                                                len(prior_sequence))])

                # Case where the prior sequence is a subsequence of the current
                else:
                    if verbosity > 0:
                        print("Aligning sequence " + str(prior_sequence_index + 1) + " to sequence " +
                              str(sequence_index + 1) + ".")
                    timestamps.append(sequence.get_timestamps())
                    timestamps[prior_sequence_index] = timestamps[sequence_index][alignment_indices[0]:
                                                                                  min(alignment_indices[0] +
                                                                                  len(prior_sequence), len(sequence))]

                aligned = True
                break

        if not aligned:
            timestamps.append(sequence.get_timestamps())

    return timestamps


# === File reading functions ===
def get_system_csv_separator():
    """Returns the separator (comma or semicolon) of the regional settings of the system; if it can't access it,
    returns a comma by default.

    .. versionadded:: 2.0

    Note
    ----
    This function detects the local decimal separator, and sets the csv delimiter on a semicolon (";") if the local
    decimal separator is a comma (","). In any other case, the csv delimiter is set on a comma (","). Note that in
    specific regions of the world, the result may not be ideal. In that case, you can force the csv delimiter to be
    the symbol of your choice by writing "csv," or "csv;" when asked for a file extension.

    Returns
    -------
    str
        The character used as separator in csv files on the current system.
    """
    try:
        import locale
        locale.setlocale(locale.LC_ALL, '')
        dec_pt_chr = locale.localeconv()['decimal_point']
        if dec_pt_chr == ",":
            separator = ";"
        else:
            separator = ","
    except OSError:
        print("Impossible to get the regional list separator, using default comma.")
        separator = ","

    return separator


# Constant
SEPARATOR = get_system_csv_separator()


def get_filetype_separator(extension):
    """Returns the separator used in specific text format files (comma, semicolon or tab). For csv, returns the
    separator used on the current system. For txt or tsv files, returns a tabulation symbol.

    .. versionadded:: 2.0

    Parameters
    ----------
    extension: str
        A file extension (csv, tsv or txt).

    Returns
    -------
    str
         The character used as separator between values in the target file extension.
    """
    if extension == "csv":  # Get the separator from local user (to get , or ;)
        separator = SEPARATOR
    else:  # For text or tsv files, tab separator
        separator = "\t"
    return separator


def read_json(path):
    """Loads and returns the content of a ``.json`` file.

    .. versionadded:: 2.0

    Parameters
    ----------
    path: str
        The path to a json file.

    Returns
    -------
    list or dict
        The content of the json file.
    """

    f = open(path, "r", encoding="utf-16-le")
    content = f.read()
    f.close()
    return json.loads(content)


def read_text_table(path):
    """Detects the encoding, loads and returns the content of a ``.csv``, ``.tsv`` or ``.txt`` file containing a table.

    .. versionadded:: 2.0

    Parameters
    ----------
    path: str
        The path to a text file.

    Returns
    -------
    list(list)
        The content of the text file, with each sub-list containing the elements of a row of the file.
    """

    rawdata = open(path, "rb").read()
    encoding = chardet.detect(rawdata)['encoding']

    file = open(path, "r", encoding=encoding)
    data = file.read().split("\n")
    file.close()

    # If the data is exported from QTM (Qualisys), convert the tsv to manageable data
    if path.split(".")[-1] == "tsv":
        if data[0].split("\t")[0] == "NO_OF_FRAMES":
            return convert_data_from_qtm(data)

    separator = get_filetype_separator(path.split(".")[-1])

    new_data = []
    for line in data:
        new_data.append(line.split(separator))

    return new_data


def read_xlsx(path):
    """Loads and returns the content of a ``.xlsx`` (Excel) file.

    .. versionadded:: 2.0

    Parameters
    ----------
    path: str
        The path to an Excel file.

    Returns
    -------
    list(list)
        The content of the Excel file, with each element of the list being a row of the Excel file.
    """
    import openpyxl as op
    workbook = op.load_workbook(path)
    data = workbook[workbook.sheetnames[0]]

    # Get the headers (timestamp and joint labels) from the first row
    table = []
    for i in range(1, len(data["A"]) + 1):
        row = []
        for j in range(1, len(data["1"]) + 1):
            row.append(data.cell(i, j).value)
        table.append(row)

    return table


# === Conversion functions ===
def convert_data_from_qtm(data, verbosity=1):
    """Processes and converts the data from a ``.tsv`` file produced by QTM, by stripping the header data,
    standardizing the name of the joint labels, and converting the distance unit from mm to m. This function then
    returns the loaded table.

    .. versionadded:: 2.0

    Parameters
    ----------
    data: list(str)
        The data from a ``.tsv`` QTM file with each line separated as the elements of a list.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    list(list(str))
        A table containing the QTM data, with each sub-list containing the elements of a row of the file.
    """
    new_data = []
    save_data = False

    if verbosity > 0:
        print("\n\tConverting data from Qualisys...")
    if verbosity == 1:
        print("\t\tStandardizing joints labels...", end=" ")
    elif verbosity > 1:
        print("\t\tStandardizing joints labels...")

    joints_conversions = load_qualisys_joint_label_conversion()

    # First, we find where the data starts
    perc = 10
    for i in range(len(data)):
        elements = data[i].split("\t")

        if elements[0] == "MARKER_NAMES":
            header = "Timestamp"
            for j in range(1, len(elements)):

                # We remove the prefix by looking for a joint name in the label:
                if elements[j] not in joints_conversions.keys():
                    for label in joints_conversions.keys():
                        if label in elements[j]:
                            if verbosity > 1:
                                print("\t\t\tChanging label name " + elements[j] + " to " +
                                      joints_conversions[label] + ".")
                            elements[j] = label
                            break

                header += "\t" + joints_conversions[elements[j]] + "_X"
                header += "\t" + joints_conversions[elements[j]] + "_Y"
                header += "\t" + joints_conversions[elements[j]] + "_Z"
            new_data.append(header)

            if verbosity > 0:
                print("\t\tLabels standardized.")

        if elements[0] == "1":
            save_data = True
            if verbosity > 0:
                print("\t\tConverting coordinates from mm to m...", end=" ")

        if save_data:

            perc = show_progression(verbosity, i, len(data), perc)
            line = []

            for j in range(1, len(elements)):

                # We keep the timestamp in ms
                if j == 1:
                    line.append(float(elements[j]))

                # We convert the joints coordinates in m
                else:
                    line.append(float(elements[j]) / 1000)

            if line != "":
                new_data.append(line)

    if verbosity > 0:
        print("100% - Done.")

    return new_data


# === File saving functions ===
def write_text_table(table, separator, path, verbosity=1):
    """Converts a table to a string, where elements on the same row are separated by a defined separator, and each row
    is separated by a line break.

    .. versionadded:: 2.0

    Parameters
    ----------
    table: list(list)
        A two-dimensional list where each element is a table row.
    separator: str
        The character used to separate elements on the same row.
    path: str
        The complete path of the text file to save.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    str
        The table in string version.
    """

    text = ""
    perc = 10

    for i in range(len(table)):
        perc = show_progression(verbosity, i, len(table), perc)
        for j in range(len(table[i])):
            text += str(table[i][j])
            if j != len(table[i]) - 1:
                text += separator
        if i != len(table) - 1:
            text += "\n"

    with open(path, 'w', encoding="utf-16-le") as f:
        f.write(text)


def write_xlsx(table, path, verbosity=1):
    """Saves a table in a ``.xlsx`` file.

    .. versionadded:: 2.0

    Parameters
    ----------
    table: list(list)
        A list where each sublist is a row of the table.
    path: str
        The complete path of where to store the Excel file.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.
    """
    try:
        import openpyxl as op
    except ImportError:
        raise ModuleNotFoundException("scipy", "save a file in .xlsx format.")
    workbook_out = op.Workbook()
    sheet_out = workbook_out.active

    perc = 10
    for i in range(len(table)):
        perc = show_progression(verbosity, i, len(table), perc)
        for j in range(len(table[i])):
            sheet_out.cell(i + 1, j + 1, table[i][j])
    workbook_out.save(path)


# === Calculation functions ===
def resample_data(data, time_points, frequency, mode="linear", time_unit="s"):
    """Resamples non-uniform data to a uniform time series according to a specific frequency, by interpolating the data.

    .. versionadded:: 2.0

    Note
    ----
    This function is a wrapper for :func:`interpolate_data`, solely creating a new array of time points from the
    indicated frequency.

    Important
    ---------
    This function is dependent of the module `numpy <https://numpy.org/>`_.

    Parameters
    ----------
    data: list(float) or numpy.ndarray(float)
        A list or an array of values.
    time_points: list(float) or numpy.ndarray(float)
        A list or an array of the time points corresponding to the values of the data.
    frequency: int or float
        The frequency at which you wish to resample the time series.
    mode: str, optional
        The way to interpolate the data. This parameter also allows for all the values accepted for the ``kind``
        parameter in the function :func:`scipy.interpolate.interp1d`: ``"linear"``, ``"nearest"``, ``"nearest-up"``,
        ``"zero"``, ``"slinear"``, ``"quadratic"``, ``"cubic"``”, ``"previous"``, and ``"next"``. See the
        `documentation for this Python module
        <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`_ for more.
    time_unit: str, optional
        The time unit of the time points. By default, it is set on "s" (seconds).

    Returns
    -------
    numpy.ndarray(float)
        The resampled values.
    numpy.ndarray(float)
        The resampled time points, at a fixed frequency.
    """

    try:
        from numpy import arange
    except ImportError:
        raise ModuleNotFoundException("numpy", "resample data.")

    step = (1 / frequency) * UNITS[time_unit]
    resampled_time_points = arange(min(time_points), max(time_points) + step/2, step)
    if resampled_time_points[-1] > max(time_points):
        resampled_time_points = resampled_time_points[:-1]
    return interpolate_data(data, time_points, resampled_time_points, mode)


def interpolate_data(data, time_points_data, time_points_interpolation, mode="linear"):
    """Interpolates incomplete data to a linear array of values.

    .. versionadded:: 2.0

    Important
    ---------
    This function is dependent of the modules `numpy <https://numpy.org/>`_ and `scipy <https://scipy.org/>`_.

    Parameters
    ----------
    data: list(float) or numpy.ndarray(float)
        A list or an array of values.
    time_points_data: list(float) or numpy.ndarray(float)
        A list or an array of the time points corresponding to the values of the data.
    time_points_interpolation: list(float) or numpy.ndarray(float)
        A list or an array of the time points to which interpolate the data.
    mode: str, optional
        The way to interpolate the data. This parameter also allows for all the values accepted for the ``kind``
        parameter in the function :func:`scipy.interpolate.interp1d`: ``"linear"``, ``"nearest"``, ``"nearest-up"``,
        ``"zero"``, ``"slinear"``, ``"quadratic"``, ``"cubic"``”, ``"previous"``, and ``"next"``. See the
        `documentation for this Python module
        <https://docs.scipy.org/doc/scipy/reference/generated/scipy.interpolate.interp1d.html>`_ for more.

    Returns
    -------
    numpy.ndarray(float)
        The interpolated values.
    numpy.ndarray(float)
        The interpolated time points (same as the parameter ``time_points_interpolation``).
    """

    if len(data) != len(time_points_data):
        raise Exception("The length of the data array (" + str(len(data)) + ") is inconsistent with the length of " +
                        "the time points (" + str(len(time_points_data)) + ").")

    try:
        from numpy import array
    except ImportError:
        raise ModuleNotFoundException("numpy", "interpolate data.")

    try:
        from scipy import interpolate
    except ImportError:
        raise ModuleNotFoundException("numpy", "interpolate data.")

    np_data = array(data)
    np_time_points = array(time_points_data)
    np_time_points_complete = array(time_points_interpolation)
    interp = interpolate.interp1d(np_time_points, np_data, kind=mode)

    resampled_data = interp(np_time_points_complete)
    return resampled_data, np_time_points_complete


def pad(data, time_points_data, time_points_padding, padding_value=0, verbosity=1):
    """Given an array of values (``data``) and its corresponding ``time_points_data``, and a larger array
    ``time_points_padding``, pads the data array with the value specified in ``padding_value`` for the time points
    present in ``time_points_padding`` that are absent from ``time_points_data``.

    .. versionadded:: 2.0

    Parameters
    ----------
    data: list(float) or numpy.ndarray(float)
        A list or an array of values.
    time_points_data: list(float) or numpy.ndarray(float)
        A list or an array of the time points corresponding to the values of the data.
    time_points_padding: list(float) or numpy.ndarray(float)
        A list or an array of time points containing the values from ``time_points_data`` (or values equal up to the
        fifth decimal), with additional values at the beginning and/or at the end.
    padding_value: int, numpy.nan or float
        The value with which to pad the data (default: 0).
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    np.ndarray
        The original data array, padded with zeros.
    list
        The array ``time_points_padding``, passed as parameter.
    """

    if len(data) != len(time_points_data):
        raise Exception("The length of the data array (" + str(len(data)) + ") is inconsistent with the length of " +
                        "the time points (" + str(len(time_points_data)) + ").")

    if time_points_data[0] < time_points_padding[0]:
        raise Exception("The first point of the time points padding array (" + str(time_points_padding[0]) + ") must " +
                        "be inferior or equal to the first time point of the data (" + str(time_points_data[0]) + ").")

    if time_points_data[-1] > time_points_padding[-1]:
        raise Exception("The last point of the time points padding array (" + str(time_points_padding[-1]) + ") must " +
                        "be superior or equal to the last time point of the data (" + str(time_points_data[-1]) + ").")

    import numpy as np

    padded_data = np.zeros(np.shape(time_points_padding), dtype=np.float64)
    inconsistency = False

    i = 0
    while round(time_points_padding[i], 5) < round(time_points_data[0], 5):
        padded_data[i] = padding_value
        if verbosity > 1:
            print("Padding · Iteration: " + str(i) + "/" + str(len(time_points_padding)) + " · Timestamp padding: " +
                  str(time_points_padding[i]) + ".")
        i += 1

    for j in range(len(time_points_data)):
        if round(time_points_padding[i], 5) != round(time_points_data[j], 5) and not inconsistency:
            print("Warning: the original time points and the time points used for the padding seem inconsistent with "
                  "each other: " + str(round(time_points_padding[i], 5)) + " and " +
                  str(round(time_points_data[j], 5)) + ".")
            inconsistency = True
        padded_data[i] = data[j]
        if verbosity > 1:
            print("Copying · Iteration: " + str(i) + "/" + str(len(time_points_padding)) + " · Timestamp padding: " +
                  str(time_points_padding[i]) + " · Iteration original: " + str(j) + "/" + str(len(time_points_data)) +
                  " · Original timestamp: " + str(time_points_data[j]) + ".")
        i += 1
    while i < len(time_points_padding):
        padded_data[i] = padding_value
        if verbosity > 1:
            print("Padding · Iteration: " + str(i) + "/" + str(len(time_points_padding)) + " · Timestamp padding: " +
                  str(time_points_padding[i]) + ".")
        i += 1

    return padded_data, time_points_padding


def add_delay(timestamps, delay):
    """Given an array of timestamps, adds or removes a delay to each timestamp.

    .. versionadded:: 2.0

    Parameters
    ----------
    timestamps: list(float)
        An array of timestamps.
    delay: float
        The delay, positive or negative to add to each timestamp.

    Returns
    -------
    list(float)
        An array of timestamps with the delay specified as parameter.
    """

    new_timestamps = []
    for t in timestamps:
        new_timestamps.append(t + delay)

    return new_timestamps


def calculate_distance(joint1, joint2, axis=None):
    """Uses the Euclidian formula to calculate the distance between two joints. This can be used to calculate
    the distance travelled by one joint between two poses, or the distance between two joints on the same pose.
    If an axis is specified, the distance is calculated on this axis only.

    .. versionadded:: 2.0

    Parameters
    ----------
    joint1: Joint
        The first Joint object.
    joint2: Joint
        The second Joint object.
    axis: str or None, optional
        If specified, the returned distances travelled will be calculated on a single axis. If ``None`` (default),
        the distance travelled will be calculated based on the 3D coordinates.

    Returns
    -------
    float
        The absolute distance, in meters, between the two joints.
    """
    if axis == "x":
        return abs(joint2.x - joint1.x)
    elif axis == "y":
        return abs(joint2.y - joint1.y)
    elif axis == "z":
        return abs(joint2.z - joint1.z)
    elif axis is None:
        x = (joint2.x - joint1.x) ** 2
        y = (joint2.y - joint1.y) ** 2
        z = (joint2.z - joint1.z) ** 2
        return math.sqrt(x + y + z)
    else:
        raise InvalidParameterValueException("axis", axis, ["x", "y", "z"])


def calculate_velocity(pose1, pose2, joint_label):
    """Given two poses and a joint label, returns the velocity of the joint, i.e. the distance travelled between the
    two poses, divided by the time elapsed between the two poses.

    .. versionadded:: 2.0

    Parameters
    ----------
    pose1: Pose
        The first Pose object, at the beginning of the movement.
    pose2: Pose
        The second Pose object, at the end of the movement.
    joint_label: str
        The label of the joint.

    Returns
    -------
    float
        The velocity of the joint between the two poses, in m/s.
    """

    # Get the distance travelled by a joint between two poses (meters)
    dist = calculate_distance(pose1.joints[joint_label], pose2.joints[joint_label])

    # Get the time elapsed between two poses (seconds)
    delay = calculate_delay(pose1, pose2)

    # Calculate the velocity (meters per seconds)
    velocity = dist / delay

    return velocity


def calculate_acceleration(pose1, pose2, pose3, joint_label):
    """Given three poses and a joint label, returns the acceleration of the joint, i.e. the difference of velocity
    between pose 1 and pose 2, and pose 2 and pose 3, over the time elapsed between pose 2 and pose 3.

    .. versionadded:: 2.0

    Parameters
    ----------
    pose1: Pose
        The first Pose object.
    pose2: Pose
        The second Pose object.
    pose3: Pose
        The third Pose object.
    joint_label: str
        The label of the joint.

    Returns
    -------
    float
        The velocity of the joint between the two poses, in m/s.
    """

    velocity1 = calculate_velocity(pose1, pose2, joint_label)
    velocity2 = calculate_velocity(pose2, pose3, joint_label)

    acceleration = (velocity2 - velocity1) / calculate_delay(pose2, pose3)

    return acceleration


def calculate_delay(pose1, pose2):
    """Returns the delay between two poses, in seconds.

    .. versionadded:: 2.0

    Parameters
    ----------
    pose1: Pose
        The first Pose object.
    pose2: Pose
        The second Pose object.

    Returns
    -------
    float
        The time, in seconds, between the two poses.

    Note
    ----
    The time returned by the function is **not** an absolute value. In other words, if ``pose1`` has a higher timestamp
    than ``pose2``, the returned value will be negative.
    """
    time_a = pose1.get_relative_timestamp()
    time_b = pose2.get_relative_timestamp()
    return time_b - time_a


def generate_random_joints(number_of_joints, x_scale=0.2, y_scale=0.3, z_scale=0.5):
    """Creates and returns a list of Joint objects with random coordinates. The coordinates are generated following
    a uniform distribution centered around 0 and with limits defined by ``x_scale``, ``y_scale`` and ``z_scale``.

    .. versionadded:: 2.0

    Parameters
    ----------
    number_of_joints: int
        The number of random joints to generate.
    x_scale: float
        The absolute maximum value of the random uniform distribution on the x axis.
    y_scale: float
        The absolute maximum value of the random uniform distribution on the y axis.
    z_scale: float
        The absolute maximum value of the random uniform distribution on the z axis.

    Returns
    -------
    list(Joint)
        A list of joints with randomized coordinates.
    """
    random_joints = []
    for i in range(number_of_joints):
        x = random.uniform(-x_scale, x_scale)
        y = random.uniform(-y_scale, y_scale)
        z = random.uniform(-z_scale, z_scale)
        j = Joint(None, x, y, z)
        random_joints.append(j)
    return random_joints


def divide_in_windows(array, window_size, overlap=0, add_incomplete_window=True):
    """Given an array of values, divides the array in windows of a size ``window_size``, with or without an overlap.

    .. versionadded:: 2.0

    Parameters
    ----------
    array: list(int or float) or np.array(int or float)
        An array of numerical values.
    window_size: int
        The number of array elements in each window.
    overlap: int
        The number of array elements overlapping in each window.
    add_incomplete_window: bool
        If set on ``True``, the last window will be added even if its size is smaller than ``window_size``. Otherwise,
        it will be ignored.

    Returns
    -------
    list(list(int or float))
        A list where each element is a window of size ``window_size``.
    """
    if overlap >= window_size:
        raise Exception("The size of the overlap (" + str(overlap) + ") cannot be bigger than or equal to the size " +
                        "of the window (" + str(window_size) + ").")
    if overlap > len(array) or window_size > len(array):
        raise Exception("The size of the window (" + str(window_size) + ") or the overlap (" + str(overlap) + ") " +
                        "cannot be bigger than the size of the array (" + str(len(array)) + ").")

    window_start = 0
    windows = []

    while window_start + window_size <= len(array):
        windows.append(array[window_start:window_start+window_size])
        window_start += window_size - overlap

    if add_incomplete_window:
        if array[window_start:]:
            windows.append(array[window_start:])

    return windows


def order_sequences_by_date(sequences):
    pass


# === Color functions ===
def load_color_names():
    """Returns a dictionary containing the 140 lower-case, whitespaces-stripped X11 and HTML/CSS colors as keys (plus 7
    variations of grey/gray), and a tuple of their 256-level RGBA codes as values. The colors are contained in
    ``res/color_codes.txt``. The color names follow the `140 HTML/CSS color names
    <https://en.wikipedia.org/wiki/X11_color_names>`_).

    .. versionadded:: 2.0

    Returns
    -------
    dict(str: tuple(int))
        A dictionary of the color names and their RGBA values.
    """
    colors_dict = {}
    file = open("res/color_codes.txt", "r")
    content = file.read().split("\n")
    file.close()

    for line in content:
        els = line.split("\t")
        colors_dict[els[0]] = (int(els[1]), int(els[2]), int(els[3]), int(els[4]))

    return colors_dict


def load_color_schemes():
    """Returns a dictionary containing the color gradients saved in ``res/color_gradients.txt``.

    .. versionadded:: 2.0

    Returns
    -------
    dict(int: list(tuple(int, int, int, int)))
        A dictionary with the name of color gradients as keys, and a list of their corresponding RGBA colors as values.
    """
    color_schemes = {}
    colors_dict = load_color_names()
    file = open("res/color_schemes.txt", "r")
    content = file.read().split("\n")
    file.close()

    for line in content:
        elements = line.split("\t")
        colors = []
        for e in range(1, len(elements)):
            if elements[e][0] == "(":
                colors_string = elements[e][1:-1].split(",")
                color = []
                for c in colors_string:
                    color.append(int(c))
                colors.append(tuple(color))
            elif elements[e][0] == "#":
                if len(elements[e]) == 7:
                    color = list(int(elements[e][i:i + 2], 16) for i in (1, 3, 5))
                    color.append(255)
                else:
                    color = list(int(elements[e][i:i + 2], 16) for i in (1, 3, 5, 7))
                colors.append(tuple(color))
            else:
                colors.append(colors_dict[elements[e]])

        color_schemes[elements[0]] = colors

    return color_schemes


def hex_color_to_rgb(color, include_alpha=None):
    """Converts a color from its hexadecimal value to its RGB or RGBA value.

    Parameters
    ----------
    color: str
        The hexadecimal value of a color, with or without a leading number sign (``"#"``).
    include_alpha: bool, optional
        If ``True``, returns the RGB color with an alpha value. If an alpha value is not present in the hexadecimal
        value, the alpha channel will be set to 255. If set on ``None`` (default), the returned value will contain
        and alpha value only if the input color contains one.

    Returns
    -------
    tuple(int, int, int) or tuple(int, int, int, int)
        A RGB or RGBA value.
    """
    if color[0] == "#":
        expected_length = 7
    else:
        expected_length = 6

    if len(color) == expected_length:
        rgba_color = list(int(color[i:i + 2], 16) for i in (1, 3, 5))
        if include_alpha:
            rgba_color.append(255)
        rgba_color = tuple(rgba_color)
    elif len(color) == expected_length + 2:
        rgba_color = tuple(int(color[i:i + 2], 16) for i in (1, 3, 5, 7))
        if not include_alpha:
            rgba_color = rgba_color[0:3]
    else:
        raise Exception("Invalid color hexadecimal value: " + str(color))

    return rgba_color


def rgb_color_to_hex(color, include_alpha=False):
    """Converts a color from its RGB or RGBA value to its hexadecimal value.

    Parameters
    ----------
    color: tuple(int, int, int) or tuple(int, int, int, int)
        The RGB or RGBA value of a color.
    include_alpha: bool, optional
        If ``True``, returns the hexadecimal color with an alpha value. If an alpha value is not present in the
        RGB value, the alpha channel will be set to ff.

    Returns
    -------
    str
        A hexadecimal value, with a leading number sign (``"#"``).
    """
    for i in range(len(color)):
        if color[i] < 0 or color[i] > 255:
            raise Exception("Invalid color index.")

    if len(color) == 3:
        if include_alpha:
            return '#%02x%02x%02xff' % color
        else:
            return '#%02x%02x%02x' % color

    elif len(color) == 4:
        if include_alpha:
            return '#%02x%02x%02x%02x' % color
        else:
            color = color[0:3]
            return '#%02x%02x%02x' % color

    else:
        raise Exception("Invalid number of arguments in the color.")


def convert_colors(color_scheme_or_colors, color_format="RGB", include_alpha=True):
    """Converts a list of colors to the desired format.

    .. versionadded:: 2.0

    Parameters
    ----------
    color_scheme_or_colors: str or list(str)
        This parameter can take a number of forms:

        • **The name of a color scheme:** a string matching one of the color gradients available in
          :doc:`color_schemes`.
        • **A list of colors:** a list containing colors, either using:

            • Their `HTML/CSS names <https://en.wikipedia.org/wiki/X11_color_names>`_ (e.g. ``"red"`` or
              ``"blanched almond"``),
            • Their hexadecimal code, starting with a number sign (``#``, e.g. ``"#ffcc00"`` or ``"#c0ffee"``).
            • Their RGB or RGBA tuples (e.g. ``(153, 204, 0)`` or ``(77, 77, 77, 255)``).

          These different codes can be used concurrently, e.g. ``["red", (14, 18, 32), "#a1b2c3"]``.

    color_format: str, optional
        The format in which you want to convert the colors. This parameter can be ``"RGB"`` (default), ``"RGBA"`` or
        ``"HEX"``.

    include_alpha: bool, optional
        If ``True``, returns the colors with an alpha value. If an alpha value is not present in the
        original value, the alpha channel will be set to 255 (or ff).

    Returns
    -------
    tuple(int, int, int) or tuple(int, int, int, int) or string
        A RGB or RGBA value, or a hexadecimal string with a leading number sign.
    """

    if type(color_scheme_or_colors) is str:

        color_schemes = load_color_schemes()

        if color_scheme_or_colors in color_schemes.keys():
            color_scheme_or_colors = color_schemes[color_scheme_or_colors]

        else:
            colors_dict = load_color_names()

            if color_scheme_or_colors not in colors_dict.keys():
                raise Exception("Invalid color scheme: " + str(color_scheme_or_colors) + ".")
            else:
                warnings.warn("The string entered is not a valid color scheme, but is a valid color name. The value of "
                              "the color will be returned, but convert_color should be used instead of convert_colors.")

            return colors_dict[color_scheme_or_colors]

    converted_colors = []
    for color in color_scheme_or_colors:
        converted_colors.append(convert_color(color, color_format, include_alpha))

    return converted_colors


def convert_color(color, color_format="rgb", include_alpha=True):
    """Converts a color to the desired format.

    .. versionadded:: 2.0

    Parameters
    ----------
    color: str or tuple(int, int, int) or tuple(int, int, int, int)
        This parameter can take a number of forms:

            • The `HTML/CSS name of a color <https://en.wikipedia.org/wiki/X11_color_names>`_ (e.g. ``"red"`` or
              ``"blanched almond"``),
            • The hexadecimal code of a color, starting with a number sign (``#``, e.g. ``"#ffcc00"`` or ``"#c0ffee"``).
            • The RGB or RGBA tuple of a color (e.g. ``(153, 204, 0)`` or ``(77, 77, 77, 255)``).

    color_format: str, optional
        The format in which you want to convert the color. This parameter can be ``"rgb"`` (default), ``"rgba"`` or
        ``"hex"``.

    include_alpha: bool, optional
        If ``True``, returns the color with an alpha value. If an alpha value is not present in the original value,
        the alpha channel will be set to 255 (or ff).

    Returns
    -------
    list(tuple(int, int, int)) or list(tuple(int, int, int, int)) or list(string)
        A list of RGB or RGBA values, or hexadecimal strings with a leading number sign.
    """

    # Color: tuple
    if type(color) is tuple:
        if len(color) == 3:
            converted_color = color
            if include_alpha:
                converted_color.append(255)
            converted_color = tuple(converted_color)
        elif len(color) == 4:
            converted_color = list(color[:3])
            if include_alpha:
                converted_color.append(color[3])
            converted_color = tuple(converted_color)
        else:
            raise Exception("Invalid color tuple: " + str(color) + ". Its length should be 3 or 4 elements.")
        for i in range(len(converted_color)):
            if converted_color[i] < 0 or converted_color[i] > 255:
                raise Exception("Invalid color value: " + str(color) + ". The values should be comprised between 0 " +
                                "and 255.")

    # Color: string
    elif type(color) is str:
        color = color.replace(" ", "")
        if color[0] == "#":
            converted_color = hex_color_to_rgb(color, include_alpha)

        else:
            colors_dict = load_color_names()

            if color not in colors_dict.keys():
                raise Exception("Invalid color name: " + str(color) + ".")

            else:
                converted_color = colors_dict[color]
                if not include_alpha:
                    converted_color = converted_color[:3]

    else:
        raise Exception("Invalid color type: " + str(type(color)) + ". Should be tuple or string.")

    if color_format.lower() in ["rgb", "rgba"]:
        return converted_color
    elif color_format.lower() == "hex":
        return rgb_color_to_hex(converted_color, include_alpha)
    else:
        raise Exception('Invalid parameter color_format: ' + str(color_format) + '. Should be "rgb", "rgba" or "hex".')


def calculate_color_points_on_gradient(color_scheme, no_points):
    """Given a color scheme and a number of points, creates a list of equidistant colors to create a gradient.

    Parameters
    ----------
    color_scheme: str or list(tuple(int, int, int, int))
        The name of a color scheme (see :doc:`color_schemes` for the available schemes), or a list of RGBA color
        tuples.
    no_points: int
        The number of colors to get on the gradient.

    Returns
    -------
    list(tuple(int, int, int, int))
        A list of colors in a progressive gradient following the colors of the color scheme.
    """
    color_scheme = convert_colors(color_scheme, "rgb", True)

    colors = []
    for i in range(no_points + 1):
        colors.append(calculate_color_ratio(color_scheme, i / no_points))

    return colors


def calculate_color_ratio(colors, ratio, type_return="rgb", include_alpha=True):
    """Given a gradient of colors, returns a color located at a specific ratio on the gradient.

    .. versionadded:: 2.0

    Parameters
    ----------
    colors: list(tuple(int, int, int)) or list(tuple(int, int, int, int))
        A list of RGB or RGBA color tuples.
    ratio: float
        The ratio of the color, between 0 and 1, on the gradient. If this parameter is equal to or inferior to 0, the
        function will return the first color of the ``colors`` parameter. If the ratio is equal or superior to 1, the
        function will return the last color of the ``colors`` parameter.
    type_return: str, optional
        The way the color should be returned. If set on ``"rgb"`` (default) or ``"rgba"``, the color will be returned
        as a tuple of integers. If set on ``"hex"``, the color will be returned as a string containing hexadecimal
        values, with a number sign (``#``) as a leading character.
    include_alpha: bool, optional
        If ``True``, returns the hexadecimal color with an alpha value. If an alpha value is not present in the
        RGB value, the alpha channel will be set to 255.

    Returns
    -------
    list(tuple(int, int, int)) or list(tuple(int, int, int, int)) or str
        A color in RGB, RGBA or hex form.
    """

    if 0 < ratio < 1:
        color_start = colors[int((ratio / (1 / (len(colors) - 1))))]
        color_end = colors[int((ratio / (1 / (len(colors) - 1)))) + 1]
    elif ratio >= 1:
        color_start = colors[-2]
        color_end = colors[-1]
        ratio = 1
    else:
        color_start = colors[0]
        color_end = colors[1]
        ratio = 0

    if int((ratio * (len(colors) - 1))) == 0 or ratio == 1:
        new_ratio = ratio
    else:
        new_ratio = (ratio * (len(colors) - 1)) % int((ratio * (len(colors) - 1)))

    diff_r = color_end[0] - color_start[0]
    diff_g = color_end[1] - color_start[1]
    diff_b = color_end[2] - color_start[2]
    if len(color_start) == 4:
        diff_a = color_end[3] - color_start[3]
    else:
        diff_a = 0

    r = color_start[0] + diff_r * new_ratio
    g = color_start[1] + diff_g * new_ratio
    b = color_start[2] + diff_b * new_ratio

    if len(color_start) == 4 and include_alpha:
        a = color_start[3] + diff_a * new_ratio
        color = (int(r), int(g), int(b), int(a))

    else:
        color = (int(r), int(g), int(b))

    if type_return.lower() in ["rgb", "rgba"]:
        return color
    elif type_return.lower() == "hex":
        return rgb_color_to_hex(color, include_alpha)


def calculate_colors_by_values(dict_values, color_scheme="default", type_return="rgb", include_alpha=True):
    """Returns a dictionary of colors on a gradient depending on the values of a dictionary. The extreme colors of the
    color scheme will be attributed to the extreme values of the dictionary; all the intermediate values are
    attributed their corresponding intermediate color on the gradient.

    .. versionadded:: 2.0

    Parameters
    ----------
    dict_values: dict(str: float or int)
        A dictionary where keys (e.g., joint labels) are attributed to values (floats or integers).
    color_scheme: str or list
        A color scheme or a list of colors that can be parsed by the function
        :func:`tool_functions.convert_colors_rgba`.
    type_return: str, optional
        The way the color should be returned. If set on ``"rgb"`` (default) or ``"rgba"``, the color will be returned
        as a tuple of integers. If set on ``"hex"``, the color will be returned as a string containing hexadecimal
        values, with a number sign (``#``) as a leading character.
    include_alpha: bool, optional
        If ``True``, returns the hexadecimal color with an alpha value. If an alpha value is not present in the
        RGB value, the alpha channel will be set to 255.

    Returns
    -------
    dict(str: tuple(int, int, int, int))
        A dictionary where each key is matched to an RGBA color.
    """
    keys_colors = {}

    # Get the min and max quantities of movement
    values = []
    for j in dict_values.keys():
        values.append(dict_values[j])
    min_value = min(values)  # Determines min global velocity (green)
    max_value = max(values)  # Determines max global velocity (red)

    colors = convert_colors(color_scheme, "rgb", True)

    # Apply the colors to all joints
    for j in dict_values.keys():
        ratio = (dict_values[j] - min_value) / (max_value - min_value)
        keys_colors[j] = calculate_color_ratio(colors, ratio, type_return, include_alpha)

    return keys_colors


def generate_random_color():
    """Generates a random color, and returns its RGBA value (with an alpha value of 255).

    .. versionadded:: 2.0

    Returns
    -------
    tuple(int, int, int, int)
        A tuple containing the RGBA values of a random color, with an alpha value of 255.
    """
    r = random.randrange(0, 256)
    g = random.randrange(0, 256)
    b = random.randrange(0, 256)
    a = 255
    return r, g, b, a


# === Audio functions ===
def scale_audio(audio_array, max_value, use_abs_values=False, set_lowest_at_zero=False, verbosity=1):
    """Scale an array of audio samples according to a maximum value.

    .. versionadded::2.0

    Parameters
    ----------
    audio_array: list(int or float) or numpy.array(int or float)
        An array of audio samples.
    max_value: int or float
        The value that will become the maximum value of the scaled array.
    use_abs_values: bool, optional
        If set on ``True``, the values of the samples will be converted to absolute values. If set on ``False``
        (default), the sign of the samples will be preserved.
    set_lowest_at_zero: bool, optional
        If set on ``True``, if the minimum sample value of the audio is over zero, the minimal value will be subtracted
        from all the samples. This parameter is set on ``False`` by default.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    list(float)
        An array of scaled audio samples.
    """
    new_audio_array = []

    if verbosity > 0:
        print("Scaling audio...")

    perc = 10

    max_audio_array = max(audio_array)
    min_audio_array = min(audio_array)
    min_value_scaled = min_audio_array / max_audio_array * max_value

    for i in range(len(audio_array)):
        perc = show_progression(verbosity, i, len(audio_array), perc)
        value = audio_array[i] / max_audio_array * max_value
        if use_abs_values:
            value = abs(value)
        if set_lowest_at_zero and min_audio_array > 0:
            value = value - min_value_scaled
        new_audio_array.append(value)

    if verbosity > 0:
        print("100% - Done.")

    return new_audio_array


def stereo_to_mono(audio_arrays, verbosity=1):
    """Turns the sample data into mono by averaging the samples from the audio channels.

    .. versionadded:: 2.0

    Parameters
    ----------
    audio_arrays: numpy.ndarray
        An array containing sub-lists with the sample values from all the channels as elements. Typically, the output
        of `scipy.io.wavfile.read <https://docs.scipy.org/doc/scipy/reference/generated/scipy.io.wavfile.read.html>`_.
    verbosity: int, optional
        Sets how much feedback the code will provide in the console output:

        • *0: Silent mode.* The code won’t provide any feedback, apart from error messages.
        • *1: Normal mode* (default). The code will provide essential feedback such as progression markers and
          current steps.
        • *2: Chatty mode.* The code will provide all possible information on the events happening. Note that this
          may clutter the output and slow down the execution.

    Returns
    -------
    np.ndarray(np.intp)
        An array containing the samples averaged across all channels.
    """

    if verbosity > 0:
        print("\n\tConverting audio samples from stereo to mono...", end=" ")

    new_audio_array = np.mean(audio_arrays, 1, np.intp)

    if verbosity > 0:
        print("100% - Done.")

    return new_audio_array


# === Joint labels functions ===
def load_kinect_joint_labels():
    """Loads the list of the 21 kinect joint labels from ``res/kinect_joint_labels.txt``.

    .. versionadded:: 2.0

    Returns
    -------
    list(str)
        The list of the Kinect joint labels.
    """
    file = open("res/kinect_joint_labels.txt")
    joint_labels = file.read().split("\n")
    return joint_labels


def load_qualisys_joint_labels(label_style="original"):
    """Loads the list of the 44 Qualisys joint labels from ``res/kualisys_joint_labels.txt``.

    .. versionadded:: 2.0

    Parameters
    ----------
    label_style: str, optional
        If set on ``"original"``, the returned joint labels are the original ones from the Qualisys system. If set
        on ``"krajjat"`` or ``"kualisys"``, the returned joint labels are the renamed Kualisys joint labels from the
        Krajjat toolbox. For more information, see :doc:`joint_labels`.

    Returns
    -------
    list(str)
        The list of the Qualisys joint labels.
    """
    file = open("res/kualisys_joint_labels.txt")
    joint_labels = file.read().split("\n")

    joint_labels_original = []
    joint_labels_krajjat = []

    for line in joint_labels:
        elements = line.split("\t")
        joint_labels_original.append(elements[0])
        joint_labels_krajjat.append(elements[1])

    if label_style == "original":
        return joint_labels_original
    elif label_style == "replaced":
        return joint_labels_krajjat


def load_qualisys_joint_label_conversion():
    """Returns a dictionary containing the original Qualisys joint labels as keys, and the renamed Kualisys joints
    labels as values. The dictionary is loaded from ``res/kualisys_joint_labels.txt``. For more information, see
    :doc:`joint_labels`.

    .. versionadded:: 2.0

    Returns
    -------
    dict(str: str)
        A dictionary with the original Qualisys joint labels as keys, and the Kualisys renamed joint labels as values.
    """
    file = open("res/kualisys_joint_labels.txt")
    content = file.read().split("\n")
    file.close()

    joints_conversions = {}

    for line in content:
        elements = line.split("\t")
        joints_conversions[elements[0]] = elements[1]

    return joints_conversions


def load_joint_labels(path):
    """Returns a list of joint labels from a path.

    .. versionadded:: 2.0

    Parameters
    ----------
    path: str
        The path to a file containing a joint label on each line.

    Returns
    -------
    list(str)
        A list containing joint labels.
    """
    with open("res/" + path, 'r') as f:
        content = f.read().split("\n")
    return content


def load_joints_connections(path):
    """Returns a list of joint pairs between which a line can be traced to form a skeleton.

    .. versionadded:: 2.0

    Parameters
    ----------
    path: str
        The path to a file containing joint label pairs separated by a tabulation on each line.

    Returns
    -------
    list(list(str))
        A list of sub-lists, each containing two elements (two joint labels).
    """
    with open("res/" + path, 'r') as f:
        content = f.read().split("\n")
    connections = []

    for line in content:
        connections.append(line.split("\t"))

    return connections


def load_qualisys_to_kinect():
    """Loads a dictionary containing Kinect joint labels as keys, and a series of Kualisys joint labels as values.
    The Kinect joints can be averaged from the joints in values. The dictionary is loaded from
    ``res/kualisys_to_kinect.txt``. For more information, see :doc:`joint_labels`.

    .. versionadded:: 2.0

    Returns
    -------
    dict(str: list(str))
        A dictionary of Kinect joint labels as keys, and a series of Kualisys joint labels as values.
    """
    file = open("res/kualisys_to_kinect.txt")
    content = file.read().split("\n")
    connections = {}

    for line in content:
        elements = line.split("\t")
        connections[elements[0]] = elements[1:]

    return connections


def load_joints_subplot_layout(joint_layout):
    """Returns a dictionary of the subplot positions of the joints on a skeleton graph. Loads the data from
    ``"res/kinect_joints_subplot_layout.txt"`` or ``"res/kualisys_joints_subplot_layout.txt"``.

    .. versionadded:: 2.0

    Parameters
    ----------
    joint_layout: str
        The layout to load, either ``"kinect"`` or ``"qualisys"``.

    Returns
    -------
    dict(str: int)
        A dictionary containing joint labels as keys, and subplot positions as values.
    """

    if joint_layout.lower() == "kinect":
        with open("res/kinect_joints_subplot_layout.txt") as f:
            content = f.read().split("\n")
    elif joint_layout.lower() in ["qualisys", "kualisys"]:
        with open("res/kualisys_joints_subplot_layout.txt") as f:
            content = f.read().split("\n")
    else:
        raise Exception("Wrong layout argument: should be 'kinect' or 'qualisys'.")

    joints_positions = {}

    for line in content:
        elements = line.split("\t")
        joints_positions[elements[0]] = int(elements[1])

    return joints_positions, joint_layout


def load_joints_silhouette_layout(joint_layout):
    """Returns a dictionary of the positions and radii of the joints, and a list of the order in which to plot the
    joints for the function :func:`plot_functions.plot_silhouette` Loads the data from
    ``"res/kinect_joints_silhouette_layout.txt"``, ``"res/kualisys_joints_silhouette_layout.txt"``, or a custom file.

    .. versionadded:: 2.0

    Parameters
    ----------
    joint_layout: str
        The layout to load, either ``"kinect"`` or ``"qualisys"``/``"kualisys"``. This parameter can also be the path
        to a custom layout (in .txt). In that case, each line must contain a joint label, the horizontal position,
        the vertical position and the radius of the circle to plot, all separated by a tabulation. The positions and
        radius are expected to be specified in pixels. The output list of the function, which defines the order in
        which to plot the joints, is defined by the order of the joints in the file.

    Returns
    -------
    dict(str: tuple)
        A dictionary containing joint labels as keys, and tuples containing three elements as values: the horizontal
        position, the vertical position, and the radius of the circle.

    list(str)
        The list of the joints, in the order they should be plotted.
    """

    if joint_layout.lower() == "kinect":
        with open("res/kinect_joints_silhouette_layout.txt") as f:
            content = f.read().split("\n")
    elif joint_layout.lower() in ["qualisys", "kualisys"]:
        with open("res/kualisys_joints_silhouette_layout.txt") as f:
            content = f.read().split("\n")
    else:
        try:
            with open(joint_layout) as f:
                content = f.read().split("\n")
        except FileNotFoundError:
            raise Exception("The file " + str(joint_layout) + " is not a valid file.")

    joints_positions = {}
    joint_layout = []

    for line in content:
        elements = line.split("\t")
        joints_positions[elements[0]] = (int(elements[1]), int(elements[2]), int(elements[3]))
        joint_layout.append(elements[0])

    return joints_positions, joint_layout


def load_steps_gui():
    """Loads the steps for the modification of the parameters in the GUI of the graphic functions, from the file
    ``res/steps_gui.txt``.

    .. versionadded:: 2.0
    """
    with open("res/steps_gui.txt") as f:
        lines = f.read().split("\n")

    steps = {}
    for line in lines:
        elements = line.split("\t")
        steps[elements[0]] = float(elements[1])

    return steps


# === Miscellaneous functions ===
def show_progression(verbosity, current_iteration, goal, next_percentage, step=10):
    """Shows a percentage of progression if verbosity is equal to 1.

    .. versionadded:: 2.0

    Parameters
    ----------
    verbosity: int
        This parameter must be equal to 1 for the function to print information. If the value is set on 2 (chatty mode),
        the percentage will not be displayed, in favour of more detailed information related to the ongoing operations.
    current_iteration: int or float
        The current iteration in a loop.
    goal: int or float
        The stopping value of the loop.
    next_percentage: int or float
        The next percentage at which to print the progression.
    step: int or float, optional
        The step between each print of the progression.

    Returns
    -------
    int or float
        The next percentage at which to print the progression.

    Example
    -------
    >>> next_percentage = 10
    >>> goal = 100
    >>> for i in range(goal):
    ...       next_percentage = show_progression(1, i, goal, next_percentage)
    10% 20̀% 30% 40% 50% 60% 70% 80M 90%
    """

    if verbosity == 1:
        while current_iteration / goal > next_percentage / 100:
            print(str(next_percentage) + "%", end=" ")
            next_percentage += step

    return next_percentage


def resample_images_to_frequency(images_paths, timestamps, frequency):
    """Given a series of images with specific timestamps, returns a new series of images and timestamps, resampled at
    a given frequency. For example, for images with timestamps at a rate of 20 Hz, and with a frequency set on 100 Hz,
    each image path will be copied 5 times with 5 different timestamps in the output. This can be used to obtain a
    video set at a defined framerate, regardless of the original frequency.

    .. versionadded:: 2.0

    Parameters
    ----------
    images_paths: list(str)
        A list of images paths.
    timestamps: list(float)
        A list of timestamps for each image path.
    frequency: int or float
        The frequency at which to resample the images (images per second).

    Returns
    -------
    list(str)
        A list of images paths.
    list(float)
        A list of timestamps.
    """
    number_of_frames = math.ceil(timestamps[-1] * frequency) + 1
    new_timestamps = []
    new_images = []
    t = 0
    duration = 1 / frequency

    for f in range(number_of_frames):
        new_timestamps.append(t)
        possible_times = [abs(i - t) for i in timestamps]
        index_image = possible_times.index(min(possible_times))
        new_images.append(images_paths[index_image])
        t += duration

    return new_images, new_timestamps


def get_min_max_values_from_plot_dictionary(plot_dictionary, keys_to_exclude=None):
    """Returns the minimum and maximum values of all the graphs contained in a plot dictionary.

    .. versionadded:: 2.0

    Parameters
    ----------
    plot_dictionary: dict
        A dictionary with labels as keys and Graph elements as values.
    keys_to_exclude: list(string)
        A list of keys of the dictionary to exclude from the search of the minimum and maximum values.

    Returns
    -------
    float
        The minimum value detected in all the series.
    float
        The maximum value detected in all the series.
    """
    min_value = 0
    max_value = 0

    if keys_to_exclude is None:
        keys_to_exclude = []

    for key in plot_dictionary.keys():
        if key not in keys_to_exclude:
            if plot_dictionary[key] is list:
                for series in plot_dictionary[key]:
                    local_min = min(series.y)
                    local_max = max(series.y)
                    if local_min < min_value:
                        min_value = local_min
                    if local_max > max_value:
                        max_value = local_max
            elif type(plot_dictionary[key]) is Graph:
                for plot in plot_dictionary[key].plots:
                    local_min = min(plot.y)
                    local_max = max(plot.y)
                    if local_min < min_value:
                        min_value = local_min
                    if local_max > max_value:
                        max_value = local_max
            else:
                if plot_dictionary[key] < min_value:
                    min_value = plot_dictionary[key]
                if plot_dictionary[key] > max_value:
                    max_value = plot_dictionary[key]

    return min_value, max_value


def kwargs_parser(dictionary, suffix):
    """Given a dictionary of keyword arguments and a suffix, returns a dictionary that removes the suffix from the given
    keywords. If a key without the suffix already exists in the dictionary, it is left untouched and the key with the
    suffix is removed.

    .. versionadded:: 2.0

    Parameters
    ----------
    dictionary: dict
        A dictionary of keyword arguments.
    suffix: str
        The suffix to look for in the keys of the dictionary.
    """
    new_dictionary = {}

    for key in dictionary.keys():
        if key.endswith(suffix):
            if len(key) >= len(suffix):
                key_without_suffix = key[:len(key)-len(suffix)]
                if key_without_suffix not in dictionary:
                    new_dictionary[key_without_suffix] = dictionary[key]
        else:
            new_dictionary[key] = dictionary[key]

    return new_dictionary


def format_time(time, time_unit="s", time_format="hh:mm:ss"):
    """Formats a given time in a given unit according to a time format.

    .. versionadded:: 2.0

    Parameters
    ----------
    time: int or float
        A value of time.
    time_unit: str, optional
        The unit of the ``time`` parameter. This parameter can take the following values: "ns", "1ns", "10ns", "100ns",
        "µs", "1µs", "10µs", "100µs", "ms", "1ms", "10ms", "100ms", "s", "sec", "1s", "min", "mn", "h", "hr", "d",
        "day".
    time_format: str, optional
        The format in which to return the time. Can be either "hh:mm:ss", "hh:mm:ss.ms", "hh:mm", "mm:ss" or "mm:ss.ms".
    """
    time = time / UNITS[time_unit]
    hour = int(time // 3600)
    minute = int((time // 60) % 60)
    second = int((time % 60) // 1)
    millisecond = int((time % 1) * 1000)

    if time_format.startswith("hh:mm:ss"):
        if time_format.endswith(".ms"):
            return str(hour).zfill(2) + ":" + str(minute).zfill(2) + ":" + str(second).zfill(2) + "." + \
                   str(millisecond).zfill(3)
        else:
            return str(hour).zfill(2) + ":" + str(minute).zfill(2) + ":" + str(second).zfill(2)

    elif time_format == "hh:mm":
        return str(hour).zfill(2) + ":" + str(minute).zfill(2)

    elif time_format.startswith("mm:ss"):
        minute = int(time // 60)
        if time_format.endswith(".ms"):
            return str(minute).zfill(2) + ":" + str(second).zfill(2) + "." + str(millisecond).zfill(3)
        else:
            return str(minute).zfill(2) + ":" + str(second).zfill(2)

    else:
        raise Exception('Wrong value for the time_format parameter. It should be "hh:mm:ss" or "mm:ss".')


def time_unit_to_datetime(time, time_unit="s"):
    """Turns any time unit into a ``datetime``format.

    .. versionadded:: 2.0

    Parameters
    ----------
    time: int or float
        A value of time.
    time_unit: str, optional
        The unit of the ``time`` parameter. This parameter can take the following values: "ns", "1ns", "10ns", "100ns",
        "µs", "1µs", "10µs", "100µs", "ms", "1ms", "10ms", "100ms", "s", "sec", "1s", "min", "mn", "h", "hr", "d",
        "day".
    """

    time = time / UNITS[time_unit]
    day = int(time // 86400)
    hour = int(time // 3600) % 24
    minute = int((time // 60) % 60)
    second = int((time % 60) // 1)
    microsecond = int((time % 1) * 1000000)

    return datetime.datetime(1, 1, 1+day, hour, minute, second, microsecond)